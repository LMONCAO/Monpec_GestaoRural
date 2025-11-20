# -*- coding: utf-8 -*-
"""
Views para Relatórios Obrigatórios de Rastreabilidade Bovina - PNIB
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Count, Sum
from decimal import Decimal
from datetime import datetime, date
import json

from .models import (
    Propriedade, AnimalIndividual, MovimentacaoIndividual,
    CategoriaAnimal
)

# Imports para exportação
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


@login_required
def relatorio_identificacao_individual(request, propriedade_id):
    """
    Relatório de Identificação Individual dos Animais - PNIB OBRIGATÓRIO
    """
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Filtros
    status_filtro = request.GET.get('status', 'ATIVO')
    categoria_filtro = request.GET.get('categoria', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    animais = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).select_related('categoria', 'propriedade_origem')
    
    if status_filtro:
        animais = animais.filter(status=status_filtro)
    
    if categoria_filtro:
        animais = animais.filter(categoria_id=categoria_filtro)
    
    if data_inicio:
        animais = animais.filter(data_cadastro__gte=data_inicio)
    
    if data_fim:
        animais = animais.filter(data_cadastro__lte=data_fim)
    
    animais = animais.order_by('numero_brinco')
    
    # Estatísticas
    total_animais = animais.count()
    animais_por_categoria = animais.values('categoria__nome').annotate(
        total=Count('id')
    )
    
    context = {
        'propriedade': propriedade,
        'animais': animais,
        'total_animais': total_animais,
        'animais_por_categoria': animais_por_categoria,
        'status_filtro': status_filtro,
        'categoria_filtro': categoria_filtro,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'data_emissao': date.today(),
    }
    
    return render(
        request,
        'gestao_rural/relatorios/relatorio_identificacao_individual.html',
        context
    )


@login_required
def relatorio_movimentacao_animais(request, propriedade_id):
    """
    Relatório de Movimentação de Animais - PNIB OBRIGATÓRIO
    """
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Filtros
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    numero_brinco = request.GET.get('numero_brinco', '')
    
    movimentacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade
    ).select_related(
        'animal', 'animal__categoria',
        'propriedade_origem', 'propriedade_destino'
    )
    
    if tipo_movimentacao:
        movimentacoes = movimentacoes.filter(tipo_movimentacao=tipo_movimentacao)
    
    if data_inicio:
        movimentacoes = movimentacoes.filter(data_movimentacao__gte=data_inicio)
    
    if data_fim:
        movimentacoes = movimentacoes.filter(data_movimentacao__lte=data_fim)
    
    if numero_brinco:
        movimentacoes = movimentacoes.filter(animal__numero_brinco__icontains=numero_brinco)
    
    movimentacoes = movimentacoes.order_by('-data_movimentacao', 'animal__numero_brinco')
    
    context = {
        'propriedade': propriedade,
        'movimentacoes': movimentacoes,
        'tipo_movimentacao': tipo_movimentacao,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'numero_brinco': numero_brinco,
        'data_emissao': date.today(),
    }
    
    return render(
        request,
        'gestao_rural/relatorios/relatorio_movimentacao_animais.html',
        context
    )


@login_required
def relatorio_sanitario(request, propriedade_id):
    """
    Relatório Sanitário - Vacinações e Tratamentos - PNIB OBRIGATÓRIO
    """
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Filtros
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    numero_brinco = request.GET.get('numero_brinco', '')
    
    vacinacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='VACINACAO'
    ).select_related('animal', 'animal__categoria')
    
    tratamentos = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='TRATAMENTO'
    ).select_related('animal', 'animal__categoria')
    
    if data_inicio:
        vacinacoes = vacinacoes.filter(data_movimentacao__gte=data_inicio)
        tratamentos = tratamentos.filter(data_movimentacao__gte=data_inicio)
    
    if data_fim:
        vacinacoes = vacinacoes.filter(data_movimentacao__lte=data_fim)
        tratamentos = tratamentos.filter(data_movimentacao__lte=data_fim)
    
    if numero_brinco:
        vacinacoes = vacinacoes.filter(animal__numero_brinco__icontains=numero_brinco)
        tratamentos = tratamentos.filter(animal__numero_brinco__icontains=numero_brinco)
    
    vacinacoes = vacinacoes.order_by('-data_movimentacao', 'animal__numero_brinco')
    tratamentos = tratamentos.order_by('-data_movimentacao', 'animal__numero_brinco')
    
    total_vacinacoes = vacinacoes.count()
    total_tratamentos = tratamentos.count()
    
    # Agrupar por animal
    animais_com_registro = {}
    for mov in list(vacinacoes) + list(tratamentos):
        brinco = mov.animal.numero_brinco
        if brinco not in animais_com_registro:
            animais_com_registro[brinco] = {
                'animal': mov.animal,
                'vacinacoes': [],
                'tratamentos': []
            }
        
        if mov.tipo_movimentacao == 'VACINACAO':
            animais_com_registro[brinco]['vacinacoes'].append(mov)
        else:
            animais_com_registro[brinco]['tratamentos'].append(mov)
    
    context = {
        'propriedade': propriedade,
        'vacinacoes': vacinacoes,
        'tratamentos': tratamentos,
        'animais_com_registro': animais_com_registro,
        'total_vacinacoes': total_vacinacoes,
        'total_tratamentos': total_tratamentos,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'numero_brinco': numero_brinco,
        'data_emissao': date.today(),
    }
    
    return render(
        request,
        'gestao_rural/relatorios/relatorio_sanitario.html',
        context
    )


@login_required
def relatorio_gta(request, propriedade_id, movimentacao_id=None):
    """
    Relatório de GTA (Guia de Trânsito Animal) - PNIB OBRIGATÓRIO
    """
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    if movimentacao_id:
        # GTA específica para uma movimentação
        movimentacao = get_object_or_404(
            MovimentacaoIndividual,
            pk=movimentacao_id,
            animal__propriedade=propriedade
        )
        movimentacoes = [movimentacao]
    else:
        # Lista de GTAs pendentes
        movimentacoes = MovimentacaoIndividual.objects.filter(
            animal__propriedade=propriedade,
            tipo_movimentacao__in=['VENDA', 'TRANSFERENCIA_SAIDA', 'COMPRA', 'TRANSFERENCIA_ENTRADA'],
            numero_documento__isnull=True
        ).select_related(
            'animal', 'propriedade_origem', 'propriedade_destino'
        ).order_by('-data_movimentacao')
    
    context = {
        'propriedade': propriedade,
        'movimentacoes': movimentacoes,
        'data_emissao': date.today(),
    }
    
    return render(
        request,
        'gestao_rural/relatorios/relatorio_gta.html',
        context
    )


# ===== EXPORTAÇÃO PDF =====

@login_required
def exportar_identificacao_individual_pdf(request, propriedade_id):
    """Exporta Relatório de Identificação Individual em PDF"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Busca dados
    animais = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        status='ATIVO'
    ).select_related('categoria', 'propriedade_origem').order_by('numero_brinco')
    
    # Cria PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="relatorio_identificacao_individual_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        f"Relatório de Identificação Individual dos Animais - PNIB",
        title_style
    ))
    story.append(Paragraph(
        f"Propriedade: {propriedade.nome_propriedade}",
        styles['Heading2']
    ))
    story.append(Paragraph(
        f"Data de Emissão: {date.today().strftime('%d/%m/%Y')}",
        styles['Normal']
    ))
    story.append(Spacer(1, 20))
    
    # Resumo
    resumo_data = [
        ['Total de Animais Identificados', str(animais.count())],
        ['Data de Emissão', date.today().strftime('%d/%m/%Y')],
    ]
    resumo_table = Table(resumo_data, colWidths=[10*cm, 8*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 20))
    
    # Tabela de animais
    animais_data = [[
        'Nº Brinco', 'Data Nasc.', 'Sexo', 'Raça', 'Categoria', 'Status'
    ]]
    
    for animal in animais[:50]:  # Limitar a 50 por página
        animais_data.append([
            animal.numero_brinco,
            animal.data_nascimento.strftime('%d/%m/%Y') if animal.data_nascimento else '-',
            animal.get_sexo_display(),
            animal.raca or '-',
            animal.categoria.nome,
            animal.get_status_display()
        ])
    
    animais_table = Table(
        animais_data,
        colWidths=[3*cm, 2.5*cm, 1.5*cm, 2.5*cm, 3*cm, 2.5*cm]
    )
    animais_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    story.append(animais_table)
    
    doc.build(story)
    return response


@login_required
def exportar_movimentacao_animais_pdf(request, propriedade_id):
    """Exporta Relatório de Movimentação em PDF"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    movimentacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade
    ).select_related('animal', 'propriedade_origem', 'propriedade_destino')
    
    if tipo_movimentacao:
        movimentacoes = movimentacoes.filter(tipo_movimentacao=tipo_movimentacao)
    if data_inicio:
        movimentacoes = movimentacoes.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        movimentacoes = movimentacoes.filter(data_movimentacao__lte=data_fim)
    
    movimentacoes = movimentacoes.order_by('-data_movimentacao')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="relatorio_movimentacao_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4, landscape=True)
    story = []
    styles = getSampleStyleSheet()
    
    story.append(Paragraph("Relatório de Movimentação de Animais - PNIB", styles['Heading1']))
    story.append(Paragraph(f"Propriedade: {propriedade.nome_propriedade}", styles['Heading2']))
    story.append(Spacer(1, 20))
    
    mov_data = [['Data', 'Animal', 'Tipo', 'Origem', 'Destino', 'Documento', 'Peso']]
    
    for mov in movimentacoes:
        mov_data.append([
            mov.data_movimentacao.strftime('%d/%m/%Y'),
            mov.animal.numero_brinco,
            mov.get_tipo_movimentacao_display(),
            mov.propriedade_origem.nome_propriedade if mov.propriedade_origem else '-',
            mov.propriedade_destino.nome_propriedade if mov.propriedade_destino else '-',
            mov.numero_documento or '-',
            f"{mov.peso_kg:.1f}" if mov.peso_kg else '-'
        ])
    
    mov_table = Table(mov_data, colWidths=[2.5*cm, 2.5*cm, 3*cm, 4*cm, 4*cm, 3*cm, 2*cm])
    mov_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(mov_table)
    
    doc.build(story)
    return response


@login_required
def exportar_saidas_sisbov_pdf(request, propriedade_id):
    """Exporta Comunicado de Saídas de Animais em PDF - Documento Oficial conforme IN 51/MAPA"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    busca = request.GET.get('busca', '')
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '').upper()
    
    # Buscar saídas
    tipos_disponiveis = ['VENDA', 'TRANSFERENCIA_SAIDA', 'MORTE']
    if tipo_movimentacao in tipos_disponiveis:
        tipos = [tipo_movimentacao]
    else:
        tipos = tipos_disponiveis
        tipo_movimentacao = ''
    
    saidas = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao__in=tipos
    ).select_related(
        'animal', 'animal__categoria',
        'propriedade_origem', 'propriedade_destino'
    )
    
    if data_inicio:
        saidas = saidas.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        saidas = saidas.filter(data_movimentacao__lte=data_fim)
    if busca:
        saidas = saidas.filter(animal__numero_brinco__icontains=busca)
    
    saidas = saidas.order_by('-data_movimentacao', 'animal__numero_brinco')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="Comunicado_Saidas_Animais_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4, landscape=True)
    story = []
    styles = getSampleStyleSheet()
    
    # Título principal - Exatamente como na imagem
    title_style = ParagraphStyle(
        'TituloOficial',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        textColor=colors.black,
        leading=20
    )
    
    story.append(Paragraph("COMUNICADO DE SAÍDAS DE ANIMAIS", title_style))
    
    # Subtítulo - Exatamente como na imagem
    subtitle_style = ParagraphStyle(
        'SubtituloOficial',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica',
        textColor=colors.black
    )
    story.append(Paragraph("IN 51/MAPA - SISBOV", subtitle_style))
    story.append(Spacer(1, 12))
    
    # Informações da propriedade - Tabela idêntica à imagem
    info_data = [
        ['Propriedade:', propriedade.nome_propriedade],
        ['Data de Emissão:', date.today().strftime('%d/%m/%Y')],
        ['Total de Saídas:', str(saidas.count())],
    ]
    if data_inicio:
        info_data.append(['Data Inicial:', data_inicio])
    if data_fim:
        info_data.append(['Data Final:', data_fim])
    if tipo_movimentacao:
        tipo_display = dict(MovimentacaoIndividual.TIPO_MOVIMENTACAO_CHOICES).get(tipo_movimentacao, tipo_movimentacao)
        info_data.append(['Tipo de Movimentação:', tipo_display])
    if busca:
        info_data.append(['Brinco:', busca])
    
    info_table = Table(info_data, colWidths=[4.5*cm, 20.5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 18))
    
    # Tabela principal de saídas - IDÊNTICA à imagem mostrada
    if saidas.exists():
        saidas_data = [[
            'Data', 'Brinco', 'Categoria', 'Peso', 'Destino', 'Documento', 'Observações'
        ]]
        
        for mov in saidas:
            # Formatar peso com vírgula (formato brasileiro: 350,0)
            peso_formatado = f"{mov.peso_kg:.1f}".replace('.', ',') if mov.peso_kg else '-'
            
            saidas_data.append([
                mov.data_movimentacao.strftime('%d/%m/%Y'),
                mov.animal.numero_brinco or mov.animal.codigo_sisbov or '-',
                mov.animal.categoria.nome if mov.animal.categoria else '-',
                peso_formatado,
                mov.propriedade_destino.nome_propriedade if mov.propriedade_destino else 'Abate / Baixa',
                mov.numero_documento or '-',
                mov.observacoes or '-'
            ])
        
        # Larguras das colunas - Ajustadas para corresponder exatamente à imagem
        # Total: 25cm (paisagem A4 = 29.7cm, deixando margens)
        saidas_table = Table(
            saidas_data,
            colWidths=[2.5*cm, 3.5*cm, 3.5*cm, 2*cm, 4*cm, 3*cm, 5.5*cm],
            repeatRows=1
        )
        saidas_table.setStyle(TableStyle([
            # Cabeçalho - Azul como na imagem
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            # Corpo da tabela
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (3, -1), 'CENTER'),  # Data, Brinco, Categoria, Peso centralizados
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Destino alinhado à esquerda
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Documento centralizado
            ('ALIGN', (6, 1), (6, -1), 'LEFT'),  # Observações alinhado à esquerda
            # Bordas pretas como na imagem
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            # Linhas alternadas (branco e cinza claro)
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        story.append(saidas_table)
    else:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Nenhuma saída encontrada para os filtros informados.", styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Rodapé - Formato oficial
    story.append(Paragraph(
        f"<i>Documento gerado em {date.today().strftime('%d/%m/%Y')} - Conforme Instrução Normativa Nº 51/MAPA</i>",
        styles['Normal']
    ))
    
    doc.build(story)
    return response


# ===== EXPORTAÇÃO PDF DOS ANEXOS IN 51 =====

@login_required
def exportar_anexo_i_pdf(request, propriedade_id):
    """Exporta Anexo I - Identificação Individual em PDF"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    animais = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).select_related('categoria', 'lote_atual').order_by('numero_brinco')
    
    if data_inicio:
        animais = animais.filter(data_identificacao__gte=data_inicio)
    if data_fim:
        animais = animais.filter(data_identificacao__lte=data_fim)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_I_Identificacao_Individual_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Cabeçalho
    title_style = ParagraphStyle(
        'AnexoTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e88e5')
    )
    story.append(Paragraph("ANEXO I - Identificação Individual dos Animais", title_style))
    story.append(Paragraph("IN 51/MAPA - SISBOV", styles['Heading3']))
    story.append(Spacer(1, 10))
    
    # Informações da propriedade
    info_data = [
        ['Propriedade:', propriedade.nome_propriedade],
        ['Data de Emissão:', date.today().strftime('%d/%m/%Y')],
        ['Total de Animais:', str(animais.count())],
    ]
    if data_inicio:
        info_data.append(['Data Inicial:', data_inicio])
    if data_fim:
        info_data.append(['Data Final:', data_fim])
    
    info_table = Table(info_data, colWidths=[4*cm, 14*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Tabela de animais
    animais_data = [[
        'Brinco/SISBOV', 'Categoria', 'Sexo', 'Data Nasc.', 'Data Identif.', 'Raça', 'Status', 'Lote'
    ]]
    
    for animal in animais:
        animais_data.append([
            f"{animal.numero_brinco}\n{animal.codigo_sisbov or ''}",
            animal.categoria.nome if animal.categoria else '-',
            animal.get_sexo_display() or '-',
            animal.data_nascimento.strftime('%d/%m/%Y') if animal.data_nascimento else '-',
            animal.data_identificacao.strftime('%d/%m/%Y') if animal.data_identificacao else '-',
            animal.raca or '-',
            animal.get_status_display(),
            animal.lote_atual.nome if animal.lote_atual else '-'
        ])
    
    animais_table = Table(
        animais_data,
        colWidths=[2.5*cm, 2.5*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.5*cm],
        repeatRows=1
    )
    animais_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(animais_table)
    
    # Rodapé
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<i>Documento gerado em {date.today().strftime('%d/%m/%Y')} - Conforme IN 51/MAPA</i>",
        styles['Normal']
    ))
    
    doc.build(story)
    return response


@login_required
def exportar_anexo_ii_pdf(request, propriedade_id):
    """Exporta Anexo II - Movimentação de Animais em PDF"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '')
    
    movimentacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade
    ).select_related(
        'animal', 'animal__categoria',
        'propriedade_origem', 'propriedade_destino'
    )
    
    if data_inicio:
        movimentacoes = movimentacoes.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        movimentacoes = movimentacoes.filter(data_movimentacao__lte=data_fim)
    if tipo_movimentacao:
        movimentacoes = movimentacoes.filter(tipo_movimentacao=tipo_movimentacao)
    
    movimentacoes = movimentacoes.order_by('-data_movimentacao', 'animal__numero_brinco')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_II_Movimentacao_Animais_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4, landscape=True)
    story = []
    styles = getSampleStyleSheet()
    
    # Cabeçalho
    title_style = ParagraphStyle(
        'AnexoTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e88e5')
    )
    story.append(Paragraph("ANEXO II - Movimentação de Animais", title_style))
    story.append(Paragraph("IN 51/MAPA - SISBOV", styles['Heading3']))
    story.append(Spacer(1, 10))
    
    # Informações
    info_data = [
        ['Propriedade:', propriedade.nome_propriedade],
        ['Data de Emissão:', date.today().strftime('%d/%m/%Y')],
        ['Total de Movimentações:', str(movimentacoes.count())],
    ]
    if data_inicio:
        info_data.append(['Data Inicial:', data_inicio])
    if data_fim:
        info_data.append(['Data Final:', data_fim])
    if tipo_movimentacao:
        info_data.append(['Tipo:', tipo_movimentacao])
    
    info_table = Table(info_data, colWidths=[4*cm, 20*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 15))
    
    # Tabela de movimentações
    mov_data = [[
        'Data', 'Animal', 'SISBOV', 'Tipo', 'Origem', 'Destino', 'Documento', 'Peso (kg)', 'Responsável'
    ]]
    
    for mov in movimentacoes:
        mov_data.append([
            mov.data_movimentacao.strftime('%d/%m/%Y'),
            mov.animal.numero_brinco,
            mov.animal.codigo_sisbov or '-',
            mov.get_tipo_movimentacao_display(),
            mov.propriedade_origem.nome_propriedade if mov.propriedade_origem else '-',
            mov.propriedade_destino.nome_propriedade if mov.propriedade_destino else '-',
            mov.numero_documento or '-',
            f"{mov.peso_kg:.1f}" if mov.peso_kg else '-',
            mov.responsavel.get_full_name() if mov.responsavel else '-'
        ])
    
    mov_table = Table(
        mov_data,
        colWidths=[2*cm, 2*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm, 2.5*cm, 1.5*cm, 3*cm],
        repeatRows=1
    )
    mov_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(mov_table)
    
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<i>Documento gerado em {date.today().strftime('%d/%m/%Y')} - Conforme IN 51/MAPA</i>",
        styles['Normal']
    ))
    
    doc.build(story)
    return response


@login_required
def exportar_anexo_iii_pdf(request, propriedade_id):
    """Exporta Anexo III - Registros Sanitários em PDF"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    vacinacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='VACINACAO'
    ).select_related('animal', 'animal__categoria')
    
    tratamentos = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='TRATAMENTO'
    ).select_related('animal', 'animal__categoria')
    
    if data_inicio:
        vacinacoes = vacinacoes.filter(data_movimentacao__gte=data_inicio)
        tratamentos = tratamentos.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        vacinacoes = vacinacoes.filter(data_movimentacao__lte=data_fim)
        tratamentos = tratamentos.filter(data_movimentacao__lte=data_fim)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_III_Registros_Sanitarios_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Cabeçalho
    title_style = ParagraphStyle(
        'AnexoTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e88e5')
    )
    story.append(Paragraph("ANEXO III - Registros Sanitários", title_style))
    story.append(Paragraph("IN 51/MAPA - SISBOV", styles['Heading3']))
    story.append(Spacer(1, 10))
    
    # Informações
    info_data = [
        ['Propriedade:', propriedade.nome_propriedade],
        ['Data de Emissão:', date.today().strftime('%d/%m/%Y')],
        ['Total de Vacinações:', str(vacinacoes.count())],
        ['Total de Tratamentos:', str(tratamentos.count())],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 14*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Vacinações
    story.append(Paragraph("VACINAÇÕES", styles['Heading2']))
    vac_data = [['Data', 'Animal', 'Categoria', 'Tipo de Vacina', 'Lote', 'Responsável']]
    
    for vac in vacinacoes:
        vac_data.append([
            vac.data_movimentacao.strftime('%d/%m/%Y'),
            vac.animal.numero_brinco,
            vac.animal.categoria.nome if vac.animal.categoria else '-',
            vac.motivo_detalhado or 'Vacinação',
            vac.numero_documento or '-',
            vac.responsavel.get_full_name() if vac.responsavel else '-'
        ])
    
    if len(vac_data) > 1:
        vac_table = Table(vac_data, colWidths=[2.5*cm, 2.5*cm, 3*cm, 4*cm, 2.5*cm, 3.5*cm], repeatRows=1)
        vac_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4caf50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        story.append(vac_table)
    else:
        story.append(Paragraph("<i>Nenhuma vacinação registrada no período.</i>", styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Tratamentos
    story.append(Paragraph("TRATAMENTOS", styles['Heading2']))
    trat_data = [['Data', 'Animal', 'Categoria', 'Tipo de Tratamento', 'Medicamento', 'Responsável']]
    
    for trat in tratamentos:
        trat_data.append([
            trat.data_movimentacao.strftime('%d/%m/%Y'),
            trat.animal.numero_brinco,
            trat.animal.categoria.nome if trat.animal.categoria else '-',
            trat.motivo_detalhado or 'Tratamento',
            trat.numero_documento or '-',
            trat.responsavel.get_full_name() if trat.responsavel else '-'
        ])
    
    if len(trat_data) > 1:
        trat_table = Table(trat_data, colWidths=[2.5*cm, 2.5*cm, 3*cm, 4*cm, 2.5*cm, 3.5*cm], repeatRows=1)
        trat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f44336')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        story.append(trat_table)
    else:
        story.append(Paragraph("<i>Nenhum tratamento registrado no período.</i>", styles['Normal']))
    
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<i>Documento gerado em {date.today().strftime('%d/%m/%Y')} - Conforme IN 51/MAPA</i>",
        styles['Normal']
    ))
    
    doc.build(story)
    return response


@login_required
def exportar_anexo_iv_pdf(request, propriedade_id):
    """Exporta Anexo IV - Inventário de Animais em PDF"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    animais = AnimalIndividual.objects.filter(propriedade=propriedade)
    animais_por_categoria = animais.values('categoria__nome').annotate(
        total=Count('id')
    ).order_by('categoria__nome')
    
    animais_por_status = animais.values('status').annotate(
        total=Count('id')
    ).order_by('status')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_IV_Inventario_Animais_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Cabeçalho
    title_style = ParagraphStyle(
        'AnexoTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e88e5')
    )
    story.append(Paragraph("ANEXO IV - Inventário de Animais", title_style))
    story.append(Paragraph("IN 51/MAPA - SISBOV", styles['Heading3']))
    story.append(Spacer(1, 10))
    
    # Informações
    info_data = [
        ['Propriedade:', propriedade.nome_propriedade],
        ['Data de Emissão:', date.today().strftime('%d/%m/%Y')],
        ['Total de Animais:', str(animais.count())],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 14*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Por Categoria
    story.append(Paragraph("Animais por Categoria", styles['Heading2']))
    cat_data = [['Categoria', 'Quantidade']]
    
    for item in animais_por_categoria:
        cat_data.append([
            item['categoria__nome'] or 'Sem categoria',
            str(item['total'])
        ])
    
    cat_table = Table(cat_data, colWidths=[14*cm, 4*cm], repeatRows=1)
    cat_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    story.append(cat_table)
    story.append(Spacer(1, 20))
    
    # Por Status
    story.append(Paragraph("Animais por Status", styles['Heading2']))
    status_data = [['Status', 'Quantidade']]
    status_labels = dict(AnimalIndividual.STATUS_CHOICES)
    
    for item in animais_por_status:
        status_data.append([
            status_labels.get(item['status'], item['status']),
            str(item['total'])
        ])
    
    status_table = Table(status_data, colWidths=[14*cm, 4*cm], repeatRows=1)
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    story.append(status_table)
    
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<i>Documento gerado em {date.today().strftime('%d/%m/%Y')} - Conforme IN 51/MAPA</i>",
        styles['Normal']
    ))
    
    doc.build(story)
    return response


# ===== EXPORTAÇÃO EXCEL DOS ANEXOS IN 51 =====

@login_required
def exportar_anexo_i_excel(request, propriedade_id):
    """Exporta Anexo I - Identificação Individual em Excel"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    animais = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).select_related('categoria', 'lote_atual').order_by('numero_brinco')
    
    if data_inicio:
        animais = animais.filter(data_identificacao__gte=data_inicio)
    if data_fim:
        animais = animais.filter(data_identificacao__lte=data_fim)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo I - Identificação"
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color='1e88e5', end_color='1e88e5', fill_type='solid')
    
    # Título
    ws['A1'] = 'ANEXO I - Identificação Individual dos Animais'
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = 'IN 51/MAPA - SISBOV'
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Informações
    ws['A4'] = 'Propriedade:'
    ws['B4'] = propriedade.nome_propriedade
    ws['A5'] = 'Data de Emissão:'
    ws['B5'] = date.today().strftime('%d/%m/%Y')
    ws['A6'] = 'Total de Animais:'
    ws['B6'] = animais.count()
    
    # Cabeçalho da tabela
    headers = ['Brinco/SISBOV', 'Categoria', 'Sexo', 'Data Nasc.', 'Data Identif.', 'Raça', 'Status', 'Lote']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Dados
    row = 9
    for animal in animais:
        ws.cell(row=row, column=1, value=f"{animal.numero_brinco}\n{animal.codigo_sisbov or ''}")
        ws.cell(row=row, column=2, value=animal.categoria.nome if animal.categoria else '-')
        ws.cell(row=row, column=3, value=animal.get_sexo_display() or '-')
        ws.cell(row=row, column=4, value=animal.data_nascimento.strftime('%d/%m/%Y') if animal.data_nascimento else '-')
        ws.cell(row=row, column=5, value=animal.data_identificacao.strftime('%d/%m/%Y') if animal.data_identificacao else '-')
        ws.cell(row=row, column=6, value=animal.raca or '-')
        ws.cell(row=row, column=7, value=animal.get_status_display())
        ws.cell(row=row, column=8, value=animal.lote_atual.nome if animal.lote_atual else '-')
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Rodapé
    ws.cell(row=row+1, column=1, value=f'Documento gerado em {date.today().strftime("%d/%m/%Y")} - Conforme IN 51/MAPA')
    ws.cell(row=row+1, column=1).font = Font(name='Arial', size=9, italic=True)
    ws.merge_cells(f'A{row+1}:H{row+1}')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_I_Identificacao_Individual_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def exportar_anexo_ii_excel(request, propriedade_id):
    """Exporta Anexo II - Movimentação em Excel"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '')
    
    movimentacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade
    ).select_related('animal', 'animal__categoria', 'propriedade_origem', 'propriedade_destino')
    
    if data_inicio:
        movimentacoes = movimentacoes.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        movimentacoes = movimentacoes.filter(data_movimentacao__lte=data_fim)
    if tipo_movimentacao:
        movimentacoes = movimentacoes.filter(tipo_movimentacao=tipo_movimentacao)
    
    movimentacoes = movimentacoes.order_by('-data_movimentacao', 'animal__numero_brinco')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo II - Movimentação"
    
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color='1e88e5', end_color='1e88e5', fill_type='solid')
    
    ws['A1'] = 'ANEXO II - Movimentação de Animais'
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = 'IN 51/MAPA - SISBOV'
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws.merge_cells('A2:I2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Propriedade:'
    ws['B4'] = propriedade.nome_propriedade
    ws['A5'] = 'Total de Movimentações:'
    ws['B5'] = movimentacoes.count()
    
    headers = ['Data', 'Animal', 'SISBOV', 'Tipo', 'Origem', 'Destino', 'Documento', 'Peso (kg)', 'Responsável']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 8
    for mov in movimentacoes:
        ws.cell(row=row, column=1, value=mov.data_movimentacao.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=mov.animal.numero_brinco)
        ws.cell(row=row, column=3, value=mov.animal.codigo_sisbov or '-')
        ws.cell(row=row, column=4, value=mov.get_tipo_movimentacao_display())
        ws.cell(row=row, column=5, value=mov.propriedade_origem.nome_propriedade if mov.propriedade_origem else '-')
        ws.cell(row=row, column=6, value=mov.propriedade_destino.nome_propriedade if mov.propriedade_destino else '-')
        ws.cell(row=row, column=7, value=mov.numero_documento or '-')
        ws.cell(row=row, column=8, value=f"{mov.peso_kg:.1f}" if mov.peso_kg else '-')
        ws.cell(row=row, column=9, value=mov.responsavel.get_full_name() if mov.responsavel else '-')
        row += 1
    
    # Ajustar larguras
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_II_Movimentacao_Animais_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def exportar_anexo_iii_excel(request, propriedade_id):
    """Exporta Anexo III - Registros Sanitários em Excel"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    vacinacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='VACINACAO'
    ).select_related('animal', 'animal__categoria')
    
    tratamentos = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='TRATAMENTO'
    ).select_related('animal', 'animal__categoria')
    
    if data_inicio:
        vacinacoes = vacinacoes.filter(data_movimentacao__gte=data_inicio)
        tratamentos = tratamentos.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        vacinacoes = vacinacoes.filter(data_movimentacao__lte=data_fim)
        tratamentos = tratamentos.filter(data_movimentacao__lte=data_fim)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo III - Sanidade"
    
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color='1e88e5', end_color='1e88e5', fill_type='solid')
    
    ws['A1'] = 'ANEXO III - Registros Sanitários'
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = 'IN 51/MAPA - SISBOV'
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Propriedade:'
    ws['B4'] = propriedade.nome_propriedade
    ws['A5'] = 'Total de Vacinações:'
    ws['B5'] = vacinacoes.count()
    ws['A6'] = 'Total de Tratamentos:'
    ws['B6'] = tratamentos.count()
    
    # Vacinações
    ws['A8'] = 'VACINAÇÕES'
    ws['A8'].font = Font(name='Arial', size=12, bold=True)
    
    vac_headers = ['Data', 'Animal', 'Categoria', 'Tipo de Vacina', 'Lote', 'Responsável', 'Observações']
    for col, header in enumerate(vac_headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='c8e6c9', end_color='c8e6c9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 10
    for vac in vacinacoes:
        ws.cell(row=row, column=1, value=vac.data_movimentacao.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=vac.animal.numero_brinco)
        ws.cell(row=row, column=3, value=vac.animal.categoria.nome if vac.animal.categoria else '-')
        ws.cell(row=row, column=4, value=vac.motivo_detalhado or 'Vacinação')
        ws.cell(row=row, column=5, value=vac.numero_documento or '-')
        ws.cell(row=row, column=6, value=vac.responsavel.get_full_name() if vac.responsavel else '-')
        ws.cell(row=row, column=7, value=vac.observacoes or '-')
        row += 1
    
    # Tratamentos
    row += 2
    ws.cell(row=row, column=1, value='TRATAMENTOS')
    ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    trat_headers = ['Data', 'Animal', 'Categoria', 'Tipo de Tratamento', 'Medicamento', 'Responsável', 'Observações']
    for col, header in enumerate(trat_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='ffcdd2', end_color='ffcdd2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    for trat in tratamentos:
        ws.cell(row=row, column=1, value=trat.data_movimentacao.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=trat.animal.numero_brinco)
        ws.cell(row=row, column=3, value=trat.animal.categoria.nome if trat.animal.categoria else '-')
        ws.cell(row=row, column=4, value=trat.motivo_detalhado or 'Tratamento')
        ws.cell(row=row, column=5, value=trat.numero_documento or '-')
        ws.cell(row=row, column=6, value=trat.responsavel.get_full_name() if trat.responsavel else '-')
        ws.cell(row=row, column=7, value=trat.observacoes or '-')
        row += 1
    
    # Ajustar larguras
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_III_Registros_Sanitarios_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def exportar_anexo_iv_excel(request, propriedade_id):
    """Exporta Anexo IV - Inventário em Excel"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    animais = AnimalIndividual.objects.filter(propriedade=propriedade)
    animais_por_categoria = animais.values('categoria__nome').annotate(
        total=Count('id')
    ).order_by('categoria__nome')
    
    animais_por_status = animais.values('status').annotate(
        total=Count('id')
    ).order_by('status')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo IV - Inventário"
    
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color='1e88e5', end_color='1e88e5', fill_type='solid')
    
    ws['A1'] = 'ANEXO IV - Inventário de Animais'
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = 'IN 51/MAPA - SISBOV'
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws.merge_cells('A2:B2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Propriedade:'
    ws['B4'] = propriedade.nome_propriedade
    ws['A5'] = 'Total de Animais:'
    ws['B5'] = animais.count()
    
    # Por Categoria
    ws['A7'] = 'Animais por Categoria'
    ws['A7'].font = Font(name='Arial', size=12, bold=True)
    
    cat_headers = ['Categoria', 'Quantidade']
    for col, header in enumerate(cat_headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 9
    for item in animais_por_categoria:
        ws.cell(row=row, column=1, value=item['categoria__nome'] or 'Sem categoria')
        ws.cell(row=row, column=2, value=item['total'])
        row += 1
    
    # Por Status
    row += 2
    ws.cell(row=row, column=1, value='Animais por Status')
    ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    status_headers = ['Status', 'Quantidade']
    status_labels = dict(AnimalIndividual.STATUS_CHOICES)
    for col, header in enumerate(status_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    for item in animais_por_status:
        ws.cell(row=row, column=1, value=status_labels.get(item['status'], item['status']))
        ws.cell(row=row, column=2, value=item['total'])
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="ANEXO_IV_Inventario_Animais_'
        f'{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response
