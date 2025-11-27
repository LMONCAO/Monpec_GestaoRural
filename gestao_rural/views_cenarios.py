# -*- coding: utf-8 -*-
"""
Views para análise de cenários de planejamento
Sistema completo de gerenciamento e comparação de cenários
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.urls import reverse
from decimal import Decimal
import json
from datetime import date, datetime

from .models import (
    Propriedade,
    PlanejamentoAnual, 
    CenarioPlanejamento,
    MovimentacaoProjetada,
    MetaComercialPlanejada,
    MetaFinanceiraPlanejada,
    VendaProjetada
)
from .services.gerar_vendas_projecao import gerar_vendas_do_cenario, gerar_vendas_todos_cenarios
from calendar import month_name
from collections import defaultdict


@login_required
def analise_cenarios(request, propriedade_id):
    """Página principal de análise de cenários"""
    propriedade = get_object_or_404(
        Propriedade, 
        id=propriedade_id, 
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar por código se fornecido
    codigo_busca = request.GET.get('codigo', '').strip()
    planejamento = None
    
    if codigo_busca:
        # Buscar planejamento por código
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
            messages.success(request, f'✅ Projeção encontrada: {planejamento.codigo} - {planejamento.ano}')
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o ID: {codigo_busca.upper()}. Verifique se o ID está correto.')
            planejamento = None
        except PlanejamentoAnual.MultipleObjectsReturned:
            planejamento = PlanejamentoAnual.objects.filter(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            ).first()
            if planejamento:
                messages.success(request, f'✅ Projeção encontrada: {planejamento.codigo} - {planejamento.ano}')
    
    # Se não encontrou por código, buscar planejamento atual (ano atual ou mais recente)
    if not planejamento:
        ano_atual = timezone.now().year
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-ano').first()
        
        # Se não existe planejamento, criar um para o ano atual
        if not planejamento:
            planejamento = PlanejamentoAnual.objects.create(
                propriedade=propriedade,
                ano=ano_atual,
                descricao=f"Planejamento {ano_atual}",
                status='RASCUNHO'
            )
            # Criar cenário baseline
            CenarioPlanejamento.objects.create(
                planejamento=planejamento,
                nome="Baseline",
                descricao="Cenário base de referência",
                is_baseline=True
            )
    
    # Buscar todos os planejamentos disponíveis para seleção
    planejamentos_disponiveis = PlanejamentoAnual.objects.filter(
        propriedade=propriedade
    ).order_by('-ano', '-data_criacao')
    
    # Buscar todos os cenários do planejamento
    cenarios = planejamento.cenarios.all().order_by('-is_baseline', 'nome')
    
    # Verificar se existem vendas geradas e gerar automaticamente se não existirem
    # Isso acontece quando o usuário busca por código ou acessa a página
    if planejamento and cenarios.exists():
        # Verificar se há vendas para este planejamento
        vendas_existem = VendaProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=planejamento
        ).exists()
        
        # Verificar se existem movimentações de VENDA para gerar vendas
        # Buscar movimentações do planejamento ou do ano do planejamento
        movimentacoes_venda_query = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        )
        
        # Filtrar por planejamento se existir, senão por ano
        if planejamento:
            movimentacoes_venda_query = movimentacoes_venda_query.filter(
                Q(planejamento=planejamento) | 
                Q(planejamento__isnull=True, data_movimentacao__year=planejamento.ano)
            )
        
        movimentacoes_venda_exist = movimentacoes_venda_query.exists()
        
        # Se há movimentações de VENDA mas não há vendas geradas, gerar automaticamente
        # Isso acontece quando o usuário busca por ID ou acessa a página pela primeira vez
        if movimentacoes_venda_exist and not vendas_existem:
            try:
                # Vincular movimentações não vinculadas ao planejamento
                movimentacoes_sem_vinculo = MovimentacaoProjetada.objects.filter(
                    propriedade=propriedade,
                    tipo_movimentacao='VENDA',
                    quantidade__gt=0,
                    planejamento__isnull=True,
                    data_movimentacao__year=planejamento.ano
                )
                
                if movimentacoes_sem_vinculo.exists():
                    total_para_vincular = movimentacoes_sem_vinculo.count()
                    movimentacoes_sem_vinculo.update(planejamento=planejamento)
                    messages.info(
                        request,
                        f'ℹ️ {total_para_vincular} movimentações de VENDA foram vinculadas à projeção {planejamento.codigo}.'
                    )
                
                # Gerar vendas para todos os cenários
                vendas_geradas = gerar_vendas_todos_cenarios(propriedade, planejamento)
                total_vendas = sum(len(vendas) for vendas in vendas_geradas.values())
                
                if total_vendas > 0:
                    messages.success(
                        request,
                        f'✅ Vendas geradas automaticamente para a projeção {planejamento.codigo}! '
                        f'Total: {total_vendas} vendas criadas para {len(vendas_geradas)} cenários.'
                    )
                else:
                    # Se não gerou vendas, verificar o motivo
                    movimentacoes_vinculadas = MovimentacaoProjetada.objects.filter(
                        propriedade=propriedade,
                        planejamento=planejamento,
                        tipo_movimentacao='VENDA',
                        quantidade__gt=0
                    ).count()
                    
                    if movimentacoes_vinculadas > 0:
                        messages.warning(
                            request,
                            f'⚠️ Foram encontradas {movimentacoes_vinculadas} movimentações de VENDA, mas nenhuma venda foi gerada. '
                            f'Verifique se os valores das movimentações estão preenchidos corretamente.'
                        )
                    else:
                        messages.info(
                            request,
                            f'ℹ️ Nenhuma movimentação de VENDA encontrada para gerar vendas na projeção {planejamento.codigo}. '
                            f'Gere primeiro as projeções na página "Projeções do Rebanho".'
                        )
            except Exception as e:
                import traceback
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Erro ao gerar vendas automaticamente: {traceback.format_exc()}')
                messages.error(request, f'❌ Erro ao gerar vendas automaticamente: {str(e)}. Verifique os logs para mais detalhes.')
    
    # Processar dados de cada cenário
    cenarios_data = []
    for cenario in cenarios:
        dados = calcular_metricas_cenario(planejamento, cenario)
        cenarios_data.append({
            'cenario': cenario,
            'dados': dados
        })
    
    ano_atual = timezone.now().year
    
    context = {
        'propriedade': propriedade,
        'planejamento': planejamento,
        'planejamentos_disponiveis': planejamentos_disponiveis,
        'cenarios_data': cenarios_data,
        'ano_atual': ano_atual,
        'codigo_busca': codigo_busca,
    }
    
    return render(request, 'gestao_rural/analise_cenarios.html', context)


@login_required
def criar_cenario(request, propriedade_id):
    """Criar novo cenário"""
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método não permitido'}, status=405)
    
    propriedade = get_object_or_404(
        Propriedade, 
        id=propriedade_id, 
        produtor__usuario_responsavel=request.user
    )
    
    planejamento_id = request.POST.get('planejamento_id')
    planejamento = get_object_or_404(PlanejamentoAnual, id=planejamento_id, propriedade=propriedade)
    
    nome = request.POST.get('nome', '').strip()
    if not nome:
        messages.error(request, 'Nome do cenário é obrigatório')
        return redirect('analise_cenarios', propriedade_id=propriedade.id)
    
    descricao = request.POST.get('descricao', '').strip()
    is_baseline = request.POST.get('is_baseline') == 'on'
    
    # Se marcar como baseline, desmarcar outros
    if is_baseline:
        CenarioPlanejamento.objects.filter(planejamento=planejamento, is_baseline=True).update(is_baseline=False)
    
    # Ajustes percentuais
    ajuste_preco = Decimal(request.POST.get('ajuste_preco_percentual', '0') or '0')
    ajuste_custo = Decimal(request.POST.get('ajuste_custo_percentual', '0') or '0')
    ajuste_producao = Decimal(request.POST.get('ajuste_producao_percentual', '0') or '0')
    ajuste_taxas = Decimal(request.POST.get('ajuste_taxas_percentual', '0') or '0')
    
    cenario = CenarioPlanejamento.objects.create(
        planejamento=planejamento,
        nome=nome,
        descricao=descricao,
        is_baseline=is_baseline,
        ajuste_preco_percentual=ajuste_preco,
        ajuste_custo_percentual=ajuste_custo,
        ajuste_producao_percentual=ajuste_producao,
        ajuste_taxas_percentual=ajuste_taxas
    )
    
    messages.success(request, f'Cenário "{nome}" criado com sucesso!')
    return redirect('analise_cenarios', propriedade_id=propriedade.id)


@login_required
def editar_cenario(request, propriedade_id, cenario_id):
    """Editar cenário existente"""
    propriedade = get_object_or_404(
        Propriedade, 
        id=propriedade_id, 
        produtor__usuario_responsavel=request.user
    )
    
    cenario = get_object_or_404(
        CenarioPlanejamento, 
        id=cenario_id,
        planejamento__propriedade=propriedade
    )
    
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if not nome:
            messages.error(request, 'Nome do cenário é obrigatório')
            return redirect('analise_cenarios', propriedade_id=propriedade.id)
        
        cenario.nome = nome
        cenario.descricao = request.POST.get('descricao', '').strip()
        is_baseline = request.POST.get('is_baseline') == 'on'
        
        # Se marcar como baseline, desmarcar outros
        if is_baseline and not cenario.is_baseline:
            CenarioPlanejamento.objects.filter(
                planejamento=cenario.planejamento, 
                is_baseline=True
            ).exclude(id=cenario.id).update(is_baseline=False)
        
        cenario.is_baseline = is_baseline
        cenario.ajuste_preco_percentual = Decimal(request.POST.get('ajuste_preco_percentual', '0') or '0')
        cenario.ajuste_custo_percentual = Decimal(request.POST.get('ajuste_custo_percentual', '0') or '0')
        cenario.ajuste_producao_percentual = Decimal(request.POST.get('ajuste_producao_percentual', '0') or '0')
        cenario.ajuste_taxas_percentual = Decimal(request.POST.get('ajuste_taxas_percentual', '0') or '0')
        cenario.save()
        
        messages.success(request, f'Cenário "{nome}" atualizado com sucesso!')
        return redirect('analise_cenarios', propriedade_id=propriedade.id)
    
    # GET - retornar dados do cenário em JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse({
            'id': cenario.id,
            'nome': cenario.nome,
            'descricao': cenario.descricao or '',
            'is_baseline': cenario.is_baseline,
            'ajuste_preco_percentual': str(cenario.ajuste_preco_percentual),
            'ajuste_custo_percentual': str(cenario.ajuste_custo_percentual),
            'ajuste_producao_percentual': str(cenario.ajuste_producao_percentual),
            'ajuste_taxas_percentual': str(cenario.ajuste_taxas_percentual),
        })
    
    # GET normal - redirecionar para a página principal
    return redirect('analise_cenarios', propriedade_id=propriedade.id)


@login_required
def excluir_cenario(request, propriedade_id, cenario_id):
    """Excluir cenário"""
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método não permitido'}, status=405)
    
    propriedade = get_object_or_404(
        Propriedade, 
        id=propriedade_id, 
        produtor__usuario_responsavel=request.user
    )
    
    cenario = get_object_or_404(
        CenarioPlanejamento, 
        id=cenario_id,
        planejamento__propriedade=propriedade
    )
    
    # Não permitir excluir baseline se for o único cenário
    if cenario.is_baseline:
        total_cenarios = CenarioPlanejamento.objects.filter(planejamento=cenario.planejamento).count()
        if total_cenarios == 1:
            messages.error(request, 'Não é possível excluir o cenário baseline se for o único cenário.')
            return redirect('analise_cenarios', propriedade_id=propriedade.id)
    
    nome = cenario.nome
    cenario.delete()
    
    messages.success(request, f'Cenário "{nome}" excluído com sucesso!')
    return redirect('analise_cenarios', propriedade_id=propriedade.id)


@login_required
def comparar_cenarios_api(request, propriedade_id):
    """API para comparar cenários"""
    propriedade = get_object_or_404(
        Propriedade, 
        id=propriedade_id, 
        produtor__usuario_responsavel=request.user
    )
    
    planejamento_id = request.GET.get('planejamento_id')
    if not planejamento_id:
        return JsonResponse({'erro': 'planejamento_id é obrigatório'}, status=400)
    
    planejamento = get_object_or_404(PlanejamentoAnual, id=planejamento_id, propriedade=propriedade)
    
    cenarios_ids = request.GET.getlist('cenarios_ids[]')
    if not cenarios_ids:
        cenarios = planejamento.cenarios.all()
    else:
        cenarios = CenarioPlanejamento.objects.filter(
            id__in=cenarios_ids,
            planejamento=planejamento
        )
    
    dados_comparacao = []
    for cenario in cenarios:
        metricas = calcular_metricas_cenario(planejamento, cenario)
        dados_comparacao.append({
            'id': cenario.id,
            'nome': cenario.nome,
            'is_baseline': cenario.is_baseline,
            'metricas': metricas
        })
    
    return JsonResponse({
        'cenarios': dados_comparacao
    })


def calcular_metricas_cenario(planejamento, cenario):
    """Calcula métricas financeiras e operacionais de um cenário"""
    # Buscar metas comerciais
    metas_comerciais = MetaComercialPlanejada.objects.filter(planejamento=planejamento)
    
    # Calcular receitas projetadas com ajuste de preço
    receitas_totais = Decimal('0')
    quantidade_animais = 0
    arrobas_totais = Decimal('0')
    
    for meta in metas_comerciais:
        preco_ajustado = meta.preco_medio_esperado * (1 + cenario.ajuste_preco_percentual / 100)
        quantidade_ajustada = int(meta.quantidade_animais * (1 + cenario.ajuste_producao_percentual / 100))
        
        receita_meta = preco_ajustado * quantidade_ajustada
        if meta.arrobas_totais:
            receita_meta = preco_ajustado * meta.arrobas_totais * (1 + cenario.ajuste_producao_percentual / 100)
        
        receitas_totais += receita_meta
        quantidade_animais += quantidade_ajustada
        if meta.arrobas_totais:
            arrobas_totais += meta.arrobas_totais * (1 + cenario.ajuste_producao_percentual / 100)
    
    # Buscar metas financeiras (custos)
    metas_financeiras = MetaFinanceiraPlanejada.objects.filter(planejamento=planejamento)
    
    custos_totais = Decimal('0')
    for meta in metas_financeiras:
        custo_ajustado = meta.valor_anual_previsto * (1 + cenario.ajuste_custo_percentual / 100)
        custos_totais += custo_ajustado
    
    # Calcular lucro
    lucro = receitas_totais - custos_totais
    
    # Calcular margem
    margem = (lucro / receitas_totais * 100) if receitas_totais > 0 else Decimal('0')
    
    return {
        'receitas_totais': float(receitas_totais),
        'custos_totais': float(custos_totais),
        'lucro': float(lucro),
        'margem': float(margem),
        'quantidade_animais': quantidade_animais,
        'arrobas_totais': float(arrobas_totais),
    }


@login_required
def gerar_vendas_cenario(request, propriedade_id, cenario_id):
    """Gerar vendas projetadas para um cenário específico"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    cenario = get_object_or_404(
        CenarioPlanejamento,
        id=cenario_id,
        planejamento__propriedade=propriedade
    )
    
    try:
        # Verificar se há movimentações antes de tentar gerar
        movimentacoes_venda = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=cenario.planejamento,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        )
        total_movimentacoes = movimentacoes_venda.count()
        
        if total_movimentacoes == 0:
            messages.warning(
                request,
                f'⚠️ Nenhuma movimentação de VENDA encontrada para o planejamento {cenario.planejamento.codigo}. '
                f'É necessário gerar primeiro as projeções na página "Projeções do Rebanho".'
            )
            return redirect('analise_cenarios', propriedade_id=propriedade.id)
        
        vendas_criadas = gerar_vendas_do_cenario(propriedade, cenario)
        
        if len(vendas_criadas) > 0:
            messages.success(
                request,
                f'✅ Vendas geradas com sucesso! Total: {len(vendas_criadas)} vendas criadas para o cenário "{cenario.nome}".'
            )
        else:
            # Verificar se já existem vendas
            vendas_existentes = VendaProjetada.objects.filter(
                propriedade=propriedade,
                cenario=cenario
            ).count()
            
            if vendas_existentes > 0:
                messages.info(
                    request,
                    f'ℹ️ Todas as vendas já estão geradas para o cenário "{cenario.nome}". '
                    f'Total: {vendas_existentes} vendas existentes.'
                )
            else:
                messages.warning(
                    request,
                    f'⚠️ Nenhuma venda foi gerada. Foram encontradas {total_movimentacoes} movimentações de VENDA, '
                    f'mas nenhuma venda foi criada. Verifique os logs para mais detalhes.'
                )
    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao gerar vendas: {traceback.format_exc()}')
        messages.error(
            request, 
            f'❌ Erro ao gerar vendas: {str(e)}. '
            f'Verifique os logs para mais detalhes.'
        )
    
    return redirect('analise_cenarios', propriedade_id=propriedade.id)


@login_required
def gerar_vendas_todos_cenarios_view(request, propriedade_id):
    """Gerar vendas para todos os cenários do planejamento atual"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento atual - pode vir da URL como parâmetro (codigo ou planejamento_id)
    planejamento = None
    codigo_busca = request.GET.get('codigo', '').strip()
    planejamento_id = request.GET.get('planejamento_id')
    
    # PRIORIDADE 1: Buscar por código (mais específico) - PROJEÇÃO SELECIONADA PELO USUÁRIO
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
            messages.info(request, f'✅ Gerando vendas para a projeção selecionada: {planejamento.codigo}')
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            planejamento = None
        except PlanejamentoAnual.MultipleObjectsReturned:
            planejamento = PlanejamentoAnual.objects.filter(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            ).order_by('-data_criacao').first()
            if planejamento:
                messages.info(request, f'✅ Gerando vendas para a projeção selecionada: {planejamento.codigo}')
    
    # PRIORIDADE 2: Buscar por ID
    if not planejamento and planejamento_id:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                id=planejamento_id,
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            planejamento = None
    
    # PRIORIDADE 3: Buscar o planejamento mais recente por data de criação
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.error(request, 'Nenhum planejamento encontrado. Crie um planejamento primeiro gerando uma projeção na página "Projeções do Rebanho".')
        return redirect('analise_cenarios', propriedade_id=propriedade.id)
    
    # Buscar o planejamento mais recente que tem movimentações de VENDA
    # Isso garante que usamos o planejamento correto mesmo quando há múltiplos no mesmo ano
    planejamento_com_movimentacoes = PlanejamentoAnual.objects.filter(
        propriedade=propriedade,
        movimentacoes_planejadas__tipo_movimentacao='VENDA',
        movimentacoes_planejadas__quantidade__gt=0
    ).distinct().order_by('-data_criacao').first()
    
    # Se encontrou um planejamento com movimentações, sempre usar o mais recente
    if planejamento_com_movimentacoes:
        # Verificar se o planejamento atual tem movimentações
        movimentacoes_atual = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            tipo_movimentacao='VENDA',
            quantidade__gt=0,
            planejamento=planejamento
        ).count()
        
        # Se o planejamento com movimentações é mais recente ou o atual não tem movimentações, usar ele
        if planejamento_com_movimentacoes.data_criacao > planejamento.data_criacao or movimentacoes_atual == 0:
            planejamento = planejamento_com_movimentacoes
    
    # Verificar se há movimentações de VENDA antes de tentar gerar
    # IMPORTANTE: Buscar TODAS as movimentações do planejamento (incluindo múltiplos anos)
    movimentacoes_venda = MovimentacaoProjetada.objects.filter(
        propriedade=propriedade,
        tipo_movimentacao='VENDA',
        quantidade__gt=0
    )
    
    # Filtrar por planejamento - buscar todas vinculadas ao planejamento
    total_movimentacoes = 0
    if planejamento:
        movimentacoes_venda = movimentacoes_venda.filter(planejamento=planejamento)
        total_movimentacoes = movimentacoes_venda.count()
        
        # Se não encontrou, tentar vincular movimentações sem planejamento
        # OU movimentações vinculadas a outros planejamentos recentes do mesmo período
        if total_movimentacoes == 0:
            # Buscar movimentações sem planejamento que podem pertencer a este
            anos_busca = [planejamento.ano + i for i in range(-2, 6)]
            movimentacoes_sem_planejamento = MovimentacaoProjetada.objects.filter(
                propriedade=propriedade,
                tipo_movimentacao='VENDA',
                quantidade__gt=0,
                planejamento__isnull=True,
                data_movimentacao__year__in=anos_busca
            )
            
            total_sem_planejamento = movimentacoes_sem_planejamento.count()
            if total_sem_planejamento > 0:
                # Vincular ao planejamento
                movimentacoes_sem_planejamento.update(planejamento=planejamento)
                messages.info(
                    request,
                    f'ℹ️ {total_sem_planejamento} movimentações de VENDA foram vinculadas ao planejamento {planejamento.codigo}.'
                )
                total_movimentacoes = total_sem_planejamento
                # Recarregar a query após vincular
                movimentacoes_venda = MovimentacaoProjetada.objects.filter(
                    propriedade=propriedade,
                    tipo_movimentacao='VENDA',
                    quantidade__gt=0,
                    planejamento=planejamento
                )
            
            # Se ainda não encontrou, tentar buscar movimentações de planejamentos recentes do mesmo período
            if total_movimentacoes == 0:
                from datetime import timedelta
                # Buscar planejamentos criados recentemente (últimas 24 horas)
                data_limite = timezone.now() - timedelta(hours=24)
                planejamentos_recentes = PlanejamentoAnual.objects.filter(
                    propriedade=propriedade,
                    data_criacao__gte=data_limite
                ).order_by('-data_criacao')
                
                for plan_recente in planejamentos_recentes:
                    movimentacoes_outro_planejamento = MovimentacaoProjetada.objects.filter(
                        propriedade=propriedade,
                        tipo_movimentacao='VENDA',
                        quantidade__gt=0,
                        planejamento=plan_recente
                    ).count()
                    
                    if movimentacoes_outro_planejamento > 0:
                        # Se o planejamento recente tem movimentações e é mais recente, usar ele
                        if plan_recente.data_criacao > planejamento.data_criacao:
                            planejamento = plan_recente
                            movimentacoes_venda = MovimentacaoProjetada.objects.filter(
                                propriedade=propriedade,
                                tipo_movimentacao='VENDA',
                                quantidade__gt=0,
                                planejamento=planejamento
                            )
                            total_movimentacoes = movimentacoes_venda.count()
                            messages.info(
                                request,
                                f'ℹ️ Usando o planejamento mais recente {planejamento.codigo} que possui {total_movimentacoes} movimentações de VENDA.'
                            )
                        else:
                            # Transferir movimentações do outro planejamento para este
                            MovimentacaoProjetada.objects.filter(
                                propriedade=propriedade,
                                tipo_movimentacao='VENDA',
                                quantidade__gt=0,
                                planejamento=plan_recente
                            ).update(planejamento=planejamento)
                            
                            messages.info(
                                request,
                                f'ℹ️ {movimentacoes_outro_planejamento} movimentações de VENDA foram transferidas do planejamento {plan_recente.codigo} para {planejamento.codigo}.'
                            )
                            total_movimentacoes = movimentacoes_outro_planejamento
                            # Recarregar a query após transferir
                            movimentacoes_venda = MovimentacaoProjetada.objects.filter(
                                propriedade=propriedade,
                                tipo_movimentacao='VENDA',
                                quantidade__gt=0,
                                planejamento=planejamento
                            )
                        break
    
    if total_movimentacoes == 0:
        # Verificar se há movimentações de VENDA em geral para esta propriedade
        total_vendas_geral = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        ).count()
        
        if total_vendas_geral == 0:
            messages.warning(
                request,
                f'⚠️ Nenhuma movimentação de VENDA encontrada para esta propriedade. '
                f'É necessário gerar primeiro as projeções na página "Projeções do Rebanho" '
                f'para criar movimentações de venda que possam ser convertidas em vendas projetadas.'
            )
        else:
            # Buscar TODAS as movimentações de VENDA (vinculadas a qualquer planejamento ou sem planejamento)
            # e transferir para o planejamento atual
            movimentacoes_todas = MovimentacaoProjetada.objects.filter(
                propriedade=propriedade,
                tipo_movimentacao='VENDA',
                quantidade__gt=0
            )
            
            # Buscar movimentações de outros planejamentos recentes (últimos 7 dias)
            from datetime import timedelta
            data_limite = timezone.now() - timedelta(days=7)
            planejamentos_com_movimentacoes = PlanejamentoAnual.objects.filter(
                propriedade=propriedade,
                movimentacoes_planejadas__tipo_movimentacao='VENDA',
                movimentacoes_planejadas__quantidade__gt=0
            ).distinct().order_by('-data_criacao')
            
            total_transferidas = 0
            planejamento_origem = None
            
            # Se encontrou planejamentos com movimentações, transferir do mais recente
            if planejamentos_com_movimentacoes.exists():
                planejamento_origem = planejamentos_com_movimentacoes.first()
                movimentacoes_para_transferir = MovimentacaoProjetada.objects.filter(
                    propriedade=propriedade,
                    tipo_movimentacao='VENDA',
                    quantidade__gt=0,
                    planejamento=planejamento_origem
                )
                total_transferidas = movimentacoes_para_transferir.count()
                
                if total_transferidas > 0:
                    # Transferir todas as movimentações para o planejamento atual
                    movimentacoes_para_transferir.update(planejamento=planejamento)
                    messages.success(
                        request,
                        f'✅ {total_transferidas} movimentações de VENDA foram transferidas do planejamento {planejamento_origem.codigo} para {planejamento.codigo}. '
                        f'As vendas serão geradas agora.'
                    )
                    # Recarregar a query após transferir
                    movimentacoes_venda = MovimentacaoProjetada.objects.filter(
                        propriedade=propriedade,
                        tipo_movimentacao='VENDA',
                        quantidade__gt=0,
                        planejamento=planejamento
                    )
                    total_movimentacoes = movimentacoes_venda.count()
                else:
                    # Tentar vincular movimentações sem planejamento
                    ano_atual = timezone.now().year
                    anos_busca = list(range(planejamento.ano - 2, planejamento.ano + 6))
                    
                    movimentacoes_para_vincular = MovimentacaoProjetada.objects.filter(
                        propriedade=propriedade,
                        tipo_movimentacao='VENDA',
                        quantidade__gt=0,
                        planejamento__isnull=True,
                        data_movimentacao__year__in=anos_busca
                    )
                    
                    total_para_vincular = movimentacoes_para_vincular.count()
                    
                    if total_para_vincular > 0:
                        # Vincular automaticamente
                        movimentacoes_para_vincular.update(planejamento=planejamento)
                        messages.success(
                            request,
                            f'✅ {total_para_vincular} movimentações de VENDA foram vinculadas ao planejamento {planejamento.codigo}. '
                            f'As vendas serão geradas agora.'
                        )
                        movimentacoes_venda = MovimentacaoProjetada.objects.filter(
                            propriedade=propriedade,
                            tipo_movimentacao='VENDA',
                            quantidade__gt=0,
                            planejamento=planejamento
                        )
                        total_movimentacoes = movimentacoes_venda.count()
                    else:
                        messages.warning(
                            request,
                            f'⚠️ Nenhuma movimentação de VENDA encontrada para a projeção {planejamento.codigo} (ano {planejamento.ano}). '
                            f'Foram encontradas {total_vendas_geral} movimentações de VENDA na propriedade, '
                            f'mas nenhuma está no período de busca ({anos_busca[0]}-{anos_busca[-1]}). '
                            f'Verifique se as movimentações foram geradas corretamente.'
                        )
                        return redirect('analise_cenarios', propriedade_id=propriedade.id)
            else:
                # Tentar vincular movimentações sem planejamento
                ano_atual = timezone.now().year
                anos_busca = list(range(planejamento.ano - 2, planejamento.ano + 6))
                
                movimentacoes_para_vincular = MovimentacaoProjetada.objects.filter(
                    propriedade=propriedade,
                    tipo_movimentacao='VENDA',
                    quantidade__gt=0,
                    planejamento__isnull=True,
                    data_movimentacao__year__in=anos_busca
                )
                
                total_para_vincular = movimentacoes_para_vincular.count()
                
                if total_para_vincular > 0:
                    # Vincular automaticamente
                    movimentacoes_para_vincular.update(planejamento=planejamento)
                    messages.success(
                        request,
                        f'✅ {total_para_vincular} movimentações de VENDA foram vinculadas ao planejamento {planejamento.codigo}. '
                        f'As vendas serão geradas agora.'
                    )
                    movimentacoes_venda = MovimentacaoProjetada.objects.filter(
                        propriedade=propriedade,
                        tipo_movimentacao='VENDA',
                        quantidade__gt=0,
                        planejamento=planejamento
                    )
                    total_movimentacoes = movimentacoes_venda.count()
                else:
                    messages.warning(
                        request,
                        f'⚠️ Nenhuma movimentação de VENDA encontrada para a projeção {planejamento.codigo} (ano {planejamento.ano}). '
                        f'Foram encontradas {total_vendas_geral} movimentações de VENDA na propriedade, '
                        f'mas nenhuma está no período de busca ({anos_busca[0]}-{anos_busca[-1]}). '
                        f'Verifique se as movimentações foram geradas corretamente.'
                    )
                    return redirect('analise_cenarios', propriedade_id=propriedade.id)
        
        # Se ainda não tem movimentações após tentar transferir/vincular, retornar
        if total_movimentacoes == 0:
            return redirect('analise_cenarios', propriedade_id=propriedade.id)
    
    try:
        # Verificar quantas vendas já existem antes de gerar
        vendas_existentes_antes = VendaProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=planejamento
        ).count()
        
        # Verificar se há movimentações antes de tentar gerar
        movimentacoes_finais = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            tipo_movimentacao='VENDA',
            quantidade__gt=0,
            planejamento=planejamento
        ).count()
        
        if movimentacoes_finais == 0:
            messages.error(
                request,
                f'❌ Não foi possível gerar vendas. Nenhuma movimentação de VENDA encontrada vinculada ao planejamento {planejamento.codigo}. '
                f'Verifique se as movimentações foram transferidas corretamente.'
            )
            return redirect('analise_cenarios', propriedade_id=propriedade.id)
        
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Iniciando geração de vendas para planejamento {planejamento.codigo} com {movimentacoes_finais} movimentações")
        
        # Verificar se há cenários, se não houver, criar um baseline automaticamente
        total_cenarios = planejamento.cenarios.count()
        if total_cenarios == 0:
            logger.info(f"Nenhum cenário encontrado. Criando cenário Baseline automaticamente...")
            cenario_baseline = CenarioPlanejamento.objects.create(
                planejamento=planejamento,
                nome="Baseline",
                descricao="Cenário base criado automaticamente para geração de vendas",
                is_baseline=True
            )
            messages.info(
                request,
                f'ℹ️ Um cenário "Baseline" foi criado automaticamente para a projeção {planejamento.codigo}.'
            )
            logger.info(f"Cenário Baseline criado: {cenario_baseline.id}")
        
        vendas_geradas = gerar_vendas_todos_cenarios(propriedade, planejamento)
        total_vendas = sum(len(vendas) for vendas in vendas_geradas.values())
        
        # Verificar quantas vendas existem agora
        vendas_existentes_agora = VendaProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=planejamento
        ).count()
        
        logger.info(f"Vendas geradas: {total_vendas} novas, {vendas_existentes_agora} total")
        
        if total_vendas > 0:
            messages.success(
                request,
                f'✅ Vendas geradas com sucesso para a projeção {planejamento.codigo}! '
                f'Total: {total_vendas} novas vendas criadas para {len(vendas_geradas)} cenários. '
                f'Total geral de vendas: {vendas_existentes_agora}'
            )
        else:
            # Verificar quantos cenários têm vendas
            cenarios_com_vendas = VendaProjetada.objects.filter(
                propriedade=propriedade,
                planejamento=planejamento
            ).values('cenario').distinct().count()
            
            total_cenarios = planejamento.cenarios.count()
            
            if vendas_existentes_agora > 0:
                messages.info(
                    request,
                    f'ℹ️ Todas as vendas já estão geradas! '
                    f'Total: {vendas_existentes_agora} vendas para {cenarios_com_vendas} cenários. '
                    f'Foram encontradas {movimentacoes_finais} movimentações de VENDA, '
                    f'e todas já possuem vendas projetadas geradas.'
                )
            else:
                # Verificar se há cenários no planejamento
                if total_cenarios == 0:
                    messages.warning(
                        request,
                        f'⚠️ Nenhum cenário encontrado no planejamento {planejamento.codigo}. '
                        f'Crie pelo menos um cenário antes de gerar vendas.'
                    )
                else:
                    messages.warning(
                        request,
                        f'⚠️ Nenhuma venda foi gerada. Foram encontradas {movimentacoes_finais} movimentações de VENDA, '
                        f'mas nenhuma venda foi criada. Verifique os logs para mais detalhes. '
                        f'Total de cenários: {total_cenarios}'
                    )
    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao gerar vendas: {traceback.format_exc()}')
        messages.error(
            request, 
            f'❌ Erro ao gerar vendas: {str(e)}. '
            f'Foram encontradas {total_movimentacoes} movimentações de VENDA. '
            f'Verifique os logs para mais detalhes.'
        )
    
    return redirect('analise_cenarios', propriedade_id=propriedade.id)


@login_required
def relatorio_vendas_projecao(request, propriedade_id):
    """Relatório de vendas da projeção - página principal"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento - pode vir da URL como parâmetro (codigo ou planejamento_id)
    planejamento = None
    codigo_busca = request.GET.get('codigo', '').strip()
    planejamento_id = request.GET.get('planejamento_id')
    
    # PRIORIDADE 1: Buscar por código (projeção selecionada pelo usuário)
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            planejamento = None
        except PlanejamentoAnual.MultipleObjectsReturned:
            planejamento = PlanejamentoAnual.objects.filter(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            ).order_by('-data_criacao').first()
    
    # PRIORIDADE 2: Buscar por ID
    if not planejamento and planejamento_id:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                id=planejamento_id,
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            planejamento = None
    
    # PRIORIDADE 3: Buscar o planejamento mais recente por data de criação
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.warning(request, 'Nenhum planejamento encontrado. Crie um planejamento primeiro gerando uma projeção na página "Projeções do Rebanho".')
        return redirect('analise_cenarios', propriedade_id=propriedade.id)
    
    # Filtros
    cenario_id = request.GET.get('cenario')
    ano = request.GET.get('ano')  # Removido default - se não informado, mostra todos os anos
    mes = request.GET.get('mes')
    tipo_relatorio = request.GET.get('tipo', 'geral')  # geral, mensal, anual
    
    # Buscar vendas - TODAS do planejamento (pode incluir múltiplos anos da projeção)
    vendas_query = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento
    )
    
    if cenario_id:
        vendas_query = vendas_query.filter(cenario_id=cenario_id)
    
    # Filtrar por ano apenas se especificado (se não informado, mostra todos os anos)
    if ano:
        vendas_query = vendas_query.filter(data_venda__year=int(ano))
    
    if mes:
        vendas_query = vendas_query.filter(data_venda__month=int(mes))
    
    vendas = vendas_query.order_by('-data_venda', 'categoria')
    
    # Debug: Log para entender o que está sendo buscado
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"DEBUG relatorio_vendas_projecao - Planejamento: {planejamento.codigo if planejamento else 'None'}")
    logger.info(f"DEBUG relatorio_vendas_projecao - Planejamento ID: {planejamento.id if planejamento else 'None'}")
    logger.info(f"DEBUG relatorio_vendas_projecao - Total de vendas encontradas: {vendas.count()}")
    
    # Verificar se há vendas no banco de dados para este planejamento (sem filtros)
    vendas_totais_planejamento = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento
    ).count()
    logger.info(f"DEBUG relatorio_vendas_projecao - Total de vendas no banco para este planejamento: {vendas_totais_planejamento}")
    
    # Se não há vendas, verificar se há movimentações
    if vendas_totais_planejamento == 0:
        movimentacoes_count = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=planejamento,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        ).count()
        logger.info(f"DEBUG relatorio_vendas_projecao - Movimentações VENDA encontradas: {movimentacoes_count}")
        
        cenarios_count = planejamento.cenarios.count()
        logger.info(f"DEBUG relatorio_vendas_projecao - Cenários encontrados: {cenarios_count}")
        
        if movimentacoes_count > 0 and cenarios_count == 0:
            messages.warning(
                request,
                f'⚠️ Nenhuma venda encontrada para a projeção {planejamento.codigo}. '
                f'Foram encontradas {movimentacoes_count} movimentações de VENDA, mas não há cenários. '
                f'Clique em "Gerar Vendas" para criar um cenário automaticamente e gerar as vendas.'
            )
        elif movimentacoes_count == 0:
            messages.warning(
                request,
                f'⚠️ Nenhuma movimentação de VENDA encontrada para a projeção {planejamento.codigo}. '
                f'É necessário gerar primeiro as projeções na página "Projeções do Rebanho".'
            )
    
    # Buscar range de anos disponíveis nas vendas/movimentações para exibir no filtro
    # Primeiro tenta buscar das vendas, se não houver, busca das movimentações
    if vendas.exists():
        anos_disponiveis = list(vendas.values_list('data_venda__year', flat=True).distinct().order_by('data_venda__year'))
        logger.info(f"DEBUG relatorio_vendas_projecao - Anos encontrados nas vendas: {anos_disponiveis}")
    else:
        # Buscar anos das movimentações projetadas vinculadas ao planejamento
        anos_disponiveis = list(MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=planejamento,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        ).values_list('data_movimentacao__year', flat=True).distinct().order_by('data_movimentacao__year'))
    
    if not anos_disponiveis:
        anos_disponiveis = [planejamento.ano]
    
    # Se não houver vendas e houver movimentações de VENDA, gerar automaticamente
    if not vendas.exists():
        # Verificar se existem movimentações de VENDA para gerar vendas
        # Buscar com planejamento ou sem planejamento (apenas por propriedade e ano)
        movimentacoes_venda = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        )
        
        # Filtrar por planejamento - buscar TODAS as movimentações do planejamento (todos os anos)
        if planejamento:
            movimentacoes_venda = movimentacoes_venda.filter(planejamento=planejamento)
        elif ano:
            movimentacoes_venda = movimentacoes_venda.filter(data_movimentacao__year=ano)
        
        if movimentacoes_venda.exists():
            # Gerar vendas para todos os cenários do planejamento
            from .services.gerar_vendas_projecao import gerar_vendas_todos_cenarios
            try:
                gerar_vendas_todos_cenarios(propriedade, planejamento)
                # Recarregar vendas
                vendas_query = VendaProjetada.objects.filter(
                    propriedade=propriedade
                )
                if planejamento:
                    vendas_query = vendas_query.filter(planejamento=planejamento)
                if cenario_id:
                    vendas_query = vendas_query.filter(cenario_id=cenario_id)
                if ano:
                    vendas_query = vendas_query.filter(data_venda__year=ano)
                if mes:
                    vendas_query = vendas_query.filter(data_venda__month=mes)
                vendas = vendas_query.order_by('-data_venda', 'categoria')
                messages.info(request, f'Vendas geradas automaticamente! Total: {vendas.count()} vendas criadas.')
            except Exception as e:
                import traceback
                messages.warning(request, f'Erro ao gerar vendas automaticamente: {str(e)}')
                # Log do erro para debug
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Erro ao gerar vendas: {traceback.format_exc()}')
    
    # Buscar cenários disponíveis
    cenarios = planejamento.cenarios.all().order_by('-is_baseline', 'nome')
    
    # Se não há cenários mas há movimentações, criar um cenário padrão automaticamente
    if not cenarios.exists():
        movimentacoes_existem = MovimentacaoProjetada.objects.filter(
            propriedade=propriedade,
            planejamento=planejamento,
            tipo_movimentacao='VENDA',
            quantidade__gt=0
        ).exists()
        
        if movimentacoes_existem:
            # Criar cenário baseline automaticamente
            cenario_baseline = CenarioPlanejamento.objects.create(
                planejamento=planejamento,
                nome="Baseline",
                descricao="Cenário base criado automaticamente",
                is_baseline=True
            )
            cenarios = planejamento.cenarios.all().order_by('-is_baseline', 'nome')
            messages.info(request, f'ℹ️ Um cenário "Baseline" foi criado automaticamente para a projeção {planejamento.codigo}.')
    
    # Converter queryset para lista para melhor renderização no template
    vendas_lista = list(vendas)
    
    # Agrupar vendas por lote (data, cenário, categoria) para mostrar múltiplos clientes
    vendas_agrupadas_por_lote = defaultdict(lambda: {
        'vendas': [],
        'data_venda': None,
        'cenario': None,
        'categoria': None,
        'quantidade_total': 0,
        'peso_total': Decimal('0'),
        'peso_medio_kg': None,  # Peso unitário (médio por animal)
        'valor_total': Decimal('0'),
        'clientes': [],
        'valor_por_kg': None,
        'valor_por_animal': None,
        'data_recebimento': None
    })
    
    for venda in vendas_lista:
        # Criar chave única para o lote: data + cenário + categoria
        cenario_id = venda.cenario.id if venda.cenario else None
        categoria_id = venda.categoria.id
        chave_lote = f"{venda.data_venda}_{cenario_id}_{categoria_id}"
        
        lote = vendas_agrupadas_por_lote[chave_lote]
        lote['vendas'].append(venda)
        lote['data_venda'] = venda.data_venda
        lote['cenario'] = venda.cenario
        lote['categoria'] = venda.categoria
        # Adicionar ID da primeira venda para referência
        if 'primeira_venda_id' not in lote:
            lote['primeira_venda_id'] = venda.id
        lote['quantidade_total'] += venda.quantidade
        lote['peso_total'] += venda.peso_total_kg or Decimal('0')
        lote['valor_total'] += venda.valor_total or Decimal('0')
        
        # Coletar clientes únicos
        cliente_nome = venda.cliente_nome or "Cliente não definido"
        if cliente_nome not in [c['nome'] for c in lote['clientes']]:
            lote['clientes'].append({
                'nome': cliente_nome,
                'quantidade': venda.quantidade,
                'valor': venda.valor_total or Decimal('0'),
                'peso': venda.peso_total_kg or Decimal('0'),
                'peso_medio_kg': venda.peso_medio_kg,  # Peso unitário do cliente
                'valor_por_kg': venda.valor_por_kg,  # Valor por KG do cliente
                'valor_por_animal': venda.valor_por_animal,  # Valor por animal do cliente
                'data_recebimento': venda.data_recebimento,  # Data de recebimento do cliente
                'venda_id': venda.id  # ID da venda para edição
            })
        else:
            # Se cliente já existe, somar valores (mas manter o primeiro venda_id para edição)
            for cliente_info in lote['clientes']:
                if cliente_info['nome'] == cliente_nome:
                    cliente_info['quantidade'] += venda.quantidade
                    cliente_info['valor'] += venda.valor_total or Decimal('0')
                    cliente_info['peso'] += venda.peso_total_kg or Decimal('0')
                    # Se não tinha valor_por_kg, usar o da venda atual
                    if not cliente_info.get('valor_por_kg') and venda.valor_por_kg:
                        cliente_info['valor_por_kg'] = venda.valor_por_kg
                    # Se não tinha valor_por_animal, usar o da venda atual
                    if not cliente_info.get('valor_por_animal') and venda.valor_por_animal:
                        cliente_info['valor_por_animal'] = venda.valor_por_animal
                    # Se não tinha data_recebimento, usar o da venda atual
                    if not cliente_info.get('data_recebimento') and venda.data_recebimento:
                        cliente_info['data_recebimento'] = venda.data_recebimento
                    break
        
        # Usar valores da primeira venda como referência (ou calcular média)
        if lote['valor_por_kg'] is None and venda.valor_por_kg:
            lote['valor_por_kg'] = venda.valor_por_kg
        if lote['valor_por_animal'] is None and venda.valor_por_animal:
            lote['valor_por_animal'] = venda.valor_por_animal
        if lote['data_recebimento'] is None and venda.data_recebimento:
            lote['data_recebimento'] = venda.data_recebimento
        # Peso médio: usar o primeiro valor encontrado ou calcular média ponderada
        if lote['peso_medio_kg'] is None and venda.peso_medio_kg:
            lote['peso_medio_kg'] = venda.peso_medio_kg
        elif lote['quantidade_total'] > 0 and lote['peso_total'] > 0:
            # Calcular peso médio se não existir
            lote['peso_medio_kg'] = lote['peso_total'] / lote['quantidade_total']
    
    # Converter para lista ordenada
    vendas_agrupadas_lista = []
    for chave_lote in sorted(vendas_agrupadas_por_lote.keys(), 
                            key=lambda k: (vendas_agrupadas_por_lote[k]['data_venda'], 
                                          vendas_agrupadas_por_lote[k]['categoria'].nome if vendas_agrupadas_por_lote[k]['categoria'] else ''), 
                            reverse=True):
        lote = vendas_agrupadas_por_lote[chave_lote]
        
        # Calcular peso médio (peso_total / quantidade_total) - sempre recalcular para garantir precisão
        if lote['quantidade_total'] > 0 and lote['peso_total'] > 0:
            lote['peso_medio_kg'] = lote['peso_total'] / lote['quantidade_total']
        elif lote['peso_medio_kg'] is None:
            lote['peso_medio_kg'] = None
        
        # Calcular valor por kg (valor_total / peso_total) - sempre recalcular para garantir precisão
        if lote['peso_total'] > 0 and lote['valor_total'] > 0:
            lote['valor_por_kg'] = lote['valor_total'] / lote['peso_total']
        elif lote['valor_por_kg'] is None:
            lote['valor_por_kg'] = None
        
        # Calcular valor por animal (valor_total / quantidade_total) - sempre recalcular para garantir precisão
        if lote['quantidade_total'] > 0 and lote['valor_total'] > 0:
            lote['valor_por_animal'] = lote['valor_total'] / lote['quantidade_total']
        elif lote['valor_por_animal'] is None:
            lote['valor_por_animal'] = None
        
        vendas_agrupadas_lista.append(lote)
    
    # Estatísticas gerais (usando vendas agrupadas)
    total_vendas = len(vendas_agrupadas_lista)
    total_quantidade = sum(l['quantidade_total'] for l in vendas_agrupadas_lista)
    total_valor = sum(l['valor_total'] for l in vendas_agrupadas_lista)
    total_peso = sum(l['peso_total'] for l in vendas_agrupadas_lista)
    
    logger.info(f"DEBUG relatorio_vendas_projecao - Estatísticas: Total={total_vendas}, Quantidade={total_quantidade}, Valor={total_valor}, Peso={total_peso}")
    
    # Agrupar por ANO primeiro
    vendas_por_ano = defaultdict(lambda: {
        'vendas': [],
        'total_quantidade': 0,
        'total_valor': Decimal('0'),
        'total_peso': Decimal('0'),
        'meses': defaultdict(lambda: {
            'vendas': [],
            'total_quantidade': 0,
            'total_valor': Decimal('0'),
            'total_peso': Decimal('0')
        })
    })
    
    for venda in vendas_lista:
        ano = venda.data_venda.year
        mes_num = venda.data_venda.month
        mes_ano = f"{ano}-{mes_num:02d}"
        
        # Agrupar por ano
        vendas_por_ano[ano]['vendas'].append(venda)
        vendas_por_ano[ano]['total_quantidade'] += venda.quantidade
        vendas_por_ano[ano]['total_valor'] += venda.valor_total or Decimal('0')
        if venda.peso_total_kg:
            vendas_por_ano[ano]['total_peso'] += venda.peso_total_kg
        
        # Agrupar por mês dentro do ano
        vendas_por_ano[ano]['meses'][mes_num]['vendas'].append(venda)
        vendas_por_ano[ano]['meses'][mes_num]['total_quantidade'] += venda.quantidade
        vendas_por_ano[ano]['meses'][mes_num]['total_valor'] += venda.valor_total or Decimal('0')
        if venda.peso_total_kg:
            vendas_por_ano[ano]['meses'][mes_num]['total_peso'] += venda.peso_total_kg
    
    # Converter para lista ordenada por ano
    vendas_por_ano_lista = []
    for ano in sorted(vendas_por_ano.keys(), reverse=True):
        dados_ano = vendas_por_ano[ano]
        meses_lista = []
        for mes_num in sorted(dados_ano['meses'].keys()):
            meses_lista.append({
                'mes': mes_num,
                'mes_nome': month_name[mes_num],
                **dados_ano['meses'][mes_num]
            })
        
        vendas_por_ano_lista.append({
            'ano': ano,
            'total_quantidade': dados_ano['total_quantidade'],
            'total_valor': dados_ano['total_valor'],
            'total_peso': dados_ano['total_peso'],
            'vendas': dados_ano['vendas'],
            'meses': meses_lista
        })
    
    # Manter também agrupamento mensal para compatibilidade
    vendas_por_mes = defaultdict(lambda: {
        'vendas': [],
        'total_quantidade': 0,
        'total_valor': Decimal('0'),
        'total_peso': Decimal('0')
    })
    
    for venda in vendas:
        mes_ano = f"{venda.data_venda.year}-{venda.data_venda.month:02d}"
        vendas_por_mes[mes_ano]['vendas'].append(venda)
        vendas_por_mes[mes_ano]['total_quantidade'] += venda.quantidade
        vendas_por_mes[mes_ano]['total_valor'] += venda.valor_total
        if venda.peso_total_kg:
            vendas_por_mes[mes_ano]['total_peso'] += venda.peso_total_kg
    
    # Converter para lista ordenada
    vendas_por_mes_lista = []
    for mes_ano in sorted(vendas_por_mes.keys(), reverse=True):
        ano_mes, mes_num = mes_ano.split('-')
        vendas_por_mes_lista.append({
            'mes_ano': mes_ano,
            'ano': int(ano_mes),
            'mes': int(mes_num),
            'mes_nome': month_name[int(mes_num)],
            **vendas_por_mes[mes_ano]
        })
    
    # Agrupar por categoria
    vendas_por_categoria = defaultdict(lambda: {
        'vendas': [],
        'total_quantidade': 0,
        'total_valor': Decimal('0'),
        'total_peso': Decimal('0')
    })
    
    for venda in vendas_lista:
        cat_nome = venda.categoria.nome
        vendas_por_categoria[cat_nome]['vendas'].append(venda)
        vendas_por_categoria[cat_nome]['total_quantidade'] += venda.quantidade
        vendas_por_categoria[cat_nome]['total_valor'] += venda.valor_total or Decimal('0')
        if venda.peso_total_kg:
            vendas_por_categoria[cat_nome]['total_peso'] += venda.peso_total_kg
    
    # Agrupar por cliente
    vendas_por_cliente = defaultdict(lambda: {
        'vendas': [],
        'total_quantidade': 0,
        'total_valor': Decimal('0'),
        'total_peso': Decimal('0')
    })
    
    for venda in vendas_lista:
        cliente_nome = venda.cliente_nome or "Cliente não definido"
        vendas_por_cliente[cliente_nome]['vendas'].append(venda)
        vendas_por_cliente[cliente_nome]['total_quantidade'] += venda.quantidade
        vendas_por_cliente[cliente_nome]['total_valor'] += venda.valor_total
        if venda.peso_total_kg:
            vendas_por_cliente[cliente_nome]['total_peso'] += venda.peso_total_kg
    
    context = {
        'propriedade': propriedade,
        'planejamento': planejamento,
        'cenarios': cenarios,
        'cenario_selecionado': int(cenario_id) if cenario_id else None,
        'ano_selecionado': int(ano) if ano else None,
        'mes_selecionado': int(mes) if mes else None,
        'tipo_relatorio': tipo_relatorio,
        'vendas': vendas_lista,  # Mantém para compatibilidade
        'vendas_agrupadas': vendas_agrupadas_lista,  # Nova estrutura agrupada por lote
        'total_vendas': total_vendas,
        'total_quantidade': total_quantidade,
        'total_valor': float(total_valor),
        'total_peso': float(total_peso),
        'vendas_por_mes': vendas_por_mes_lista,
        'vendas_por_ano': vendas_por_ano_lista,
        'vendas_por_categoria': dict(vendas_por_categoria),
        'vendas_por_cliente': dict(vendas_por_cliente),
        'anos_disponiveis': anos_disponiveis,
        'anos_choices': range(min(anos_disponiveis) if anos_disponiveis else planejamento.ano - 5,
                             max(anos_disponiveis) + 1 if anos_disponiveis else planejamento.ano + 5),
    }
    
    return render(request, 'gestao_rural/relatorio_vendas_projecao.html', context)


@login_required
def relatorio_vendas_mensal(request, propriedade_id, ano, mes):
    """Relatório mensal detalhado de vendas"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento - pode vir da URL como parâmetro (codigo ou planejamento_id)
    planejamento = None
    codigo_busca = request.GET.get('codigo', '').strip()
    planejamento_id = request.GET.get('planejamento_id')
    
    # PRIORIDADE 1: Buscar por código (projeção selecionada pelo usuário)
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            planejamento = None
        except PlanejamentoAnual.MultipleObjectsReturned:
            planejamento = PlanejamentoAnual.objects.filter(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            ).order_by('-data_criacao').first()
    
    # PRIORIDADE 2: Buscar por ID
    if not planejamento and planejamento_id:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                id=planejamento_id,
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            planejamento = None
    
    # PRIORIDADE 3: Buscar o planejamento mais recente por data de criação
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.warning(request, f'Nenhum planejamento encontrado.')
        return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    cenario_id = request.GET.get('cenario')
    
    # Buscar vendas do mês
    vendas_query = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento,
        data_venda__year=ano,
        data_venda__month=mes
    )
    
    if cenario_id:
        vendas_query = vendas_query.filter(cenario_id=cenario_id)
    
    vendas = vendas_query.order_by('data_venda', 'categoria')
    
    # Converter queryset para lista
    vendas_lista = list(vendas)
    
    # Estatísticas do mês
    total_quantidade = sum(v.quantidade for v in vendas_lista)
    total_valor = sum(v.valor_total for v in vendas_lista)
    total_peso = sum(v.peso_total_kg or Decimal('0') for v in vendas_lista)
    
    # Agrupar por categoria
    vendas_por_categoria = defaultdict(lambda: {
        'vendas': [],
        'quantidade': 0,
        'valor': Decimal('0'),
        'peso': Decimal('0')
    })
    
    for venda in vendas_lista:
        cat_nome = venda.categoria.nome
        vendas_por_categoria[cat_nome]['vendas'].append(venda)
        vendas_por_categoria[cat_nome]['quantidade'] += venda.quantidade
        vendas_por_categoria[cat_nome]['valor'] += venda.valor_total
        if venda.peso_total_kg:
            vendas_por_categoria[cat_nome]['peso'] += venda.peso_total_kg
    
    context = {
        'propriedade': propriedade,
        'planejamento': planejamento,
        'ano': ano,
        'mes': mes,
        'mes_nome': month_name[mes],
        'vendas': vendas_lista,
        'vendas_por_categoria': dict(vendas_por_categoria),
        'total_quantidade': total_quantidade,
        'total_valor': float(total_valor),
        'total_peso': float(total_peso),
        'cenario_selecionado': int(cenario_id) if cenario_id else None,
    }
    
    return render(request, 'gestao_rural/relatorio_vendas_mensal.html', context)


@login_required
def relatorio_vendas_anual(request, propriedade_id, ano):
    """Relatório anual consolidado de vendas"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento - pode vir da URL como parâmetro (codigo ou planejamento_id)
    planejamento = None
    codigo_busca = request.GET.get('codigo', '').strip()
    planejamento_id = request.GET.get('planejamento_id')
    
    # PRIORIDADE 1: Buscar por código (projeção selecionada pelo usuário)
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            planejamento = None
        except PlanejamentoAnual.MultipleObjectsReturned:
            planejamento = PlanejamentoAnual.objects.filter(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            ).order_by('-data_criacao').first()
    
    # PRIORIDADE 2: Buscar por ID
    if not planejamento and planejamento_id:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                id=planejamento_id,
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            planejamento = None
    
    # PRIORIDADE 3: Buscar o planejamento mais recente por data de criação
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.warning(request, f'Nenhum planejamento encontrado.')
        return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    cenario_id = request.GET.get('cenario')
    
    # Buscar vendas do ano - TODAS do planejamento (pode incluir múltiplos anos da projeção)
    vendas_query = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento,
        data_venda__year=ano
    )
    
    if cenario_id:
        vendas_query = vendas_query.filter(cenario_id=cenario_id)
    
    vendas = vendas_query.order_by('data_venda', 'categoria')
    
    # Converter queryset para lista para melhor renderização no template
    vendas_lista = list(vendas)
    
    # Estatísticas do ano
    total_quantidade = sum(v.quantidade for v in vendas_lista)
    total_valor = sum(v.valor_total or Decimal('0') for v in vendas_lista)
    total_peso = sum(v.peso_total_kg or Decimal('0') for v in vendas_lista)
    
    # Agrupar por mês
    vendas_por_mes = defaultdict(lambda: {
        'vendas': [],
        'quantidade': 0,
        'valor': Decimal('0'),
        'peso': Decimal('0')
    })
    
    for venda in vendas_lista:
        mes_num = venda.data_venda.month
        vendas_por_mes[mes_num]['vendas'].append(venda)
        vendas_por_mes[mes_num]['quantidade'] += venda.quantidade
        vendas_por_mes[mes_num]['valor'] += venda.valor_total or Decimal('0')
        if venda.peso_total_kg:
            vendas_por_mes[mes_num]['peso'] += venda.peso_total_kg
    
    # Converter para lista ordenada
    vendas_por_mes_lista = []
    for mes_num in sorted(vendas_por_mes.keys()):
        vendas_por_mes_lista.append({
            'mes': mes_num,
            'mes_nome': month_name[mes_num],
            **vendas_por_mes[mes_num]
        })
    
    # Agrupar por categoria
    vendas_por_categoria = defaultdict(lambda: {
        'quantidade': 0,
        'valor': Decimal('0'),
        'peso': Decimal('0')
    })
    
    for venda in vendas_lista:
        cat_nome = venda.categoria.nome
        vendas_por_categoria[cat_nome]['quantidade'] += venda.quantidade
        vendas_por_categoria[cat_nome]['valor'] += venda.valor_total or Decimal('0')
        if venda.peso_total_kg:
            vendas_por_categoria[cat_nome]['peso'] += venda.peso_total_kg
    
    context = {
        'propriedade': propriedade,
        'planejamento': planejamento,
        'ano': ano,
        'vendas': vendas_lista,
        'vendas_por_mes': vendas_por_mes_lista,
        'vendas_por_categoria': dict(vendas_por_categoria),
        'total_quantidade': total_quantidade,
        'total_valor': float(total_valor),
        'total_peso': float(total_peso),
        'cenario_selecionado': int(cenario_id) if cenario_id else None,
    }
    
    return render(request, 'gestao_rural/relatorio_vendas_anual.html', context)


@login_required
def editar_cliente_venda(request, propriedade_id, venda_id):
    """View para editar cliente de uma venda projetada"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    venda = get_object_or_404(
        VendaProjetada,
        id=venda_id,
        propriedade=propriedade
    )
    
    if request.method == 'POST':
        cliente_nome = request.POST.get('cliente_nome', '').strip()
        quantidade_cliente = request.POST.get('quantidade_cliente', venda.quantidade)
        
        try:
            quantidade_cliente = int(quantidade_cliente)
        except (ValueError, TypeError):
            quantidade_cliente = venda.quantidade
        
        # Se quantidade for menor que a venda total, dividir a venda
        if quantidade_cliente < venda.quantidade and quantidade_cliente > 0:
            # Criar nova venda com o restante
            quantidade_restante = venda.quantidade - quantidade_cliente
            
            # Calcular valores proporcionais
            proporcao = Decimal(quantidade_cliente) / Decimal(venda.quantidade)
            valor_cliente = venda.valor_total * proporcao
            peso_cliente = (venda.peso_total_kg or Decimal('0')) * proporcao
            
            proporcao_restante = Decimal(quantidade_restante) / Decimal(venda.quantidade)
            valor_restante = venda.valor_total * proporcao_restante
            peso_restante = (venda.peso_total_kg or Decimal('0')) * proporcao_restante
            
            # Atualizar venda atual
            venda.quantidade = quantidade_cliente
            venda.cliente_nome = cliente_nome
            venda.valor_total = valor_cliente
            venda.peso_total_kg = peso_cliente
            if venda.peso_total_kg and venda.quantidade:
                venda.peso_medio_kg = venda.peso_total_kg / venda.quantidade
            if venda.valor_total and venda.quantidade:
                venda.valor_por_animal = venda.valor_total / venda.quantidade
            if venda.peso_total_kg and venda.valor_total:
                venda.valor_por_kg = venda.valor_total / venda.peso_total_kg
            venda.save()
            
            # Salvar cliente original antes de atualizar
            cliente_original = venda.cliente_nome
            
            # Atualizar venda atual
            venda.quantidade = quantidade_cliente
            venda.cliente_nome = cliente_nome
            venda.valor_total = valor_cliente
            venda.peso_total_kg = peso_cliente
            if venda.peso_total_kg and venda.quantidade:
                venda.peso_medio_kg = venda.peso_total_kg / venda.quantidade
            if venda.valor_total and venda.quantidade:
                venda.valor_por_animal = venda.valor_total / venda.quantidade
            if venda.peso_total_kg and venda.valor_total:
                venda.valor_por_kg = venda.valor_total / venda.peso_total_kg
            venda.save()
            
            # Criar nova venda com o restante (mantém cliente original)
            VendaProjetada.objects.create(
                propriedade=venda.propriedade,
                planejamento=venda.planejamento,
                cenario=venda.cenario,
                movimentacao_projetada=venda.movimentacao_projetada,
                data_venda=venda.data_venda,
                categoria=venda.categoria,
                quantidade=quantidade_restante,
                cliente_nome=cliente_original,  # Mantém o cliente original
                peso_total_kg=peso_restante,
                peso_medio_kg=venda.peso_medio_kg,
                valor_por_kg=venda.valor_por_kg,
                valor_por_animal=venda.valor_por_animal,
                valor_total=valor_restante,
                data_recebimento=venda.data_recebimento,
                prazo_pagamento_dias=venda.prazo_pagamento_dias,
                observacoes=venda.observacoes
            )
        else:
            # Atualizar venda completa
            venda.cliente_nome = cliente_nome
            venda.save()
        
        messages.success(request, f'✅ Cliente atualizado com sucesso!')
        
        # Redirecionar de volta para o relatório
        codigo = request.GET.get('codigo', '')
        cenario = request.GET.get('cenario', '')
        params = []
        if codigo:
            params.append(f'codigo={codigo}')
        if cenario:
            params.append(f'cenario={cenario}')
        query_string = '?' + '&'.join(params) if params else ''
        
        return redirect(f"{reverse('relatorio_vendas_projecao', args=[propriedade_id])}{query_string}")
    
    # Buscar clientes cadastrados
    try:
        from .models_cadastros import Cliente
        clientes = Cliente.objects.filter(
            propriedade=propriedade,
            ativo=True
        ).order_by('nome')
    except:
        clientes = []
    
    context = {
        'propriedade': propriedade,
        'venda': venda,
        'clientes': clientes,
    }
    
    return render(request, 'gestao_rural/editar_cliente_venda.html', context)


@login_required
def buscar_clientes_ajax(request, propriedade_id):
    """View AJAX para buscar clientes cadastrados"""
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    termo = request.GET.get('q', '').strip()
    
    try:
        from .models_cadastros import Cliente
        clientes = Cliente.objects.filter(
            propriedade=propriedade,
            ativo=True
        )
        
        if termo:
            clientes = clientes.filter(
                Q(nome__icontains=termo) | 
                Q(nome_fantasia__icontains=termo) |
                Q(cpf_cnpj__icontains=termo)
            )
        
        clientes = clientes.order_by('nome')[:20]
        
        resultados = [{
            'id': cliente.id,
            'nome': cliente.nome,
            'nome_fantasia': cliente.nome_fantasia or '',
            'cpf_cnpj': cliente.cpf_cnpj,
            'tipo_cliente': cliente.get_tipo_cliente_display() if hasattr(cliente, 'get_tipo_cliente_display') else ''
        } for cliente in clientes]
        
        return JsonResponse({'clientes': resultados})
    except Exception as e:
        return JsonResponse({'clientes': [], 'erro': str(e)})


@login_required
def atualizar_peso_unitario_venda(request, propriedade_id, venda_id):
    """View AJAX para atualizar o peso unitário (peso médio) de uma venda"""
    from django.http import JsonResponse
    from decimal import Decimal, InvalidOperation
    
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    venda = get_object_or_404(
        VendaProjetada,
        id=venda_id,
        propriedade=propriedade
    )
    
    if request.method == 'POST':
        try:
            peso_unitario_str = request.POST.get('peso_unitario', '').strip()
            
            if not peso_unitario_str:
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Peso unitário não informado'
                }, status=400)
            
            # Converter para Decimal, substituindo vírgula por ponto
            peso_unitario_str = peso_unitario_str.replace(',', '.')
            
            try:
                peso_unitario = Decimal(peso_unitario_str)
            except (InvalidOperation, ValueError):
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Valor inválido para peso unitário'
                }, status=400)
            
            if peso_unitario < 0:
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Peso unitário não pode ser negativo'
                }, status=400)
            
            # Atualizar peso médio
            venda.peso_medio_kg = peso_unitario
            
            # Recalcular peso total se necessário
            if venda.quantidade > 0:
                venda.peso_total_kg = peso_unitario * venda.quantidade
            
            # Recalcular valor total se houver valor por kg
            if venda.valor_por_kg and venda.peso_total_kg:
                venda.valor_total = venda.valor_por_kg * venda.peso_total_kg
            elif venda.valor_por_animal and venda.quantidade:
                venda.valor_total = venda.valor_por_animal * venda.quantidade
            
            venda.save()
            
            return JsonResponse({
                'sucesso': True,
                'peso_unitario': str(venda.peso_medio_kg),
                'peso_total': str(venda.peso_total_kg) if venda.peso_total_kg else None,
                'valor_total': str(venda.valor_total) if venda.valor_total else None
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                'sucesso': False,
                'erro': f'Erro ao atualizar peso unitário: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'sucesso': False,
        'erro': 'Método não permitido'
    }, status=405)


@login_required
def atualizar_valor_por_kg_venda(request, propriedade_id, venda_id):
    """View AJAX para atualizar o valor por KG de uma venda"""
    from django.http import JsonResponse
    from decimal import Decimal, InvalidOperation
    
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    venda = get_object_or_404(
        VendaProjetada,
        id=venda_id,
        propriedade=propriedade
    )
    
    if request.method == 'POST':
        try:
            valor_por_kg_str = request.POST.get('valor_por_kg', '').strip()
            
            if not valor_por_kg_str:
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Valor por KG não informado'
                }, status=400)
            
            # Converter para Decimal, substituindo vírgula por ponto e removendo R$
            valor_por_kg_str = valor_por_kg_str.replace('R$', '').replace(' ', '').replace(',', '.')
            
            try:
                valor_por_kg = Decimal(valor_por_kg_str)
            except (InvalidOperation, ValueError):
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Valor inválido para valor por KG'
                }, status=400)
            
            if valor_por_kg < 0:
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Valor por KG não pode ser negativo'
                }, status=400)
            
            # Atualizar valor por kg
            venda.valor_por_kg = valor_por_kg
            
            # Recalcular peso total se necessário (usando peso médio)
            if not venda.peso_total_kg and venda.peso_medio_kg and venda.quantidade:
                venda.peso_total_kg = venda.peso_medio_kg * venda.quantidade
            
            # Recalcular valor total baseado no peso total
            if venda.peso_total_kg:
                venda.valor_total = valor_por_kg * venda.peso_total_kg
            # Se não houver peso total, usar quantidade e valor por animal como fallback
            elif venda.valor_por_animal and venda.quantidade:
                venda.valor_total = venda.valor_por_animal * venda.quantidade
            # Se não houver nenhum, calcular com peso médio
            elif venda.peso_medio_kg and venda.quantidade:
                peso_total_calculado = venda.peso_medio_kg * venda.quantidade
                venda.peso_total_kg = peso_total_calculado
                venda.valor_total = valor_por_kg * peso_total_calculado
            
            # Recalcular valor por animal automaticamente
            # valor_por_animal = valor_por_kg * peso_medio_kg
            if venda.peso_medio_kg:
                venda.valor_por_animal = valor_por_kg * venda.peso_medio_kg
            # Se não houver peso médio, calcular a partir do peso total e quantidade
            elif venda.peso_total_kg and venda.quantidade > 0:
                peso_medio_calculado = venda.peso_total_kg / venda.quantidade
                venda.peso_medio_kg = peso_medio_calculado
                venda.valor_por_animal = valor_por_kg * peso_medio_calculado
            
            venda.save()
            
            return JsonResponse({
                'sucesso': True,
                'valor_por_kg': str(venda.valor_por_kg),
                'valor_por_animal': str(venda.valor_por_animal) if venda.valor_por_animal else None,
                'valor_total': str(venda.valor_total) if venda.valor_total else None,
                'peso_total': str(venda.peso_total_kg) if venda.peso_total_kg else None
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                'sucesso': False,
                'erro': f'Erro ao atualizar valor por KG: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'sucesso': False,
        'erro': 'Método não permitido'
    }, status=405)


@login_required
def relatorio_vendas_consolidado_pdf(request, propriedade_id):
    """Relatório consolidado de vendas em PDF (paisagem) - Ano a ano e mensal até outubro/2025"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from datetime import datetime
    from calendar import month_name
    
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento
    codigo_busca = request.GET.get('codigo', '').strip()
    planejamento = None
    
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.warning(request, 'Nenhum planejamento encontrado.')
        return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    # Buscar vendas até outubro/2025
    vendas = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento
    ).filter(
        data_venda__lte=datetime(2025, 10, 31).date()
    ).order_by('data_venda__year', 'data_venda__month', 'categoria')
    
    # Agrupar por ano e mês
    vendas_por_ano_mes = defaultdict(lambda: defaultdict(lambda: {
        'vendas': [],
        'quantidade': 0,
        'valor': Decimal('0'),
        'peso': Decimal('0')
    }))
    
    for venda in vendas:
        ano = venda.data_venda.year
        mes = venda.data_venda.month
        vendas_por_ano_mes[ano][mes]['vendas'].append(venda)
        vendas_por_ano_mes[ano][mes]['quantidade'] += venda.quantidade
        vendas_por_ano_mes[ano][mes]['valor'] += venda.valor_total or Decimal('0')
        vendas_por_ano_mes[ano][mes]['peso'] += venda.peso_total_kg or Decimal('0')
    
    # Calcular totais por ano
    totais_por_ano = {}
    for ano in sorted(vendas_por_ano_mes.keys()):
        totais_por_ano[ano] = {
            'quantidade': sum(dados['quantidade'] for dados in vendas_por_ano_mes[ano].values()),
            'valor': sum(dados['valor'] for dados in vendas_por_ano_mes[ano].values()),
            'peso': sum(dados['peso'] for dados in vendas_por_ano_mes[ano].values())
        }
    
    # Total geral
    total_geral = {
        'quantidade': sum(t['quantidade'] for t in totais_por_ano.values()),
        'valor': sum(t['valor'] for t in totais_por_ano.values()),
        'peso': sum(t['peso'] for t in totais_por_ano.values())
    }
    
    # Criar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_vendas_consolidado_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    # Usar paisagem (landscape)
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    # Estilos customizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=TA_CENTER,
        spaceAfter=20,
        spaceBefore=10
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['BodyText'],
        fontSize=12,
        textColor=colors.HexColor('#4a5568'),
        alignment=TA_CENTER,
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Função para formatar moeda
    def formatar_moeda(valor):
        if valor is None or valor == 0:
            return 'R$ 0,00'
        try:
            valor_decimal = Decimal(str(valor))
            parte_inteira = int(abs(valor_decimal))
            parte_decimal = abs(valor_decimal) - parte_inteira
            parte_inteira_formatada = f'{parte_inteira:,}'.replace(',', '.')
            parte_decimal_formatada = f'{parte_decimal:.2f}'.split('.')[1]
            valor_formatado = f'{parte_inteira_formatada},{parte_decimal_formatada}'
            if valor_decimal < 0:
                valor_formatado = f'-{valor_formatado}'
            return f'R$ {valor_formatado}'
        except:
            return 'R$ 0,00'
    
    # Função para formatar número
    def formatar_numero(valor, decimais=0):
        if valor is None or valor == 0:
            return '0' if decimais == 0 else '0,00'
        try:
            valor_decimal = Decimal(str(valor))
            if decimais > 0:
                parte_inteira = int(abs(valor_decimal))
                parte_inteira_formatada = f'{parte_inteira:,}'.replace(',', '.')
                formato_decimal = f'{{:.{decimais}f}}'
                parte_decimal = formato_decimal.format(abs(valor_decimal) - parte_inteira).split('.')[1]
                return f'{parte_inteira_formatada},{parte_decimal}'
            else:
                parte_inteira = int(abs(valor_decimal))
                return f'{parte_inteira:,}'.replace(',', '.')
        except:
            return '0'
    
    # Elementos do PDF
    elements = []
    
    # Título
    elements.append(Paragraph(
        f"RELATÓRIO CONSOLIDADO DE VENDAS<br/>FAZENDA CANTA GALO",
        title_style
    ))
    
    elements.append(Paragraph(
        f"Período: Até Outubro/2025 | Projeção: {planejamento.codigo}",
        subtitle_style
    ))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Tabela consolidada por ano
    elements.append(Paragraph("CONSOLIDAÇÃO POR ANO", section_style))
    
    dados_consolidacao = [['Ano', 'Quantidade', 'Peso Total (kg)', 'Valor Total (R$)']]
    
    for ano in sorted(totais_por_ano.keys()):
        dados_consolidacao.append([
            str(ano),
            formatar_numero(totais_por_ano[ano]['quantidade']),
            formatar_numero(totais_por_ano[ano]['peso'], 2),
            formatar_moeda(totais_por_ano[ano]['valor'])
        ])
    
    # Linha de total geral
    dados_consolidacao.append([
        '<b>TOTAL GERAL</b>',
        f'<b>{formatar_numero(total_geral["quantidade"])}</b>',
        f'<b>{formatar_numero(total_geral["peso"], 2)}</b>',
        f'<b>{formatar_moeda(total_geral["valor"])}</b>'
    ])
    
    tabela_consolidacao = Table(dados_consolidacao, colWidths=[3*cm, 4*cm, 5*cm, 6*cm])
    tabela_consolidacao.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), TA_CENTER),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1e3a8a')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(tabela_consolidacao)
    elements.append(Spacer(1, 1*cm))
    
    # Tabela detalhada mensal por ano
    elements.append(Paragraph("DETALHAMENTO MENSAL POR ANO", section_style))
    
    # Cabeçalho da tabela mensal
    meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    for ano in sorted(vendas_por_ano_mes.keys()):
        # Título do ano
        elements.append(Paragraph(
            f"<b>ANO {ano}</b>",
            ParagraphStyle(
                'AnoStyle',
                parent=styles['Heading3'],
                fontSize=13,
                textColor=colors.HexColor('#1e3a8a'),
                spaceAfter=8,
                spaceBefore=10
            )
        ))
        
        # Tabela mensal do ano
        dados_mensal = [['Mês', 'Quantidade', 'Peso Total (kg)', 'Valor Total (R$)']]
        
        meses_ano = sorted(vendas_por_ano_mes[ano].keys())
        for mes in meses_ano:
            if mes <= 10:  # Apenas até outubro
                dados_mensal.append([
                    meses_nomes[mes - 1],
                    formatar_numero(vendas_por_ano_mes[ano][mes]['quantidade']),
                    formatar_numero(vendas_por_ano_mes[ano][mes]['peso'], 2),
                    formatar_moeda(vendas_por_ano_mes[ano][mes]['valor'])
                ])
        
        # Total do ano
        dados_mensal.append([
            '<b>TOTAL</b>',
            f'<b>{formatar_numero(totais_por_ano[ano]["quantidade"])}</b>',
            f'<b>{formatar_numero(totais_por_ano[ano]["peso"], 2)}</b>',
            f'<b>{formatar_moeda(totais_por_ano[ano]["valor"])}</b>'
        ])
        
        tabela_mensal = Table(dados_mensal, colWidths=[3*cm, 4*cm, 5*cm, 6*cm])
        tabela_mensal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), TA_CENTER),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -2), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#c8e6c9')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#2e7d32')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        elements.append(KeepTogether(tabela_mensal))
        elements.append(Spacer(1, 0.8*cm))
    
    # Rodapé
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle(
            'RodapeStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
    ))
    
    # Gerar PDF
    doc.build(elements)
    
    return response


@login_required
def exportar_relatorio_vendas_pdf(request, propriedade_id):
    """Exporta relatório completo de vendas para PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from datetime import datetime
    from decimal import Decimal
    
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento
    codigo_busca = request.GET.get('codigo', '').strip()
    cenario_id = request.GET.get('cenario')
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')
    
    planejamento = None
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.warning(request, 'Nenhum planejamento encontrado.')
        return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    # Buscar vendas com os mesmos filtros da view principal
    vendas_query = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento
    )
    
    if cenario_id:
        vendas_query = vendas_query.filter(cenario_id=cenario_id)
    if ano:
        vendas_query = vendas_query.filter(data_venda__year=int(ano))
    if mes:
        vendas_query = vendas_query.filter(data_venda__month=int(mes))
    
    vendas_lista = list(vendas_query.order_by('-data_venda', 'categoria'))
    
    # Agrupar vendas (mesma lógica da view principal)
    vendas_agrupadas_por_lote = defaultdict(lambda: {
        'vendas': [],
        'data_venda': None,
        'cenario': None,
        'categoria': None,
        'quantidade_total': 0,
        'peso_total': Decimal('0'),
        'peso_medio_kg': None,
        'valor_total': Decimal('0'),
        'clientes': [],
        'valor_por_kg': None,
        'valor_por_animal': None,
        'data_recebimento': None
    })
    
    for venda in vendas_lista:
        cenario_id_v = venda.cenario.id if venda.cenario else None
        categoria_id = venda.categoria.id
        chave_lote = f"{venda.data_venda}_{cenario_id_v}_{categoria_id}"
        
        lote = vendas_agrupadas_por_lote[chave_lote]
        lote['vendas'].append(venda)
        lote['data_venda'] = venda.data_venda
        lote['cenario'] = venda.cenario
        lote['categoria'] = venda.categoria
        if 'primeira_venda_id' not in lote:
            lote['primeira_venda_id'] = venda.id
        lote['quantidade_total'] += venda.quantidade
        lote['peso_total'] += venda.peso_total_kg or Decimal('0')
        lote['valor_total'] += venda.valor_total or Decimal('0')
        
        cliente_nome = venda.cliente_nome or "Cliente não definido"
        if cliente_nome not in [c['nome'] for c in lote['clientes']]:
            lote['clientes'].append({
                'nome': cliente_nome,
                'quantidade': venda.quantidade,
                'valor': venda.valor_total or Decimal('0'),
                'peso': venda.peso_total_kg or Decimal('0'),
                'peso_medio_kg': venda.peso_medio_kg,
                'venda_id': venda.id
            })
        
        if lote['valor_por_kg'] is None and venda.valor_por_kg:
            lote['valor_por_kg'] = venda.valor_por_kg
        if lote['valor_por_animal'] is None and venda.valor_por_animal:
            lote['valor_por_animal'] = venda.valor_por_animal
        if lote['data_recebimento'] is None and venda.data_recebimento:
            lote['data_recebimento'] = venda.data_recebimento
        if lote['peso_medio_kg'] is None and venda.peso_medio_kg:
            lote['peso_medio_kg'] = venda.peso_medio_kg
        elif lote['quantidade_total'] > 0 and lote['peso_total'] > 0:
            lote['peso_medio_kg'] = lote['peso_total'] / lote['quantidade_total']
    
    # Converter para lista ordenada
    vendas_agrupadas_lista = []
    for chave_lote in sorted(vendas_agrupadas_por_lote.keys(), 
                            key=lambda k: (vendas_agrupadas_por_lote[k]['data_venda'], 
                                          vendas_agrupadas_por_lote[k]['categoria'].nome if vendas_agrupadas_por_lote[k]['categoria'] else ''), 
                            reverse=True):
        lote = vendas_agrupadas_por_lote[chave_lote]
        if lote['peso_total'] > 0 and lote['valor_por_kg'] is None:
            lote['valor_por_kg'] = lote['valor_total'] / lote['peso_total']
        if lote['quantidade_total'] > 0 and lote['valor_por_animal'] is None:
            lote['valor_por_animal'] = lote['valor_total'] / lote['quantidade_total']
        if lote['peso_medio_kg'] is None and lote['quantidade_total'] > 0 and lote['peso_total'] > 0:
            lote['peso_medio_kg'] = lote['peso_total'] / lote['quantidade_total']
        vendas_agrupadas_lista.append(lote)
    
    # Função para formatar moeda
    def formatar_moeda(valor):
        if valor is None or valor == 0:
            return 'R$ 0,00'
        try:
            valor_decimal = Decimal(str(valor))
            parte_inteira = int(abs(valor_decimal))
            parte_inteira_formatada = f'{parte_inteira:,}'.replace(',', '.')
            parte_decimal_formatada = f'{abs(valor_decimal) - parte_inteira:.2f}'.split('.')[1]
            valor_formatado = f'{parte_inteira_formatada},{parte_decimal_formatada}'
            if valor_decimal < 0:
                valor_formatado = f'-{valor_formatado}'
            return f'R$ {valor_formatado}'
        except:
            return 'R$ 0,00'
    
    # Função para formatar número
    def formatar_numero(valor, decimais=0):
        if valor is None or valor == 0:
            return '0' if decimais == 0 else '0,00'
        try:
            valor_decimal = Decimal(str(valor))
            if decimais > 0:
                parte_inteira = int(abs(valor_decimal))
                parte_inteira_formatada = f'{parte_inteira:,}'.replace(',', '.')
                parte_decimal = f'{abs(valor_decimal) - parte_inteira:.{decimais}f}'.split('.')[1]
                return f'{parte_inteira_formatada},{parte_decimal}'
            else:
                parte_inteira = int(abs(valor_decimal))
                return f'{parte_inteira:,}'.replace(',', '.')
        except:
            return '0'
    
    # Criar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_vendas_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1*cm,
        rightMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=TA_CENTER,
        spaceAfter=15
    )
    
    elements = []
    
    # Título
    elements.append(Paragraph(
        f"RELATÓRIO DE VENDAS DA PROJEÇÃO<br/>{propriedade.nome_propriedade}",
        title_style
    ))
    
    if planejamento:
        elements.append(Paragraph(
            f"Projeção: {planejamento.codigo} | Ano: {planejamento.ano}",
            ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER, spaceAfter=15)
        ))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Tabela de vendas
    dados_tabela = [[
        'Data', 'Cenário', 'Qtd', 'Categoria', 'Cliente', 
        'Peso Unit. (kg)', 'Peso Total (kg)', 'Vlr/kg (R$)', 
        'Vlr/Animal (R$)', 'Valor Total (R$)', 'Data Receb.'
    ]]
    
    for lote in vendas_agrupadas_lista:
        if len(lote['clientes']) > 1:
            # Lote com múltiplos clientes
            dados_tabela.append([
                str(lote['data_venda'].strftime('%d/%m/%Y')) if lote['data_venda'] else '',
                str(lote['cenario'].nome) if lote['cenario'] else '',
                str(formatar_numero(lote['quantidade_total'])),
                str(lote['categoria'].nome) if lote['categoria'] else '',
                f"{len(lote['clientes'])} cliente(s)",
                str(formatar_numero(lote['peso_medio_kg'], 2)) if lote['peso_medio_kg'] else '-',
                str(formatar_numero(lote['peso_total'], 2)),
                str(formatar_moeda(lote['valor_por_kg'])) if lote['valor_por_kg'] else '-',
                str(formatar_moeda(lote['valor_por_animal'])) if lote['valor_por_animal'] else '-',
                str(formatar_moeda(lote['valor_total'])),
                str(lote['data_recebimento'].strftime('%d/%m/%Y')) if lote['data_recebimento'] else ''
            ])
            # Adicionar linhas dos clientes individuais
            for cliente in lote['clientes']:
                # Buscar a venda correspondente para obter todos os dados
                venda_cliente = None
                valor_por_kg_cliente = None
                valor_por_animal_cliente = None
                data_recebimento_cliente = None
                
                for venda in lote['vendas']:
                    if venda.cliente_nome == cliente['nome']:
                        venda_cliente = venda
                        valor_por_kg_cliente = venda.valor_por_kg
                        valor_por_animal_cliente = venda.valor_por_animal
                        data_recebimento_cliente = venda.data_recebimento
                        break
                
                dados_tabela.append([
                    '', '', 
                    str(formatar_numero(cliente['quantidade'])), 
                    '', 
                    str(cliente['nome']),
                    str(formatar_numero(cliente['peso_medio_kg'], 2)) if cliente.get('peso_medio_kg') else '-',
                    str(formatar_numero(cliente['peso'], 2)),
                    str(formatar_moeda(valor_por_kg_cliente)) if valor_por_kg_cliente else '-',
                    str(formatar_moeda(valor_por_animal_cliente)) if valor_por_animal_cliente else '-', 
                    str(formatar_moeda(cliente['valor'])), 
                    str(data_recebimento_cliente.strftime('%d/%m/%Y')) if data_recebimento_cliente else ''
                ])
        else:
            # Lote com um único cliente
            cliente_nome = lote['clientes'][0]['nome'] if lote['clientes'] else 'Cliente não definido'
            dados_tabela.append([
                str(lote['data_venda'].strftime('%d/%m/%Y')) if lote['data_venda'] else '',
                str(lote['cenario'].nome) if lote['cenario'] else '',
                str(formatar_numero(lote['quantidade_total'])),
                str(lote['categoria'].nome) if lote['categoria'] else '',
                str(cliente_nome),
                str(formatar_numero(lote['peso_medio_kg'], 2)) if lote['peso_medio_kg'] else '-',
                str(formatar_numero(lote['peso_total'], 2)),
                str(formatar_moeda(lote['valor_por_kg'])) if lote['valor_por_kg'] else '-',
                str(formatar_moeda(lote['valor_por_animal'])) if lote['valor_por_animal'] else '-',
                str(formatar_moeda(lote['valor_total'])),
                str(lote['data_recebimento'].strftime('%d/%m/%Y')) if lote['data_recebimento'] else ''
            ])
    
    # Garantir que todos os valores sejam strings válidas (nunca None)
    for linha in dados_tabela:
        for i, valor in enumerate(linha):
            if valor is None:
                linha[i] = ''
            else:
                linha[i] = str(valor)
    
    # Calcular larguras das colunas
    col_widths = [2.5*cm, 2*cm, 1.5*cm, 3.5*cm, 3*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.5*cm, 2*cm]
    
    # Número de linhas (incluindo cabeçalho)
    num_linhas = len(dados_tabela)
    
    if num_linhas == 0:
        # Se não houver dados, adicionar mensagem
        elements.append(Paragraph("Nenhuma venda encontrada para os filtros selecionados.", styles['Normal']))
        doc.build(elements)
        return response
    
    tabela = Table(dados_tabela, colWidths=col_widths, repeatRows=1)
    
    # Criar estilo da tabela - evitar usar -1 para evitar erros
    num_cols = 11  # Número de colunas fixo
    ultima_col = num_cols - 1
    
    # Criar estilo básico - usar strings para alinhamento para evitar erros
    if num_linhas > 1:
        ultima_linha = num_linhas - 1
        
        # Criar lista de estilos - usar strings 'LEFT', 'RIGHT', 'CENTER' em vez de constantes
        estilos = [
            ('BACKGROUND', (0, 0), (ultima_col, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (ultima_col, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (ultima_col, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (ultima_col, 0), 9),
            ('BOTTOMPADDING', (0, 0), (ultima_col, 0), 10),
            ('ALIGN', (0, 0), (ultima_col, 0), 'CENTER'),  # Cabeçalho centralizado
            ('BACKGROUND', (0, 1), (ultima_col, ultima_linha), colors.white),
            ('GRID', (0, 0), (ultima_col, ultima_linha), 0.5, colors.grey),
            ('VALIGN', (0, 0), (ultima_col, ultima_linha), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (ultima_col, ultima_linha), 8),
            ('ALIGN', (0, 1), (4, ultima_linha), 'LEFT'),  # Colunas de texto à esquerda
            ('ALIGN', (5, 1), (9, ultima_linha), 'RIGHT'),  # Colunas numéricas à direita
            ('ALIGN', (10, 1), (10, ultima_linha), 'CENTER'),  # Data recebimento centralizada
        ]
        
        # Adicionar cores alternadas
        for i in range(1, num_linhas):
            if i % 2 == 0:
                estilos.append(('BACKGROUND', (0, i), (ultima_col, i), colors.HexColor('#f5f5f5')))
        
        estilo_tabela = TableStyle(estilos)
    else:
        # Se não houver dados além do cabeçalho
        estilo_tabela = TableStyle([
            ('BACKGROUND', (0, 0), (ultima_col, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (ultima_col, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (ultima_col, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (ultima_col, 0), 9),
            ('BOTTOMPADDING', (0, 0), (ultima_col, 0), 10),
            ('ALIGN', (0, 0), (ultima_col, 0), 'CENTER'),
        ])
    
    tabela.setStyle(estilo_tabela)
    
    elements.append(tabela)
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle('Rodape', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)
    ))
    
    doc.build(elements)
    return response


@login_required
def exportar_relatorio_vendas_excel(request, propriedade_id):
    """Exporta relatório completo de vendas para Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    from decimal import Decimal
    
    propriedade = get_object_or_404(
        Propriedade,
        id=propriedade_id,
        produtor__usuario_responsavel=request.user
    )
    
    # Buscar planejamento
    codigo_busca = request.GET.get('codigo', '').strip()
    cenario_id = request.GET.get('cenario')
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')
    
    planejamento = None
    if codigo_busca:
        try:
            planejamento = PlanejamentoAnual.objects.get(
                codigo=codigo_busca.upper(),
                propriedade=propriedade
            )
        except PlanejamentoAnual.DoesNotExist:
            messages.error(request, f'❌ Nenhuma projeção encontrada com o código: {codigo_busca.upper()}')
            return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    if not planejamento:
        planejamento = PlanejamentoAnual.objects.filter(
            propriedade=propriedade
        ).order_by('-data_criacao', '-ano').first()
    
    if not planejamento:
        messages.warning(request, 'Nenhum planejamento encontrado.')
        return redirect('relatorio_vendas_projecao', propriedade_id=propriedade.id)
    
    # Buscar vendas
    vendas_query = VendaProjetada.objects.filter(
        propriedade=propriedade,
        planejamento=planejamento
    )
    
    if cenario_id:
        vendas_query = vendas_query.filter(cenario_id=cenario_id)
    if ano:
        vendas_query = vendas_query.filter(data_venda__year=int(ano))
    if mes:
        vendas_query = vendas_query.filter(data_venda__month=int(mes))
    
    vendas_lista = list(vendas_query.order_by('-data_venda', 'categoria'))
    
    # Agrupar vendas (mesma lógica)
    vendas_agrupadas_por_lote = defaultdict(lambda: {
        'vendas': [],
        'data_venda': None,
        'cenario': None,
        'categoria': None,
        'quantidade_total': 0,
        'peso_total': Decimal('0'),
        'peso_medio_kg': None,
        'valor_total': Decimal('0'),
        'clientes': [],
        'valor_por_kg': None,
        'valor_por_animal': None,
        'data_recebimento': None
    })
    
    for venda in vendas_lista:
        cenario_id_v = venda.cenario.id if venda.cenario else None
        categoria_id = venda.categoria.id
        chave_lote = f"{venda.data_venda}_{cenario_id_v}_{categoria_id}"
        
        lote = vendas_agrupadas_por_lote[chave_lote]
        lote['vendas'].append(venda)
        lote['data_venda'] = venda.data_venda
        lote['cenario'] = venda.cenario
        lote['categoria'] = venda.categoria
        lote['quantidade_total'] += venda.quantidade
        lote['peso_total'] += venda.peso_total_kg or Decimal('0')
        lote['valor_total'] += venda.valor_total or Decimal('0')
        
        cliente_nome = venda.cliente_nome or "Cliente não definido"
        if cliente_nome not in [c['nome'] for c in lote['clientes']]:
            lote['clientes'].append({
                'nome': cliente_nome,
                'quantidade': venda.quantidade,
                'valor': venda.valor_total or Decimal('0'),
                'peso': venda.peso_total_kg or Decimal('0'),
                'peso_medio_kg': venda.peso_medio_kg,
                'venda_id': venda.id
            })
        
        if lote['valor_por_kg'] is None and venda.valor_por_kg:
            lote['valor_por_kg'] = venda.valor_por_kg
        if lote['valor_por_animal'] is None and venda.valor_por_animal:
            lote['valor_por_animal'] = venda.valor_por_animal
        if lote['data_recebimento'] is None and venda.data_recebimento:
            lote['data_recebimento'] = venda.data_recebimento
        if lote['peso_medio_kg'] is None and venda.peso_medio_kg:
            lote['peso_medio_kg'] = venda.peso_medio_kg
        elif lote['quantidade_total'] > 0 and lote['peso_total'] > 0:
            lote['peso_medio_kg'] = lote['peso_total'] / lote['quantidade_total']
    
    # Converter para lista ordenada
    vendas_agrupadas_lista = []
    for chave_lote in sorted(vendas_agrupadas_por_lote.keys(), 
                            key=lambda k: (vendas_agrupadas_por_lote[k]['data_venda'], 
                                          vendas_agrupadas_por_lote[k]['categoria'].nome if vendas_agrupadas_por_lote[k]['categoria'] else ''), 
                            reverse=True):
        lote = vendas_agrupadas_por_lote[chave_lote]
        if lote['peso_total'] > 0 and lote['valor_por_kg'] is None:
            lote['valor_por_kg'] = lote['valor_total'] / lote['peso_total']
        if lote['quantidade_total'] > 0 and lote['valor_por_animal'] is None:
            lote['valor_por_animal'] = lote['valor_total'] / lote['quantidade_total']
        if lote['peso_medio_kg'] is None and lote['quantidade_total'] > 0 and lote['peso_total'] > 0:
            lote['peso_medio_kg'] = lote['peso_total'] / lote['quantidade_total']
        vendas_agrupadas_lista.append(lote)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas Projetadas"
    
    # Estilos
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho
    ws['A1'] = 'Data'
    ws['B1'] = 'Cenário'
    ws['C1'] = 'Quantidade'
    ws['D1'] = 'Categoria'
    ws['E1'] = 'Cliente'
    ws['F1'] = 'Peso Unit. (kg)'
    ws['G1'] = 'Peso Total (kg)'
    ws['H1'] = 'Vlr/kg (R$)'
    ws['I1'] = 'Vlr/Animal (R$)'
    ws['J1'] = 'Valor Total (R$)'
    ws['K1'] = 'Data Recebimento'
    
    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Preencher dados
    row = 2
    for lote in vendas_agrupadas_lista:
        if len(lote['clientes']) > 1:
            # Lote com múltiplos clientes
            ws[f'A{row}'] = lote['data_venda'].strftime('%d/%m/%Y') if lote['data_venda'] else ''
            ws[f'B{row}'] = lote['cenario'].nome if lote['cenario'] else ''
            ws[f'C{row}'] = lote['quantidade_total']
            ws[f'D{row}'] = lote['categoria'].nome if lote['categoria'] else ''
            ws[f'E{row}'] = f"{len(lote['clientes'])} cliente(s)"
            ws[f'F{row}'] = float(lote['peso_medio_kg']) if lote['peso_medio_kg'] else None
            ws[f'G{row}'] = float(lote['peso_total'])
            ws[f'H{row}'] = float(lote['valor_por_kg']) if lote['valor_por_kg'] else None
            ws[f'I{row}'] = float(lote['valor_por_animal']) if lote['valor_por_animal'] else None
            ws[f'J{row}'] = float(lote['valor_total'])
            ws[f'K{row}'] = lote['data_recebimento'].strftime('%d/%m/%Y') if lote['data_recebimento'] else ''
            
            # Formatação
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{col}{row}'].border = border
            
            row += 1
            
            # Adicionar linhas dos clientes individuais
            for cliente in lote['clientes']:
                ws[f'C{row}'] = cliente['quantidade']
                ws[f'E{row}'] = cliente['nome']
                ws[f'F{row}'] = float(cliente['peso_medio_kg']) if cliente.get('peso_medio_kg') else None
                ws[f'G{row}'] = float(cliente['peso'])
                ws[f'J{row}'] = float(cliente['valor'])
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    ws[f'{col}{row}'].border = border
                row += 1
        else:
            # Lote com um único cliente
            cliente_nome = lote['clientes'][0]['nome'] if lote['clientes'] else 'Cliente não definido'
            ws[f'A{row}'] = lote['data_venda'].strftime('%d/%m/%Y') if lote['data_venda'] else ''
            ws[f'B{row}'] = lote['cenario'].nome if lote['cenario'] else ''
            ws[f'C{row}'] = lote['quantidade_total']
            ws[f'D{row}'] = lote['categoria'].nome if lote['categoria'] else ''
            ws[f'E{row}'] = cliente_nome
            ws[f'F{row}'] = float(lote['peso_medio_kg']) if lote['peso_medio_kg'] else None
            ws[f'G{row}'] = float(lote['peso_total'])
            ws[f'H{row}'] = float(lote['valor_por_kg']) if lote['valor_por_kg'] else None
            ws[f'I{row}'] = float(lote['valor_por_animal']) if lote['valor_por_animal'] else None
            ws[f'J{row}'] = float(lote['valor_total'])
            ws[f'K{row}'] = lote['data_recebimento'].strftime('%d/%m/%Y') if lote['data_recebimento'] else ''
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{col}{row}'].border = border
            row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    
    # Formatação de números
    for row_num in range(2, row):
        ws[f'F{row_num}'].number_format = '#,##0.00'
        ws[f'G{row_num}'].number_format = '#,##0.00'
        ws[f'H{row_num}'].number_format = 'R$ #,##0.00'
        ws[f'I{row_num}'].number_format = 'R$ #,##0.00'
        ws[f'J{row_num}'].number_format = 'R$ #,##0.00'
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_vendas_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response
