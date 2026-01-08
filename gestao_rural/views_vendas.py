# -*- coding: utf-8 -*-
"""
Views do Módulo de VENDAS
Inclui: Dashboard, Notas Fiscais de Saída, Emissão de NF-e, Relatórios
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.forms import inlineformset_factory
from django.http import JsonResponse, HttpResponse, Http404
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from decimal import Decimal
from datetime import date, datetime, timedelta
import logging
import os

logger = logging.getLogger(__name__)

from .models import Propriedade
from .models_cadastros import Cliente
from .decorators import obter_propriedade_com_permissao
from .models_compras_financeiro import NotaFiscal, ItemNotaFiscal, NumeroSequencialNFE
from .forms_vendas import VendaForm, ItemVendaForm, ConfigurarSerieNFeForm, ParametrosVendaPorCategoriaForm, BulkVendaPorCategoriaForm


# ============================================================================
# DASHBOARD DE VENDAS
# ============================================================================

@login_required
def vendas_dashboard(request, propriedade_id):
    """Dashboard do módulo de vendas"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)

    # Estatísticas gerais
    hoje = date.today()
    mes_atual = date(hoje.year, hoje.month, 1)

    try:
        notas_vendas = NotaFiscal.objects.filter(
            propriedade=propriedade,
            tipo='SAIDA'
        )
    except Exception as e:
        print(f'Erro ao buscar notas fiscais de vendas: {e}')
        notas_vendas = NotaFiscal.objects.none()

    # Valores com tratamento de erro
    try:
        valor_total_vendas = notas_vendas.filter(status='AUTORIZADA').aggregate(
            total=Sum('valor_total')
        )['total'] or Decimal('0.00')
    except Exception as e:
        print(f'Erro ao calcular valor total vendas: {e}')
        valor_total_vendas = Decimal('0.00')

    try:
        valor_vendas_mes = notas_vendas.filter(
            status='AUTORIZADA',
            data_emissao__gte=mes_atual
        ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0.00')
    except Exception as e:
        print(f'Erro ao calcular valor vendas mês: {e}')
        valor_vendas_mes = Decimal('0.00')

    # Contagens com tratamento de erro
    try:
        total_notas = notas_vendas.count()
        notas_autorizadas = notas_vendas.filter(status='AUTORIZADA').count()
        notas_pendentes = notas_vendas.filter(status='PENDENTE').count()
        notas_canceladas = notas_vendas.filter(status='CANCELADA').count()
    except Exception as e:
        print(f'Erro ao contar notas: {e}')
        total_notas = notas_autorizadas = notas_pendentes = notas_canceladas = 0

    # Notas recentes com tratamento de erro
    try:
        notas_recentes = notas_vendas.select_related('cliente').order_by('-data_emissao')[:10]
    except Exception as e:
        print(f'Erro ao buscar notas recentes: {e}')
        notas_recentes = []
    
    context = {
        'propriedade': propriedade,
        'valor_total_vendas': valor_total_vendas,
        'valor_vendas_mes': valor_vendas_mes,
        'total_notas': total_notas,
        'notas_autorizadas': notas_autorizadas,
        'notas_pendentes': notas_pendentes,
        'notas_canceladas': notas_canceladas,
        'notas_recentes': notas_recentes,
    }
    
    return render(request, 'gestao_rural/vendas_dashboard.html', context)


# ============================================================================
# LISTAGEM DE NOTAS FISCAIS DE VENDAS
# ============================================================================

@login_required
def vendas_notas_fiscais_lista(request, propriedade_id):
    """Lista de notas fiscais de SAÍDA (vendas)"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    # Filtros
    status_filter = request.GET.get('status', '')
    busca = request.GET.get('busca', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Query base
    notas = NotaFiscal.objects.filter(
        propriedade=propriedade,
        tipo='SAIDA'
    ).select_related('cliente').order_by('-data_emissao', '-numero')
    
    if status_filter:
        notas = notas.filter(status=status_filter)
    if busca:
        notas = notas.filter(
            Q(numero__icontains=busca) |
            Q(chave_acesso__icontains=busca) |
            Q(cliente__nome__icontains=busca) |
            Q(cliente__cpf_cnpj__icontains=busca)
        )
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__gte=data_inicio_obj)
        except ValueError:
            pass
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__lte=data_fim_obj)
        except ValueError:
            pass
    
    context = {
        'propriedade': propriedade,
        'notas': notas,
        'status_filter': status_filter,
        'busca': busca,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }
    
    return render(request, 'gestao_rural/vendas_notas_fiscais_lista.html', context)


# ============================================================================
# EMISSÃO DE NOTA FISCAL
# ============================================================================

@login_required
def vendas_nota_fiscal_emitir(request, propriedade_id):
    """Emitir nota fiscal de saída (venda)"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    from .forms_completos import NotaFiscalSaidaForm, ItemNotaFiscalForm
    
    # Criar formset para itens
    ItemNotaFiscalFormSet = inlineformset_factory(
        NotaFiscal,
        ItemNotaFiscal,
        form=ItemNotaFiscalForm,
        extra=1,
        can_delete=True,
        min_num=1,
        validate_min=True
    )
    
    serie_padrao = '1'
    
    if request.method == 'POST':
        form = NotaFiscalSaidaForm(request.POST, propriedade=propriedade)
        formset = ItemNotaFiscalFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            # Obter próximo número E INCREMENTAR (apenas ao salvar)
            from .services_nfe_utils import obter_proximo_numero_nfe
            proximo_numero = obter_proximo_numero_nfe(propriedade, serie_padrao)
            
            nota = form.save(commit=False)
            nota.propriedade = propriedade
            nota.tipo = 'SAIDA'
            nota.numero = str(proximo_numero)
            nota.serie = serie_padrao
            nota.status = 'PENDENTE'
            nota.save()
            
            # Salvar itens
            formset.instance = nota
            itens = formset.save(commit=False)
            
            # Calcular valor_total de cada item antes de salvar
            for item in itens:
                if not item.valor_total or item.valor_total == 0:
                    item.valor_total = item.quantidade * item.valor_unitario
                item.save()
            
            # Recalcular valores totais da nota fiscal
            valor_total_itens = sum(item.valor_total for item in nota.itens.all())
            nota.valor_produtos = valor_total_itens
            nota.valor_total = valor_total_itens
            nota.save()
            
            # Emitir NF-e
            try:
                from .services_nfe import emitir_nfe
                from .security_avancado import registrar_log_auditoria, obter_ip_address
                from .models_auditoria import LogAuditoria
                
                resultado = emitir_nfe(nota)
                
                if resultado and resultado.get('sucesso'):
                    nota.status = 'AUTORIZADA'
                    nota.chave_acesso = resultado.get('chave_acesso', '')
                    nota.protocolo_autorizacao = resultado.get('protocolo', '')
                    nota.data_autorizacao = timezone.now()
                    if resultado.get('xml'):
                        from django.core.files.base import ContentFile
                        nota.arquivo_xml.save(f'nfe_{nota.numero}.xml', ContentFile(resultado['xml']), save=False)
                    nota.save()
                    
                    # Registrar log de auditoria
                    registrar_log_auditoria(
                        tipo_acao=LogAuditoria.TipoAcao.EMITIR_NFE,
                        descricao=f'NF-e {nota.numero}/{nota.serie} emitida com sucesso. Chave: {nota.chave_acesso}',
                        usuario=request.user,
                        ip_address=obter_ip_address(request),
                        user_agent=request.META.get('HTTP_USER_AGENT', ''),
                        nivel_severidade=LogAuditoria.NivelSeveridade.ALTO,
                        sucesso=True,
                        metadata={
                            'nota_fiscal_id': nota.id,
                            'numero': nota.numero,
                            'serie': nota.serie,
                            'chave_acesso': nota.chave_acesso,
                            'protocolo': nota.protocolo_autorizacao,
                            'propriedade_id': propriedade.id,
                            'cliente_id': nota.cliente.id if nota.cliente else None,
                            'valor_total': str(nota.valor_total),
                        }
                    )
                    
                    # Criar lançamento financeiro se forma de recebimento foi informada
                    forma_recebimento = form.cleaned_data.get('forma_recebimento')
                    if forma_recebimento:
                        try:
                            from .models_financeiro import LancamentoFinanceiro, CategoriaFinanceira
                            from django.db import transaction
                            
                            conta_destino = form.cleaned_data.get('conta_destino')
                            categoria_receita = form.cleaned_data.get('categoria_receita')
                            data_vencimento_recebimento = form.cleaned_data.get('data_vencimento_recebimento')
                            
                            # Se não informado, buscar padrão
                            if not conta_destino:
                                from .models_financeiro import ContaFinanceira
                                conta_destino = ContaFinanceira.objects.filter(
                                    propriedade=propriedade,
                                    ativa=True
                                ).first()
                            
                            if not categoria_receita:
                                categoria_receita = CategoriaFinanceira.objects.filter(
                                    propriedade=propriedade,
                                    tipo=CategoriaFinanceira.TIPO_RECEITA,
                                    ativa=True
                                ).first()
                                if not categoria_receita:
                                    categoria_receita = CategoriaFinanceira.objects.create(
                                        propriedade=propriedade,
                                        nome='Vendas de Animais',
                                        tipo=CategoriaFinanceira.TIPO_RECEITA,
                                        descricao='Receitas provenientes de vendas',
                                        ativa=True
                                    )
                            
                            # Se tem conta e categoria, criar lançamento
                            if conta_destino and categoria_receita:
                                cliente_nome = nota.cliente.nome if nota.cliente else 'Cliente não informado'
                                data_vencimento = data_vencimento_recebimento or nota.data_emissao
                                
                                # Para PIX/Dinheiro, marcar como quitado na data de emissão
                                # Para Boleto/Transferência, deixar pendente
                                formas_quitadas_imediato = ['PIX', 'DINHEIRO']
                                status_lancamento = LancamentoFinanceiro.STATUS_QUITADO if forma_recebimento in formas_quitadas_imediato else LancamentoFinanceiro.STATUS_PENDENTE
                                data_quitacao = nota.data_emissao if forma_recebimento in formas_quitadas_imediato else None
                                
                                with transaction.atomic():
                                    lancamento = LancamentoFinanceiro.objects.create(
                                        propriedade=propriedade,
                                        categoria=categoria_receita,
                                        conta_destino=conta_destino,
                                        tipo=CategoriaFinanceira.TIPO_RECEITA,
                                        descricao=f'NF-e {nota.numero}/{nota.serie} - {cliente_nome}',
                                        valor=nota.valor_total,
                                        data_competencia=nota.data_emissao,
                                        data_vencimento=data_vencimento,
                                        data_quitacao=data_quitacao,
                                        forma_pagamento=forma_recebimento,
                                        status=status_lancamento,
                                        documento_referencia=f'NF-e {nota.numero}',
                                        observacoes=f'Gerado automaticamente da NF-e {nota.numero}. Cliente: {cliente_nome}'
                                    )
                                    logger.info(f'Lançamento financeiro criado: ID={lancamento.id} para NF-e {nota.numero}')
                        except Exception as e:
                            logger.error(f'Erro ao criar lançamento financeiro para NF-e {nota.numero}: {str(e)}', exc_info=True)
                            # Não falhar a emissão da NF-e se houver erro ao criar lançamento
                    
                    # Se for requisição AJAX, retornar JSON
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'sucesso': True,
                            'nota_id': nota.id,
                            'chave_acesso': nota.chave_acesso or '',
                            'numero': nota.numero,
                            'mensagem': f'NF-e {nota.numero} emitida e autorizada com sucesso!'
                        })
                    
                    messages.success(request, f'NF-e {nota.numero} emitida e autorizada com sucesso!')
                else:
                    erro_msg = resultado.get("erro", "Erro desconhecido") if resultado else "Resposta inválida da API"
                    
                    # Se a API não está configurada, criar NF-e com status PENDENTE
                    if 'API de NF-e não configurada' in erro_msg:
                        nota.status = 'PENDENTE'
                        nota.save()
                        
                        # Criar lançamento financeiro mesmo com status PENDENTE, se forma_recebimento foi informada
                        forma_recebimento = form.cleaned_data.get('forma_recebimento')
                        if forma_recebimento:
                            try:
                                from .models_financeiro import LancamentoFinanceiro, CategoriaFinanceira
                                from django.db import transaction
                                
                                conta_destino = form.cleaned_data.get('conta_destino')
                                categoria_receita = form.cleaned_data.get('categoria_receita')
                                data_vencimento_recebimento = form.cleaned_data.get('data_vencimento_recebimento')
                                
                                # Se não informado, buscar padrão
                                if not conta_destino:
                                    from .models_financeiro import ContaFinanceira
                                    conta_destino = ContaFinanceira.objects.filter(
                                        propriedade=propriedade,
                                        ativa=True
                                    ).first()
                                
                                if not categoria_receita:
                                    categoria_receita = CategoriaFinanceira.objects.filter(
                                        propriedade=propriedade,
                                        tipo=CategoriaFinanceira.TIPO_RECEITA,
                                        ativa=True
                                    ).first()
                                    if not categoria_receita:
                                        categoria_receita = CategoriaFinanceira.objects.create(
                                            propriedade=propriedade,
                                            nome='Vendas de Animais',
                                            tipo=CategoriaFinanceira.TIPO_RECEITA,
                                            descricao='Receitas provenientes de vendas',
                                            ativa=True
                                        )
                                
                                # Se tem conta e categoria, criar lançamento
                                if conta_destino and categoria_receita:
                                    cliente_nome = nota.cliente.nome if nota.cliente else 'Cliente não informado'
                                    data_vencimento = data_vencimento_recebimento or nota.data_emissao
                                    
                                    # Para PIX/Dinheiro, marcar como quitado na data de emissão
                                    # Para Boleto/Transferência, deixar pendente
                                    formas_quitadas_imediato = ['PIX', 'DINHEIRO']
                                    status_lancamento = LancamentoFinanceiro.STATUS_QUITADO if forma_recebimento in formas_quitadas_imediato else LancamentoFinanceiro.STATUS_PENDENTE
                                    data_quitacao = nota.data_emissao if forma_recebimento in formas_quitadas_imediato else None
                                    
                                    with transaction.atomic():
                                        lancamento = LancamentoFinanceiro.objects.create(
                                            propriedade=propriedade,
                                            categoria=categoria_receita,
                                            conta_destino=conta_destino,
                                            tipo=CategoriaFinanceira.TIPO_RECEITA,
                                            descricao=f'NF-e {nota.numero}/{nota.serie} - {cliente_nome}',
                                            valor=nota.valor_total,
                                            data_competencia=nota.data_emissao,
                                            data_vencimento=data_vencimento,
                                            data_quitacao=data_quitacao,
                                            forma_pagamento=forma_recebimento,
                                            status=status_lancamento,
                                            documento_referencia=f'NF-e {nota.numero}',
                                            observacoes=f'Gerado automaticamente da NF-e {nota.numero}. Cliente: {cliente_nome}'
                                        )
                                        logger.info(f'Lançamento financeiro criado: ID={lancamento.id} para NF-e {nota.numero} (PENDENTE)')
                            except Exception as e:
                                logger.error(f'Erro ao criar lançamento financeiro para NF-e {nota.numero}: {str(e)}', exc_info=True)
                        
                        # Se for requisição AJAX, retornar JSON com sucesso mas status PENDENTE
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({
                                'sucesso': True,
                                'nota_id': nota.id,
                                'chave_acesso': '',
                                'numero': nota.numero,
                                'mensagem': f'NF-e {nota.numero} criada com status PENDENTE. Configure a API de NF-e para emissão automática.',
                                'status': 'PENDENTE'
                            })
                        
                        messages.warning(request, f'NF-e {nota.numero} criada com status PENDENTE. Configure a API de NF-e para emissão automática.')
                    else:
                        # Outros erros: retornar erro
                        # Se for requisição AJAX, retornar JSON
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({
                                'sucesso': False,
                                'erro': erro_msg
                            }, status=400)
                        
                        messages.error(request, f'Erro ao emitir NF-e: {erro_msg}')
            except Exception as e:
                logger.error(f'Erro ao emitir NF-e: {str(e)}', exc_info=True)
                erro_msg = str(e)
                
                # Se for requisição AJAX, retornar JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'sucesso': False,
                        'erro': erro_msg
                    }, status=500)
                
                messages.error(request, f'Erro ao emitir NF-e: {erro_msg}')
            
            # Se for requisição AJAX e chegou aqui, retornar JSON de erro
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Erro desconhecido ao processar a NF-e'
                }, status=500)
            
            return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
        else:
            # Se for requisição AJAX e formulário inválido, retornar JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {}
                if not form.is_valid():
                    errors.update(form.errors)
                if not formset.is_valid():
                    errors['formset'] = formset.errors
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Por favor, corrija os erros no formulário.',
                    'errors': errors
                }, status=400)
            
            messages.error(request, 'Por favor, corrija os erros no formulário.')
            # Em caso de erro, obter próximo número SEM INCREMENTAR (para mostrar no formulário)
            from .services_nfe_utils import visualizar_proximo_numero_nfe
            proximo_numero = visualizar_proximo_numero_nfe(propriedade, serie_padrao)
    else:
        form = NotaFiscalSaidaForm(propriedade=propriedade)
        form.initial['data_emissao'] = date.today()
        form.initial['data_entrada'] = date.today()
        formset = ItemNotaFiscalFormSet(instance=NotaFiscal())
        
        # Obter próximo número SEM INCREMENTAR (apenas para visualização)
        from .services_nfe_utils import visualizar_proximo_numero_nfe
        proximo_numero = visualizar_proximo_numero_nfe(propriedade, serie_padrao)
    
    clientes = Cliente.objects.filter(
        Q(propriedade=propriedade) | Q(propriedade__isnull=True),
        ativo=True
    ).order_by('nome')
    
    context = {
        'propriedade': propriedade,
        'form': form,
        'formset': formset,
        'clientes': clientes,
        'proximo_numero': proximo_numero,
        'serie_padrao': serie_padrao,
    }
    
    return render(request, 'gestao_rural/vendas_nota_fiscal_emitir.html', context)


# ============================================================================
# DETALHES DA NOTA FISCAL
# ============================================================================

@login_required
def vendas_nota_fiscal_detalhes(request, propriedade_id, nota_id):
    """Detalhes de uma nota fiscal de saída"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    
    itens = nota.itens.all().select_related('produto')
    
    context = {
        'propriedade': propriedade,
        'nota': nota,
        'itens': itens,
    }
    
    return render(request, 'gestao_rural/vendas_nota_fiscal_detalhes.html', context)


# ============================================================================
# CANCELAR NOTA FISCAL
# ============================================================================

@login_required
def vendas_nota_fiscal_cancelar(request, propriedade_id, nota_id):
    """Cancelar uma nota fiscal de saída"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    
    if nota.status != 'AUTORIZADA':
        messages.error(request, 'Apenas NF-e autorizadas podem ser canceladas.')
        return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    if request.method == 'POST':
        justificativa = request.POST.get('justificativa', '').strip()
        
        if len(justificativa) < 15:
            messages.error(request, 'A justificativa deve ter no mínimo 15 caracteres.')
            return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
        
        try:
            from .services_nfe import cancelar_nfe
            from requests.exceptions import RequestException, Timeout, ConnectionError
            from .security_avancado import registrar_log_auditoria, obter_ip_address
            from .models_auditoria import LogAuditoria
            
            # Validar se nota tem chave de acesso
            if not nota.chave_acesso:
                messages.error(request, 'NF-e não possui chave de acesso. Não é possível cancelar.')
                return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
            
            resultado = cancelar_nfe(nota, justificativa)
            
            if resultado and resultado.get('sucesso'):
                nota.status = 'CANCELADA'
                nota.data_cancelamento = timezone.now()
                nota.justificativa_cancelamento = justificativa
                nota.save()
                messages.success(request, 'NF-e cancelada com sucesso!')
                
                # Registrar log de auditoria
                registrar_log_auditoria(
                    tipo_acao=LogAuditoria.TipoAcao.CANCELAR_NFE,
                    descricao=f'NF-e {nota.numero}/{nota.serie} cancelada. Chave: {nota.chave_acesso}',
                    usuario=request.user,
                    ip_address=obter_ip_address(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    nivel_severidade=LogAuditoria.NivelSeveridade.CRITICO,
                    sucesso=True,
                    metadata={
                        'nota_fiscal_id': nota.id,
                        'numero': nota.numero,
                        'serie': nota.serie,
                        'chave_acesso': nota.chave_acesso,
                        'justificativa': justificativa[:100],  # Limitar tamanho
                        'propriedade_id': propriedade.id,
                    }
                )
            else:
                erro_msg = resultado.get("erro", "Erro desconhecido") if resultado else "Resposta inválida da API"
                messages.error(request, f'Erro ao cancelar NF-e: {erro_msg}')
                
                # Registrar log de auditoria para falha
                registrar_log_auditoria(
                    tipo_acao=LogAuditoria.TipoAcao.CANCELAR_NFE,
                    descricao=f'Falha ao cancelar NF-e {nota.numero}/{nota.serie}',
                    usuario=request.user,
                    ip_address=obter_ip_address(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    nivel_severidade=LogAuditoria.NivelSeveridade.ALTO,
                    sucesso=False,
                    erro=erro_msg,
                    metadata={
                        'nota_fiscal_id': nota.id,
                        'numero': nota.numero,
                        'serie': nota.serie,
                        'chave_acesso': nota.chave_acesso,
                        'propriedade_id': propriedade.id,
                    }
                )
                
        except (Timeout, ConnectionError) as e:
            logger.error(f'Erro de conexão ao cancelar NF-e: {str(e)}', exc_info=True)
            messages.error(request, 'Erro de conexão com a API. Verifique sua conexão com a internet e tente novamente.')
        except RequestException as e:
            logger.error(f'Erro na requisição ao cancelar NF-e: {str(e)}', exc_info=True)
            messages.error(request, f'Erro na comunicação com a API de NF-e: {str(e)}')
        except ValueError as e:
            logger.error(f'Erro de validação ao cancelar NF-e: {str(e)}', exc_info=True)
            messages.error(request, f'Erro de validação: {str(e)}')
        except Exception as e:
            logger.error(f'Erro inesperado ao cancelar NF-e: {str(e)}', exc_info=True)
            messages.error(request, f'Erro inesperado ao cancelar NF-e. Entre em contato com o suporte se o problema persistir.')
        
        return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    context = {
        'propriedade': propriedade,
        'nota': nota,
    }
    
    return render(request, 'gestao_rural/vendas_nota_fiscal_detalhes.html', context)


# ============================================================================
# CONSULTAR STATUS DA NOTA FISCAL
# ============================================================================

@login_required
def vendas_nota_fiscal_consultar_status(request, propriedade_id, nota_id):
    """Consultar status de uma nota fiscal na SEFAZ"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    
    if not nota.chave_acesso:
        return JsonResponse({'sucesso': False, 'erro': 'NF-e não possui chave de acesso'}, status=400)
    
    try:
        from .services_nfe import consultar_status_nfe
        from .security_avancado import registrar_log_auditoria, obter_ip_address
        from .models_auditoria import LogAuditoria
        
        resultado = consultar_status_nfe(nota)
        
        if resultado and resultado.get('sucesso'):
            # Atualizar status se necessário
            novo_status = resultado.get('status')
            if novo_status and novo_status != nota.status:
                nota.status = novo_status
                nota.save(update_fields=['status'])
            
            # Registrar log de auditoria
            registrar_log_auditoria(
                tipo_acao=LogAuditoria.TipoAcao.CONSULTAR_STATUS_NFE,
                descricao=f'Consulta de status da NF-e {nota.numero}/{nota.serie}. Status: {nota.status}',
                usuario=request.user,
                ip_address=obter_ip_address(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                nivel_severidade=LogAuditoria.NivelSeveridade.MEDIO,
                sucesso=True,
                metadata={
                    'nota_fiscal_id': nota.id,
                    'numero': nota.numero,
                    'serie': nota.serie,
                    'chave_acesso': nota.chave_acesso,
                    'status': nota.status,
                    'propriedade_id': propriedade.id,
                }
            )
            
            return JsonResponse({
                'sucesso': True,
                'status': nota.status,
                'status_display': nota.get_status_display(),
                'mensagem': resultado.get('mensagem', 'Status consultado com sucesso')
            })
        else:
            erro_msg = resultado.get("erro", "Erro desconhecido") if resultado else "Resposta inválida da API"
            return JsonResponse({'sucesso': False, 'erro': erro_msg}, status=400)
    except Exception as e:
        logger.error(f'Erro ao consultar status da NF-e: {str(e)}', exc_info=True)
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


# ============================================================================
# ATUALIZAR STATUS PARA AUTORIZADA (TEMPORÁRIO PARA TESTES)
# ============================================================================

@login_required
def vendas_nota_fiscal_autorizar_teste(request, propriedade_id, nota_id):
    """
    Atualizar status de NF-e para AUTORIZADA (temporário para testes)
    
    ⚠️ ATENÇÃO: Esta função só está disponível em ambiente de desenvolvimento (DEBUG=True).
    Em produção, esta função é bloqueada para evitar problemas fiscais.
    """
    from django.conf import settings
    from django.core.exceptions import PermissionDenied
    
    # Bloquear em produção
    if not settings.DEBUG:
        raise PermissionDenied(
            "Esta função só está disponível em ambiente de desenvolvimento. "
            "Em produção, use a emissão real através da SEFAZ ou APIs autorizadas."
        )
    
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    
    if nota.status == 'AUTORIZADA':
        messages.info(request, 'NF-e já está autorizada.')
        return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    try:
        nota.status = 'AUTORIZADA'
        nota.protocolo_autorizacao = '123456789012345'
        nota.data_autorizacao = timezone.now()
        if not nota.chave_acesso:
            # Gerar chave de acesso única baseada no número da nota e série
            import random
            import time
            chave_base = f'3520011234567800019055001000000000{nota.numero:08d}{nota.serie:02d}'
            # Adicionar dígitos aleatórios e timestamp para garantir unicidade
            max_tentativas = 100
            tentativa = 0
            chave_gerada = False
            
            while not chave_gerada and tentativa < max_tentativas:
                # Usar timestamp e random para garantir unicidade
                timestamp_suffix = str(int(time.time() * 1000))[-6:]  # Últimos 6 dígitos do timestamp
                random_suffix = f'{random.randint(1000, 9999):04d}'
                chave_sufixo = (timestamp_suffix + random_suffix)[:4]
                chave_candidata = chave_base[:40] + chave_sufixo[:4]
                
                # Verificar se a chave já existe
                if not NotaFiscal.objects.filter(chave_acesso=chave_candidata).exclude(id=nota.id).exists():
                    nota.chave_acesso = chave_candidata
                    chave_gerada = True
                else:
                    tentativa += 1
                    time.sleep(0.001)  # Pequeno delay para evitar colisões
            
            if not chave_gerada:
                # Se não conseguiu gerar uma chave única, usar UUID como fallback
                import uuid
                chave_fallback = f'3520011234567800019055001000000000{nota.numero:08d}{nota.serie:02d}{str(uuid.uuid4())[:4].replace("-", "")}'
                nota.chave_acesso = chave_fallback[:44]
        
        # Tentar salvar com tratamento de IntegrityError
        try:
            nota.save()
        except Exception as ie:
            from django.db import IntegrityError
            # Se ainda houver erro de unicidade, tentar novamente com nova chave
            if isinstance(ie, IntegrityError) and 'chave_acesso' in str(ie).lower():
                import uuid
                chave_fallback = f'3520011234567800019055001000000000{nota.numero:08d}{nota.serie:02d}{str(uuid.uuid4())[:4].replace("-", "")}'
                nota.chave_acesso = chave_fallback[:44]
                nota.save()
            else:
                raise
        
        messages.success(request, f'✅ NF-e {nota.numero}/{nota.serie} atualizada para AUTORIZADA com sucesso!')
        logger.info(f'NF-e {nota.id} atualizada para AUTORIZADA manualmente por {request.user.username}')
    except Exception as e:
        logger.error(f'Erro ao atualizar status da NF-e: {str(e)}', exc_info=True)
        messages.error(request, f'Erro ao atualizar status: {str(e)}')
    
    return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)


@login_required
def vendas_nota_fiscal_autorizar_teste_numero(request, propriedade_id, numero, serie='1'):
    """Atualizar status de NF-e para AUTORIZADA por número (temporário para testes)"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    nota = get_object_or_404(NotaFiscal, numero=numero, serie=serie, propriedade=propriedade, tipo='SAIDA')
    
    if nota.status == 'AUTORIZADA':
        messages.info(request, 'NF-e já está autorizada.')
        return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    try:
        nota.status = 'AUTORIZADA'
        nota.protocolo_autorizacao = '123456789012345'
        nota.data_autorizacao = timezone.now()
        if not nota.chave_acesso:
            # Gerar chave de acesso única baseada no número da nota e série
            import random
            import time
            chave_base = f'3520011234567800019055001000000000{nota.numero:08d}{nota.serie:02d}'
            # Adicionar dígitos aleatórios e timestamp para garantir unicidade
            max_tentativas = 100
            tentativa = 0
            chave_gerada = False
            
            while not chave_gerada and tentativa < max_tentativas:
                # Usar timestamp e random para garantir unicidade
                timestamp_suffix = str(int(time.time() * 1000))[-6:]  # Últimos 6 dígitos do timestamp
                random_suffix = f'{random.randint(1000, 9999):04d}'
                chave_sufixo = (timestamp_suffix + random_suffix)[:4]
                chave_candidata = chave_base[:40] + chave_sufixo[:4]
                
                # Verificar se a chave já existe
                if not NotaFiscal.objects.filter(chave_acesso=chave_candidata).exclude(id=nota.id).exists():
                    nota.chave_acesso = chave_candidata
                    chave_gerada = True
                else:
                    tentativa += 1
                    time.sleep(0.001)  # Pequeno delay para evitar colisões
            
            if not chave_gerada:
                # Se não conseguiu gerar uma chave única, usar UUID como fallback
                import uuid
                chave_fallback = f'3520011234567800019055001000000000{nota.numero:08d}{nota.serie:02d}{str(uuid.uuid4())[:4].replace("-", "")}'
                nota.chave_acesso = chave_fallback[:44]
        
        # Tentar salvar com tratamento de IntegrityError
        try:
            nota.save()
        except Exception as ie:
            from django.db import IntegrityError
            # Se ainda houver erro de unicidade, tentar novamente com nova chave
            if isinstance(ie, IntegrityError) and 'chave_acesso' in str(ie).lower():
                import uuid
                chave_fallback = f'3520011234567800019055001000000000{nota.numero:08d}{nota.serie:02d}{str(uuid.uuid4())[:4].replace("-", "")}'
                nota.chave_acesso = chave_fallback[:44]
                nota.save()
            else:
                raise
        
        messages.success(request, f'✅ NF-e {nota.numero}/{nota.serie} atualizada para AUTORIZADA com sucesso!')
        logger.info(f'NF-e {nota.id} (número {nota.numero}/{nota.serie}) atualizada para AUTORIZADA manualmente por {request.user.username}')
    except Exception as e:
        logger.error(f'Erro ao atualizar status da NF-e: {str(e)}', exc_info=True)
        messages.error(request, f'Erro ao atualizar status: {str(e)}')
    
    return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)


# ============================================================================
# ENVIAR NF-E POR E-MAIL
# ============================================================================

@login_required
@require_http_methods(["POST"])
def vendas_nota_fiscal_enviar_email(request, propriedade_id, nota_id):
    """Enviar NF-e por e-mail"""
    import json
    try:
        propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
        nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    except Http404 as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=404)
    
    try:
        data = json.loads(request.body)
        email = data.get('email')
        
        if not email:
            return JsonResponse({'sucesso': False, 'erro': 'E-mail não fornecido'}, status=400)
        
        # Validar formato do e-mail
        import re
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return JsonResponse({'sucesso': False, 'erro': 'E-mail inválido'}, status=400)
        
        # Preparar dados do e-mail
        from django.core.mail import EmailMessage
        from django.conf import settings
        
        # Obter dados do cliente se disponível
        cliente_nome = nota.cliente.nome if nota.cliente else 'Cliente'
        cliente_cpf_cnpj = nota.cliente.cpf_cnpj if nota.cliente and nota.cliente.cpf_cnpj else ''
        cliente_email = nota.cliente.email if nota.cliente and nota.cliente.email else None
        
        # Buscar boletos relacionados à nota fiscal (ANTES de criar o corpo do email)
        anexos_boleto = []
        try:
            from .models_financeiro import LancamentoFinanceiro, AnexoLancamentoFinanceiro
            from decimal import Decimal
            from django.db.models import Q
            
            # Buscar LancamentoFinanceiro relacionado (por documento_referencia contendo número da NF)
            numero_nota_refs = []
            if nota.numero:
                numero_nota_refs.append(f"NF {nota.numero}")
                numero_nota_refs.append(f"NF-e {nota.numero}")
                numero_nota_refs.append(f"NFe {nota.numero}")
                numero_nota_refs.append(f"{nota.numero}")
            
            # Buscar por documento_referencia
            lancamentos_relacionados = LancamentoFinanceiro.objects.none()
            if numero_nota_refs:
                query = Q()
                for ref in numero_nota_refs:
                    query |= Q(documento_referencia__icontains=ref)
                
                lancamentos_relacionados = LancamentoFinanceiro.objects.filter(
                    propriedade=propriedade,
                    forma_pagamento=LancamentoFinanceiro.FORMA_BOLETO,
                    status=LancamentoFinanceiro.STATUS_PENDENTE
                ).filter(query)
            
            # Se não encontrou por referência, buscar por cliente e valor (tolerância de 2%)
            if not lancamentos_relacionados.exists() and nota.cliente:
                valor_min = nota.valor_total * Decimal('0.98')
                valor_max = nota.valor_total * Decimal('1.02')
                
                lancamentos_relacionados = LancamentoFinanceiro.objects.filter(
                    propriedade=propriedade,
                    descricao__icontains=cliente_nome,
                    valor__gte=valor_min,
                    valor__lte=valor_max,
                    forma_pagamento=LancamentoFinanceiro.FORMA_BOLETO,
                    status=LancamentoFinanceiro.STATUS_PENDENTE
                )
            
            # Buscar anexos PDF que possam ser boletos
            for lancamento in lancamentos_relacionados:
                anexos = AnexoLancamentoFinanceiro.objects.filter(
                    lancamento=lancamento
                )
                for anexo in anexos:
                    if anexo.arquivo and anexo.arquivo.name:
                        # Verificar se é PDF e se o nome contém "boleto"
                        if anexo.arquivo.name.lower().endswith('.pdf'):
                            nome_lower = anexo.arquivo.name.lower()
                            if 'boleto' in nome_lower or 'boleto' in (anexo.nome_original or '').lower():
                                anexos_boleto.append(anexo)
        except Exception as e:
            logger.warning(f'Erro ao buscar boletos relacionados: {str(e)}')
        
        # Corpo do e-mail em HTML
        mensagem_boleto = f'<p><strong>Boleto Bancário:</strong> O boleto relacionado a esta NF-e também está anexado.</p>' if anexos_boleto else ''
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #198754;">Nota Fiscal Eletrônica - NF-e {nota.numero}</h2>
                <p>Prezado(a) {cliente_nome},</p>
                <p>Segue em anexo a Nota Fiscal Eletrônica número <strong>{nota.numero}</strong> emitida em {nota.data_emissao.strftime('%d/%m/%Y')}.</p>
                
                <div style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <h3 style="margin-top: 0;">Dados da NF-e:</h3>
                    <p><strong>Número:</strong> {nota.numero}/{nota.serie}</p>
                    <p><strong>Data de Emissão:</strong> {nota.data_emissao.strftime('%d/%m/%Y')}</p>
                    <p><strong>Valor Total:</strong> R$ {nota.valor_total:.2f}</p>
                    {f'<p><strong>Chave de Acesso:</strong> {nota.chave_acesso}</p>' if nota.chave_acesso else ''}
                    {f'<p><strong>Cliente:</strong> {cliente_nome}</p>' if cliente_nome else ''}
                    {f'<p><strong>CPF/CNPJ:</strong> {cliente_cpf_cnpj}</p>' if cliente_cpf_cnpj else ''}
                </div>
                
                <p>Os arquivos PDF (DANFE) e XML estão anexados a este e-mail.</p>
                {mensagem_boleto}
                <p>Em caso de dúvidas, entre em contato conosco.</p>
                
                <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                <p style="color: #666; font-size: 12px;">
                    Este é um e-mail automático. Por favor, não responda diretamente a esta mensagem.
                </p>
            </div>
        </body>
        </html>
        """
        
        # Criar mensagem de e-mail
        email_msg = EmailMessage(
            subject=f'NF-e {nota.numero}/{nota.serie} - {propriedade.nome_propriedade}',
            body=html_body,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else settings.EMAIL_HOST_USER if hasattr(settings, 'EMAIL_HOST_USER') else 'noreply@monpec.com.br',
            to=[email],
        )
        email_msg.content_subtype = "html"  # Definir como HTML
        
        # Anexar PDF e XML se existirem
        arquivos_anexados = []
        if nota.arquivo_pdf and nota.arquivo_pdf.name:
            try:
                email_msg.attach_file(nota.arquivo_pdf.path, mimetype='application/pdf')
                arquivos_anexados.append('PDF (DANFE)')
            except Exception as e:
                logger.warning(f'Erro ao anexar PDF ao e-mail: {str(e)}')
        
        if nota.arquivo_xml and nota.arquivo_xml.name:
            try:
                email_msg.attach_file(nota.arquivo_xml.path, mimetype='application/xml')
                arquivos_anexados.append('XML')
            except Exception as e:
                logger.warning(f'Erro ao anexar XML ao e-mail: {str(e)}')
        
        # Anexar boletos se encontrados
        for anexo_boleto in anexos_boleto:
            try:
                email_msg.attach_file(anexo_boleto.arquivo.path, mimetype='application/pdf')
                arquivos_anexados.append('Boleto')
            except Exception as e:
                logger.warning(f'Erro ao anexar boleto ao e-mail: {str(e)}')
        
        # Verificar se há arquivos para anexar
        if not arquivos_anexados:
            return JsonResponse({
                'sucesso': False,
                'erro': 'A NF-e não possui arquivos PDF ou XML para anexar. Gere os arquivos antes de enviar por e-mail.'
            }, status=400)
        
        # Enviar e-mail
        try:
            email_msg.send()
            logger.info(f'NF-e {nota.numero} enviada por e-mail para {email} com arquivos: {", ".join(arquivos_anexados)}')
            
            arquivos_msg = f" com {', '.join(arquivos_anexados)}" if arquivos_anexados else ""
            return JsonResponse({
                'sucesso': True,
                'mensagem': f'NF-e enviada para {email} com sucesso!{arquivos_msg}.'
            })
        except Exception as e:
            logger.error(f'Erro ao enviar e-mail: {str(e)}', exc_info=True)
            return JsonResponse({
                'sucesso': False,
                'erro': f'Erro ao enviar e-mail: {str(e)}. Verifique as configurações de e-mail do servidor.'
            }, status=500)
    except Exception as e:
        logger.error(f'Erro ao enviar NF-e por e-mail: {str(e)}', exc_info=True)
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


# ============================================================================
# ENVIAR NF-E POR WHATSAPP
# ============================================================================

@login_required
@require_http_methods(["POST"])
def vendas_nota_fiscal_enviar_whatsapp(request, propriedade_id, nota_id):
    """Enviar NF-e por WhatsApp"""
    import json
    try:
        propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
        nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    except Http404 as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=404)
    
    try:
        data = json.loads(request.body)
        telefone = data.get('telefone')
        
        if not telefone:
            return JsonResponse({'sucesso': False, 'erro': 'Telefone não fornecido'}, status=400)
        
        # Limpar telefone (remover caracteres não numéricos)
        telefone = ''.join(filter(str.isdigit, telefone))
        
        # Validar telefone (deve ter pelo menos 10 dígitos)
        if len(telefone) < 10:
            return JsonResponse({'sucesso': False, 'erro': 'Telefone inválido. Deve conter pelo menos 10 dígitos.'}, status=400)
        
        # Obter dados do cliente
        cliente_nome = nota.cliente.nome if nota.cliente else 'Cliente'
        
        # Preparar mensagem para WhatsApp
        mensagem = f"""*Nota Fiscal Eletrônica - NF-e {nota.numero}*

Olá {cliente_nome}!

Segue a Nota Fiscal Eletrônica número *{nota.numero}/{nota.serie}* emitida em {nota.data_emissao.strftime('%d/%m/%Y')}.

*Dados da NF-e:*
• Número: {nota.numero}/{nota.serie}
• Data: {nota.data_emissao.strftime('%d/%m/%Y')}
• Valor Total: R$ {nota.valor_total:.2f}
{f'• Chave de Acesso: {nota.chave_acesso}' if nota.chave_acesso else ''}

Os arquivos PDF e XML serão enviados em seguida.

_Propriedade: {propriedade.nome_propriedade}_"""
        
        # Codificar mensagem para URL
        import urllib.parse
        mensagem_encoded = urllib.parse.quote(mensagem)
        
        # Verificar se há API do WhatsApp configurada
        from django.conf import settings
        
        # Opção 1: WhatsApp Business API (Twilio, Evolution API, etc.)
        whatsapp_api_url = getattr(settings, 'WHATSAPP_API_URL', None)
        whatsapp_api_token = getattr(settings, 'WHATSAPP_API_TOKEN', None)
        whatsapp_api_instance = getattr(settings, 'WHATSAPP_API_INSTANCE', None)
        
        if whatsapp_api_url and whatsapp_api_token:
            # Enviar via API do WhatsApp
            try:
                try:
                    import requests
                except ImportError:
                    logger.warning('Biblioteca requests não instalada. Instale com: pip install requests')
                    raise Exception('Biblioteca requests não está instalada. Execute: pip install requests')
                
                # Preparar dados para envio
                payload = {
                    'number': telefone,
                    'message': mensagem,
                    'token': whatsapp_api_token
                }
                
                # Se tiver instância configurada
                if whatsapp_api_instance:
                    payload['instance'] = whatsapp_api_instance
                
                # Enviar mensagem de texto
                response = requests.post(
                    f'{whatsapp_api_url}/send-message',
                    json=payload,
                    timeout=10
                )
                
                if response.status_code == 200:
                    # Tentar enviar arquivos se disponíveis
                    arquivos_enviados = []
                    
                    # Enviar PDF
                    if nota.arquivo_pdf and nota.arquivo_pdf.name:
                        try:
                            with open(nota.arquivo_pdf.path, 'rb') as pdf_file:
                                files = {'file': (f'NF-e_{nota.numero}.pdf', pdf_file, 'application/pdf')}
                                data = {
                                    'number': telefone,
                                    'token': whatsapp_api_token
                                }
                                if whatsapp_api_instance:
                                    data['instance'] = whatsapp_api_instance
                                
                                file_response = requests.post(
                                    f'{whatsapp_api_url}/send-file',
                                    files=files,
                                    data=data,
                                    timeout=30
                                )
                                if file_response.status_code == 200:
                                    arquivos_enviados.append('PDF')
                        except Exception as e:
                            logger.warning(f'Erro ao enviar PDF via WhatsApp API: {str(e)}')
                    
                    # Enviar XML
                    if nota.arquivo_xml and nota.arquivo_xml.name:
                        try:
                            with open(nota.arquivo_xml.path, 'rb') as xml_file:
                                files = {'file': (f'NF-e_{nota.numero}.xml', xml_file, 'application/xml')}
                                data = {
                                    'number': telefone,
                                    'token': whatsapp_api_token
                                }
                                if whatsapp_api_instance:
                                    data['instance'] = whatsapp_api_instance
                                
                                file_response = requests.post(
                                    f'{whatsapp_api_url}/send-file',
                                    files=files,
                                    data=data,
                                    timeout=30
                                )
                                if file_response.status_code == 200:
                                    arquivos_enviados.append('XML')
                        except Exception as e:
                            logger.warning(f'Erro ao enviar XML via WhatsApp API: {str(e)}')
                    
                    logger.info(f'NF-e {nota.numero} enviada via WhatsApp API para {telefone}')
                    
                    return JsonResponse({
                        'sucesso': True,
                        'mensagem': f'NF-e enviada via WhatsApp para {telefone} com sucesso!',
                        'arquivos_enviados': arquivos_enviados
                    })
                else:
                    # Se API falhar, usar link do WhatsApp Web como fallback
                    logger.warning(f'Erro na API do WhatsApp: {response.status_code} - {response.text}')
                    raise Exception('API do WhatsApp não disponível')
                    
            except Exception as e:
                logger.warning(f'Erro ao usar API do WhatsApp, usando link do WhatsApp Web: {str(e)}')
                # Continuar para usar link do WhatsApp Web
        
        # Opção 2: Link do WhatsApp Web (fallback ou quando API não está configurada)
        whatsapp_url = f'https://wa.me/{telefone}?text={mensagem_encoded}'
        
        # Se tiver arquivos, criar um link alternativo que abre o WhatsApp Web
        # Nota: WhatsApp Web não permite anexar arquivos diretamente via URL
        # Será necessário que o usuário anexe manualmente
        
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Link do WhatsApp gerado para {telefone}. Clique para abrir o WhatsApp Web.',
            'whatsapp_url': whatsapp_url,
            'nota': {
                'numero': nota.numero,
                'serie': nota.serie,
                'valor_total': str(nota.valor_total),
                'data_emissao': nota.data_emissao.strftime('%d/%m/%Y'),
                'chave_acesso': nota.chave_acesso or '',
            },
            'tem_pdf': bool(nota.arquivo_pdf and nota.arquivo_pdf.name),
            'tem_xml': bool(nota.arquivo_xml and nota.arquivo_xml.name),
            'observacao': 'Se os arquivos PDF e XML estiverem disponíveis, você precisará anexá-los manualmente no WhatsApp.'
        })
    except Exception as e:
        logger.error(f'Erro ao enviar NF-e por WhatsApp: {str(e)}', exc_info=True)
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


# ============================================================================
# BAIXAR ARQUIVOS DA NF-E (PDF E XML)
# ============================================================================

@login_required
def vendas_nota_fiscal_baixar_arquivos(request, propriedade_id, nota_id):
    """Baixar PDF e XML da NF-e em um arquivo ZIP (apenas se ambos existirem)"""
    import zipfile
    import io
    from django.http import HttpResponse, FileResponse
    
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    nota = get_object_or_404(NotaFiscal, id=nota_id, propriedade=propriedade, tipo='SAIDA')
    
    # Verificar quais arquivos existem
    tem_pdf = nota.arquivo_pdf and nota.arquivo_pdf.name
    tem_xml = nota.arquivo_xml and nota.arquivo_xml.name
    
    # Se houver apenas um arquivo, baixar diretamente
    if tem_pdf and not tem_xml:
        try:
            return FileResponse(nota.arquivo_pdf.open(), as_attachment=True, filename=f'NF-e_{nota.numero}.pdf')
        except Exception as e:
            logger.error(f'Erro ao baixar PDF: {str(e)}', exc_info=True)
            messages.error(request, f'Erro ao baixar PDF: {str(e)}')
            return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    if tem_xml and not tem_pdf:
        try:
            return FileResponse(nota.arquivo_xml.open(), as_attachment=True, filename=f'NF-e_{nota.numero}.xml')
        except Exception as e:
            logger.error(f'Erro ao baixar XML: {str(e)}', exc_info=True)
            messages.error(request, f'Erro ao baixar XML: {str(e)}')
            return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    # Se houver ambos, baixar em ZIP
    if tem_pdf and tem_xml:
        try:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                try:
                    pdf_path = nota.arquivo_pdf.path
                    zip_file.write(pdf_path, f'NF-e_{nota.numero}.pdf')
                except Exception as e:
                    logger.warning(f'Erro ao adicionar PDF ao ZIP: {str(e)}')
                
                try:
                    xml_path = nota.arquivo_xml.path
                    zip_file.write(xml_path, f'NF-e_{nota.numero}.xml')
                except Exception as e:
                    logger.warning(f'Erro ao adicionar XML ao ZIP: {str(e)}')
            
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="NF-e_{nota.numero}.zip"'
            return response
        except Exception as e:
            logger.error(f'Erro ao criar arquivo ZIP da NF-e: {str(e)}', exc_info=True)
            messages.error(request, f'Erro ao baixar arquivos: {str(e)}')
            return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
    
    # Se não houver nenhum arquivo
    messages.warning(request, 'Esta NF-e não possui arquivos PDF ou XML para download.')
    return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)


# ============================================================================
# NOVA VENDA COM EMISSÃO AUTOMÁTICA DE NF-E
# ============================================================================

@login_required
def vendas_venda_nova(request, propriedade_id):
    """Criar nova venda e emitir NF-e automaticamente"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    # Criar formset para itens
    ItemVendaFormSet = inlineformset_factory(
        NotaFiscal,
        ItemNotaFiscal,
        form=ItemVendaForm,
        extra=1,
        can_delete=True,
        min_num=1,
        validate_min=True
    )
    
    serie_padrao = '1'
    
    if request.method == 'POST':
        form = VendaForm(request.POST, propriedade=propriedade)
        formset = ItemVendaFormSet(request.POST)
        emitir_nfe = request.POST.get('emitir_nfe', '') == 'on'
        
        if form.is_valid() and formset.is_valid():
            # Obter próximo número E INCREMENTAR (apenas ao salvar)
            from .services_nfe_utils import obter_proximo_numero_nfe
            proximo_numero = obter_proximo_numero_nfe(propriedade, serie_padrao)
            
            nota = form.save(commit=False)
            nota.propriedade = propriedade
            nota.tipo = 'SAIDA'
            nota.numero = str(proximo_numero)
            nota.serie = serie_padrao
            nota.status = 'PENDENTE'
            nota.valor_produtos = Decimal('0.00')
            nota.valor_frete = Decimal('0.00')
            nota.valor_seguro = Decimal('0.00')
            nota.valor_desconto = Decimal('0.00')
            nota.valor_outros = Decimal('0.00')
            nota.valor_total = Decimal('0.00')
            
            nota.save()
            
            # Salvar itens
            formset.instance = nota
            itens = formset.save(commit=False)
            
            # Calcular valor_total de cada item antes de salvar
            for item in itens:
                if not item.valor_total or item.valor_total == 0:
                    item.valor_total = item.quantidade * item.valor_unitario
                item.save()
            
            # Recalcular valores totais da nota fiscal
            valor_total_itens = sum(item.valor_total for item in nota.itens.all())
            nota.valor_produtos = valor_total_itens
            nota.valor_total = valor_total_itens
            nota.save()
            
            # Se solicitado, emitir NF-e automaticamente
            if emitir_nfe:
                try:
                    from .services_nfe import emitir_nfe
                    resultado = emitir_nfe(nota)
                    
                    if resultado and resultado.get('sucesso'):
                        nota.status = 'AUTORIZADA'
                        nota.chave_acesso = resultado.get('chave_acesso', '')
                        nota.protocolo_autorizacao = resultado.get('protocolo', '')
                        nota.data_autorizacao = timezone.now()
                        if resultado.get('xml'):
                            from django.core.files.base import ContentFile
                            nota.arquivo_xml.save(f'nfe_{nota.numero}.xml', ContentFile(resultado['xml']), save=False)
                        nota.save()
                        messages.success(request, f'Venda registrada e NF-e {nota.numero} emitida e autorizada com sucesso!')
                    else:
                        erro_msg = resultado.get("erro", "Erro desconhecido") if resultado else "Resposta inválida da API"
                        messages.warning(request, f'Venda registrada, mas NF-e não foi autorizada: {erro_msg}')
                except Exception as e:
                    logger.error(f'Erro ao emitir NF-e automaticamente: {str(e)}', exc_info=True)
                    messages.warning(request, f'Venda registrada, mas houve erro ao emitir NF-e: {str(e)}')
            else:
                messages.success(request, f'Venda registrada com sucesso! Você pode emitir a NF-e depois.')
            
            return redirect('vendas_nota_fiscal_detalhes', propriedade_id=propriedade.id, nota_id=nota.id)
        else:
            messages.error(request, 'Por favor, corrija os erros no formulário.')
            # Em caso de erro, obter próximo número SEM INCREMENTAR (para mostrar no formulário)
            from .services_nfe_utils import visualizar_proximo_numero_nfe
            proximo_numero = visualizar_proximo_numero_nfe(propriedade, serie_padrao)
    else:
        form = VendaForm(propriedade=propriedade)
        # Definir data padrão como hoje
        form.initial['data_emissao'] = date.today()
        formset = ItemVendaFormSet(instance=NotaFiscal())
        
        # Obter próximo número SEM INCREMENTAR (apenas para visualização)
        from .services_nfe_utils import visualizar_proximo_numero_nfe
        proximo_numero = visualizar_proximo_numero_nfe(propriedade, serie_padrao)
    
    clientes = Cliente.objects.filter(
        Q(propriedade=propriedade) | Q(propriedade__isnull=True),
        ativo=True
    ).order_by('nome')
    
    context = {
        'propriedade': propriedade,
        'form': form,
        'formset': formset,
        'clientes': clientes,
        'proximo_numero': proximo_numero,
        'serie_padrao': serie_padrao,
    }
    
    return render(request, 'gestao_rural/vendas_venda_nova.html', context)


# ============================================================================
# SINCRONIZAR NF-E RECEBIDAS
# ============================================================================

@login_required
def vendas_sincronizar_nfe_recebidas(request, propriedade_id):
    """Sincronizar NF-e recebidas via API"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    if request.method == 'POST':
        try:
            from datetime import timedelta
            from .services_nfe_consulta import consultar_nfe_recebidas, baixar_xml_nfe, importar_nfe_do_xml
            from django.conf import settings
            from django.db import transaction
            
            dias = int(request.POST.get('dias', 30))
            data_fim = date.today()
            data_inicio = data_fim - timedelta(days=dias)
            
            # Consultar NF-e recebidas
            resultado = consultar_nfe_recebidas(
                propriedade=propriedade,
                data_inicio=data_inicio,
                data_fim=data_fim,
                limite=100
            )
            
            if not resultado.get('sucesso'):
                messages.error(request, f'Erro ao consultar NF-e recebidas: {resultado.get("erro", "Erro desconhecido")}')
                return redirect('vendas_dashboard', propriedade_id=propriedade.id)
            
            notas_encontradas = resultado.get('notas', [])
            notas_importadas = 0
            notas_ja_existentes = 0
            
            api_config = getattr(settings, 'API_NFE', {})
            
            for nota_data in notas_encontradas:
                chave_acesso = nota_data.get('chave_acesso', '')
                if not chave_acesso:
                    continue
                
                # Verificar se já existe
                if NotaFiscal.objects.filter(chave_acesso=chave_acesso).exists():
                    notas_ja_existentes += 1
                    continue
                
                try:
                    with transaction.atomic():
                        # Baixar XML
                        resultado_xml = baixar_xml_nfe(chave_acesso, api_config)
                        if not resultado_xml.get('sucesso'):
                            continue
                        
                        xml_content = resultado_xml.get('xml')
                        if not xml_content:
                            continue
                        
                        # Importar
                        resultado_importacao = importar_nfe_do_xml(
                            xml_content=xml_content,
                            propriedade=propriedade,
                            usuario=request.user
                        )
                        
                        if resultado_importacao.get('sucesso'):
                            notas_importadas += 1
                except Exception as e:
                    logger.error(f'Erro ao importar NF-e {chave_acesso}: {str(e)}', exc_info=True)
                    continue
            
            messages.success(
                request,
                f'Sincronização concluída: {notas_importadas} nota(s) importada(s), '
                f'{notas_ja_existentes} já existente(s)'
            )
        except Exception as e:
            logger.error(f'Erro ao sincronizar NF-e recebidas: {str(e)}', exc_info=True)
            messages.error(request, f'Erro ao sincronizar: {str(e)}')
    
    return redirect('vendas_dashboard', propriedade_id=propriedade.id)


# ============================================================================
# EXPORTAR NOTAS FISCAIS PARA EXCEL
# ============================================================================

@login_required
def vendas_notas_fiscais_exportar_excel(request, propriedade_id):
    """Exporta notas fiscais de saída para Excel"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Biblioteca openpyxl não instalada. Execute: pip install openpyxl')
        return redirect('vendas_notas_fiscais_lista', propriedade_id=propriedade.id)
    
    # Aplicar mesmos filtros da listagem
    status_filter = request.GET.get('status', '')
    busca = request.GET.get('busca', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    notas = NotaFiscal.objects.filter(
        propriedade=propriedade,
        tipo='SAIDA'
    ).select_related('cliente').order_by('-data_emissao', '-numero')
    
    if status_filter:
        notas = notas.filter(status=status_filter)
    if busca:
        notas = notas.filter(
            Q(numero__icontains=busca) |
            Q(chave_acesso__icontains=busca) |
            Q(cliente__nome__icontains=busca) |
            Q(cliente__cpf_cnpj__icontains=busca)
        )
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__gte=data_inicio_obj)
        except ValueError:
            pass
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__lte=data_fim_obj)
        except ValueError:
            pass
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais de Saída"
    
    # Estilos
    title_font = Font(name='Calibri', size=16, bold=True, color='1F497D')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # Título
    ws.merge_cells('A1:K1')
    ws['A1'] = f'Relatório de Notas Fiscais de Saída - {propriedade.nome_propriedade}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Data de geração
    ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws.merge_cells('A2:K2')
    
    # Cabeçalhos
    headers = [
        'Número', 'Série', 'Data Emissão', 'Cliente', 'CPF/CNPJ',
        'Valor Total', 'Status', 'Chave de Acesso', 'Protocolo', 'Data Autorização', 'Observações'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    row = 5
    total_geral = Decimal('0.00')
    for nota in notas:
        ws.cell(row=row, column=1, value=nota.numero)
        ws.cell(row=row, column=2, value=nota.serie)
        ws.cell(row=row, column=3, value=nota.data_emissao.strftime('%d/%m/%Y') if nota.data_emissao else '')
        ws.cell(row=row, column=4, value=nota.cliente.nome if nota.cliente else '')
        ws.cell(row=row, column=5, value=nota.cliente.cpf_cnpj if nota.cliente else '')
        ws.cell(row=row, column=6, value=float(nota.valor_total))
        ws.cell(row=row, column=7, value=nota.get_status_display())
        ws.cell(row=row, column=8, value=nota.chave_acesso or '')
        ws.cell(row=row, column=9, value=nota.protocolo_autorizacao or '')
        ws.cell(row=row, column=10, value=nota.data_autorizacao.strftime('%d/%m/%Y %H:%M') if nota.data_autorizacao else '')
        ws.cell(row=row, column=11, value=nota.observacoes or '')
        
        if nota.status == 'AUTORIZADA':
            total_geral += nota.valor_total
        row += 1
    
    # Totais
    ws.cell(row=row + 1, column=5, value='TOTAL AUTORIZADO:')
    ws.cell(row=row + 1, column=6, value=float(total_geral)).font = Font(bold=True)
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['D'].width = 30  # Cliente
    ws.column_dimensions['H'].width = 45  # Chave de Acesso
    ws.column_dimensions['K'].width = 40  # Observações
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_nfe_saida_{propriedade.id}.xlsx"'
    wb.save(response)
    
    return response


# ============================================================================
# RELATÓRIO DE CONTABILIDADE
# ============================================================================

@login_required
def vendas_relatorio_contabilidade(request, propriedade_id):
    """Relatório de notas fiscais para contabilidade"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    # Filtros
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    status_filter = request.GET.get('status', 'AUTORIZADA')  # Por padrão, apenas autorizadas
    
    # Query base
    notas = NotaFiscal.objects.filter(
        propriedade=propriedade,
        tipo='SAIDA'
    ).select_related('cliente').order_by('-data_emissao', '-numero')
    
    if status_filter:
        notas = notas.filter(status=status_filter)
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__gte=data_inicio_obj)
        except ValueError:
            messages.error(request, 'Formato de Data Início inválido.')
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__lte=data_fim_obj)
        except ValueError:
            messages.error(request, 'Formato de Data Fim inválido.')
    
    # Agrupar por mês/ano para resumo
    from django.db.models.functions import TruncMonth
    resumo_por_mes = notas.annotate(
        mes_ano=TruncMonth('data_emissao')
    ).values('mes_ano').annotate(
        total_notas=Count('id'),
        valor_total_mes=Sum('valor_total')
    ).order_by('-mes_ano')
    
    context = {
        'propriedade': propriedade,
        'notas': notas,
        'resumo_por_mes': resumo_por_mes,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'status_filter': status_filter,
    }
    return render(request, 'gestao_rural/vendas_relatorio_contabilidade.html', context)


@login_required
def vendas_relatorio_contabilidade_exportar_excel(request, propriedade_id):
    """Exporta relatório de contabilidade para Excel"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Biblioteca openpyxl não instalada. Execute: pip install openpyxl')
        return redirect('vendas_relatorio_contabilidade', propriedade_id=propriedade.id)
    
    # Mesmos filtros do relatório
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    status_filter = request.GET.get('status', 'AUTORIZADA')
    
    notas = NotaFiscal.objects.filter(
        propriedade=propriedade,
        tipo='SAIDA'
    ).select_related('cliente').order_by('data_emissao', 'numero')
    
    if status_filter:
        notas = notas.filter(status=status_filter)
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__gte=data_inicio_obj)
        except ValueError:
            pass
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            notas = notas.filter(data_emissao__lte=data_fim_obj)
        except ValueError:
            pass
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Contabilidade"
    
    # Estilos
    title_font = Font(name='Calibri', size=16, bold=True, color='1F497D')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f'Relatório de Contabilidade - {propriedade.nome_propriedade}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Cabeçalhos
    headers = [
        'Data Emissão', 'Número', 'Série', 'Cliente', 'CPF/CNPJ',
        'Valor Total', 'Status', 'Chave de Acesso', 'Protocolo', 'Observações'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    row = 4
    total_geral = Decimal('0.00')
    for nota in notas:
        ws.cell(row=row, column=1, value=nota.data_emissao.strftime('%d/%m/%Y') if nota.data_emissao else '')
        ws.cell(row=row, column=2, value=nota.numero)
        ws.cell(row=row, column=3, value=nota.serie)
        ws.cell(row=row, column=4, value=nota.cliente.nome if nota.cliente else '')
        ws.cell(row=row, column=5, value=nota.cliente.cpf_cnpj if nota.cliente else '')
        ws.cell(row=row, column=6, value=float(nota.valor_total))
        ws.cell(row=row, column=7, value=nota.get_status_display())
        ws.cell(row=row, column=8, value=nota.chave_acesso or '')
        ws.cell(row=row, column=9, value=nota.protocolo_autorizacao or '')
        ws.cell(row=row, column=10, value=nota.observacoes or '')
        
        if nota.status == 'AUTORIZADA':
            total_geral += nota.valor_total
        row += 1
    
    # Totais
    ws.cell(row=row + 1, column=5, value='TOTAL AUTORIZADO:')
    ws.cell(row=row + 1, column=6, value=float(total_geral)).font = Font(bold=True)
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['D'].width = 30  # Cliente
    ws.column_dimensions['H'].width = 45  # Chave de Acesso
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_contabilidade_{propriedade.id}.xlsx"'
    wb.save(response)
    
    return response


# ============================================================================
# CONFIGURAR SÉRIES DE NF-E
# ============================================================================

@login_required
def vendas_configurar_series_nfe(request, propriedade_id):
    """Configura as séries de NF-e e seus próximos números para a propriedade."""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    from .services_nfe_utils import configurar_serie_nfe
    
    series_existentes = NumeroSequencialNFE.objects.filter(propriedade=propriedade).order_by('serie')
    
    if request.method == 'POST':
        form = ConfigurarSerieNFeForm(request.POST)
        if form.is_valid():
            serie = form.cleaned_data['serie']
            proximo_numero = form.cleaned_data['proximo_numero']
            
            try:
                configurar_serie_nfe(propriedade, serie, proximo_numero)
                messages.success(request, f'Série {serie} configurada com sucesso!')
                return redirect('vendas_configurar_series_nfe', propriedade_id=propriedade.id)
            except Exception as e:
                logger.error(f'Erro ao configurar série NF-e: {str(e)}', exc_info=True)
                messages.error(request, f'Erro ao configurar série NF-e: {str(e)}')
        else:
            messages.error(request, 'Por favor, corrija os erros no formulário.')
    else:
        form = ConfigurarSerieNFeForm()
    
    context = {
        'propriedade': propriedade,
        'form': form,
        'series_existentes': series_existentes,
    }
    return render(request, 'gestao_rural/vendas_configurar_series_nfe.html', context)


@login_required
@require_http_methods(["POST"])
def vendas_excluir_serie_nfe(request, propriedade_id, serie_id):
    """Exclui uma configuração de série de NF-e."""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    serie_obj = get_object_or_404(NumeroSequencialNFE, id=serie_id, propriedade=propriedade)
    
    try:
        serie_obj.delete()
        messages.success(request, f'Série {serie_obj.serie} excluída com sucesso.')
    except Exception as e:
        logger.error(f'Erro ao excluir série NF-e: {str(e)}', exc_info=True)
        messages.error(request, f'Erro ao excluir série {serie_obj.serie}: {str(e)}')
    
    return redirect('vendas_configurar_series_nfe', propriedade_id=propriedade.id)


# ============================================================================
# VIEWS DE VENDAS POR CATEGORIA
# ============================================================================

@login_required
def vendas_por_categoria_lista(request, propriedade_id):
    """Lista parâmetros de venda por categoria"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    from .models import ParametrosVendaPorCategoria
    parametros = ParametrosVendaPorCategoria.objects.filter(propriedade=propriedade).order_by('categoria__nome')
    
    context = {
        'propriedade': propriedade,
        'parametros': parametros,
    }
    return render(request, 'gestao_rural/vendas_por_categoria_lista.html', context)


@login_required
def vendas_por_categoria_novo(request, propriedade_id):
    """Criar novo parâmetro de venda por categoria"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    if request.method == 'POST':
        form = ParametrosVendaPorCategoriaForm(request.POST, propriedade=propriedade)
        if form.is_valid():
            parametro = form.save(commit=False)
            parametro.propriedade = propriedade
            parametro.save()
            messages.success(request, 'Parâmetro de venda criado com sucesso!')
            return redirect('vendas_por_categoria_lista', propriedade_id=propriedade.id)
    else:
        form = ParametrosVendaPorCategoriaForm(propriedade=propriedade)
    
    context = {
        'propriedade': propriedade,
        'form': form,
    }
    return render(request, 'gestao_rural/vendas_por_categoria_form.html', context)


@login_required
def vendas_por_categoria_editar(request, propriedade_id, parametro_id):
    """Editar parâmetro de venda por categoria"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    from .models import ParametrosVendaPorCategoria
    parametro = get_object_or_404(ParametrosVendaPorCategoria, id=parametro_id, propriedade=propriedade)
    
    if request.method == 'POST':
        form = ParametrosVendaPorCategoriaForm(request.POST, instance=parametro, propriedade=propriedade)
        if form.is_valid():
            form.save()
            messages.success(request, 'Parâmetro de venda atualizado com sucesso!')
            return redirect('vendas_por_categoria_lista', propriedade_id=propriedade.id)
    else:
        form = ParametrosVendaPorCategoriaForm(instance=parametro, propriedade=propriedade)
    
    context = {
        'propriedade': propriedade,
        'form': form,
        'parametro': parametro,
    }
    return render(request, 'gestao_rural/vendas_por_categoria_form.html', context)


@login_required
def vendas_por_categoria_bulk(request, propriedade_id):
    """Atualização em lote de parâmetros de venda"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    
    if request.method == 'POST':
        form = BulkVendaPorCategoriaForm(request.POST)
        if form.is_valid():
            # Implementar lógica de atualização em lote
            messages.success(request, 'Parâmetros atualizados em lote com sucesso!')
            return redirect('vendas_por_categoria_lista', propriedade_id=propriedade.id)
    else:
        form = BulkVendaPorCategoriaForm()
    
    context = {
        'propriedade': propriedade,
        'form': form,
    }
    return render(request, 'gestao_rural/vendas_por_categoria_bulk.html', context)


@login_required
def vendas_por_categoria_excluir(request, propriedade_id, parametro_id):
    """Excluir parâmetro de venda por categoria"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    from .models import ParametrosVendaPorCategoria
    parametro = get_object_or_404(ParametrosVendaPorCategoria, id=parametro_id, propriedade=propriedade)
    
    if request.method == 'POST':
        parametro.delete()
        messages.success(request, 'Parâmetro de venda excluído com sucesso!')
        return redirect('vendas_por_categoria_lista', propriedade_id=propriedade.id)
    
    context = {
        'propriedade': propriedade,
        'parametro': parametro,
    }
    return render(request, 'gestao_rural/vendas_por_categoria_excluir.html', context)


@login_required
def vendas_por_categoria_toggle_status(request, propriedade_id, parametro_id):
    """Alternar status ativo/inativo do parâmetro"""
    propriedade = obter_propriedade_com_permissao(request.user, propriedade_id)
    from .models import ParametrosVendaPorCategoria
    parametro = get_object_or_404(ParametrosVendaPorCategoria, id=parametro_id, propriedade=propriedade)
    
    parametro.ativo = not parametro.ativo
    parametro.save(update_fields=['ativo'])
    
    status_msg = 'ativado' if parametro.ativo else 'desativado'
    messages.success(request, f'Parâmetro {status_msg} com sucesso!')
    
    return redirect('vendas_por_categoria_lista', propriedade_id=propriedade.id)


@login_required
def validar_certificado_digital_produtor(request, produtor_id):
    """Valida o certificado digital de um produtor"""
    from .models import ProdutorRural
    from django.http import JsonResponse
    from datetime import date
    
    try:
        produtor = get_object_or_404(ProdutorRural, id=produtor_id)
        
        # Verificar se o usuário tem permissão
        if not request.user.is_superuser and produtor.usuario_responsavel != request.user:
            return JsonResponse({'erro': 'Sem permissão para acessar este produtor'}, status=403)
        
        # Validar certificado
        tem_certificado = produtor.certificado_digital is not None
        certificado_valido = produtor.tem_certificado_valido() if tem_certificado else False
        
        resultado = {
            'produtor_id': produtor.id,
            'produtor_nome': produtor.nome,
            'tem_certificado': tem_certificado,
            'certificado_valido': certificado_valido,
            'certificado_valido_ate': produtor.certificado_valido_ate.strftime('%d/%m/%Y') if produtor.certificado_valido_ate else None,
            'certificado_expirado': produtor.certificado_valido_ate < date.today() if produtor.certificado_valido_ate else False,
        }
        
        return JsonResponse(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao validar certificado digital: {str(e)}")
        return JsonResponse({'erro': f'Erro ao validar certificado: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def validar_certificado_upload(request):
    """
    Valida e extrai informações de um certificado digital enviado via upload
    Retorna: CNPJ, data de validade, emissor, etc.
    """
    from .models import ProdutorRural
    from django.http import JsonResponse
    from datetime import datetime
    import re
    
    try:
        if 'arquivo' not in request.FILES:
            return JsonResponse({'sucesso': False, 'erro': 'Nenhum arquivo enviado'}, status=400)
        
        arquivo = request.FILES['arquivo']
        senha = request.POST.get('senha', '')
        
        # Validar extensão
        if not arquivo.name.lower().endswith(('.p12', '.pfx')):
            return JsonResponse({
                'sucesso': False, 
                'erro': 'Arquivo deve ser .p12 ou .pfx'
            }, status=400)
        
        # Tentar validar com OpenSSL/PyOpenSSL
        try:
            from OpenSSL import crypto
            
            # Ler arquivo
            arquivo.seek(0)
            conteudo = arquivo.read()
            
            # Tentar carregar certificado
            try:
                p12 = crypto.load_pkcs12(conteudo, senha.encode() if senha else b'')
                cert = p12.get_certificate()
                
                # Extrair informações
                subject = cert.get_subject()
                issuer = cert.get_issuer()
                
                # Extrair CNPJ do subject (pode estar em CN ou OU)
                cnpj = None
                for attr in ['CN', 'OU', 'O']:
                    valor = getattr(subject, attr, None)
                    if valor:
                        # Procurar CNPJ no formato XX.XXX.XXX/XXXX-XX
                        cnpj_match = re.search(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})', str(valor))
                        if cnpj_match:
                            cnpj = cnpj_match.group(1)
                            break
                        # Procurar CNPJ sem formatação (14 dígitos)
                        cnpj_match = re.search(r'(\d{14})', str(valor))
                        if cnpj_match:
                            cnpj_raw = cnpj_match.group(1)
                            # Formatar CNPJ
                            cnpj = f"{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}/{cnpj_raw[8:12]}-{cnpj_raw[12:]}"
                            break
                
                # Extrair data de validade
                not_after = cert.get_notAfter()
                if not_after:
                    # Formato: b'20251231235959Z'
                    data_str = not_after.decode('utf-8').replace('Z', '')
                    data_validade = datetime.strptime(data_str, '%Y%m%d%H%M%S').date()
                else:
                    data_validade = None
                
                # Verificar se está expirado
                expirado = data_validade < date.today() if data_validade else False
                
                # Extrair nome/razão social
                razao_social = getattr(subject, 'CN', None) or getattr(subject, 'O', None) or ''
                
                # Emissor
                emissor = getattr(issuer, 'CN', None) or getattr(issuer, 'O', None) or 'Desconhecido'
                
                return JsonResponse({
                    'sucesso': True,
                    'cnpj': cnpj,
                    'razao_social': str(razao_social),
                    'validade_ate': data_validade.strftime('%d/%m/%Y') if data_validade else None,
                    'validade_ate_iso': data_validade.isoformat() if data_validade else None,
                    'expirado': expirado,
                    'emissor': str(emissor),
                    'valido': not expirado,
                })
                
            except crypto.Error as e:
                if 'mac verify failure' in str(e).lower() or 'bad decrypt' in str(e).lower():
                    return JsonResponse({
                        'sucesso': False,
                        'erro': 'Senha incorreta. Verifique a senha do certificado.'
                    }, status=400)
                else:
                    return JsonResponse({
                        'sucesso': False,
                        'erro': f'Erro ao processar certificado: {str(e)}'
                    }, status=400)
                
        except ImportError:
            # Se PyOpenSSL não estiver instalado, retornar erro informativo
            return JsonResponse({
                'sucesso': False,
                'erro': 'Biblioteca PyOpenSSL não instalada. Instale com: pip install pyopenssl'
            }, status=500)
            
    except Exception as e:
        logger.error(f"Erro ao validar certificado upload: {str(e)}", exc_info=True)
        return JsonResponse({
            'sucesso': False,
            'erro': f'Erro ao processar certificado: {str(e)}'
        }, status=500)
