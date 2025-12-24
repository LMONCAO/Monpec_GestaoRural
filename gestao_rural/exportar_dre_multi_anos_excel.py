# -*- coding: utf-8 -*-
"""
Módulo para exportar DRE com múltiplos anos lado a lado (comparativo)
Layout idêntico ao sistema
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal


def formatar_moeda(valor):
    """Formata valor como moeda brasileira"""
    if valor is None:
        valor = Decimal('0.00')
    if isinstance(valor, Decimal):
        valor = float(valor)
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def verificar_pessoa_fisica(produtor):
    """Verifica se o produtor é pessoa física (CPF) ou jurídica (CNPJ)"""
    if not produtor or not produtor.cpf_cnpj:
        return False
    cpf_cnpj_limpo = produtor.cpf_cnpj.replace(".", "").replace("-", "").replace("/", "")
    return len(cpf_cnpj_limpo) == 11  # CPF tem 11 dígitos


def criar_planilha_dre_multi_anos(workbook, dre_por_ano, produtor, propriedades_selecionadas, is_pessoa_fisica=False):
    """
    Cria planilha de DRE com múltiplos anos lado a lado
    Layout idêntico ao sistema
    """
    ws = workbook.create_sheet("DRE Comparativo", 0)
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    subtitulo_font = Font(name='Arial', size=11, bold=True)
    cabecalho_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=9)
    negrito_font = Font(name='Arial', size=9, bold=True)
    
    titulo_fill = PatternFill(start_color='1a5490', end_color='1a5490', fill_type='solid')
    subtitulo_fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    cabecalho_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    receita_fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
    receita_liquida_fill = PatternFill(start_color='c8e6c9', end_color='c8e6c9', fill_type='solid')
    despesa_fill = PatternFill(start_color='f1f3f5', end_color='f1f3f5', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    anos = sorted(dre_por_ano.keys())
    num_cols = 2 + len(anos)  # Código + Descrição + Anos
    
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO"
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = f"CONSOLIDADO - {produtor.nome.upper()}"
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    ws.merge_cells(f'A3:{get_column_letter(num_cols)}3')
    cell = ws['A3']
    cell.value = f"CPF/CNPJ: {produtor.cpf_cnpj}"
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho da tabela
    row = 5
    ws['A5'] = 'CÓDIGO'
    ws['B5'] = 'DESCRIÇÃO'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        ws[f'{col}5'] = f'{ano} (R$)'
        cell = ws[f'{col}5']
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        col_idx += 1
    
    # Aplicar estilo ao cabeçalho
    for col in ['A', 'B']:
        cell = ws[f'{col}5']
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal='left' if col == 'B' else 'center', vertical='center')
        cell.border = border
    
    row = 6
    
    # 1. RECEITA BRUTA DE VENDAS
    ws[f'A{row}'] = '3.01.01.01.01'
    ws[f'B{row}'] = 'RECEITA BRUTA DE VENDAS'
    ws[f'A{row}'].fill = receita_fill
    ws[f'B{row}'].fill = receita_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('receita_bruta', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.fill = receita_fill
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    ws[f'A{row}'] = '3.01.01.01.01.0001'
    ws[f'B{row}'] = 'Vendas Mercadorias Produção Própria'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('receita_bruta', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 2. DEDUÇÕES DA RECEITA BRUTA
    ws[f'A{row}'] = '3.01.01.01.02'
    ws[f'B{row}'] = 'DEDUÇÕES DA RECEITA BRUTA'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('total_deducoes', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.fill = cabecalho_fill
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Funrural
    ws[f'A{row}'] = '3.01.01.01.02.0004'
    ws[f'B{row}'] = '  Funrural s/Vendas'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('funrural_vendas', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # ICMS
    ws[f'A{row}'] = '3.01.01.01.02.0005'
    ws[f'B{row}'] = '  ICMS s/Vendas'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('icms_vendas', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Outros Impostos
    ws[f'A{row}'] = '3.01.01.01.02.0006'
    ws[f'B{row}'] = '  Outros Impostos s/Vendas'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('outros_impostos_vendas', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Total Deduções
    ws[f'A{row}'] = '3.01.01.01.02'
    ws[f'B{row}'] = 'Total Deduções'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('total_deducoes', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.fill = cabecalho_fill
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 3. RECEITA LÍQUIDA
    ws[f'A{row}'] = '3.01.01.01.03'
    ws[f'B{row}'] = 'RECEITA LÍQUIDA'
    ws[f'A{row}'].fill = receita_liquida_fill
    ws[f'B{row}'].fill = receita_liquida_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('receita_liquida', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.fill = receita_liquida_fill
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 4. CUSTOS MERCADORIAS VENDIDAS
    ws[f'A{row}'] = '3.01.01.01.03.'
    ws[f'B{row}'] = 'CUSTOS MERCADORIAS VENDIDAS'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    row += 1
    
    ws[f'A{row}'] = '3.01.01.01.03.0001'
    ws[f'B{row}'] = '  Custos Mercadorias Produção Própria Vendidas'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('cpv', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 5. LUCRO BRUTO
    ws[f'A{row}'] = '3.01.01.01.04'
    ws[f'B{row}'] = 'LUCRO BRUTO'
    ws[f'A{row}'].fill = receita_liquida_fill
    ws[f'B{row}'].fill = receita_liquida_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('lucro_bruto', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.fill = receita_liquida_fill
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 6. DESPESAS OPERACIONAIS
    ws[f'A{row}'] = '3.01.01.07.'
    ws[f'B{row}'] = 'DESPESAS OPERACIONAIS'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('despesas_operacionais', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.fill = cabecalho_fill
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Despesas Diversas
    ws[f'A{row}'] = '3.01.01.07.01.'
    ws[f'B{row}'] = '  DESPESAS DIVERSAS'
    ws[f'A{row}'].fill = despesa_fill
    ws[f'B{row}'].fill = despesa_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    row += 1
    
    # Retirada Labore
    ws[f'A{row}'] = '3.01.01.07.01.0001'
    ws[f'B{row}'] = '    Retirada Labore'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('retirada_labore', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Depreciação
    ws[f'A{row}'] = '3.01.01.07.01.0013'
    ws[f'B{row}'] = '    Despesas Encargos Depreciação'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('depreciacao_amortizacao', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Total Despesas Diversas
    ws[f'A{row}'] = '3.01.01.07.01'
    ws[f'B{row}'] = '  Total Despesas Diversas'
    ws[f'A{row}'].fill = despesa_fill
    ws[f'B{row}'].fill = despesa_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        # Somar despesas detalhadas
        retirada = dre_por_ano[ano].get('retirada_labore', Decimal('0.00'))
        deprec = dre_por_ano[ano].get('depreciacao_amortizacao', Decimal('0.00'))
        total = retirada + deprec
        ws[f'{col}{row}'] = formatar_moeda(-total) if total > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.fill = despesa_fill
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Total Despesas Operacionais
    ws[f'A{row}'] = '3.01.01.07'
    ws[f'B{row}'] = 'Total Despesas Operacionais'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('despesas_operacionais', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.fill = cabecalho_fill
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 7. RESULTADO OPERACIONAL
    ws[f'A{row}'] = '3.01.01.01.01'
    ws[f'B{row}'] = 'RESULTADO OPERACIONAL'
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('resultado_operacional', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 8. DESPESAS E RECEITAS NÃO OPERACIONAIS
    ws[f'A{row}'] = '3.01.01.08.'
    ws[f'B{row}'] = 'DESPESAS E RECEITAS NÃO OPERACIONAIS'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    row += 1
    
    ws[f'A{row}'] = '3.01.01.08.0001'
    ws[f'B{row}'] = '  Despesas Juros e Multas'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('despesas_financeiras', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Resultado Não Operacional
    ws[f'A{row}'] = '3.01.01.08'
    ws[f'B{row}'] = 'Resultado Não Operacional'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        despesas_fin = dre_por_ano[ano].get('despesas_financeiras', Decimal('0.00'))
        receitas_fin = dre_por_ano[ano].get('receitas_financeiras', Decimal('0.00'))
        resultado = receitas_fin - despesas_fin
        ws[f'{col}{row}'] = formatar_moeda(resultado)
        cell = ws[f'{col}{row}']
        cell.fill = cabecalho_fill
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        if resultado < 0:
            cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 9. RESULTADO ANTES DO IMPOSTO DE RENDA (LAIR)
    ws[f'A{row}'] = '3.01.01.09'
    ws[f'B{row}'] = 'RESULTADO ANTES DO IMPOSTO DE RENDA (LAIR)'
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('lair', dre_por_ano[ano].get('resultado_antes_ir', Decimal('0.00')))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.font = negrito_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 10. PROVISÃO DE IMPOSTOS
    ws[f'A{row}'] = '3.01.01.10.'
    ws[f'B{row}'] = 'PROVISÃO DE IMPOSTOS'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    row += 1
    
    # Usar "IR" para pessoa física, "IRPJ" para pessoa jurídica
    nome_imposto = 'IR' if is_pessoa_fisica else 'IRPJ'
    
    ws[f'A{row}'] = '3.01.01.10.0001'
    ws[f'B{row}'] = f'  {nome_imposto}'
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('irpj', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Arial', size=9, color='FF0000')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # Total Provisão de Impostos
    ws[f'A{row}'] = '3.01.01.10'
    ws[f'B{row}'] = 'Total Provisão de Impostos'
    ws[f'A{row}'].fill = cabecalho_fill
    ws[f'B{row}'].fill = cabecalho_fill
    ws[f'A{row}'].font = negrito_font
    ws[f'B{row}'].font = negrito_font
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('irpj', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(-valor) if valor > 0 else formatar_moeda(Decimal('0.00'))
        cell = ws[f'{col}{row}']
        cell.fill = cabecalho_fill
        cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    row += 1
    
    # 11. RESULTADO LÍQUIDO DO EXERCÍCIO
    ws[f'A{row}'] = '3.01'
    ws[f'B{row}'] = 'RESULTADO LÍQUIDO DO EXERCÍCIO'
    ws[f'A{row}'].fill = subtitulo_fill
    ws[f'B{row}'].fill = subtitulo_fill
    ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    col_idx = 3
    for ano in anos:
        col = get_column_letter(col_idx)
        valor = dre_por_ano[ano].get('resultado_liquido', Decimal('0.00'))
        ws[f'{col}{row}'] = formatar_moeda(valor)
        cell = ws[f'{col}{row}']
        cell.fill = subtitulo_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        col_idx += 1
    
    # Aplicar bordas em todas as células
    for r in range(5, row + 1):
        for col_idx in range(1, num_cols + 1):
            col = get_column_letter(col_idx)
            cell = ws[f'{col}{r}']
            if not cell.border:
                cell.border = border
            if not cell.font:
                cell.font = normal_font
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    for col_idx in range(3, num_cols + 1):
        col = get_column_letter(col_idx)
        ws.column_dimensions[col].width = 18


def exportar_dre_multi_anos_excel(dre_por_ano, propriedades_selecionadas, produtor):
    """
    Exporta DRE com múltiplos anos lado a lado para Excel
    Verifica se é pessoa física para usar "IR" ao invés de "IRPJ"
    """
    # Verificar se é pessoa física
    is_pessoa_fisica = verificar_pessoa_fisica(produtor)
    
    # Ajustar CSLL para zero se for pessoa física
    if is_pessoa_fisica:
        for ano in dre_por_ano:
            dre_por_ano[ano]['csll'] = Decimal('0.00')
            dre_por_ano[ano]['total_impostos'] = dre_por_ano[ano].get('irpj', Decimal('0.00'))
    
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remover sheet padrão
    
    # Criar planilha DRE comparativo (passa is_pessoa_fisica)
    criar_planilha_dre_multi_anos(workbook, dre_por_ano, produtor, propriedades_selecionadas, is_pessoa_fisica)
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    anos_str = '_'.join(map(str, sorted(dre_por_ano.keys())))
    filename = f"DRE_Comparativo_{produtor.nome.replace(' ', '_')}_{anos_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response

