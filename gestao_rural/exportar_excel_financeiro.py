# -*- coding: utf-8 -*-
"""
Módulo para exportar relatórios financeiros (DRE e Balanço) para Excel
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime


def formatar_moeda(valor):
    """Formata valor como moeda brasileira"""
    if valor is None:
        valor = Decimal('0.00')
    if isinstance(valor, Decimal):
        valor = float(valor)
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def verificar_pessoa_fisica(produtor):
    """Verifica se o produtor é pessoa física (CPF) ou jurídica (CNPJ)"""
    if not produtor or not produtor.cpf_cnpj:
        return False
    cpf_cnpj_limpo = produtor.cpf_cnpj.replace(".", "").replace("-", "").replace("/", "")
    return len(cpf_cnpj_limpo) == 11  # CPF tem 11 dígitos


def criar_planilha_dre(workbook, dre_data, produtor, ano):
    """
    Cria planilha de DRE no workbook
    """
    ws = ws = workbook.create_sheet("DRE", 0)
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    subtitulo_font = Font(name='Arial', size=12, bold=True)
    cabecalho_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    valor_font = Font(name='Arial', size=10)
    
    titulo_fill = PatternFill(start_color='1a5490', end_color='1a5490', fill_type='solid')
    subtitulo_fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    cabecalho_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:C1')
    cell = ws['A1']
    cell.value = f"DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO - {ano}"
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    ws.merge_cells('A2:C2')
    cell = ws['A2']
    cell.value = f"{produtor.nome.upper()}"
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # Cabeçalho da tabela
    row = 4
    ws['A4'] = 'CÓDIGO'
    ws['B4'] = 'DESCRIÇÃO'
    ws['C4'] = 'VALOR (R$)'
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}4']
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal='left' if col == 'B' else 'center', vertical='center')
        cell.border = border
    
    # Dados do DRE
    row = 5
    
    # 1. RECEITA BRUTA
    ws[f'A{row}'] = '3.01.01.01.01'
    ws[f'B{row}'] = 'RECEITA BRUTA DE VENDAS'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('receita_bruta', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 2. DEDUÇÕES
    ws[f'A{row}'] = '3.01.01.01.02'
    ws[f'B{row}'] = 'DEDUÇÕES DA RECEITA BRUTA'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('total_deducoes', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    ws[f'A{row}'] = '3.01.01.01.02.0004'
    ws[f'B{row}'] = '  Funrural s/Vendas'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('funrural_vendas', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    ws[f'A{row}'] = '3.01.01.01.02.0005'
    ws[f'B{row}'] = '  ICMS s/Vendas'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('icms_vendas', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    ws[f'A{row}'] = '3.01.01.01.02.0006'
    ws[f'B{row}'] = '  Outros Impostos s/Vendas'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('outros_impostos_vendas', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 3. RECEITA LÍQUIDA
    ws[f'A{row}'] = '3.01.01.01.03'
    ws[f'B{row}'] = 'RECEITA LÍQUIDA'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('receita_liquida', 0))
    ws[f'C{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 4. CPV
    ws[f'A{row}'] = '3.01.01.01.03.'
    ws[f'B{row}'] = 'CUSTOS MERCADORIAS VENDIDAS'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('cpv', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 5. LUCRO BRUTO
    ws[f'A{row}'] = '3.01.01.01.04'
    ws[f'B{row}'] = 'LUCRO BRUTO'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('lucro_bruto', 0))
    ws[f'C{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 6. DESPESAS OPERACIONAIS
    ws[f'A{row}'] = '3.01.01.07.'
    ws[f'B{row}'] = 'DESPESAS OPERACIONAIS'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('despesas_operacionais', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 7. RESULTADO OPERACIONAL
    ws[f'A{row}'] = '3.01.01.01.01'
    ws[f'B{row}'] = 'RESULTADO OPERACIONAL'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('resultado_operacional', 0))
    ws[f'C{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 8. RESULTADO NÃO OPERACIONAL
    ws[f'A{row}'] = '3.01.01.08.'
    ws[f'B{row}'] = 'DESPESAS E RECEITAS NÃO OPERACIONAIS'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('resultado_nao_operacional', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 9. LAIR
    ws[f'A{row}'] = '3.01.01.09'
    ws[f'B{row}'] = 'RESULTADO ANTES DO IMPOSTO DE RENDA (LAIR)'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('resultado_antes_ir', 0))
    ws[f'C{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 10. IMPOSTOS
    # Verificar se é pessoa física
    is_pessoa_fisica = verificar_pessoa_fisica(produtor)
    nome_imposto = 'IR' if is_pessoa_fisica else 'IRPJ'
    
    ws[f'A{row}'] = '3.01.01.10.'
    ws[f'B{row}'] = 'PROVISÃO DE IMPOSTOS'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('total_impostos', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    ws[f'A{row}'] = '3.01.01.10.0001'
    ws[f'B{row}'] = f'  {nome_imposto}'
    ws[f'C{row}'] = formatar_moeda(-dre_data.get('irpj', 0))
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    row += 1
    
    # 11. RESULTADO LÍQUIDO
    ws[f'A{row}'] = '3.01'
    ws[f'B{row}'] = 'RESULTADO LÍQUIDO DO EXERCÍCIO'
    ws[f'C{row}'] = formatar_moeda(dre_data.get('resultado_liquido', 0))
    ws[f'C{row}'].font = Font(name='Arial', size=11, bold=True)
    ws[f'C{row}'].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    
    # Aplicar bordas
    for r in range(4, row + 1):
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{r}']
            cell.border = border
            if r > 4:
                cell.font = normal_font
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20


def criar_planilha_balanco(workbook, balanco_data, produtor, ano):
    """
    Cria planilha de Balanço Patrimonial no workbook
    """
    ws = workbook.create_sheet("Balanço Patrimonial")
    
    # Estilos (mesmos do DRE)
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    subtitulo_font = Font(name='Arial', size=12, bold=True)
    cabecalho_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    titulo_fill = PatternFill(start_color='1a5490', end_color='1a5490', fill_type='solid')
    subtitulo_fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    cabecalho_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = f"BALANÇO PATRIMONIAL - {ano}"
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    ws.merge_cells('A2:D2')
    cell = ws['A2']
    cell.value = f"{produtor.nome.upper()}"
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # Cabeçalho
    row = 4
    ws['A4'] = 'CÓDIGO'
    ws['B4'] = 'DESCRIÇÃO'
    ws['C4'] = 'ATIVO'
    ws['D4'] = 'PASSIVO + PL'
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}4']
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal='left' if col == 'B' else 'center', vertical='center')
        cell.border = border
    
    # ATIVO
    row = 5
    ws[f'A{row}'] = 'ATIVO'
    ws[f'B{row}'] = 'ATIVO TOTAL'
    ws[f'C{row}'] = formatar_moeda(balanco_data.get('ativo_total', 0))
    ws[f'C{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    
    # PASSIVO + PATRIMÔNIO LÍQUIDO
    ws[f'D{row}'] = formatar_moeda(balanco_data.get('passivo_pl_total', 0))
    ws[f'D{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    
    # Aplicar bordas
    for r in range(4, row + 1):
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{r}']
            cell.border = border
            if r > 4:
                cell.font = normal_font
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def criar_planilha_faturamento(workbook, propriedades, ano):
    """
    Cria planilha de Faturamento Contábil (receitas mensais)
    """
    ws = workbook.create_sheet("Faturamento Contábil")
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    cabecalho_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    titulo_fill = PatternFill(start_color='1a5490', end_color='1a5490', fill_type='solid')
    cabecalho_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = f"FATURAMENTO CONTÁBIL - {ano}"
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    meses = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez', 'Total']
    ws['A3'] = 'Propriedade'
    for idx, mes in enumerate(meses, start=1):
        col = get_column_letter(idx + 1)
        ws[f'{col}3'] = mes
        cell = ws[f'{col}3']
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    from gestao_rural.models_financeiro import LancamentoFinanceiro, CategoriaFinanceira
    
    categoria_receita, _ = CategoriaFinanceira.objects.get_or_create(
        nome='Vendas de Gado',
        tipo=CategoriaFinanceira.TIPO_RECEITA
    )
    
    row = 4
    for propriedade in propriedades:
        ws[f'A{row}'] = propriedade.nome_propriedade
        total_prop = Decimal('0.00')
        
        for mes in range(1, 13):
            receitas_mes = LancamentoFinanceiro.objects.filter(
                propriedade=propriedade,
                categoria=categoria_receita,
                data_competencia__year=ano,
                data_competencia__month=mes,
                tipo=CategoriaFinanceira.TIPO_RECEITA,
                status=LancamentoFinanceiro.STATUS_QUITADO
            )
            valor_mes = sum(r.valor for r in receitas_mes)
            total_prop += valor_mes
            
            col = get_column_letter(mes + 1)
            ws[f'{col}{row}'] = formatar_moeda(valor_mes)
            ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
        
        # Total
        col_total = get_column_letter(14)
        ws[f'{col_total}{row}'] = formatar_moeda(total_prop)
        ws[f'{col_total}{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'{col_total}{row}'].alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Total geral
    ws[f'A{row}'] = 'TOTAL GERAL'
    ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    
    for mes in range(1, 13):
        col = get_column_letter(mes + 1)
        # Calcular total da coluna
        total_col = Decimal('0.00')
        for r in range(4, row):
            cell_val = ws[f'{col}{r}'].value
            if cell_val:
                # Extrair valor da string formatada
                try:
                    valor_str = cell_val.replace('R$', '').replace('.', '').replace(',', '.')
                    total_col += Decimal(valor_str)
                except:
                    pass
        ws[f'{col}{row}'] = formatar_moeda(total_col)
        ws[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
    
    # Aplicar bordas
    for r in range(3, row + 1):
        for col_idx in range(1, 15):
            col = get_column_letter(col_idx)
            cell = ws[f'{col}{r}']
            cell.border = border
            if r > 3:
                cell.font = normal_font
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, 15):
        col = get_column_letter(col_idx)
        ws.column_dimensions[col].width = 15


def exportar_dre_balanco_excel(dre_data, balanco_data, propriedades, produtor, ano):
    """
    Exporta DRE e Balanço para Excel
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remover sheet padrão
    
    # Criar planilhas
    criar_planilha_dre(workbook, dre_data, produtor, ano)
    criar_planilha_balanco(workbook, balanco_data, produtor, ano)
    criar_planilha_faturamento(workbook, propriedades, ano)
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"DRE_Balanco_{produtor.nome.replace(' ', '_')}_{ano}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response

