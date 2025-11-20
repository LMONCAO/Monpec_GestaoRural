# -*- coding: utf-8 -*-
"""
Gerador Dinâmico de Relatórios
Gera relatórios baseados em configurações customizadas
"""

from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
from decimal import Decimal
from datetime import datetime, date
import json

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class GeradorRelatoriosDinamico:
    """Gera relatórios dinamicamente baseado em configurações"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._criar_estilos()
    
    def _criar_estilos(self):
        """Cria estilos customizados para PDFs"""
        self.styles.add(ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            alignment=TA_CENTER,
            spaceAfter=12
        ))
        
        self.styles.add(ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c5282'),
            spaceBefore=12,
            spaceAfter=8
        ))
    
    def gerar_dados(self, relatorio, propriedade, filtros_adicionais=None):
        """
        Gera os dados do relatório baseado na configuração
        Retorna um dicionário com os dados processados
        """
        filtros_adicionais = filtros_adicionais or {}
        
        # Combinar filtros do relatório com filtros adicionais
        filtros = relatorio.get_filtros_dict().copy()
        filtros.update(filtros_adicionais)
        
        # Obter dados baseado no módulo
        dados = self._obter_dados_modulo(relatorio.modulo, propriedade, filtros)
        
        # Aplicar filtros
        dados = self._aplicar_filtros(dados, filtros)
        
        # Aplicar agrupamentos
        if relatorio.get_agrupamentos_list():
            dados = self._aplicar_agrupamentos(dados, relatorio.get_agrupamentos_list())
        
        # Aplicar ordenação
        if relatorio.get_ordenacao_list():
            dados = self._aplicar_ordenacao(dados, relatorio.get_ordenacao_list())
        
        # Selecionar apenas campos escolhidos
        campos_selecionados = relatorio.get_campos_selecionados_list()
        if campos_selecionados:
            dados = self._selecionar_campos(dados, campos_selecionados)
        
        return {
            'dados': dados,
            'total_registros': len(dados) if isinstance(dados, list) else 0,
            'relatorio': relatorio,
            'propriedade': propriedade,
        }
    
    def gerar_pdf(self, relatorio, propriedade, filtros_adicionais=None):
        """Gera PDF do relatório"""
        dados_dict = self.gerar_dados(relatorio, propriedade, filtros_adicionais)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{relatorio.id}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm)
        story = []
        
        # Título
        story.append(Paragraph(relatorio.nome, self.styles['CustomTitle']))
        if relatorio.descricao:
            story.append(Paragraph(relatorio.descricao, self.styles['Normal']))
        story.append(Paragraph(f"Propriedade: {propriedade.nome_propriedade}", self.styles['Normal']))
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self.styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        
        # Tabela de dados
        if dados_dict['dados']:
            dados = dados_dict['dados']
            campos = relatorio.get_campos_selecionados_list()
            
            if campos:
                headers = [campo.get('label', campo.get('nome', '')) for campo in campos]
            else:
                # Se não há campos selecionados, usar todas as chaves do primeiro registro
                headers = list(dados[0].keys()) if dados else []
            
            table_data = [headers]
            
            for registro in dados:
                linha = []
                if campos:
                    for campo in campos:
                        valor = registro.get(campo.get('nome', ''), '')
                        linha.append(self._formatar_valor(valor, campo.get('tipo', 'texto')))
                else:
                    linha = [self._formatar_valor(v, 'texto') for v in registro.values()]
                table_data.append(linha)
            
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(table)
        
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Total de registros: {dados_dict['total_registros']}", self.styles['Normal']))
        
        doc.build(story)
        return response
    
    def gerar_excel(self, relatorio, propriedade, filtros_adicionais=None):
        """Gera Excel do relatório"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está disponível")
        
        dados_dict = self.gerar_dados(relatorio, propriedade, filtros_adicionais)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Título
        ws['A1'] = relatorio.nome
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Propriedade: {propriedade.nome_propriedade}"
        ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Cabeçalhos
        dados = dados_dict['dados']
        campos = relatorio.get_campos_selecionados_list()
        
        if campos:
            headers = [campo.get('label', campo.get('nome', '')) for campo in campos]
        else:
            headers = list(dados[0].keys()) if dados else []
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dados
        for row_idx, registro in enumerate(dados, start=6):
            if campos:
                for col_idx, campo in enumerate(campos, start=1):
                    valor = registro.get(campo.get('nome', ''), '')
                    ws.cell(row=row_idx, column=col_idx, value=valor)
            else:
                for col_idx, valor in enumerate(registro.values(), start=1):
                    ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_{relatorio.id}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
    
    def _obter_dados_modulo(self, modulo, propriedade, filtros):
        """Obtém dados do módulo especificado"""
        dados = []
        
        if modulo == 'PECUARIA':
            from .models import InventarioRebanho
            queryset = InventarioRebanho.objects.filter(propriedade=propriedade)
            
            if filtros.get('data_inicio'):
                queryset = queryset.filter(data_inventario__gte=filtros['data_inicio'])
            if filtros.get('data_fim'):
                queryset = queryset.filter(data_inventario__lte=filtros['data_fim'])
            
            for item in queryset.select_related('categoria'):
                dados.append({
                    'numero_brinco': '-',
                    'categoria': item.categoria.nome if item.categoria else '',
                    'quantidade': item.quantidade,
                    'valor_por_cabeca': item.valor_por_cabeca or 0,
                    'valor_total': item.valor_total or 0,
                    'data_inventario': item.data_inventario,
                })
        
        elif modulo == 'IATF':
            try:
                from .models_iatf_completo import IATFIndividual
                queryset = IATFIndividual.objects.filter(propriedade=propriedade)
                
                if filtros.get('data_inicio'):
                    queryset = queryset.filter(data_iatf__gte=filtros['data_inicio'])
                if filtros.get('data_fim'):
                    queryset = queryset.filter(data_iatf__lte=filtros['data_fim'])
                
                for item in queryset.select_related('animal_individual', 'protocolo', 'touro_semen'):
                    dados.append({
                        'animal': item.animal_individual.numero_brinco if item.animal_individual else '-',
                        'protocolo': item.protocolo.nome if item.protocolo else '-',
                        'data_iatf': item.data_iatf,
                        'resultado': item.get_resultado_display(),
                        'taxa_prenhez': '-',
                        'custo_total': item.custo_total or 0,
                    })
            except ImportError:
                pass
        
        # Adicionar mais módulos conforme necessário
        
        return dados
    
    def _aplicar_filtros(self, dados, filtros):
        """Aplica filtros aos dados"""
        # Implementar lógica de filtros
        # Por enquanto, retorna dados sem filtros adicionais
        return dados
    
    def _aplicar_agrupamentos(self, dados, agrupamentos):
        """Aplica agrupamentos aos dados"""
        # Implementar lógica de agrupamentos
        return dados
    
    def _aplicar_ordenacao(self, dados, ordenacao):
        """Aplica ordenação aos dados"""
        # Implementar lógica de ordenação
        return dados
    
    def _selecionar_campos(self, dados, campos_selecionados):
        """Seleciona apenas os campos especificados"""
        campos_nomes = [c.get('nome') for c in campos_selecionados]
        dados_filtrados = []
        
        for registro in dados:
            novo_registro = {k: v for k, v in registro.items() if k in campos_nomes}
            dados_filtrados.append(novo_registro)
        
        return dados_filtrados
    
    def _formatar_valor(self, valor, tipo):
        """Formata valor baseado no tipo"""
        if valor is None:
            return '-'
        
        if tipo == 'moeda':
            if isinstance(valor, (int, float, Decimal)):
                return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            return str(valor)
        elif tipo == 'percentual':
            if isinstance(valor, (int, float, Decimal)):
                return f"{float(valor):.2f}%".replace('.', ',')
            return str(valor)
        elif tipo == 'data':
            if isinstance(valor, (date, datetime)):
                return valor.strftime('%d/%m/%Y')
            return str(valor)
        elif tipo == 'numero':
            if isinstance(valor, (int, float, Decimal)):
                return f"{float(valor):,.0f}".replace(',', '.')
            return str(valor)
        else:
            return str(valor)






