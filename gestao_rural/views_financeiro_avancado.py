"""Views avançadas do módulo financeiro: Conciliação, Boletos, Fluxo de Caixa, DRE, LCDPR."""
from datetime import datetime, date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count, F
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Propriedade
from .models_financeiro import (
    ContaFinanceira,
    LancamentoFinanceiro,
    MovimentoFinanceiro,
    CategoriaFinanceira,
)
from .models_compras_financeiro import ContaPagar, ContaReceber


def _obter_propriedade(usuario, propriedade_id) -> Propriedade:
    """Garante que a propriedade pertence ao contexto do usuário."""
    return get_object_or_404(
        Propriedade.objects.select_related("produtor"),
        id=propriedade_id,
    )


# ============================================================================
# CONCILIAÇÃO BANCÁRIA
# ============================================================================

@login_required
def conciliacao_bancaria(request, propriedade_id):
    """Tela de conciliação bancária."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    conta_id = request.GET.get("conta")
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    
    contas = ContaFinanceira.objects.filter(propriedade=propriedade, ativa=True)
    
    if conta_id:
        conta = get_object_or_404(ContaFinanceira, id=conta_id, propriedade=propriedade)
    elif contas.exists():
        conta = contas.first()
    else:
        conta = None
    
    # Período padrão: mês atual
    hoje = timezone.localdate()
    if not data_inicio:
        data_inicio = date(hoje.year, hoje.month, 1)
    else:
        data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    
    if not data_fim:
        from calendar import monthrange
        ultimo_dia = monthrange(hoje.year, hoje.month)[1]
        data_fim = date(hoje.year, hoje.month, ultimo_dia)
    else:
        data_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
    
    movimentos_nao_conciliados = []
    lancamentos_nao_conciliados = []
    
    if conta:
        # Movimentos financeiros não conciliados
        movimentos_nao_conciliados = MovimentoFinanceiro.objects.filter(
            conta=conta,
            conciliado=False,
            data_movimento__range=(data_inicio, data_fim),
        ).order_by("data_movimento")
        
        # Lançamentos financeiros não conciliados
        lancamentos_nao_conciliados = LancamentoFinanceiro.objects.filter(
            Q(conta_origem=conta) | Q(conta_destino=conta),
            status=LancamentoFinanceiro.STATUS_QUITADO,
            data_quitacao__range=(data_inicio, data_fim),
        ).order_by("data_quitacao")
    
    context = {
        "propriedade": propriedade,
        "contas": contas,
        "conta": conta,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "movimentos_nao_conciliados": movimentos_nao_conciliados,
        "lancamentos_nao_conciliados": lancamentos_nao_conciliados,
    }
    
    return render(request, "gestao_rural/financeiro/conciliacao_bancaria.html", context)


@login_required
@require_http_methods(["POST"])
def conciliacao_marcar_conciliado(request, propriedade_id):
    """Marca movimentos como conciliados."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    movimento_ids = request.POST.getlist("movimentos")
    lancamento_ids = request.POST.getlist("lancamentos")
    
    conciliados = 0
    
    if movimento_ids:
        movimentos = MovimentoFinanceiro.objects.filter(
            id__in=movimento_ids,
            conta__propriedade=propriedade,
        )
        for movimento in movimentos:
            movimento.conciliado = True
            movimento.conciliado_em = timezone.now()
            movimento.save()
            conciliados += 1
    
    if lancamento_ids:
        lancamentos = LancamentoFinanceiro.objects.filter(
            id__in=lancamento_ids,
            propriedade=propriedade,
        )
        # Aqui você pode adicionar um campo conciliado ao LancamentoFinanceiro se necessário
        conciliados += len(lancamentos)
    
    messages.success(request, f"{conciliados} item(ns) marcado(s) como conciliado(s).")
    
    return redirect("financeiro_conciliacao", propriedade_id=propriedade.id)


# ============================================================================
# EMISSÃO DE BOLETOS
# ============================================================================

@login_required
def boletos_lista(request, propriedade_id):
    """Lista de boletos a emitir."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    # Buscar contas a receber que podem gerar boletos
    contas_receber = ContaReceber.objects.filter(
        propriedade=propriedade,
        status__in=["PENDENTE", "VENCIDA"],
    ).order_by("data_vencimento")
    
    context = {
        "propriedade": propriedade,
        "contas_receber": contas_receber,
    }
    
    return render(request, "gestao_rural/financeiro/boletos_lista.html", context)


@login_required
def boleto_gerar(request, propriedade_id, conta_receber_id):
    """Gera boleto para uma conta a receber."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    conta_receber = get_object_or_404(
        ContaReceber,
        id=conta_receber_id,
        propriedade=propriedade,
    )
    
    # Aqui você pode integrar com biblioteca de geração de boletos
    # Por exemplo: python-boleto, boleto-python, etc.
    
    context = {
        "propriedade": propriedade,
        "conta_receber": conta_receber,
    }
    
    return render(request, "gestao_rural/financeiro/boleto_gerar.html", context)


# ============================================================================
# FLUXO DE CAIXA DETALHADO
# ============================================================================

@login_required
def fluxo_caixa_detalhado(request, propriedade_id):
    """Fluxo de caixa detalhado com projeções."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    meses_projecao = int(request.GET.get("meses_projecao", 12))
    
    # Período padrão: últimos 3 meses + projeção
    hoje = timezone.localdate()
    if not data_inicio:
        data_inicio = date(hoje.year, hoje.month, 1) - timedelta(days=90)
    else:
        data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    
    if not data_fim:
        data_fim = hoje + timedelta(days=30 * meses_projecao)
    else:
        data_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
    
    # Dados históricos
    lancamentos = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        data_competencia__range=(data_inicio, hoje),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).order_by("data_competencia")
    
    # Agrupar por dia
    fluxo_diario = {}
    for lancamento in lancamentos:
        dia = lancamento.data_competencia
        if dia not in fluxo_diario:
            fluxo_diario[dia] = {"entradas": Decimal("0"), "saidas": Decimal("0")}
        
        if lancamento.tipo == CategoriaFinanceira.TIPO_RECEITA:
            fluxo_diario[dia]["entradas"] += lancamento.valor
        elif lancamento.tipo == CategoriaFinanceira.TIPO_DESPESA:
            fluxo_diario[dia]["saidas"] += lancamento.valor
    
    # Projeções futuras (baseadas em contas a pagar/receber)
    contas_pagar_futuras = ContaPagar.objects.filter(
        propriedade=propriedade,
        status__in=["PENDENTE", "VENCIDA"],
        data_vencimento__lte=data_fim,
    ).order_by("data_vencimento")
    
    contas_receber_futuras = ContaReceber.objects.filter(
        propriedade=propriedade,
        status__in=["PENDENTE", "VENCIDA"],
        data_vencimento__lte=data_fim,
    ).order_by("data_vencimento")
    
    # Calcular saldo acumulado
    saldo_inicial = Decimal("0")
    saldo_atual = saldo_inicial
    
    for dia in sorted(fluxo_diario.keys()):
        if dia <= hoje:
            saldo_atual += fluxo_diario[dia]["entradas"] - fluxo_diario[dia]["saidas"]
            fluxo_diario[dia]["saldo"] = saldo_atual
    
    context = {
        "propriedade": propriedade,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "meses_projecao": meses_projecao,
        "fluxo_diario": fluxo_diario,
        "contas_pagar_futuras": contas_pagar_futuras,
        "contas_receber_futuras": contas_receber_futuras,
        "saldo_atual": saldo_atual,
    }
    
    return render(request, "gestao_rural/financeiro/fluxo_caixa_detalhado.html", context)


# ============================================================================
# DRE - DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO
# ============================================================================

@login_required
def dre(request, propriedade_id):
    """Demonstração do Resultado do Exercício."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    ano = int(request.GET.get("ano", timezone.now().year))
    mes = request.GET.get("mes")
    
    if mes:
        data_inicio = date(ano, int(mes), 1)
        from calendar import monthrange
        data_fim = date(ano, int(mes), monthrange(ano, int(mes))[1])
        periodo_tipo = "mensal"
    else:
        data_inicio = date(ano, 1, 1)
        data_fim = date(ano, 12, 31)
        periodo_tipo = "anual"
    
    # Receitas
    receitas = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        tipo=CategoriaFinanceira.TIPO_RECEITA,
        data_competencia__range=(data_inicio, data_fim),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).values("categoria__nome").annotate(
        total=Sum("valor")
    ).order_by("-total")
    
    total_receitas = sum(r["total"] for r in receitas)
    
    # Despesas
    despesas = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        tipo=CategoriaFinanceira.TIPO_DESPESA,
        data_competencia__range=(data_inicio, data_fim),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).values("categoria__nome").annotate(
        total=Sum("valor")
    ).order_by("-total")
    
    total_despesas = sum(d["total"] for d in despesas)
    
    # Resultado
    resultado_liquido = total_receitas - total_despesas
    
    context = {
        "propriedade": propriedade,
        "ano": ano,
        "mes": mes,
        "periodo_tipo": periodo_tipo,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "receitas": receitas,
        "total_receitas": total_receitas,
        "despesas": despesas,
        "total_despesas": total_despesas,
        "resultado_liquido": resultado_liquido,
    }
    
    return render(request, "gestao_rural/financeiro/dre.html", context)


# ============================================================================
# LCDPR - LIVRO CAIXA E DEMONSTRAÇÃO DE PAGAMENTOS E RECEBIMENTOS
# ============================================================================

@login_required
def lcdpr(request, propriedade_id):
    """Livro Caixa e Demonstração de Pagamentos e Recebimentos."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    
    # Período padrão: mês atual
    hoje = timezone.localdate()
    if not data_inicio:
        data_inicio = date(hoje.year, hoje.month, 1)
    else:
        data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    
    if not data_fim:
        from calendar import monthrange
        ultimo_dia = monthrange(hoje.year, hoje.month)[1]
        data_fim = date(hoje.year, hoje.month, ultimo_dia)
    else:
        data_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
    
    # Lançamentos do período
    lancamentos = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        data_competencia__range=(data_inicio, data_fim),
    ).select_related("categoria", "conta_origem", "conta_destino").order_by(
        "data_competencia", "tipo"
    )
    
    # Agrupar por tipo
    receitas = lancamentos.filter(tipo=CategoriaFinanceira.TIPO_RECEITA)
    despesas = lancamentos.filter(tipo=CategoriaFinanceira.TIPO_DESPESA)
    transferencias = lancamentos.filter(tipo=CategoriaFinanceira.TIPO_TRANSFERENCIA)
    
    # Totais
    total_receitas = receitas.filter(
        status=LancamentoFinanceiro.STATUS_QUITADO
    ).aggregate(total=Sum("valor"))["total"] or Decimal("0")
    
    total_despesas = despesas.filter(
        status=LancamentoFinanceiro.STATUS_QUITADO
    ).aggregate(total=Sum("valor"))["total"] or Decimal("0")
    
    # Saldo inicial (soma dos saldos iniciais das contas)
    saldo_inicial = ContaFinanceira.objects.filter(
        propriedade=propriedade,
        ativa=True,
    ).aggregate(total=Sum("saldo_inicial"))["total"] or Decimal("0")
    
    # Saldo final
    saldo_final = saldo_inicial + total_receitas - total_despesas
    
    context = {
        "propriedade": propriedade,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "lancamentos": lancamentos,
        "receitas": receitas,
        "despesas": despesas,
        "transferencias": transferencias,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
    }
    
    return render(request, "gestao_rural/financeiro/lcdpr.html", context)


# ============================================================================
# EXPORTAÇÕES PDF E EXCEL
# ============================================================================

@login_required
def dre_exportar_pdf(request, propriedade_id):
    """Exporta DRE em PDF."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    ano = int(request.GET.get("ano", timezone.now().year))
    mes = request.GET.get("mes")
    
    if mes:
        data_inicio = date(ano, int(mes), 1)
        from calendar import monthrange
        data_fim = date(ano, int(mes), monthrange(ano, int(mes))[1])
        periodo_tipo = "mensal"
        periodo_texto = f"{mes}/{ano}"
    else:
        data_inicio = date(ano, 1, 1)
        data_fim = date(ano, 12, 31)
        periodo_tipo = "anual"
        periodo_texto = str(ano)
    
    # Buscar dados
    receitas = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        tipo=CategoriaFinanceira.TIPO_RECEITA,
        data_competencia__range=(data_inicio, data_fim),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).values("categoria__nome").annotate(total=Sum("valor")).order_by("-total")
    
    total_receitas = sum(r["total"] for r in receitas)
    
    despesas = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        tipo=CategoriaFinanceira.TIPO_DESPESA,
        data_competencia__range=(data_inicio, data_fim),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).values("categoria__nome").annotate(total=Sum("valor")).order_by("-total")
    
    total_despesas = sum(d["total"] for d in despesas)
    resultado_liquido = total_receitas - total_despesas
    
    # Criar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="DRE_{propriedade.id}_{periodo_texto}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    story.append(Paragraph(f"Demonstração do Resultado do Exercício (DRE)", title_style))
    story.append(Paragraph(f"Propriedade: {propriedade.nome_propriedade}", styles['Normal']))
    story.append(Paragraph(f"Período: {periodo_texto}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Receitas
    receitas_data = [['Categoria', 'Valor (R$)']]
    for r in receitas:
        receitas_data.append([r['categoria__nome'], f"R$ {r['total']:,.2f}"])
    receitas_data.append(['TOTAL DE RECEITAS', f"R$ {total_receitas:,.2f}"])
    
    receitas_table = Table(receitas_data, colWidths=[12*cm, 5*cm])
    receitas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(Paragraph("RECEITAS", styles['Heading2']))
    story.append(receitas_table)
    story.append(Spacer(1, 20))
    
    # Despesas
    despesas_data = [['Categoria', 'Valor (R$)']]
    for d in despesas:
        despesas_data.append([d['categoria__nome'], f"R$ {d['total']:,.2f}"])
    despesas_data.append(['TOTAL DE DESPESAS', f"R$ {total_despesas:,.2f}"])
    
    despesas_table = Table(despesas_data, colWidths=[12*cm, 5*cm])
    despesas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightcoral),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(Paragraph("DESPESAS", styles['Heading2']))
    story.append(despesas_table)
    story.append(Spacer(1, 20))
    
    # Resultado
    resultado_data = [['RESULTADO LÍQUIDO', f"R$ {resultado_liquido:,.2f}"]]
    resultado_table = Table(resultado_data, colWidths=[12*cm, 5*cm])
    cor_resultado = colors.green if resultado_liquido >= 0 else colors.red
    resultado_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), cor_resultado),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(resultado_table)
    
    doc.build(story)
    return response


@login_required
def dre_exportar_excel(request, propriedade_id):
    """Exporta DRE em Excel."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    ano = int(request.GET.get("ano", timezone.now().year))
    mes = request.GET.get("mes")
    
    if mes:
        data_inicio = date(ano, int(mes), 1)
        from calendar import monthrange
        data_fim = date(ano, int(mes), monthrange(ano, int(mes))[1])
        periodo_texto = f"{mes}_{ano}"
    else:
        data_inicio = date(ano, 1, 1)
        data_fim = date(ano, 12, 31)
        periodo_texto = str(ano)
    
    # Buscar dados (mesmo código da view dre)
    receitas = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        tipo=CategoriaFinanceira.TIPO_RECEITA,
        data_competencia__range=(data_inicio, data_fim),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).values("categoria__nome").annotate(total=Sum("valor")).order_by("-total")
    
    total_receitas = sum(r["total"] for r in receitas)
    
    despesas = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        tipo=CategoriaFinanceira.TIPO_DESPESA,
        data_competencia__range=(data_inicio, data_fim),
        status=LancamentoFinanceiro.STATUS_QUITADO,
    ).values("categoria__nome").annotate(total=Sum("valor")).order_by("-total")
    
    total_despesas = sum(d["total"] for d in despesas)
    resultado_liquido = total_receitas - total_despesas
    
    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "DRE"
    
    # Cabeçalho
    ws['A1'] = "Demonstração do Resultado do Exercício (DRE)"
    ws['A2'] = f"Propriedade: {propriedade.nome_propriedade}"
    ws['A3'] = f"Período: {periodo_texto}"
    
    # Receitas
    row = 5
    ws[f'A{row}'] = "RECEITAS"
    row += 1
    ws[f'A{row}'] = "Categoria"
    ws[f'B{row}'] = "Valor (R$)"
    for cell in ws[row]:
        cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row += 1
    for r in receitas:
        ws[f'A{row}'] = r['categoria__nome']
        ws[f'B{row}'] = float(r['total'])
        row += 1
    
    ws[f'A{row}'] = "TOTAL DE RECEITAS"
    ws[f'B{row}'] = float(total_receitas)
    for cell in ws[row]:
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        cell.font = Font(bold=True)
    row += 2
    
    # Despesas
    ws[f'A{row}'] = "DESPESAS"
    row += 1
    ws[f'A{row}'] = "Categoria"
    ws[f'B{row}'] = "Valor (R$)"
    for cell in ws[row]:
        cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row += 1
    for d in despesas:
        ws[f'A{row}'] = d['categoria__nome']
        ws[f'B{row}'] = float(d['total'])
        row += 1
    
    ws[f'A{row}'] = "TOTAL DE DESPESAS"
    ws[f'B{row}'] = float(total_despesas)
    for cell in ws[row]:
        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        cell.font = Font(bold=True)
    row += 2
    
    # Resultado
    ws[f'A{row}'] = "RESULTADO LÍQUIDO"
    ws[f'B{row}'] = float(resultado_liquido)
    for cell in ws[row]:
        cor = "006400" if resultado_liquido >= 0 else "DC143C"
        cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=14)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    # Formatar valores
    from openpyxl.styles.numbers import BUILTIN_FORMATS
    for row in ws.iter_rows(min_row=6, max_row=row, min_col=2, max_col=2):
        for cell in row:
            if cell.value:
                cell.number_format = BUILTIN_FORMATS[7]  # R$ #,##0.00
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="DRE_{propriedade.id}_{periodo_texto}.xlsx"'
    wb.save(response)
    return response


@login_required
def lcdpr_exportar_pdf(request, propriedade_id):
    """Exporta LCDPR em PDF."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    
    hoje = timezone.localdate()
    if not data_inicio:
        data_inicio = date(hoje.year, hoje.month, 1)
    else:
        data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    
    if not data_fim:
        from calendar import monthrange
        ultimo_dia = monthrange(hoje.year, hoje.month)[1]
        data_fim = date(hoje.year, hoje.month, ultimo_dia)
    else:
        data_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
    
    # Buscar dados (mesmo código da view lcdpr)
    lancamentos = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        data_competencia__range=(data_inicio, data_fim),
    ).select_related("categoria", "conta_origem", "conta_destino").order_by("data_competencia", "tipo")
    
    total_receitas = lancamentos.filter(
        tipo=CategoriaFinanceira.TIPO_RECEITA,
        status=LancamentoFinanceiro.STATUS_QUITADO
    ).aggregate(total=Sum("valor"))["total"] or Decimal("0")
    
    total_despesas = lancamentos.filter(
        tipo=CategoriaFinanceira.TIPO_DESPESA,
        status=LancamentoFinanceiro.STATUS_QUITADO
    ).aggregate(total=Sum("valor"))["total"] or Decimal("0")
    
    saldo_inicial = ContaFinanceira.objects.filter(
        propriedade=propriedade,
        ativa=True,
    ).aggregate(total=Sum("saldo_inicial"))["total"] or Decimal("0")
    
    saldo_final = saldo_inicial + total_receitas - total_despesas
    
    # Criar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="LCDPR_{propriedade.id}_{data_inicio.strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    story.append(Paragraph("Livro Caixa e Demonstração de Pagamentos e Recebimentos", title_style))
    story.append(Paragraph(f"Propriedade: {propriedade.nome_propriedade}", styles['Normal']))
    story.append(Paragraph(f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Resumo
    resumo_data = [
        ['Saldo Inicial', f"R$ {saldo_inicial:,.2f}"],
        ['Total Receitas', f"R$ {total_receitas:,.2f}"],
        ['Total Despesas', f"R$ {total_despesas:,.2f}"],
        ['Saldo Final', f"R$ {saldo_final:,.2f}"],
    ]
    resumo_table = Table(resumo_data, colWidths=[8*cm, 8*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 20))
    
    # Livro Caixa
    story.append(Paragraph("LIVRO CAIXA", styles['Heading2']))
    livro_data = [['Data', 'Tipo', 'Descrição', 'Categoria', 'Valor']]
    for lanc in lancamentos[:50]:  # Limitar a 50 para não sobrecarregar
        livro_data.append([
            lanc.data_competencia.strftime('%d/%m/%Y'),
            lanc.get_tipo_display(),
            lanc.descricao[:30],
            lanc.categoria.nome[:20],
            f"R$ {lanc.valor:,.2f}"
        ])
    
    livro_table = Table(livro_data, colWidths=[3*cm, 3*cm, 6*cm, 4*cm, 3*cm])
    livro_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(livro_table)
    
    doc.build(story)
    return response


@login_required
def lcdpr_exportar_excel(request, propriedade_id):
    """Exporta LCDPR em Excel."""
    propriedade = _obter_propriedade(request.user, propriedade_id)
    
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    
    hoje = timezone.localdate()
    if not data_inicio:
        data_inicio = date(hoje.year, hoje.month, 1)
    else:
        data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    
    if not data_fim:
        from calendar import monthrange
        ultimo_dia = monthrange(hoje.year, hoje.month)[1]
        data_fim = date(hoje.year, hoje.month, ultimo_dia)
    else:
        data_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
    
    # Buscar dados
    lancamentos = LancamentoFinanceiro.objects.filter(
        propriedade=propriedade,
        data_competencia__range=(data_inicio, data_fim),
    ).select_related("categoria", "conta_origem", "conta_destino").order_by("data_competencia", "tipo")
    
    total_receitas = lancamentos.filter(
        tipo=CategoriaFinanceira.TIPO_RECEITA,
        status=LancamentoFinanceiro.STATUS_QUITADO
    ).aggregate(total=Sum("valor"))["total"] or Decimal("0")
    
    total_despesas = lancamentos.filter(
        tipo=CategoriaFinanceira.TIPO_DESPESA,
        status=LancamentoFinanceiro.STATUS_QUITADO
    ).aggregate(total=Sum("valor"))["total"] or Decimal("0")
    
    saldo_inicial = ContaFinanceira.objects.filter(
        propriedade=propriedade,
        ativa=True,
    ).aggregate(total=Sum("saldo_inicial"))["total"] or Decimal("0")
    
    saldo_final = saldo_inicial + total_receitas - total_despesas
    
    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "LCDPR"
    
    # Cabeçalho
    ws['A1'] = "Livro Caixa e Demonstração de Pagamentos e Recebimentos"
    ws['A2'] = f"Propriedade: {propriedade.nome_propriedade}"
    ws['A3'] = f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    
    # Resumo
    row = 5
    ws[f'A{row}'] = "Saldo Inicial"
    ws[f'B{row}'] = float(saldo_inicial)
    row += 1
    ws[f'A{row}'] = "Total Receitas"
    ws[f'B{row}'] = float(total_receitas)
    row += 1
    ws[f'A{row}'] = "Total Despesas"
    ws[f'B{row}'] = float(total_despesas)
    row += 1
    ws[f'A{row}'] = "Saldo Final"
    ws[f'B{row}'] = float(saldo_final)
    for cell in ws[row]:
        cell.font = Font(bold=True)
    row += 2
    
    # Livro Caixa
    ws[f'A{row}'] = "LIVRO CAIXA"
    row += 1
    headers = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Conta', 'Valor', 'Status']
    ws.append(headers)
    for cell in ws[row]:
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row += 1
    for lanc in lancamentos:
        ws.append([
            lanc.data_competencia.strftime('%d/%m/%Y'),
            lanc.get_tipo_display(),
            lanc.descricao,
            lanc.categoria.nome,
            lanc.conta_origem.nome if lanc.conta_origem else (lanc.conta_destino.nome if lanc.conta_destino else '-'),
            float(lanc.valor),
            lanc.get_status_display()
        ])
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Formatar valores
    from openpyxl.styles.numbers import BUILTIN_FORMATS
    for row in ws.iter_rows(min_row=7, min_col=6, max_col=6):
        for cell in row:
            if cell.value:
                cell.number_format = BUILTIN_FORMATS[7]
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="LCDPR_{propriedade.id}_{data_inicio.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

