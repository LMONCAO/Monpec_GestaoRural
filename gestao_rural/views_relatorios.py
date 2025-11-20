from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Sum, Avg, Count
from decimal import Decimal
from datetime import datetime, timedelta
import json
from .models import Propriedade, InventarioRebanho, MovimentacaoProjetada
from .models import CustoFixo, CustoVariavel, Financiamento, IndicadorFinanceiro
from .views_analise import calcular_indicadores_basicos
from .relatorios_modulos import (
    gerar_relatorio_pecuaria_pdf,
    gerar_relatorio_nutricao_pdf,
    gerar_relatorio_operacoes_pdf,
    gerar_relatorio_compras_pdf,
    gerar_relatorio_projetos_bancarios_pdf
)

# Imports para exportação
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


@login_required
def relatorios_dashboard(request, propriedade_id):
    """Dashboard do módulo de relatórios"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca dados para resumo
    resumo_dados = gerar_resumo_propriedade(propriedade)
    
    context = {
        'propriedade': propriedade,
        'resumo_dados': resumo_dados,
    }
    
    return render(request, 'gestao_rural/relatorios_dashboard.html', context)


@login_required
def relatorio_final(request, propriedade_id):
    """Relatório final consolidado simples (inventário + indicadores básicos)."""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)

    # Inventário mais recente
    from django.db.models import Max
    data_max = InventarioRebanho.objects.filter(propriedade=propriedade).aggregate(Max('data_inventario'))['data_inventario__max']
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade)
    if data_max:
        inventario = inventario.filter(data_inventario=data_max)
    inventario = inventario.select_related('categoria').order_by('categoria__nome')

    total_animais = sum(i.quantidade for i in inventario)
    valor_total_inventario = sum((i.valor_total or 0) for i in inventario)

    indicadores = calcular_indicadores_basicos(propriedade)

    context = {
        'propriedade': propriedade,
        'inventario': inventario,
        'total_animais': total_animais,
        'valor_total_inventario': valor_total_inventario,
        'indicadores': indicadores,
    }

    return render(request, 'gestao_rural/relatorio_final.html', context)


@login_required
def relatorio_inventario(request, propriedade_id):
    """Relatório de inventário do rebanho"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca inventário atual
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade).order_by('categoria__nome')
    
    # Calcula totais
    total_animais = sum(item.quantidade for item in inventario)
    valor_total_rebanho = sum(item.valor_total for item in inventario if item.valor_total)
    
    # Agrupa por categoria
    inventario_por_categoria = {}
    for item in inventario:
        categoria = item.categoria.nome
        if categoria not in inventario_por_categoria:
            inventario_por_categoria[categoria] = {
                'quantidade': 0,
                'valor_total': Decimal('0.00'),
                'itens': []
            }
        inventario_por_categoria[categoria]['quantidade'] += item.quantidade
        inventario_por_categoria[categoria]['valor_total'] += item.valor_total or Decimal('0.00')
        inventario_por_categoria[categoria]['itens'].append(item)
    
    context = {
        'propriedade': propriedade,
        'inventario': inventario,
        'inventario_por_categoria': inventario_por_categoria,
        'total_animais': total_animais,
        'valor_total_rebanho': valor_total_rebanho,
        'data_relatorio': datetime.now().date(),
    }
    
    return render(request, 'gestao_rural/relatorio_inventario.html', context)


@login_required
def relatorio_financeiro(request, propriedade_id):
    """Relatório financeiro completo"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Dados financeiros
    dados_financeiros = gerar_dados_financeiros(propriedade)
    
    # Indicadores financeiros recentes
    indicadores = IndicadorFinanceiro.objects.filter(
        propriedade=propriedade
    ).order_by('-data_referencia')[:20]
    
    context = {
        'propriedade': propriedade,
        'dados_financeiros': dados_financeiros,
        'indicadores': indicadores,
        'data_relatorio': datetime.now().date(),
    }
    
    return render(request, 'gestao_rural/relatorio_financeiro.html', context)


@login_required
def relatorio_custos(request, propriedade_id):
    """Relatório de custos de produção"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Custos fixos
    custos_fixos = CustoFixo.objects.filter(propriedade=propriedade, ativo=True)
    total_custos_fixos_mes = sum(custo.valor_mensal for custo in custos_fixos if custo.valor_mensal)
    total_custos_fixos_ano = sum(custo.custo_anual for custo in custos_fixos)
    
    # Custos variáveis
    custos_variaveis = CustoVariavel.objects.filter(propriedade=propriedade, ativo=True)
    total_custos_variaveis_ano = sum(custo.custo_anual_por_cabeca for custo in custos_variaveis)
    
    # Agrupa custos fixos por tipo
    custos_fixos_por_tipo = {}
    for custo in custos_fixos:
        tipo = custo.tipo_custo
        if tipo not in custos_fixos_por_tipo:
            custos_fixos_por_tipo[tipo] = {
                'total_mensal': Decimal('0.00'),
                'total_anual': Decimal('0.00'),
                'custos': []
            }
        custos_fixos_por_tipo[tipo]['total_mensal'] += custo.valor_mensal or Decimal('0.00')
        custos_fixos_por_tipo[tipo]['total_anual'] += custo.custo_anual
        custos_fixos_por_tipo[tipo]['custos'].append(custo)
    
    context = {
        'propriedade': propriedade,
        'custos_fixos': custos_fixos,
        'custos_variaveis': custos_variaveis,
        'custos_fixos_por_tipo': custos_fixos_por_tipo,
        'total_custos_fixos_mes': total_custos_fixos_mes,
        'total_custos_fixos_ano': total_custos_fixos_ano,
        'total_custos_variaveis_ano': total_custos_variaveis_ano,
        'data_relatorio': datetime.now().date(),
    }
    
    return render(request, 'gestao_rural/relatorio_custos.html', context)


@login_required
def relatorio_endividamento(request, propriedade_id):
    """Relatório de endividamento"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Financiamentos ativos
    financiamentos = Financiamento.objects.filter(propriedade=propriedade, ativo=True)
    
    # Calcula totais
    total_financiado = sum(f.valor_principal for f in financiamentos)
    total_parcelas_mes = sum(f.valor_parcela for f in financiamentos)
    total_parcelas_ano = total_parcelas_mes * 12
    
    # Agrupa por tipo
    financiamentos_por_tipo = {}
    for financiamento in financiamentos:
        tipo = financiamento.tipo.nome
        if tipo not in financiamentos_por_tipo:
            financiamentos_por_tipo[tipo] = {
                'total_principal': Decimal('0.00'),
                'total_parcelas_mes': Decimal('0.00'),
                'financiamentos': []
            }
        financiamentos_por_tipo[tipo]['total_principal'] += financiamento.valor_principal
        financiamentos_por_tipo[tipo]['total_parcelas_mes'] += financiamento.valor_parcela
        financiamentos_por_tipo[tipo]['financiamentos'].append(financiamento)
    
    context = {
        'propriedade': propriedade,
        'financiamentos': financiamentos,
        'financiamentos_por_tipo': financiamentos_por_tipo,
        'total_financiado': total_financiado,
        'total_parcelas_mes': total_parcelas_mes,
        'total_parcelas_ano': total_parcelas_ano,
        'data_relatorio': datetime.now().date(),
    }
    
    return render(request, 'gestao_rural/relatorio_endividamento.html', context)


@login_required
def relatorio_consolidado(request, propriedade_id):
    """Relatório consolidado geral"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca todos os dados
    resumo_dados = gerar_resumo_propriedade(propriedade)
    dados_financeiros = gerar_dados_financeiros(propriedade)
    
    # Inventário
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade)
    valor_total_rebanho = sum(item.valor_total for item in inventario if item.valor_total)
    
    # Custos
    custos_fixos = CustoFixo.objects.filter(propriedade=propriedade, ativo=True)
    total_custos_fixos_ano = sum(custo.custo_anual for custo in custos_fixos)
    
    # Financiamentos
    financiamentos = Financiamento.objects.filter(propriedade=propriedade, ativo=True)
    total_parcelas_ano = sum(f.valor_parcela for f in financiamentos) * 12
    
    context = {
        'propriedade': propriedade,
        'resumo_dados': resumo_dados,
        'dados_financeiros': dados_financeiros,
        'valor_total_rebanho': valor_total_rebanho,
        'total_custos_fixos_ano': total_custos_fixos_ano,
        'total_parcelas_ano': total_parcelas_ano,
        'data_relatorio': datetime.now().date(),
    }
    
    return render(request, 'gestao_rural/relatorio_consolidado.html', context)


def gerar_resumo_propriedade(propriedade):
    """Gera resumo geral da propriedade"""
    resumo = {}
    
    try:
        # Inventário
        inventario = InventarioRebanho.objects.filter(propriedade=propriedade)
        resumo['total_animais'] = sum(item.quantidade for item in inventario)
        resumo['valor_rebanho'] = sum(item.valor_total for item in inventario if item.valor_total)
        
        # Custos
        custos_fixos = CustoFixo.objects.filter(propriedade=propriedade, ativo=True)
        resumo['custos_fixos_mes'] = sum(custo.valor_mensal for custo in custos_fixos if custo.valor_mensal)
        
        # Financiamentos
        financiamentos = Financiamento.objects.filter(propriedade=propriedade, ativo=True)
        resumo['total_financiado'] = sum(f.valor_principal for f in financiamentos)
        resumo['parcelas_mes'] = sum(f.valor_parcela for f in financiamentos)
        
    except Exception as e:
        print(f"Erro ao gerar resumo: {e}")
    
    return resumo


def gerar_dados_financeiros(propriedade):
    """Gera dados financeiros para relatórios"""
    dados = {}
    
    try:
        # Receitas (simuladas baseadas no rebanho)
        inventario = InventarioRebanho.objects.filter(propriedade=propriedade)
        dados['receita_potencial_ano'] = sum(item.valor_total for item in inventario if item.valor_total) * Decimal('0.15')  # 15% ao ano
        
        # Custos
        custos_fixos = CustoFixo.objects.filter(propriedade=propriedade, ativo=True)
        dados['custos_fixos_ano'] = sum(custo.custo_anual for custo in custos_fixos)
        
        # Financiamentos
        financiamentos = Financiamento.objects.filter(propriedade=propriedade, ativo=True)
        dados['parcelas_ano'] = sum(f.valor_parcela for f in financiamentos) * 12
        
        # Resultado
        dados['lucro_estimado'] = dados['receita_potencial_ano'] - dados['custos_fixos_ano'] - dados['parcelas_ano']
        
    except Exception as e:
        print(f"Erro ao gerar dados financeiros: {e}")
    
    return dados


# ===== FUNÇÕES DE EXPORTAÇÃO =====

@login_required
def exportar_relatorio_inventario_pdf(request, propriedade_id):
    """Exporta relatório de inventário em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca dados do inventário
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade).order_by('categoria__nome')
    total_animais = sum(item.quantidade for item in inventario)
    valor_total_rebanho = sum(item.valor_total for item in inventario if item.valor_total)
    
    # Cria o PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_inventario_{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    story.append(Paragraph(f"Relatório de Inventário - {propriedade.nome_propriedade}", title_style))
    story.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Resumo
    resumo_data = [
        ['Total de Animais', str(total_animais)],
        ['Valor Total do Rebanho', f"R$ {valor_total_rebanho:.2f}"]
    ]
    resumo_table = Table(resumo_data, colWidths=[8*cm, 8*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 20))
    
    # Tabela de inventário
    inventario_data = [['Categoria', 'Quantidade', 'Valor por Cabeça', 'Valor Total']]
    for item in inventario:
        valor_por_cabeca = item.valor_por_cabeca or Decimal('0.00')
        valor_total = item.valor_total or Decimal('0.00')
        inventario_data.append([
            item.categoria.nome,
            str(item.quantidade),
            f"R$ {valor_por_cabeca:.2f}",
            f"R$ {valor_total:.2f}"
        ])
    
    inventario_table = Table(inventario_data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
    inventario_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(inventario_table)
    
    doc.build(story)
    return response


@login_required
def exportar_relatorio_inventario_excel(request, propriedade_id):
    """Exporta relatório de inventário em Excel"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca dados do inventário
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade).order_by('categoria__nome')
    total_animais = sum(item.quantidade for item in inventario)
    valor_total_rebanho = sum(item.valor_total for item in inventario if item.valor_total)
    
    # Cria o workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário do Rebanho"
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Título
    ws['A1'] = f"Relatório de Inventário - {propriedade.nome_propriedade}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].font = normal_font
    
    # Resumo
    ws['A4'] = "Total de Animais:"
    ws['B4'] = total_animais
    ws['A5'] = "Valor Total do Rebanho:"
    ws['B5'] = f"R$ {valor_total_rebanho:.2f}"
    
    # Cabeçalho da tabela
    headers = ['Categoria', 'Quantidade', 'Valor por Cabeça', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
    
    # Dados do inventário
    row = 8
    for item in inventario:
        ws.cell(row=row, column=1, value=item.categoria.nome)
        ws.cell(row=row, column=2, value=item.quantidade)
        ws.cell(row=row, column=3, value=f"R$ {(item.valor_por_cabeca or Decimal('0.00')):.2f}")
        ws.cell(row=row, column=4, value=f"R$ {(item.valor_total or Decimal('0.00')):.2f}")
        row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_inventario_{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def exportar_relatorio_financeiro_pdf(request, propriedade_id):
    """Exporta relatório financeiro em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca dados financeiros
    dados_financeiros = gerar_dados_financeiros(propriedade)
    indicadores = IndicadorFinanceiro.objects.filter(propriedade=propriedade).order_by('-data_referencia')[:10]
    
    # Cria o PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_financeiro_{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    story.append(Paragraph(f"Relatório Financeiro - {propriedade.nome_propriedade}", title_style))
    story.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Resumo financeiro
    resumo_data = [
        ['Receita Potencial/Ano', f"R$ {dados_financeiros.get('receita_potencial_ano', 0):.2f}"],
        ['Custos Fixos/Ano', f"R$ {dados_financeiros.get('custos_fixos_ano', 0):.2f}"],
        ['Parcelas/Ano', f"R$ {dados_financeiros.get('parcelas_ano', 0):.2f}"],
        ['Lucro Estimado', f"R$ {dados_financeiros.get('lucro_estimado', 0):.2f}"]
    ]
    resumo_table = Table(resumo_data, colWidths=[8*cm, 8*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 20))
    
    # Indicadores recentes
    if indicadores:
        story.append(Paragraph("Indicadores Financeiros Recentes", styles['Heading2']))
        indicadores_data = [['Data', 'Receita Bruta', 'Custos Totais', 'Lucro Líquido', 'Margem (%)']]
        for indicador in indicadores:
            margem = (indicador.lucro_liquido / indicador.receita_bruta * 100) if indicador.receita_bruta > 0 else 0
            indicadores_data.append([
                indicador.data_referencia.strftime('%d/%m/%Y'),
                f"R$ {indicador.receita_bruta:.2f}",
                f"R$ {indicador.custos_totais:.2f}",
                f"R$ {indicador.lucro_liquido:.2f}",
                f"{margem:.1f}%"
            ])
        
        indicadores_table = Table(indicadores_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        indicadores_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(indicadores_table)
    
    doc.build(story)
    return response


@login_required
def exportar_relatorio_financeiro_excel(request, propriedade_id):
    """Exporta relatório financeiro em Excel"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    # Busca dados financeiros
    dados_financeiros = gerar_dados_financeiros(propriedade)
    indicadores = IndicadorFinanceiro.objects.filter(propriedade=propriedade).order_by('-data_referencia')[:10]
    
    # Cria o workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Financeiro"
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Título
    ws['A1'] = f"Relatório Financeiro - {propriedade.nome_propriedade}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].font = normal_font
    
    # Resumo financeiro
    ws['A4'] = "Receita Potencial/Ano:"
    ws['B4'] = f"R$ {dados_financeiros.get('receita_potencial_ano', 0):.2f}"
    ws['A5'] = "Custos Fixos/Ano:"
    ws['B5'] = f"R$ {dados_financeiros.get('custos_fixos_ano', 0):.2f}"
    ws['A6'] = "Parcelas/Ano:"
    ws['B6'] = f"R$ {dados_financeiros.get('parcelas_ano', 0):.2f}"
    ws['A7'] = "Lucro Estimado:"
    ws['B7'] = f"R$ {dados_financeiros.get('lucro_estimado', 0):.2f}"
    
    # Indicadores recentes
    if indicadores:
        ws['A9'] = "Indicadores Financeiros Recentes"
        ws['A9'].font = header_font
        
        headers = ['Data', 'Receita Bruta', 'Custos Totais', 'Lucro Líquido', 'Margem (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = header_font
        
        row = 11
        for indicador in indicadores:
            margem = (indicador.lucro_liquido / indicador.receita_bruta * 100) if indicador.receita_bruta > 0 else 0
            ws.cell(row=row, column=1, value=indicador.data_referencia.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=f"R$ {indicador.receita_bruta:.2f}")
            ws.cell(row=row, column=3, value=f"R$ {indicador.custos_totais:.2f}")
            ws.cell(row=row, column=4, value=f"R$ {indicador.lucro_liquido:.2f}")
            ws.cell(row=row, column=5, value=f"{margem:.1f}%")
            row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # Resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_financeiro_{propriedade.nome_propriedade}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


# Funções similares para os outros relatórios...
@login_required
def exportar_relatorio_custos_pdf(request, propriedade_id):
    """Exporta relatório de custos em PDF"""
    # Implementação similar ao inventário
    pass


@login_required
def exportar_relatorio_custos_excel(request, propriedade_id):
    """Exporta relatório de custos em Excel"""
    # Implementação similar ao inventário
    pass


@login_required
def exportar_relatorio_endividamento_pdf(request, propriedade_id):
    """Exporta relatório de endividamento em PDF"""
    # Implementação similar ao inventário
    pass


@login_required
def exportar_relatorio_endividamento_excel(request, propriedade_id):
    """Exporta relatório de endividamento em Excel"""
    # Implementação similar ao inventário
    pass


@login_required
def exportar_relatorio_consolidado_pdf(request, propriedade_id):
    """Exporta relatório consolidado em PDF"""
    # Implementação similar ao inventário
    pass


@login_required
def exportar_relatorio_consolidado_excel(request, propriedade_id):
    """Exporta relatório consolidado em Excel"""
    # Implementação similar ao inventário
    pass


# ===== RELATÓRIOS POR MÓDULO =====

@login_required
def relatorio_pecuaria_pdf(request, propriedade_id):
    """Gera relatório de Pecuária em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    try:
        pdf_buffer = gerar_relatorio_pecuaria_pdf(propriedade)
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_pecuaria_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Erro ao gerar relatório: {str(e)}')
        return render(request, 'gestao_rural/relatorios_dashboard.html', {'propriedade': propriedade})


@login_required
def relatorio_nutricao_pdf(request, propriedade_id):
    """Gera relatório de Nutrição em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    try:
        pdf_buffer = gerar_relatorio_nutricao_pdf(propriedade)
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_nutricao_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Erro ao gerar relatório: {str(e)}')
        return render(request, 'gestao_rural/relatorios_dashboard.html', {'propriedade': propriedade})


@login_required
def relatorio_operacoes_pdf(request, propriedade_id):
    """Gera relatório de Operações em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    try:
        pdf_buffer = gerar_relatorio_operacoes_pdf(propriedade)
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_operacoes_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Erro ao gerar relatório: {str(e)}')
        return render(request, 'gestao_rural/relatorios_dashboard.html', {'propriedade': propriedade})


@login_required
def relatorio_compras_pdf(request, propriedade_id):
    """Gera relatório de Compras em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    try:
        pdf_buffer = gerar_relatorio_compras_pdf(propriedade)
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_compras_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Erro ao gerar relatório: {str(e)}')
        return render(request, 'gestao_rural/relatorios_dashboard.html', {'propriedade': propriedade})


@login_required
def relatorio_projetos_bancarios_pdf(request, propriedade_id):
    """Gera relatório de Projetos Bancários em PDF"""
    propriedade = get_object_or_404(Propriedade, id=propriedade_id)
    
    try:
        pdf_buffer = gerar_relatorio_projetos_bancarios_pdf(propriedade)
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_projetos_bancarios_{propriedade.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Erro ao gerar relatório: {str(e)}')
        return render(request, 'gestao_rural/relatorios_dashboard.html', {'propriedade': propriedade})

