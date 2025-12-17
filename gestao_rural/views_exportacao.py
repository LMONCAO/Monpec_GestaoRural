# -*- coding: utf-8 -*-
"""
Views para exportação de dados
Exportação para Excel, PDF, etc.
"""

from datetime import datetime

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from .relatorios_iatf import coletar_dados_iatf, IATFModuloIndisponivel


def calcular_altura_tabela(table_data, font_size=10, padding=8):
    """
    Calcula a altura aproximada de uma tabela baseada no número de linhas e fonte
    """
    num_linhas = len(table_data)
    altura_linha = font_size + padding + 4  # Adicionar margem extra
    altura_total = num_linhas * altura_linha
    return altura_total


def ajustar_tabela_para_pagina(table_data, col_widths, max_height_cm=25):
    """
    Ajusta uma tabela para caber na página, dividindo se necessário
    """
    altura_tabela = calcular_altura_tabela(table_data)
    altura_maxima_pt = max_height_cm * 28.35  # Converter cm para pontos
    
    if altura_tabela <= altura_maxima_pt:
        return [table_data]  # Tabela cabe em uma página
    
    # Dividir a tabela em partes menores
    linhas_por_pagina = int(len(table_data) * (altura_maxima_pt / altura_tabela))
    linhas_por_pagina = max(8, linhas_por_pagina)  # Mínimo de 8 linhas por página
    
    tabelas_divididas = []
    for i in range(0, len(table_data), linhas_por_pagina):
        parte_tabela = table_data[i:i + linhas_por_pagina]
        tabelas_divididas.append(parte_tabela)
    
    return tabelas_divididas


def ajustar_fonte_por_conteudo(table_data, fonte_base=10):
    """
    Ajusta o tamanho da fonte baseado na quantidade de conteúdo
    """
    num_linhas = len(table_data)
    
    if num_linhas <= 10:
        return fonte_base  # Fonte normal
    elif num_linhas <= 20:
        return fonte_base - 1  # Fonte um pouco menor
    elif num_linhas <= 30:
        return fonte_base - 2  # Fonte menor
    else:
        return fonte_base - 3  # Fonte bem menor


def calcular_larguras_dinamicas(table_data, num_colunas, largura_total_cm=27.7):
    """
    Calcula larguras de colunas dinamicamente baseado no conteúdo e número de colunas
    """
    # Percentuais de largura por número de colunas
    if num_colunas == 6:
        # Tabela de inventário: 6 colunas
        percentuais = [25, 10, 12, 12, 12, 21]  # Categoria 25%, valores 10-12%, Total 21%
    elif num_colunas == 11:
        # Tabela de evolução: 11 colunas
        percentuais = [15, 8, 8, 8, 8, 10, 8, 8, 8, 8, 15]  # Categoria 15%, Total 15%
    elif num_colunas == 12:
        # Tabela de projeção por ano: 12 colunas
        percentuais = [14, 7, 7, 7, 7, 9, 7, 7, 7, 7, 8, 13]  # Categoria 14%, Total 13%
    else:
        # Distribuição padrão para outras tabelas
        percentuais = [100 / num_colunas] * num_colunas
        if num_colunas >= 6:
            # Primeira coluna maior
            percentuais[0] = (100 / num_colunas) * 1.8
            # Redistribuir o resto
            ajuste = (percentuais[0] - (100 / num_colunas)) / (num_colunas - 1)
            for i in range(1, num_colunas):
                percentuais[i] -= ajuste
    
    # Converter para cm e depois para unidades do ReportLab
    larguras = [(largura_total_cm * p / 100) for p in percentuais]
    
    return [largura * cm for largura in larguras]


@login_required
def exportar_inventario_excel(request, propriedade_id):
    """Exporta inventário para Excel"""
    from .models import InventarioRebanho, Propriedade
    
    propriedade = get_object_or_404(Propriedade, id=propriedade_id, produtor__usuario_responsavel=request.user)
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade).select_related('categoria')
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário"
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = 'Categoria'
    ws['B1'] = 'Quantidade'
    ws['C1'] = 'Valor por Cabeça'
    ws['D1'] = 'Valor Total'
    ws['E1'] = 'Data Inventário'
    
    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Preencher dados
    row = 2
    total_geral = Decimal('0')
    for item in inventario:
        ws[f'A{row}'] = item.categoria.nome
        ws[f'B{row}'] = item.quantidade
        ws[f'C{row}'] = float(item.valor_por_cabeca)
        valor_total = float(item.valor_total) if item.valor_total else 0
        ws[f'D{row}'] = valor_total
        total_geral += Decimal(str(valor_total))
        ws[f'E{row}'] = item.data_inventario.strftime('%d/%m/%Y')
        row += 1
    
    # Adicionar total
    ws[f'C{row}'] = 'TOTAL:'
    ws[f'D{row}'] = float(total_geral)
    ws[f'D{row}'].font = Font(bold=True)
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventario_{propriedade.id}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_projecao_excel(request, propriedade_id):
    """Exporta projeção para Excel"""
    from .models import MovimentacaoProjetada, Propriedade
    from decimal import Decimal
    
    propriedade = get_object_or_404(Propriedade, id=propriedade_id, produtor__usuario_responsavel=request.user)
    movimentacoes = MovimentacaoProjetada.objects.filter(
        propriedade=propriedade
    ).select_related('categoria').order_by('data_movimentacao')
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Projeção"
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = 'Data'
    ws['B1'] = 'Tipo Movimentação'
    ws['C1'] = 'Categoria'
    ws['D1'] = 'Quantidade'
    ws['E1'] = 'Valor Unitário'
    ws['F1'] = 'Valor Total'
    ws['G1'] = 'Observação'
    
    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Preencher dados
    row = 2
    for mov in movimentacoes:
        ws[f'A{row}'] = mov.data_movimentacao.strftime('%d/%m/%Y')
        ws[f'B{row}'] = mov.get_tipo_movimentacao_display()
        ws[f'C{row}'] = mov.categoria.nome
        ws[f'D{row}'] = mov.quantidade
        ws[f'E{row}'] = float(mov.valor_por_cabeca) if mov.valor_por_cabeca else 0
        ws[f'F{row}'] = float(mov.valor_total) if mov.valor_total else 0
        ws[f'G{row}'] = mov.observacao or ''
        row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="projecao_{propriedade.id}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_projecao_pdf(request, propriedade_id):
    """Exporta projeção para PDF em modo paisagem com todas as tabelas"""
    from .models import MovimentacaoProjetada, Propriedade, InventarioRebanho, ParametrosProjecaoRebanho
    from .views import gerar_resumo_projecao_por_ano, gerar_evolucao_detalhada_rebanho
    from datetime import datetime
    from collections import defaultdict
    from decimal import Decimal
    
    propriedade = get_object_or_404(Propriedade, id=propriedade_id, produtor__usuario_responsavel=request.user)
    
    # Buscar dados
    movimentacoes = list(MovimentacaoProjetada.objects.filter(
        propriedade=propriedade
    ).select_related('categoria').order_by('data_movimentacao'))
    
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade).select_related('categoria')
    parametros = ParametrosProjecaoRebanho.objects.filter(propriedade=propriedade).first()
    
    if not movimentacoes:
        from django.contrib import messages
        messages.error(request, 'Nenhuma projeção encontrada. Gere uma projeção primeiro.')
        from django.shortcuts import redirect
        return redirect('pecuaria_projecao', propriedade_id=propriedade.id)
    
    # Gerar resumo por ano
    resumo_projecao_por_ano = gerar_resumo_projecao_por_ano(movimentacoes, inventario)
    evolucao_detalhada = gerar_evolucao_detalhada_rebanho(movimentacoes, inventario)
    
    # Criar resposta PDF
    response = HttpResponse(content_type='application/pdf')
    data_atual = datetime.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="projecao_{propriedade.id}_{data_atual}.pdf"'
    
    # Criar documento PDF em paisagem
    from reportlab.lib.pagesizes import landscape, A4
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    story = []
    
    # Estilos profissionais
    styles = getSampleStyleSheet()
    
    # Título principal mais elegante
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=8,
        spaceBefore=0,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=26
    )
    
    # Seções principais mais destacadas
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold',
        leading=20,
        borderWidth=0,
        borderColor=colors.HexColor('#1e3a8a'),
        borderPadding=2
    )
    
    # Subseções mais elegantes
    subsection_style = ParagraphStyle(
        'SubSectionStyle',
        parent=styles['Heading3'],
        fontSize=14,
        textColor=colors.HexColor('#6495ed'),
        spaceAfter=6,
        spaceBefore=10,
        fontName='Helvetica-Bold',
        leading=18
    )
    
    # Estilo para texto normal mais legível
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        spaceAfter=4,
        spaceBefore=2,
        fontName='Helvetica',
        leading=14
    )
    
    # Cabeçalho profissional
    story.append(Spacer(1, 0.5*cm))
    
    # Logo/Identificação da empresa
    logo_text = Paragraph("MONPEC", ParagraphStyle(
        'LogoStyle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=0,
        spaceBefore=0,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    ))
    story.append(logo_text)
    
    subtitle_text = Paragraph("Sistema de Gestão Rural Inteligente", ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#666666'),
        spaceAfter=10,
        spaceBefore=0,
        alignment=TA_CENTER,
        fontName='Helvetica'
    ))
    story.append(subtitle_text)
    
    # Linha divisória
    story.append(Spacer(1, 0.2*cm))
    line = Table([['']], colWidths=[27.7*cm])
    line.setStyle(TableStyle([('LINEBELOW', (0, 0), (0, 0), 2, colors.HexColor('#1e3a8a'))]))
    story.append(line)
    story.append(Spacer(1, 0.3*cm))
    
    # Título principal
    story.append(Paragraph("Relatório de Projeção do Rebanho", title_style))
    story.append(Spacer(1, 0.4*cm))
    
    # Informações detalhadas da propriedade
    info_table_data = [
        ['Propriedade:', propriedade.nome_propriedade, 'Período:', f"{min(resumo_projecao_por_ano.keys()) if resumo_projecao_por_ano else 'N/A'} - {max(resumo_projecao_por_ano.keys()) if resumo_projecao_por_ano else 'N/A'}"],
        ['Produtor:', propriedade.produtor.nome if propriedade.produtor else 'N/A', 'Data:', datetime.now().strftime('%d/%m/%Y %H:%M')],
        ['Localização:', f"{getattr(propriedade, 'cidade', 'N/A')}, {getattr(propriedade, 'estado', 'N/A')}" if getattr(propriedade, 'cidade', None) and getattr(propriedade, 'estado', None) else 'N/A', 'Anos Projetados:', str(len(resumo_projecao_por_ano)) if resumo_projecao_por_ano else '0']
    ]
    
    info_table = Table(info_table_data, colWidths=[3.5*cm, 8.5*cm, 3.5*cm, 8.5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#1e3a8a')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e0e0e0')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),  # Borda externa grossa
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.8*cm))
    
    # Resumo Executivo
    story.append(Paragraph("RESUMO EXECUTIVO", section_style))
    story.append(Spacer(1, 0.3*cm))
    
    # Calcular indicadores principais
    total_animais_inicial = sum(item.quantidade for item in inventario) if inventario.exists() else 0
    valor_total_rebanho = sum(float(item.valor_total or 0) for item in inventario) if inventario.exists() else 0
    
    # Calcular principais movimentações
    total_nascimentos = sum(dados.get('nascimentos', 0) for dados in evolucao_detalhada.values()) if evolucao_detalhada else 0
    total_vendas = sum(dados.get('vendas', 0) for dados in evolucao_detalhada.values()) if evolucao_detalhada else 0
    total_compras = sum(dados.get('compras', 0) for dados in evolucao_detalhada.values()) if evolucao_detalhada else 0
    
    # Card de indicadores principais
    indicadores_data = [
        ['INDICADORES PRINCIPAIS', '', '', ''],
        ['Total de Animais Inicial:', f"{total_animais_inicial:,}".replace(',', '.'), 'Valor Total do Rebanho:', f"R$ {valor_total_rebanho:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Total de Nascimentos:', f"{total_nascimentos:,}".replace(',', '.'), 'Total de Vendas:', f"{total_vendas:,}".replace(',', '.')],
        ['Total de Compras:', f"{total_compras:,}".replace(',', '.'), 'Anos Projetados:', str(len(resumo_projecao_por_ano)) if resumo_projecao_por_ano else '0'],
    ]
    
    indicadores_table = Table(indicadores_data, colWidths=[6.5*cm, 4.5*cm, 6.5*cm, 4.5*cm])
    indicadores_table.setStyle(TableStyle([
        # Cabeçalho da tabela
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Dados da tabela
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (2, 1), (2, -1), colors.HexColor('#1e3a8a')),
        
        # Bordas e espaçamento
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),  # Borda externa grossa
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e3a8a')),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        
        # Alternância de cores nas linhas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(indicadores_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Índice/Sumário
    story.append(Paragraph("ÍNDICE", section_style))
    story.append(Spacer(1, 0.3*cm))
    
    indice_data = [['Seção', 'Página']]
    pagina_atual = 2  # Começando da página 2 (após resumo executivo)
    
    if inventario.exists():
        indice_data.append(['1. Inventário Inicial', str(pagina_atual)])
        pagina_atual += 1
    
    if evolucao_detalhada:
        indice_data.append(['2. Evolução Detalhada do Rebanho', str(pagina_atual)])
        pagina_atual += 1
    
    if resumo_projecao_por_ano:
        indice_data.append(['3. Projeção por Ano', str(pagina_atual)])
        for ano in sorted(resumo_projecao_por_ano.keys()):
            indice_data.append([f'   3.{ano}. Projeção {ano}', str(pagina_atual)])
            pagina_atual += 1
    
    # Adicionar análise financeira ao índice
    indice_data.append(['4. Análise Financeira', str(pagina_atual)])
    
    indice_table = Table(indice_data, colWidths=[21*cm, 3*cm])
    indice_table.setStyle(TableStyle([
        # Cabeçalho da tabela
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Dados da tabela
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        
        # Bordas e espaçamento
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),  # Borda externa grossa
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e3a8a')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        
        # Alternância de cores nas linhas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(indice_table)
    story.append(PageBreak())
    
    # Adicionar Inventário Inicial se existir
    if inventario.exists():
        story.append(PageBreak())  # Nova página para Inventário Inicial
        story.append(Paragraph("1. INVENTÁRIO INICIAL", section_style))
        story.append(Spacer(1, 0.5*cm))
        table_data = [['Categoria', 'Quantidade', 'Fêmeas', 'Machos', 'Valor/Cabeça', 'Valor Total']]
        
        for item in inventario:
            femeas = item.quantidade if any(t in item.categoria.nome for t in ['Fêmea', 'Vaca', 'Bezerra', 'Novilha']) else 0
            machos = item.quantidade if any(t in item.categoria.nome for t in ['Macho', 'Boi', 'Bezerro']) else 0
            
            table_data.append([
                item.categoria.nome,  # Nome completo sem truncamento
                str(item.quantidade),
                str(femeas),
                str(machos),
                f"R$ {float(item.valor_por_cabeca):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {float(item.valor_total or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        # Calcular totais
        total_qtde = sum(item.quantidade for item in inventario)
        total_femeas = sum(item.quantidade if any(t in item.categoria.nome for t in ['Fêmea', 'Vaca']) else 0 for item in inventario)
        total_machos = sum(item.quantidade if any(t in item.categoria.nome for t in ['Macho', 'Boi']) else 0 for item in inventario)
        total_valor = sum(float(item.valor_total or 0) for item in inventario)
        
        table_data.append([
            'TOTAL',
            str(total_qtde),
            str(total_femeas),
            str(total_machos),
            '',
            f"R$ {total_valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        ])
        
        # Calcular larguras dinamicamente
        col_widths = calcular_larguras_dinamicas(table_data, 6)
        
        # Ajustar fonte baseado no conteúdo
        fonte_tabela = ajustar_fonte_por_conteudo(table_data, 11)
        fonte_cabecalho = fonte_tabela + 2
        
        # Dividir tabela se necessário
        tabelas_divididas = ajustar_tabela_para_pagina(table_data, col_widths)
        
        for i, parte_tabela in enumerate(tabelas_divididas):
            table = Table(parte_tabela, colWidths=col_widths)
    table.setStyle(TableStyle([
                # Cabeçalho da tabela
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), fonte_cabecalho),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                
                # Dados da tabela
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), fonte_tabela),
                ('FONTNAME', (0, 1), (0, -2), 'Helvetica-Bold'),  # Categoria em negrito
                ('ALIGN', (0, 1), (0, -2), 'LEFT'),  # Categoria à esquerda
                ('ALIGN', (1, 1), (-1, -2), 'CENTER'),  # Demais dados centralizados
                ('VALIGN', (0, 1), (-1, -2), 'MIDDLE'),
                
                # Linha de totais (apenas na última parte)
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), fonte_tabela + 1),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4fd')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1e3a8a')),
                ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
                
                # Bordas e espaçamento
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),  # Borda externa grossa
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e3a8a')),  # Linha forte abaixo do cabeçalho
                ('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#1e3a8a')),  # Linha forte acima dos totais
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                
                # Alternância de cores nas linhas
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    
    story.append(table)
    
    # Adicionar quebra de página apenas se não for a última parte
    if i < len(tabelas_divididas) - 1:
        story.append(PageBreak())

        # Quebra de página após Inventário Inicial
    story.append(PageBreak())
    
    # Adicionar Evolução Detalhada do Rebanho
    if evolucao_detalhada:
        story.append(PageBreak())  # Nova página para Evolução Detalhada
        story.append(Paragraph("2. EVOLUÇÃO DETALHADA DO REBANHO", section_style))
        story.append(Spacer(1, 0.5*cm))
        table_data = [['Categoria', 'Inicial', 'Nasc', 'Comp', 'Vend', 'Trans', 'Mort', 'Final', 'Peso (kg)', 'R$/Cabeça', 'Total (R$)']]
        
        for categoria, dados in evolucao_detalhada.items():
            trans = f"+{dados.get('transferencias_entrada', 0)}/{dados.get('transferencias_saida', 0)}" if dados.get('transferencias_entrada', 0) > 0 or dados.get('transferencias_saida', 0) > 0 else "0"
            
            table_data.append([
                categoria,  # Nome completo sem truncamento
                str(dados.get('saldo_inicial', 0) or 0),
                str(dados.get('nascimentos', 0) or 0),
                str(dados.get('compras', 0) or 0),
                str(dados.get('vendas', 0) or 0),
                trans,
                str(dados.get('mortes', 0) or 0),
                str(dados.get('saldo_final', 0) or 0),
                f"{float(dados.get('peso_medio_kg', 0) or 0):,.1f}",
                f"R$ {float(dados.get('valor_unitario', 0) or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {float(dados.get('valor_total', 0) or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        # Calcular larguras dinamicamente
        col_widths_evolucao = calcular_larguras_dinamicas(table_data, 11)
        
        # Ajustar fonte baseado no conteúdo
        fonte_tabela = ajustar_fonte_por_conteudo(table_data, 10)
        fonte_cabecalho = fonte_tabela + 2
        
        # Dividir tabela se necessário
        tabelas_divididas = ajustar_tabela_para_pagina(table_data, col_widths_evolucao)
        
        for i, parte_tabela in enumerate(tabelas_divididas):
            table = Table(parte_tabela, colWidths=col_widths_evolucao)
            table.setStyle(TableStyle([
                # Cabeçalho da tabela
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), fonte_cabecalho),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                
                # Dados da tabela
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), fonte_tabela),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Categoria em negrito
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Categoria à esquerda
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Demais dados centralizados
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                
                # Bordas e espaçamento
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),  # Borda externa grossa
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e3a8a')),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                
                # Alternância de cores nas linhas
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            
            story.append(table)
            
            # Adicionar quebra de página apenas se não for a última parte            if i < len(tabelas_divididas) - 1:        story.append(PageBreak())
        
        # Quebra de página após Evolução Detalhada        story.append(PageBreak())
    
    # Adicionar tabelas por ano
    if resumo_projecao_por_ano:
        # Não adicionar título geral aqui, cada ano terá sua própria página
        # story.append(Paragraph("3. PROJEÇÃO POR ANO", section_style))
        # story.append(Spacer(1, 0.3*cm))
        
        for ano, dados_ano in sorted(resumo_projecao_por_ano.items()):
            story.append(PageBreak())  # Nova página para cada ano
            story.append(Spacer(1, 0.5*cm))
            story.append(Paragraph(f"3.{ano}. PROJEÇÃO ANO {ano}", section_style))
            story.append(Spacer(1, 0.3*cm))
            
            table_data = [['Categoria', 'Inicial', 'Nasc', 'Comp', 'Vend', 'Trans', 'Mort', 'Evol', 'Final', 'Peso', 'R$/Cabeça', 'Total']]
            
            for categoria, dados in dados_ano.items():
                trans = f"+{dados.get('transferencias_entrada', 0)}/{dados.get('transferencias_saida', 0)}" if dados.get('transferencias_entrada', 0) > 0 or dados.get('transferencias_saida', 0) > 0 else "0"
                
        table_data.append([
                    categoria,  # Nome completo sem truncamento
                    str(dados.get('saldo_inicial', 0) or 0),
                    str(dados.get('nascimentos', 0) or 0),
                    str(dados.get('compras', 0) or 0),
                    str(dados.get('vendas', 0) or 0),
                    trans,
                    str(dados.get('mortes', 0) or 0),
                    str(dados.get('evolucao_categoria', '-') or '-'),
                    str(dados.get('saldo_final', 0) or 0),
                    f"{float(dados.get('peso_medio_kg', 0) or 0):,.1f}",
                    f"R$ {float(dados.get('valor_unitario', 0) or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {float(dados.get('valor_total', 0) or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
        
        # Calcular larguras dinamicamente
        col_widths_projecao = calcular_larguras_dinamicas(table_data, 12)
        
        # Ajustar fonte baseado no conteúdo
        fonte_tabela = ajustar_fonte_por_conteudo(table_data, 9)
        fonte_cabecalho = fonte_tabela + 2
        
        # Dividir tabela se necessário
        tabelas_divididas = ajustar_tabela_para_pagina(table_data, col_widths_projecao)
        
        for j, parte_tabela in enumerate(tabelas_divididas):
            table = Table(parte_tabela, colWidths=col_widths_projecao)
            table.setStyle(TableStyle([
                    # Cabeçalho da tabela
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6495ed')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), fonte_cabecalho),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    
                    # Dados da tabela
                    ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -2), fonte_tabela),
                    ('FONTNAME', (0, 1), (0, -2), 'Helvetica-Bold'),  # Categoria em negrito
                    ('ALIGN', (0, 1), (0, -2), 'LEFT'),  # Categoria à esquerda
                    ('ALIGN', (1, 1), (-1, -2), 'CENTER'),  # Demais dados centralizados
                    ('VALIGN', (0, 1), (-1, -2), 'MIDDLE'),
                    
                    # Linha de totais (apenas na última parte)
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, -1), (-1, -1), fonte_tabela + 1),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4fd')),
                    ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#6495ed')),
                    ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
                    
                    # Bordas e espaçamento
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                    ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#6495ed')),  # Borda externa grossa
                    ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#6495ed')),
                    ('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#6495ed')),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    
                    # Alternância de cores nas linhas
                    ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
                ]))
            
            story.append(table)
            
            # Adicionar quebra de página apenas se não for a última parte da última tabela
            if j < len(tabelas_divididas) - 1:
                story.append(PageBreak())
            
            # Quebra de página após cada ano (exceto o último)
            if ano != max(resumo_projecao_por_ano.keys()):
                story.append(PageBreak())
    
    # Análise Financeira
    story.append(PageBreak())  # Nova página para Análise Financeira
    story.append(Paragraph("4. ANÁLISE FINANCEIRA", section_style))
    story.append(Spacer(1, 0.3*cm))
    
    # Calcular receitas e despesas por ano
    analise_financeira_data = [['Ano', 'Receitas (R$)', 'Despesas (R$)', 'Lucro (R$)', 'Margem (%)', 'ROI (%)']]
    
    receita_total_periodo = 0
    despesa_total_periodo = 0
    
    for ano, dados_ano in sorted(resumo_projecao_por_ano.items()):
        receitas_ano = sum(float(dados.get('valor_total', 0) or 0) for dados in dados_ano.values() if dados.get('vendas', 0) > 0)
        despesas_ano = sum(float(dados.get('valor_total', 0) or 0) for dados in dados_ano.values() if dados.get('compras', 0) > 0)
        lucro_ano = receitas_ano - despesas_ano
        margem_ano = (lucro_ano / receitas_ano * 100) if receitas_ano > 0 else 0
        roi_ano = (lucro_ano / valor_total_rebanho * 100) if valor_total_rebanho > 0 else 0
        
        receita_total_periodo += receitas_ano
        despesa_total_periodo += despesas_ano
        
        analise_financeira_data.append([
            str(ano),
            f"R$ {receitas_ano:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {despesas_ano:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {lucro_ano:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"{margem_ano:.1f}%",
            f"{roi_ano:.1f}%"
        ])
    
    # Adicionar linha de totais
    lucro_total = receita_total_periodo - despesa_total_periodo
    margem_total = (lucro_total / receita_total_periodo * 100) if receita_total_periodo > 0 else 0
    roi_total = (lucro_total / valor_total_rebanho * 100) if valor_total_rebanho > 0 else 0
    
    analise_financeira_data.append([
        'TOTAL PERÍODO',
        f"R$ {receita_total_periodo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
        f"R$ {despesa_total_periodo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
        f"R$ {lucro_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
        f"{margem_total:.1f}%",
        f"{roi_total:.1f}%"
    ])
    
    analise_table = Table(analise_financeira_data, colWidths=[3.5*cm, 5.5*cm, 5.5*cm, 5.5*cm, 3.5*cm, 3.5*cm])
    analise_table.setStyle(TableStyle([
        # Cabeçalho da tabela
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Dados da tabela
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 11),
        ('ALIGN', (0, 1), (0, -2), 'CENTER'),  # Ano centralizado
        ('ALIGN', (1, 1), (-1, -2), 'RIGHT'),  # Valores à direita
        ('VALIGN', (0, 1), (-1, -2), 'MIDDLE'),
        
        # Linha de totais
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4fd')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1e3a8a')),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
        
        # Bordas e espaçamento
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a8a')),  # Borda externa grossa
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e3a8a')),
        ('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#1e3a8a')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        
        # Alternância de cores nas linhas
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(analise_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Resumo dos indicadores de rentabilidade
    story.append(Paragraph("INDICADORES DE RENTABILIDADE", subsection_style))
    story.append(Spacer(1, 0.2*cm))
    
    indicadores_rentabilidade_data = [
        ['Indicador', 'Valor', 'Interpretação'],
        ['Margem Bruta Média', f"{margem_total:.1f}%", 'Excelente' if margem_total > 20 else 'Boa' if margem_total > 10 else 'Regular' if margem_total > 0 else 'Negativa'],
        ['ROI Médio', f"{roi_total:.1f}%", 'Excelente' if roi_total > 15 else 'Bom' if roi_total > 8 else 'Regular' if roi_total > 0 else 'Negativo'],
        ['Receita Total Projetada', f"R$ {receita_total_periodo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), f"Em {len(resumo_projecao_por_ano)} anos"],
        ['Lucro Total Projetado', f"R$ {lucro_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), f"Período {min(resumo_projecao_por_ano.keys())}-{max(resumo_projecao_por_ano.keys())}"],
    ]
    
    indicadores_table = Table(indicadores_rentabilidade_data, colWidths=[7*cm, 5*cm, 7*cm])
    indicadores_table.setStyle(TableStyle([
        # Cabeçalho da tabela
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6495ed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Dados da tabela
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Indicador à esquerda
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),  # Valor à direita
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Interpretação à esquerda
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        
        # Bordas e espaçamento
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#6495ed')),  # Borda externa grossa
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#6495ed')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        
        # Alternância de cores nas linhas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(indicadores_table)
    story.append(PageBreak())
    
    # Função para adicionar rodapé com numeração
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        
        # Rodapé esquerdo - Propriedade
        canvas.drawString(1*cm, 0.5*cm, f"Propriedade: {propriedade.nome_propriedade}")
        
        # Rodapé direito - Data e página
        canvas.drawRightString(28.7*cm, 0.5*cm, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        canvas.drawRightString(28.7*cm, 0.2*cm, f"Página {doc.page}")
        
        canvas.restoreState()
    
    # Construir PDF com rodapé
    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    return response


@login_required
def exportar_iatf_excel(request, propriedade_id):
    """Exporta o relatório completo de IATF em formato Excel."""
    from .models import Propriedade

    propriedade = get_object_or_404(
        Propriedade, id=propriedade_id, produtor__usuario_responsavel=request.user
    )

    filtros = {
        'data_inicio': request.GET.get('data_inicio'),
        'data_final': request.GET.get('data_final'),
        'diagnostico_ate': request.GET.get('diagnostico_ate'),
        'lote_id': request.GET.get('lote_id'),
        'resultado': request.GET.get('resultado'),
    }

    try:
        dados = coletar_dados_iatf(propriedade, filtros)
    except IATFModuloIndisponivel as exc:
        return HttpResponse(str(exc), status=400, content_type='text/plain; charset=utf-8')

    resumo = dados['resumo']
    filtros_aplicados = dados['filtros_aplicados']

    def _to_float(value):
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _aplicar_header(cells):
        for cell in cells:
            cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _escrever_tabela(ws, headers, linhas):
        ws.append(headers)
        _aplicar_header(ws[ws.max_row])
        for linha in linhas:
            ws.append(linha)

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    ws_resumo['A1'] = "Relatório detalhado de IATF"
    ws_resumo['A1'].font = Font(bold=True, size=16)
    ws_resumo['A2'] = "Propriedade"
    ws_resumo['B2'] = getattr(propriedade, 'nome_propriedade', str(propriedade))
    ws_resumo['A3'] = "Gerado em"
    ws_resumo['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')

    filtros_texto = []
    if filtros_aplicados.get('data_inicio'):
        filtros_texto.append(f"IATF ≥ {filtros_aplicados['data_inicio'].strftime('%d/%m/%Y')}")
    if filtros_aplicados.get('data_final'):
        filtros_texto.append(f"IATF ≤ {filtros_aplicados['data_final'].strftime('%d/%m/%Y')}")
    if filtros_aplicados.get('diagnostico_ate'):
        filtros_texto.append(f"Diagnóstico ≤ {filtros_aplicados['diagnostico_ate'].strftime('%d/%m/%Y')}")
    if filtros_aplicados.get('lote_id'):
        filtros_texto.append(f"Lote #{filtros_aplicados['lote_id']}")
    if filtros_aplicados.get('resultado'):
        filtros_texto.append(f"Resultado = {filtros_aplicados['resultado']}")

    ws_resumo['A4'] = "Filtros"
    ws_resumo['B4'] = ', '.join(filtros_texto) if filtros_texto else 'Sem filtros adicionais'

    ws_resumo.append([])
    ws_resumo.append(['Indicador', 'Valor'])
    _aplicar_header(ws_resumo[ws_resumo.max_row])

    metricas = [
        ("Total de IATFs avaliadas", resumo['total_animais']),
        ("Prenhezes confirmadas", resumo['total_prenhezes']),
        ("Taxa de prenhez (%)", _to_float(resumo['taxa_prenhez'])),
        ("Vazias", resumo['total_vazias']),
        ("Abortos", resumo['total_abortos']),
        ("Repetições de cio", resumo['total_repeticoes']),
        ("Diagnósticos pendentes", resumo['total_pendentes']),
        ("Dias médios até diagnóstico", _to_float(resumo['media_dias_diagnostico'])),
        ("Custo total (R$)", _to_float(resumo['custo_total'])),
        ("Custo médio por animal (R$)", _to_float(resumo['custo_medio_animal'])),
        ("Custo médio por prenhez (R$)", _to_float(resumo['custo_medio_prenhez'])),
    ]
    for label, valor in metricas:
        ws_resumo.append([label, valor if valor is not None else '-'])

    for row in ws_resumo.iter_rows(min_row=7, min_col=2, max_col=2, max_row=ws_resumo.max_row):
        celula = row[0]
        if isinstance(celula.value, (int, float)):
            indicador = ws_resumo[f"A{celula.row}"].value or ''
            if 'Custo' in indicador:
                celula.number_format = 'R$ #,##0.00'
            elif 'Taxa' in indicador:
                celula.number_format = '0.00'
            elif 'Dias' in indicador:
                celula.number_format = '0.00'

    ws_resumo.append([])
    ws_resumo.append(['Resultados', 'Quantidade', 'Taxa (%)'])
    _aplicar_header(ws_resumo[ws_resumo.max_row])
    for item in dados['status_counts']:
        ws_resumo.append([
            item['rotulo'],
            item['total'],
            _to_float(item['taxa']),
        ])
    for row in ws_resumo.iter_rows(
        min_row=ws_resumo.max_row - len(dados['status_counts']) + 1,
        min_col=3,
        max_col=3,
        max_row=ws_resumo.max_row,
    ):
        row[0].number_format = '0.00'

    ws_resumo.column_dimensions['A'].width = 42
    ws_resumo.column_dimensions['B'].width = 28
    ws_resumo.freeze_panes = 'A7'

    def _planilha_agrupamento(nome, dados_agrupados, incluir_custos=True):
        ws = wb.create_sheet(nome)
        headers = [
            nome.split(' ', 1)[-1],
            'Total',
            'Prenhezes',
            'Taxa Prenhez (%)',
            'Vazias',
            'Abortos',
            'Repetições',
            'Dias médios diagnóstico',
        ]
        if incluir_custos:
            headers.append('Custo médio/prenhez (R$)')

        linhas = []
        for item in dados_agrupados:
            linha = [
                item['rotulo'],
                item['total'],
                item['prenhezes'],
                _to_float(item['taxa_prenhez']),
                item['vazias'],
                item['abortos'],
                item['repeticoes'],
                _to_float(item['media_dias_diagnostico']),
            ]
            if incluir_custos:
                linha.append(_to_float(item['custo_medio_prenhez']))
            linhas.append(linha)

        _escrever_tabela(ws, headers, linhas)

        if linhas:
            ws.freeze_panes = 'A2'
            ws.column_dimensions['A'].width = 32
            for col in range(2, ws.max_column + 1):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.auto_filter.ref = ws.dimensions
            taxa_col = 4
            dias_col = 8
            custo_col = 9 if incluir_custos else None
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if _to_float(row[taxa_col - 1].value) is not None:
                    row[taxa_col - 1].number_format = '0.00'
                if _to_float(row[dias_col - 1].value) is not None:
                    row[dias_col - 1].number_format = '0.00'
                if custo_col and _to_float(row[custo_col - 1].value) is not None:
                    row[custo_col - 1].number_format = 'R$ #,##0.00'

            if nome == "Por Touro" and len(linhas) >= 1:
                chart = BarChart()
                chart.title = "Prenhezes por touro"
                data_ref = Reference(ws, min_col=3, min_row=1, max_row=len(linhas) + 1)
                categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(linhas) + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(categories_ref)
                chart.y_axis.title = "Prenhezes"
                chart.x_axis.title = "Touros"
                ws.add_chart(chart, f'{chr(65 + ws.max_column + 1)}2')

    _planilha_agrupamento("Por Touro", dados['por_touro'])
    _planilha_agrupamento("Por Inseminador", dados['por_inseminador'])
    _planilha_agrupamento("Por Invernada", dados['por_invernada'], incluir_custos=False)
    _planilha_agrupamento("Por Lote", dados['por_lote'])
    _planilha_agrupamento("Por Protocolo", dados['por_protocolo'])

    ws_diagnosticos = wb.create_sheet("Diagnosticos")
    linhas_diagnosticos = [
        [
            item['data_diagnostico'],
            item['total'],
            item['prenhezes'],
            _to_float(item['taxa_prenhez']),
            item['vazias'],
        ]
        for item in dados['diagnosticos_por_data']
    ]
    _escrever_tabela(
        ws_diagnosticos,
        ['Data do diagnóstico', 'Total', 'Prenhezes', 'Taxa Prenhez (%)', 'Vazias'],
        linhas_diagnosticos,
    )
    if linhas_diagnosticos:
        ws_diagnosticos.freeze_panes = 'A2'
        ws_diagnosticos.column_dimensions['A'].width = 18
        for row in ws_diagnosticos.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws_diagnosticos.max_row):
            row[0].number_format = '0.00'

    ws_detalhes = wb.create_sheet("Detalhes")
    ws_detalhes.append([
        'Animal',
        'Categoria',
        'Invernada',
        'Protocolo',
        'Lote IATF',
        'Touro',
        'Inseminador',
        'Resultado',
        'Status',
        'Data IATF',
        'Data diagnóstico',
        'Dias até diagnóstico',
        'Custo total (R$)',
        'Observações',
    ])
    _aplicar_header(ws_detalhes[1])

    for item in dados['detalhes']:
        ws_detalhes.append([
            item['animal'],
            item['categoria'],
            item['invernada'],
            item['protocolo'],
            item['lote_iatf'],
            item['touro'],
            item['inseminador'],
            item['resultado'],
            item['status'],
            item['data_iatf'],
            item['data_diagnostico'],
            item['dias_para_diagnostico'],
            _to_float(item['custo_total']),
            item['observacoes'],
        ])

    ws_detalhes.freeze_panes = 'A2'
    col_widths = {
        'A': 16,
        'B': 20,
        'C': 24,
        'D': 22,
        'E': 22,
        'F': 26,
        'G': 22,
        'H': 18,
        'I': 18,
        'J': 14,
        'K': 18,
        'L': 18,
        'M': 18,
        'N': 48,
    }
    for coluna, largura in col_widths.items():
        ws_detalhes.column_dimensions[coluna].width = largura
    for row in ws_detalhes.iter_rows(min_row=2, min_col=13, max_col=13, max_row=ws_detalhes.max_row):
        row[0].number_format = 'R$ #,##0.00'

    nome_arquivo = f"iatf_{propriedade.id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response


@login_required
def exportar_iatf_pdf(request, propriedade_id):
    """Exporta o relatório completo de IATF em PDF com tabelas analíticas."""
    from .models import Propriedade

    propriedade = get_object_or_404(
        Propriedade, id=propriedade_id, produtor__usuario_responsavel=request.user
    )

    filtros = {
        'data_inicio': request.GET.get('data_inicio'),
        'data_final': request.GET.get('data_final'),
        'diagnostico_ate': request.GET.get('diagnostico_ate'),
        'lote_id': request.GET.get('lote_id'),
        'resultado': request.GET.get('resultado'),
    }

    try:
        dados = coletar_dados_iatf(propriedade, filtros)
    except IATFModuloIndisponivel as exc:
        return HttpResponse(str(exc), status=400, content_type='text/plain; charset=utf-8')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="iatf_{propriedade.id}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=1.7 * cm,
        rightMargin=1.7 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'IATFTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    subtitle_style = ParagraphStyle(
        'IATFSubtitle',
        parent=styles['BodyText'],
        fontSize=11,
        textColor=colors.HexColor('#4a5568'),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    section_style = ParagraphStyle(
        'IATFSection',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=8,
        spaceBefore=12,
    )

    def format_percent(valor):
        if valor is None:
            return '-'
        return f"{float(valor):.2f}%".replace('.', ',')

    def format_decimal(valor, casas=2):
        if valor is None:
            return '-'
        return f"{float(valor):.{casas}f}".replace('.', ',')

    def format_currency(valor):
        if valor is None:
            return '-'
        texto = f"{float(valor):,.2f}"
        texto = texto.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {texto}"

    def format_date(valor):
        if not valor:
            return '-'
        return valor.strftime('%d/%m/%Y')

    story = []
    story.append(Paragraph("Relatório detalhado de IATF", title_style))
    story.append(Paragraph(getattr(propriedade, 'nome_propriedade', str(propriedade)), subtitle_style))
    story.append(Paragraph(datetime.now().strftime('%d/%m/%Y %H:%M'), subtitle_style))

    filtros_aplicados = dados['filtros_aplicados']
    filtros_texto = []
    if filtros_aplicados.get('data_inicio'):
        filtros_texto.append(f"IATFs a partir de {format_date(filtros_aplicados['data_inicio'])}")
    if filtros_aplicados.get('data_final'):
        filtros_texto.append(f"IATFs até {format_date(filtros_aplicados['data_final'])}")
    if filtros_aplicados.get('diagnostico_ate'):
        filtros_texto.append(f"Diagnóstico até {format_date(filtros_aplicados['diagnostico_ate'])}")
    if filtros_aplicados.get('resultado'):
        filtros_texto.append(f"Resultado filtrado: {filtros_aplicados['resultado']}")
    if filtros_aplicados.get('lote_id'):
        filtros_texto.append(f"Lote específico #{filtros_aplicados['lote_id']}")

    if filtros_texto:
        story.append(Paragraph(' | '.join(filtros_texto), ParagraphStyle(
            'Filtros',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#718096'),
            alignment=TA_CENTER,
            spaceAfter=12,
        )))

    resumo = dados['resumo']
    tabela_resumo = [
        ['Indicador', 'Valor'],
        ['Total de IATFs avaliadas', resumo['total_animais']],
        ['Prenhezes confirmadas', resumo['total_prenhezes']],
        ['Taxa de prenhez', format_percent(resumo['taxa_prenhez'])],
        ['Vazias', resumo['total_vazias']],
        ['Abortos', resumo['total_abortos']],
        ['Repetições de cio', resumo['total_repeticoes']],
        ['Diagnósticos pendentes', resumo['total_pendentes']],
        ['Dias médios até diagnóstico', format_decimal(resumo['media_dias_diagnostico'])],
        ['Custo total', format_currency(resumo['custo_total'])],
        ['Custo médio por animal', format_currency(resumo['custo_medio_animal'])],
        ['Custo médio por prenhez', format_currency(resumo['custo_medio_prenhez'])],
    ]
    tabela_resumo_obj = Table(tabela_resumo, colWidths=[9 * cm, 7 * cm], hAlign='LEFT')
    tabela_resumo_obj.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5F5')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tabela_resumo_obj)
    story.append(Spacer(1, 0.4 * cm))

    if dados['status_counts']:
        story.append(Paragraph("Distribuição por resultado", section_style))
        tabela_status = [['Resultado', 'Quantidade', 'Taxa']]
        for item in dados['status_counts']:
            tabela_status.append([
                item['rotulo'],
                item['total'],
                format_percent(item['taxa']),
            ])
        tabela_status_obj = Table(tabela_status, hAlign='LEFT')
        tabela_status_obj.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5F5')),
        ]))
        story.append(tabela_status_obj)
        story.append(Spacer(1, 0.4 * cm))

    def _adicionar_secao_tabela(titulo, dados_agrupados, incluir_custos=True):
        if not dados_agrupados:
            return
        story.append(Paragraph(titulo, section_style))
        headers = [
            titulo.split(' ', 1)[-1],
            'Total',
            'Prenhezes',
            'Taxa',
            'Vazias',
            'Abortos',
            'Repetições',
            'Dias médios',
        ]
        if incluir_custos:
            headers.append('Custo médio/prenhez')
        tabela = [headers]
        for item in dados_agrupados:
            linha = [
                item['rotulo'],
                item['total'],
                item['prenhezes'],
                format_percent(item['taxa_prenhez']),
                item['vazias'],
                item['abortos'],
                item['repeticoes'],
                format_decimal(item['media_dias_diagnostico']),
            ]
            if incluir_custos:
                linha.append(format_currency(item['custo_medio_prenhez']))
            tabela.append(linha)

        colunas = len(headers)
        largura_base = [6.5 * cm] + [2.3 * cm] * (colunas - 1)
        tabela_obj = Table(tabela, colWidths=largura_base, repeatRows=1, hAlign='LEFT')
        tabela_obj.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5F5')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(tabela_obj)
        story.append(Spacer(1, 0.3 * cm))

    _adicionar_secao_tabela("Desempenho por touro", dados['por_touro'])
    _adicionar_secao_tabela("Desempenho por inseminador", dados['por_inseminador'])
    _adicionar_secao_tabela("Desempenho por invernada", dados['por_invernada'], incluir_custos=False)
    _adicionar_secao_tabela("Desempenho por lote", dados['por_lote'])
    _adicionar_secao_tabela("Desempenho por protocolo", dados['por_protocolo'])

    if dados['diagnosticos_por_data']:
        story.append(Paragraph("Diagnósticos ao longo do tempo", section_style))
        tabela_diag = [['Data', 'Total', 'Prenhezes', 'Taxa', 'Vazias']]
        for item in dados['diagnosticos_por_data']:
            tabela_diag.append([
                format_date(item['data_diagnostico']),
                item['total'],
                item['prenhezes'],
                format_percent(item['taxa_prenhez']),
                item['vazias'],
            ])
        tabela_diag_obj = Table(tabela_diag, repeatRows=1, hAlign='LEFT')
        tabela_diag_obj.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5F5')),
        ]))
        story.append(tabela_diag_obj)
        story.append(Spacer(1, 0.3 * cm))

    detalhes = dados['detalhes']
    if detalhes:
        story.append(Paragraph("Detalhamento por animal", section_style))
        limite = 40
        if len(detalhes) > limite:
            aviso = Paragraph(
                f"Mostrando {limite} de {len(detalhes)} registros. Utilize a exportação Excel para obter a lista completa.",
                ParagraphStyle('Aviso', parent=styles['Italic'], textColor=colors.HexColor('#718096'), spaceAfter=6),
            )
            story.append(aviso)
        tabela_detalhes = [['Animal', 'Resultado', 'Touro', 'Lote', 'Diagnóstico', 'Dias', 'Custo', 'Observações']]
        for item in detalhes[:limite]:
            tabela_detalhes.append([
                item['animal'],
                item['resultado'],
                item['touro'],
                item['lote_iatf'],
                format_date(item['data_diagnostico']),
                item['dias_para_diagnostico'] if item['dias_para_diagnostico'] is not None else '-',
                format_currency(item['custo_total']),
                (item['observacoes'] or '')[:160],
            ])
        tabela_detalhes_obj = Table(
            tabela_detalhes,
            colWidths=[3.2 * cm, 3 * cm, 3.5 * cm, 3.5 * cm, 3 * cm, 1.8 * cm, 3 * cm, 5 * cm],
            repeatRows=1,
            hAlign='LEFT',
        )
        tabela_detalhes_obj.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (1, 1), (-2, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5F5')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(tabela_detalhes_obj)

    doc.build(story)
    return response


@login_required
def exportar_inventario_pdf(request, propriedade_id):
    """Exporta inventário para PDF"""
    from .models import InventarioRebanho, Propriedade
    from datetime import datetime
    
    propriedade = get_object_or_404(Propriedade, id=propriedade_id, produtor__usuario_responsavel=request.user)
    inventario = InventarioRebanho.objects.filter(propriedade=propriedade).select_related('categoria')
    
    # Criar resposta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventario_{propriedade.id}.pdf"'
    
    # Criar documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Título
    story.append(Paragraph("Inventário do Rebanho", title_style))
    story.append(Paragraph(f"<b>Propriedade:</b> {propriedade.nome_propriedade}", styles['Normal']))
    story.append(Paragraph(f"<b>Data:</b> {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Tabela de inventário
    table_data = [['Categoria', 'Quantidade', 'Valor por Cabeça', 'Valor Total']]
    total_geral = Decimal('0')
    
    for item in inventario:
        valor_total = float(item.valor_total) if item.valor_total else 0
        total_geral += Decimal(str(valor_total))
        table_data.append([
            item.categoria.nome,
            str(item.quantidade),
            f"R$ {float(item.valor_por_cabeca):,.2f}",
            f"R$ {valor_total:,.2f}"
        ])
    
    # Adicionar total
    table_data.append(['TOTAL', '', '', f"R$ {float(total_geral):,.2f}"])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d3d3d3')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    return response

