# -*- coding: utf-8 -*-
"""
Views para Sistema de Rastreabilidade Bovina - PNIB
"""

import re
import traceback
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, Max
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from collections import defaultdict

from .models import (
    Propriedade, CategoriaAnimal, AnimalIndividual,
    MovimentacaoIndividual, BrincoAnimal, CurralLote
)


def _normalizar_codigo(codigo: str) -> str:
    """Remove caracteres não numéricos e devolve o código limpo."""
    if not codigo:
        return ''
    return re.sub(r'\D', '', codigo)


def _normalizar_codigo_sisbov(codigo):
    """
    Normaliza código SISBOV para formato padrão BR + 13 dígitos.
    Remove espaços, caracteres especiais e garante formato consistente.
    """
    if not codigo:
        return None
    
    codigo_str = str(codigo).strip().upper()
    
    # Remover todos os caracteres não numéricos (exceto BR no início)
    # Primeiro, verificar se começa com BR
    tem_br = codigo_str.startswith('BR')
    codigo_limpo = re.sub(r'\D', '', codigo_str)
    
    # Se tem 13 dígitos, adicionar BR se não tiver
    if len(codigo_limpo) == 13:
        return f"BR{codigo_limpo}"
    # Se tem 15 dígitos, remover os 2 primeiros e adicionar BR
    elif len(codigo_limpo) == 15:
        return f"BR{codigo_limpo[2:]}"
    # Se tem menos de 13 dígitos mas começa com BR, tentar usar como está
    elif tem_br and len(codigo_limpo) >= 11:
        # Se já tem BR e tem pelo menos 11 dígitos, tentar normalizar
        if len(codigo_limpo) == 13:
            return f"BR{codigo_limpo}"
        elif len(codigo_limpo) == 15:
            return f"BR{codigo_limpo[2:]}"
    # Se já está no formato BR + 13 dígitos (string com BR)
    elif tem_br and len(codigo_limpo) == 13:
        return f"BR{codigo_limpo}"
    
    # Se não conseguiu normalizar, retornar None
    return None


def _processar_animal_importacao(animal_data, propriedade, resultados_detalhados):
    """
    Processa um animal da importação, cria/atualiza e classifica status de conformidade.
    Retorna (animal, criado, atualizado, status_conformidade, divergencias)
    """
    codigo_sisbov_raw = animal_data.get('codigo_sisbov', '').strip()
    if not codigo_sisbov_raw:
        return None, False, False, None, []
    
    # Normalizar código SISBOV
    codigo_normalizado = _normalizar_codigo_sisbov(codigo_sisbov_raw)
    if not codigo_normalizado:
        return None, False, False, None, []
    
    numero_brinco = animal_data.get('numero_brinco', '').strip()
    raca = animal_data.get('raca', '').strip() or None
    sexo = animal_data.get('sexo')
    if sexo and sexo not in ['M', 'F']:
        sexo = None
    data_nascimento = animal_data.get('data_nascimento')
    peso_atual_kg = animal_data.get('peso_kg')
    
    # Buscar animal existente
    animal_existente = AnimalIndividual.objects.filter(
        codigo_sisbov=codigo_normalizado,
        propriedade=propriedade
    ).first()
    
    # Criar ou atualizar
    animal, criado = AnimalIndividual.objects.get_or_create(
        codigo_sisbov=codigo_normalizado,
        propriedade=propriedade,
        defaults={
            'numero_brinco': numero_brinco or codigo_normalizado[-6:] if codigo_normalizado else '',
            'raca': raca or '',
            'sexo': sexo or 'I',
            'data_nascimento': data_nascimento,
            'peso_atual_kg': peso_atual_kg,
            'status': 'ATIVO',
            'status_sanitario': 'INDEFINIDO'
        }
    )
    
    status_conformidade = None
    divergencias = []
    atualizado = False
    
    if criado:
        status_conformidade = 'NOVO'
        resultados_detalhados['animais_criados_lista'].append({
            'animal_id': animal.id,
            'codigo_sisbov': codigo_normalizado,
            'numero_brinco': animal.numero_brinco or '',
            'status': 'NOVO',
        })
    else:
        # Verificar conformidade
        tem_divergencia = False
        
        # Comparar sexo
        sexo_sistema = animal.sexo if animal.sexo else 'I'
        sexo_arquivo = sexo or 'I'
        if sexo_sistema != sexo_arquivo and sexo_arquivo != 'I':
            divergencias.append('sexo')
            tem_divergencia = True
        
        # Comparar raça
        raca_sistema = animal.raca or ''
        raca_arquivo = raca or ''
        if raca_sistema and raca_arquivo and raca_sistema.upper().strip() != raca_arquivo.upper().strip():
            divergencias.append('raca')
            tem_divergencia = True
        
        # Comparar data de nascimento
        if data_nascimento and animal.data_nascimento:
            if isinstance(data_nascimento, str):
                try:
                    data_arquivo_obj = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
                except:
                    try:
                        data_arquivo_obj = datetime.strptime(data_nascimento, '%d/%m/%Y').date()
                    except:
                        data_arquivo_obj = None
            else:
                data_arquivo_obj = data_nascimento
            
            if data_arquivo_obj and animal.data_nascimento != data_arquivo_obj:
                divergencias.append('data_nascimento')
                tem_divergencia = True
        
        # Atualizar dados se necessário
        if raca and not animal.raca:
            animal.raca = raca
            atualizado = True
        if sexo and sexo != 'I' and animal.sexo == 'I':
            animal.sexo = sexo
            atualizado = True
        if data_nascimento and not animal.data_nascimento:
            if isinstance(data_nascimento, str):
                try:
                    animal.data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
                except:
                    try:
                        animal.data_nascimento = datetime.strptime(data_nascimento, '%d/%m/%Y').date()
                    except:
                        pass
            else:
                animal.data_nascimento = data_nascimento
            atualizado = True
        if peso_atual_kg:
            animal.peso_atual_kg = peso_atual_kg
            atualizado = True
        
        if atualizado:
            animal.save()
        
        # Classificar status
        if tem_divergencia:
            status_conformidade = 'DIVERGENTE'
            resultados_detalhados['animais_divergentes_lista'].append({
                'animal_id': animal.id,
                'codigo_sisbov': codigo_normalizado,
                'numero_brinco': animal.numero_brinco or '',
                'divergencias': divergencias,
                'atualizado': atualizado,
            })
        else:
            status_conformidade = 'CONFORME'
            resultados_detalhados['animais_conformes_lista'].append({
                'animal_id': animal.id,
                'codigo_sisbov': codigo_normalizado,
                'numero_brinco': animal.numero_brinco or '',
                'atualizado': atualizado,
            })
        
        if atualizado:
            resultados_detalhados['animais_atualizados_lista'].append({
                'animal_id': animal.id,
                'codigo_sisbov': codigo_normalizado,
                'numero_brinco': animal.numero_brinco or '',
                'status': status_conformidade,
            })
    
    return animal, criado, atualizado, status_conformidade, divergencias


def _calcular_era(data_nascimento, categoria=None):
    """
    Calcula a Era (faixa etária) do animal baseado na data de nascimento.
    Retorna faixas de idade em meses: '0-12', '12-24', '24-36', '+36'
    """
    if data_nascimento:
        hoje = date.today()
        idade_meses = (hoje.year - data_nascimento.year) * 12 + (hoje.month - data_nascimento.month)
        
        if idade_meses < 12:
            return '0-12'
        elif idade_meses < 24:
            return '12-24'
        elif idade_meses < 36:
            return '24-36'
        else:
            return '+36'
    
    # Se não tem data, usar categoria como fallback aproximado
    if categoria:
        nome_cat = categoria.nome.lower() if hasattr(categoria, 'nome') else str(categoria).lower()
        if 'bezerro' in nome_cat:
            return '0-12'
        elif 'novilho' in nome_cat or 'novilha' in nome_cat:
            return '12-24'
        elif 'touro' in nome_cat or 'vaca' in nome_cat:
            # Pode ser qualquer faixa acima de 24 meses
            return '24-36'
    
    return 'Indefinido'


@login_required
def rastreabilidade_dashboard(request, propriedade_id):
    """Dashboard principal de rastreabilidade bovina"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Estatísticas gerais
    total_animais = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        status='ATIVO'
    ).count()
    
    animais_por_categoria = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        status='ATIVO'
    ).values('categoria__nome').annotate(
        total=Count('id')
    ).order_by('-total')
    
    animais_por_status = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).values('status').annotate(
        total=Count('id')
    ).order_by('status')
    
    # Movimentações recentes
    movimentacoes_recentes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade
    ).select_related('animal', 'animal__categoria').order_by('-data_movimentacao')[:10]
    
    # Brincos disponíveis
    brincos_disponiveis = BrincoAnimal.objects.filter(
        propriedade=propriedade,
        status='DISPONIVEL'
    ).count()
    
    brincos_em_uso = BrincoAnimal.objects.filter(
        propriedade=propriedade,
        status='EM_USO'
    ).count()
    
    context = {
        'propriedade': propriedade,
        'total_animais': total_animais,
        'animais_por_categoria': animais_por_categoria,
        'animais_por_status': animais_por_status,
        'movimentacoes_recentes': movimentacoes_recentes,
        'brincos_disponiveis': brincos_disponiveis,
        'brincos_em_uso': brincos_em_uso,
        'status_labels': dict(AnimalIndividual.STATUS_CHOICES),
        'status_sanitario_labels': dict(AnimalIndividual.STATUS_SANITARIO_CHOICES),
    }
    
    return render(request, 'gestao_rural/rastreabilidade_dashboard.html', context)


@login_required
def preview_importacao_bnd_sisbov(request, propriedade_id):
    """Preview da importação - retorna comparação entre sistema e arquivo"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método não permitido'}, status=405)
    
    try:
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return JsonResponse({'erro': 'Nenhum arquivo enviado'}, status=400)
        
        nome_arquivo = arquivo.name.lower()
        animais_arquivo = []
        
        # Processar arquivo conforme tipo
        if nome_arquivo.endswith('.xlsx') or nome_arquivo.endswith('.xls'):
            import openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
            
            # Encontrar cabeçalho
            cabecalho = None
            linha_cabecalho = 0
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if any(cell and isinstance(cell, str) and 'sisbov' in cell.lower() for cell in row):
                    cabecalho = [str(cell).lower().strip() if cell else '' for cell in row]
                    linha_cabecalho = idx
                    break
            
            if not cabecalho:
                return JsonResponse({'erro': 'Cabeçalho não encontrado no arquivo'}, status=400)
            
            # Mapear colunas
            col_indices = {}
            colunas_esperadas = {
                'sisbov': ['sisbov', 'codigo_sisbov', 'código sisbov', 'codigo sisbov'],
                'brinco': ['brinco', 'numero_brinco', 'número brinco', 'numero brinco'],
                'raca': ['raca', 'raça', 'raça do animal'],
                'sexo': ['sexo', 'genero', 'gênero'],
                'data_nascimento': ['data_nascimento', 'data nascimento', 'nascimento', 'dt_nasc'],
                'peso': ['peso', 'peso_atual', 'peso atual', 'peso_kg']
            }
            
            for campo, variacoes in colunas_esperadas.items():
                for idx, col_name in enumerate(cabecalho):
                    if any(var in col_name for var in variacoes):
                        col_indices[campo] = idx
                        break
            
            # Processar linhas
            for idx, row in enumerate(ws.iter_rows(min_row=linha_cabecalho + 1, values_only=True), start=linha_cabecalho + 1):
                if not any(row):
                    continue
                
                codigo_sisbov_raw = str(row[col_indices['sisbov']]).strip() if col_indices.get('sisbov') is not None and row[col_indices['sisbov']] else None
                if not codigo_sisbov_raw or codigo_sisbov_raw.lower() in ['none', 'null', '']:
                    continue
                
                # Normalizar código SISBOV
                codigo_sisbov = _normalizar_codigo_sisbov(codigo_sisbov_raw)
                if not codigo_sisbov:
                    continue  # Código inválido, pular
                
                sexo = None
                if 'sexo' in col_indices and row[col_indices['sexo']]:
                    sexo_str = str(row[col_indices['sexo']]).strip().upper()
                    if sexo_str in ['M', 'MACHO', 'MALE']:
                        sexo = 'M'
                    elif sexo_str in ['F', 'FEMEA', 'FÊMEA', 'FEMALE']:
                        sexo = 'F'
                
                data_nascimento = None
                if 'data_nascimento' in col_indices and row[col_indices['data_nascimento']]:
                    data_val = row[col_indices['data_nascimento']]
                    if isinstance(data_val, datetime):
                        data_nascimento = data_val.date()
                    elif isinstance(data_val, date):
                        data_nascimento = data_val
                    elif isinstance(data_val, str):
                        try:
                            data_nascimento = datetime.strptime(data_val, '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                data_nascimento = datetime.strptime(data_val, '%d/%m/%Y').date()
                            except ValueError:
                                pass
                
                animais_arquivo.append({
                    'codigo_sisbov': codigo_sisbov,
                    'sexo': sexo or 'I',
                    'data_nascimento': data_nascimento.isoformat() if data_nascimento else None,
                })
            
            wb.close()
        
        elif nome_arquivo.endswith('.csv'):
            import csv
            import io
            
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            conteudo = None
            
            for encoding in encodings:
                try:
                    arquivo.seek(0)
                    conteudo = arquivo.read().decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if not conteudo:
                return JsonResponse({'erro': 'Erro ao ler arquivo CSV'}, status=400)
            
            csv_reader = csv.DictReader(io.StringIO(conteudo))
            
            for row in csv_reader:
                codigo_sisbov_raw = row.get('sisbov') or row.get('codigo_sisbov') or row.get('Código SISBOV') or ''
                codigo_sisbov_raw = str(codigo_sisbov_raw).strip()
                
                if not codigo_sisbov_raw:
                    continue
                
                # Normalizar código SISBOV
                codigo_sisbov = _normalizar_codigo_sisbov(codigo_sisbov_raw)
                if not codigo_sisbov:
                    continue  # Código inválido, pular
                
                sexo_str = (row.get('sexo') or row.get('gênero') or '').strip().upper()
                sexo = None
                if sexo_str in ['M', 'MACHO', 'MALE']:
                    sexo = 'M'
                elif sexo_str in ['F', 'FEMEA', 'FÊMEA', 'FEMALE']:
                    sexo = 'F'
                
                data_nascimento = None
                data_str = (row.get('data_nascimento') or row.get('nascimento') or '').strip()
                if data_str:
                    try:
                        data_nascimento = datetime.strptime(data_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            data_nascimento = datetime.strptime(data_str, '%d/%m/%Y').date()
                        except ValueError:
                            pass
                
                animais_arquivo.append({
                    'codigo_sisbov': codigo_sisbov,
                    'sexo': sexo or 'I',
                    'data_nascimento': data_nascimento.isoformat() if data_nascimento else None,
                })
        
        elif nome_arquivo.endswith('.pdf'):
            try:
                import PyPDF2
                import pdfplumber
            except ImportError as e:
                biblioteca_faltante = str(e).split("'")[1] if "'" in str(e) else "PyPDF2 ou pdfplumber"
                return JsonResponse({
                    'erro': f'Biblioteca {biblioteca_faltante} não está instalada. Para instalar, execute: pip install PyPDF2 pdfplumber'
                }, status=400)
            
            from .bnd_sisbov_parser import BNDSisbovParser
            
            parser = BNDSisbovParser()
            dados_extraidos = parser.extrair_dados_pdf(arquivo)
            animais_extraidos = dados_extraidos.get('animais', [])
            
            for animal_data in animais_extraidos:
                codigo_sisbov = animal_data.get('codigo_sisbov', '').strip()
                if not codigo_sisbov:
                    continue
                
                # Normalizar código SISBOV usando função padronizada
                codigo_normalizado = _normalizar_codigo_sisbov(codigo_sisbov)
                if not codigo_normalizado:
                    continue  # Código inválido, pular
                
                sexo = animal_data.get('sexo')
                if sexo and sexo not in ['M', 'F']:
                    sexo = 'I'
                
                data_nascimento = animal_data.get('data_nascimento')
                
                animais_arquivo.append({
                    'codigo_sisbov': codigo_normalizado,
                    'sexo': sexo or 'I',
                    'data_nascimento': data_nascimento.isoformat() if isinstance(data_nascimento, date) else (data_nascimento if data_nascimento else None),
                })
        
        else:
            return JsonResponse({'erro': 'Formato de arquivo não suportado'}, status=400)
        
        # Buscar animais do sistema
        animais_sistema = AnimalIndividual.objects.filter(
            propriedade=propriedade,
            status='ATIVO'
        ).select_related('categoria')
        
        # Agrupar por Sexo e Era
        sistema_por_sexo_era = defaultdict(int)
        arquivo_por_sexo_era = defaultdict(int)
        
        # Processar animais do sistema
        for animal in animais_sistema:
            sexo = animal.sexo if animal.sexo else 'I'
            era = _calcular_era(animal.data_nascimento, animal.categoria)
            chave = f"{sexo}_{era}"
            sistema_por_sexo_era[chave] += 1
        
        # Processar animais do arquivo (normalizando códigos SISBOV)
        animais_arquivo_normalizados = []
        for animal_data in animais_arquivo:
            codigo_original = animal_data.get('codigo_sisbov')
            if codigo_original:
                codigo_normalizado = _normalizar_codigo_sisbov(codigo_original)
                if codigo_normalizado:
                    animal_data['codigo_sisbov'] = codigo_normalizado
                    animais_arquivo_normalizados.append(animal_data)
        
        animais_arquivo = animais_arquivo_normalizados
        
        # Processar animais do arquivo para tabela comparativa
        for animal_data in animais_arquivo:
            sexo = animal_data.get('sexo', 'I')
            data_nasc = animal_data.get('data_nascimento')
            data_nasc_obj = None
            if data_nasc:
                try:
                    if isinstance(data_nasc, str):
                        data_nasc_obj = datetime.strptime(data_nasc, '%Y-%m-%d').date()
                    else:
                        data_nasc_obj = data_nasc
                except:
                    pass
            era = _calcular_era(data_nasc_obj)
            chave = f"{sexo}_{era}"
            arquivo_por_sexo_era[chave] += 1
        
        # Criar tabela comparativa
        todas_chaves = set(list(sistema_por_sexo_era.keys()) + list(arquivo_por_sexo_era.keys()))
        tabela_comparativa = []
        
        for chave in sorted(todas_chaves):
            sexo, era = chave.split('_', 1)
            qtd_sistema = sistema_por_sexo_era.get(chave, 0)
            qtd_arquivo = arquivo_por_sexo_era.get(chave, 0)
            diferenca = qtd_arquivo - qtd_sistema
            
            sexo_display = 'Macho' if sexo == 'M' else ('Fêmea' if sexo == 'F' else 'Indefinido')
            
            tabela_comparativa.append({
                'sexo': sexo_display,
                'era': era,
                'quantidade_sistema': qtd_sistema,
                'quantidade_arquivo': qtd_arquivo,
                'diferenca': diferenca,
            })
        
        # Estatísticas gerais
        total_sistema = animais_sistema.count()
        total_arquivo = len(animais_arquivo)
        
        # Criar dicionários para busca rápida (normalizando códigos SISBOV)
        animais_sistema_dict = {}
        for animal in animais_sistema:
            if animal.codigo_sisbov:
                codigo_normalizado = _normalizar_codigo_sisbov(animal.codigo_sisbov)
                if codigo_normalizado:
                    animais_sistema_dict[codigo_normalizado] = animal
        
        animais_arquivo_dict = {}
        for animal_data in animais_arquivo:
            codigo = animal_data.get('codigo_sisbov')
            if codigo:
                codigo_normalizado = _normalizar_codigo_sisbov(codigo)
                if codigo_normalizado:
                    animais_arquivo_dict[codigo_normalizado] = animal_data
        
        # Classificar animais do sistema
        animais_conformes = []  # Encontrado no BND e dados idênticos
        animais_divergentes = []  # Encontrado no BND mas com dados diferentes
        animais_nao_conformes = []  # No sistema mas NÃO no BND
        
        codigos_sistema = set(animais_sistema_dict.keys())
        codigos_arquivo = set(animais_arquivo_dict.keys())
        
        # Verificar cada animal do sistema
        for codigo_sisbov, animal_sistema in animais_sistema_dict.items():
            if codigo_sisbov in codigos_arquivo:
                # Animal encontrado no BND - verificar se está conforme
                animal_arquivo = animais_arquivo_dict[codigo_sisbov]
                
                # Comparar dados
                divergencias = []
                
                # Comparar sexo
                sexo_sistema = animal_sistema.sexo if animal_sistema.sexo else 'I'
                sexo_arquivo = animal_arquivo.get('sexo', 'I')
                if sexo_sistema != sexo_arquivo and sexo_arquivo != 'I':
                    divergencias.append('sexo')
                
                # Comparar raça
                raca_sistema = animal_sistema.raca or ''
                raca_arquivo = animal_arquivo.get('raca', '') or ''
                if raca_sistema and raca_arquivo and raca_sistema.upper().strip() != raca_arquivo.upper().strip():
                    divergencias.append('raca')
                
                # Comparar data de nascimento
                data_sistema = animal_sistema.data_nascimento
                data_arquivo = animal_arquivo.get('data_nascimento')
                if data_arquivo:
                    try:
                        if isinstance(data_arquivo, str):
                            data_arquivo_obj = datetime.strptime(data_arquivo, '%Y-%m-%d').date()
                        else:
                            data_arquivo_obj = data_arquivo
                        
                        if data_sistema and data_sistema != data_arquivo_obj:
                            divergencias.append('data_nascimento')
                    except:
                        pass
                
                if divergencias:
                    animais_divergentes.append({
                        'codigo_sisbov': codigo_sisbov,
                        'numero_brinco': animal_sistema.numero_brinco or '',
                        'divergencias': divergencias,
                        'sexo_sistema': sexo_sistema,
                        'sexo_arquivo': sexo_arquivo,
                        'raca_sistema': raca_sistema,
                        'raca_arquivo': raca_arquivo,
                    })
                else:
                    animais_conformes.append({
                        'codigo_sisbov': codigo_sisbov,
                        'numero_brinco': animal_sistema.numero_brinco or '',
                    })
            else:
                # Animal no sistema mas NÃO no BND - não conforme
                animais_nao_conformes.append({
                    'codigo_sisbov': codigo_sisbov,
                    'numero_brinco': animal_sistema.numero_brinco or '',
                    'sexo': animal_sistema.sexo or 'I',
                    'raca': animal_sistema.raca or '',
                })
        
        # Animais que serão criados (estão no BND mas não no sistema)
        animais_que_serao_criados = len(codigos_arquivo - codigos_sistema)
        
        # Estatísticas de divergências
        divergencias_sexo = sum(1 for a in animais_divergentes if 'sexo' in a['divergencias'])
        divergencias_raca = sum(1 for a in animais_divergentes if 'raca' in a['divergencias'])
        divergencias_data = sum(1 for a in animais_divergentes if 'data_nascimento' in a['divergencias'])
        
        return JsonResponse({
            'sucesso': True,
            'total_sistema': total_sistema,
            'total_arquivo': total_arquivo,
            'tabela_comparativa': tabela_comparativa,
            'resumo': {
                'animais_que_serao_criados': animais_que_serao_criados,
                'animais_conformes': len(animais_conformes),
                'animais_divergentes': len(animais_divergentes),
                'animais_nao_conformes': len(animais_nao_conformes),
                'divergencias_sexo': divergencias_sexo,
                'divergencias_raca': divergencias_raca,
                'divergencias_data': divergencias_data,
                'total_divergencias': len(animais_divergentes),
            },
            'detalhes': {
                'conformes': animais_conformes[:50],  # Limitar para não sobrecarregar
                'divergentes': animais_divergentes[:50],
                'nao_conformes': animais_nao_conformes[:50],
            }
        })
    
    except Exception as e:
        import logging
        logging.error(f"Erro no preview importação: {traceback.format_exc()}")
        return JsonResponse({'erro': f'Erro ao processar arquivo: {str(e)}'}, status=500)


@login_required
def resultado_importacao_bnd_sisbov(request, propriedade_id):
    """Página de resultados da importação BND/SISBOV"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Buscar resultados da sessão
    resultados = request.session.get('resultados_importacao_bnd', None)
    
    if not resultados or resultados.get('propriedade_id') != propriedade_id:
        messages.warning(request, 'Nenhum resultado de importação encontrado. Realize uma importação primeiro.')
        return redirect('importar_bnd_sisbov', propriedade_id=propriedade_id)
    
    # Limpar resultados da sessão após exibir
    if 'resultados_importacao_bnd' in request.session:
        del request.session['resultados_importacao_bnd']
    
    detalhes = resultados.get('resultados_detalhados', {})
    
    context = {
        'propriedade': propriedade,
        'resultados': resultados,
        'animais_criados': detalhes.get('animais_criados_lista', []),
        'animais_atualizados': detalhes.get('animais_atualizados_lista', []),
        'animais_conformes': detalhes.get('animais_conformes_lista', []),
        'animais_divergentes': detalhes.get('animais_divergentes_lista', []),
        'animais_nao_conformes': detalhes.get('animais_nao_conformes_lista', []),
    }
    
    return render(request, 'gestao_rural/resultado_importacao_bnd_sisbov.html', context)


@login_required
def importar_bnd_sisbov(request, propriedade_id):
    """Importação de dados BND/SISBOV de arquivo Excel ou CSV - VERSÃO CORRIGIDA"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    if request.method == 'POST':
        try:
            arquivo = request.FILES.get('arquivo')
            if not arquivo:
                messages.error(request, 'Por favor, selecione um arquivo para importar.')
                return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                    'propriedade': propriedade
                })
            
            nome_arquivo = arquivo.name.lower()
            animais_criados = 0
            animais_atualizados = 0
            animais_erros = []
            linhas_processadas = 0
            
            # Armazenar resultados detalhados da importação
            resultados_detalhados = {
                'animais_criados_lista': [],
                'animais_atualizados_lista': [],
                'animais_conformes_lista': [],
                'animais_divergentes_lista': [],
                'animais_nao_conformes_lista': [],
            }
            
            # Detectar tipo de arquivo
            if nome_arquivo.endswith('.xlsx') or nome_arquivo.endswith('.xls'):
                # Processar Excel
                try:
                    import openpyxl
                    from openpyxl import load_workbook
                    
                    # Carregar workbook
                    wb = load_workbook(arquivo, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # Encontrar cabeçalho
                    cabecalho = None
                    linha_cabecalho = 0
                    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if any(cell and isinstance(cell, str) and 'sisbov' in cell.lower() for cell in row):
                            cabecalho = [str(cell).lower().strip() if cell else '' for cell in row]
                            linha_cabecalho = idx
                            break
                    
                    if not cabecalho:
                        messages.error(request, 'Cabeçalho não encontrado no arquivo. Verifique se há uma coluna "SISBOV" ou "Código SISBOV".')
                        return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                            'propriedade': propriedade
                        })
                    
                    # Mapear colunas
                    col_indices = {}
                    colunas_esperadas = {
                        'sisbov': ['sisbov', 'codigo_sisbov', 'código sisbov', 'codigo sisbov'],
                        'brinco': ['brinco', 'numero_brinco', 'número brinco', 'numero brinco'],
                        'raca': ['raca', 'raça', 'raça do animal'],
                        'sexo': ['sexo', 'genero', 'gênero'],
                        'data_nascimento': ['data_nascimento', 'data nascimento', 'nascimento', 'dt_nasc'],
                        'peso': ['peso', 'peso_atual', 'peso atual', 'peso_kg']
                    }
                    
                    for campo, variacoes in colunas_esperadas.items():
                        for idx, col_name in enumerate(cabecalho):
                            if any(var in col_name for var in variacoes):
                                col_indices[campo] = idx
                                break
                    
                    if 'sisbov' not in col_indices:
                        messages.error(request, 'Coluna SISBOV não encontrada no arquivo.')
                        return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                            'propriedade': propriedade
                        })
                    
                    # Processar linhas
                    for idx, row in enumerate(ws.iter_rows(min_row=linha_cabecalho + 1, values_only=True), start=linha_cabecalho + 1):
                        if not any(row):  # Linha vazia
                            continue
                        
                        linhas_processadas += 1
                        try:
                            # Extrair dados
                            codigo_sisbov_raw = str(row[col_indices['sisbov']]).strip() if col_indices.get('sisbov') is not None and row[col_indices['sisbov']] else None
                            if not codigo_sisbov_raw or codigo_sisbov_raw.lower() in ['none', 'null', '']:
                                continue
                            
                            numero_brinco = None
                            if 'brinco' in col_indices and row[col_indices['brinco']]:
                                numero_brinco = str(row[col_indices['brinco']]).strip()
                            
                            raca = None
                            if 'raca' in col_indices and row[col_indices['raca']]:
                                raca = str(row[col_indices['raca']]).strip()
                            
                            sexo = None
                            if 'sexo' in col_indices and row[col_indices['sexo']]:
                                sexo_str = str(row[col_indices['sexo']]).strip().upper()
                                if sexo_str in ['M', 'MACHO', 'MALE']:
                                    sexo = 'M'
                                elif sexo_str in ['F', 'FEMEA', 'FÊMEA', 'FEMALE']:
                                    sexo = 'F'
                            
                            data_nascimento = None
                            if 'data_nascimento' in col_indices and row[col_indices['data_nascimento']]:
                                data_val = row[col_indices['data_nascimento']]
                                if isinstance(data_val, datetime):
                                    data_nascimento = data_val.date()
                                elif isinstance(data_val, date):
                                    data_nascimento = data_val
                                elif isinstance(data_val, str):
                                    try:
                                        data_nascimento = datetime.strptime(data_val, '%Y-%m-%d').date()
                                    except ValueError:
                                        try:
                                            data_nascimento = datetime.strptime(data_val, '%d/%m/%Y').date()
                                        except ValueError:
                                            pass
                            
                            peso_atual_kg = None
                            if 'peso' in col_indices and row[col_indices['peso']]:
                                try:
                                    peso_val = row[col_indices['peso']]
                                    if isinstance(peso_val, (int, float)):
                                        peso_atual_kg = Decimal(str(peso_val))
                                    elif isinstance(peso_val, str):
                                        peso_atual_kg = Decimal(peso_val.replace(',', '.'))
                                except (ValueError, InvalidOperation):
                                    pass
                            
                            # Processar animal usando função auxiliar
                            animal_data_dict = {
                                'codigo_sisbov': codigo_sisbov_raw,
                                'numero_brinco': numero_brinco,
                                'raca': raca,
                                'sexo': sexo,
                                'data_nascimento': data_nascimento.isoformat() if data_nascimento else None,
                                'peso_kg': peso_atual_kg,
                            }
                            
                            animal, criado, atualizado, status, divergencias = _processar_animal_importacao(
                                animal_data_dict, propriedade, resultados_detalhados
                            )
                            
                            if animal:
                                if criado:
                                    animais_criados += 1
                                elif atualizado:
                                    animais_atualizados += 1
                        
                        except Exception as e:
                            animais_erros.append(f"Linha {idx}: {str(e)}")
                    
                    wb.close()
                    
                except ImportError:
                    messages.error(request, 'Biblioteca openpyxl não está instalada. Execute: pip install openpyxl')
                    return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                        'propriedade': propriedade
                    })
                except Exception as e:
                    messages.error(request, f'Erro ao processar arquivo Excel: {str(e)}')
                    import logging
                    logging.error(f"Erro ao importar BND/SISBOV: {traceback.format_exc()}")
                    return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                        'propriedade': propriedade
                    })
            
            elif nome_arquivo.endswith('.csv'):
                # Processar CSV
                import csv
                import io
                
                try:
                    # Tentar diferentes encodings
                    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    conteudo = None
                    
                    for encoding in encodings:
                        try:
                            arquivo.seek(0)
                            conteudo = arquivo.read().decode(encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                    
                    if not conteudo:
                        messages.error(request, 'Erro ao ler arquivo CSV. Encoding não suportado.')
                        return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                            'propriedade': propriedade
                        })
                    
                    # Ler CSV
                    csv_reader = csv.DictReader(io.StringIO(conteudo))
                    
                    for idx, row in enumerate(csv_reader, start=2):
                        linhas_processadas += 1
                        try:
                            codigo_sisbov = row.get('sisbov') or row.get('codigo_sisbov') or row.get('Código SISBOV') or ''
                            codigo_sisbov = str(codigo_sisbov).strip()
                            
                            if not codigo_sisbov:
                                continue
                            
                            numero_brinco = (row.get('brinco') or row.get('numero_brinco') or '').strip()
                            raca = (row.get('raca') or row.get('raça') or '').strip()
                            sexo_str = (row.get('sexo') or row.get('gênero') or '').strip().upper()
                            sexo = None
                            if sexo_str in ['M', 'MACHO', 'MALE']:
                                sexo = 'M'
                            elif sexo_str in ['F', 'FEMEA', 'FÊMEA', 'FEMALE']:
                                sexo = 'F'
                            
                            data_nascimento = None
                            data_str = (row.get('data_nascimento') or row.get('nascimento') or '').strip()
                            if data_str:
                                try:
                                    data_nascimento = datetime.strptime(data_str, '%Y-%m-%d').date()
                                except ValueError:
                                    try:
                                        data_nascimento = datetime.strptime(data_str, '%d/%m/%Y').date()
                                    except ValueError:
                                        pass
                            
                            peso_atual_kg = None
                            peso_str = (row.get('peso') or row.get('peso_atual') or '').strip()
                            if peso_str:
                                try:
                                    peso_atual_kg = Decimal(peso_str.replace(',', '.'))
                                except (ValueError, InvalidOperation):
                                    pass
                            
                            # Processar animal usando função auxiliar
                            animal_data_dict = {
                                'codigo_sisbov': codigo_sisbov,
                                'numero_brinco': numero_brinco,
                                'raca': raca,
                                'sexo': sexo,
                                'data_nascimento': data_nascimento.isoformat() if data_nascimento else None,
                                'peso_kg': peso_atual_kg,
                            }
                            
                            animal, criado, atualizado, status, divergencias = _processar_animal_importacao(
                                animal_data_dict, propriedade, resultados_detalhados
                            )
                            
                            if animal:
                                if criado:
                                    animais_criados += 1
                                elif atualizado:
                                    animais_atualizados += 1
                        
                        except Exception as e:
                            animais_erros.append(f"Linha {idx}: {str(e)}")
                
                except Exception as e:
                    messages.error(request, f'Erro ao processar arquivo CSV: {str(e)}')
                    import logging
                    import traceback
                    logging.error(f"Erro ao importar CSV BND/SISBOV: {traceback.format_exc()}")
                    return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                        'propriedade': propriedade
                    })
            
            elif nome_arquivo.endswith('.pdf'):
                # Processar PDF BND SISBOV
                try:
                    # Verificar se bibliotecas necessárias estão instaladas
                    try:
                        import PyPDF2
                        import pdfplumber
                    except ImportError as e:
                        messages.error(request, f'Bibliotecas necessárias não estão instaladas: {str(e)}. Execute: pip install PyPDF2 pdfplumber')
                        return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                            'propriedade': propriedade
                        })
                    
                    # Importar parser
                    from .bnd_sisbov_parser import BNDSisbovParser
                    
                    # Criar parser e extrair dados
                    parser = BNDSisbovParser()
                    dados_extraidos = parser.extrair_dados_pdf(arquivo)
                    
                    animais_extraidos = dados_extraidos.get('animais', [])
                    
                    if not animais_extraidos:
                        messages.warning(request, 'Nenhum animal foi encontrado no PDF. Verifique se o arquivo contém dados válidos do SISBOV.')
                        return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                            'propriedade': propriedade
                        })
                    
                    # Processar cada animal extraído
                    for idx, animal_data in enumerate(animais_extraidos, start=1):
                        linhas_processadas += 1
                        try:
                            codigo_sisbov_raw = animal_data.get('codigo_sisbov', '').strip()
                            if not codigo_sisbov_raw:
                                continue
                            
                            numero_brinco = animal_data.get('numero_brinco', '').strip()
                            if not numero_brinco:
                                # Tentar usar número de manejo como brinco
                                numero_manejo = animal_data.get('numero_manejo', '')
                                if numero_manejo:
                                    numero_brinco = numero_manejo
                            
                            raca = animal_data.get('raca', '').strip() or None
                            
                            sexo = animal_data.get('sexo')
                            if sexo and sexo not in ['M', 'F']:
                                sexo = None
                            
                            data_nascimento = animal_data.get('data_nascimento')
                            peso_atual_kg = animal_data.get('peso_kg')
                            
                            # Processar animal usando função auxiliar
                            animal_data_dict = {
                                'codigo_sisbov': codigo_sisbov_raw,
                                'numero_brinco': numero_brinco,
                                'raca': raca,
                                'sexo': sexo,
                                'data_nascimento': data_nascimento.isoformat() if isinstance(data_nascimento, date) else (data_nascimento if data_nascimento else None),
                                'peso_kg': peso_atual_kg,
                            }
                            
                            animal, criado, atualizado, status, divergencias = _processar_animal_importacao(
                                animal_data_dict, propriedade, resultados_detalhados
                            )
                            
                            if animal:
                                if criado:
                                    animais_criados += 1
                                elif atualizado:
                                    animais_atualizados += 1
                        
                        except Exception as e:
                            animais_erros.append(f"Animal {idx}: {str(e)}")
                    
                    # Log do relatório de extração
                    relatorio = parser.gerar_relatorio_extracao()
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.info(f"Importação BND SISBOV PDF - {relatorio}")
                
                except ImportError:
                    messages.error(request, 'Bibliotecas necessárias não estão instaladas. Execute: pip install PyPDF2 pdfplumber')
                    return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                        'propriedade': propriedade
                    })
                except Exception as e:
                    messages.error(request, f'Erro ao processar arquivo PDF: {str(e)}')
                    import logging
                    logging.error(f"Erro ao importar PDF BND/SISBOV: {traceback.format_exc()}")
                    return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                        'propriedade': propriedade
                    })
            
            else:
                messages.error(request, 'Formato de arquivo não suportado. Use Excel (.xlsx, .xls), CSV (.csv) ou PDF (.pdf).')
                return render(request, 'gestao_rural/importar_bnd_sisbov.html', {
                    'propriedade': propriedade
                })
            
            # Mensagens de resultado
            if animais_criados > 0:
                messages.success(request, f'✅ {animais_criados} animais criados com sucesso!')
            if animais_atualizados > 0:
                messages.info(request, f'ℹ️ {animais_atualizados} animais atualizados.')
            if animais_erros:
                total_erros = len(animais_erros)
                if total_erros <= 5:
                    erros_msg = '; '.join(animais_erros[:5])
                    messages.warning(request, f'⚠️ {total_erros} erros encontrados: {erros_msg}')
                else:
                    messages.warning(request, f'⚠️ {total_erros} erros encontrados durante a importação.')
            
            if linhas_processadas == 0:
                messages.warning(request, 'Nenhuma linha foi processada. Verifique o formato do arquivo.')
            
            # Identificar animais não conformes (no sistema mas não no arquivo)
            # Coletar todos os códigos SISBOV que foram importados
            codigos_importados = set()
            for lista in resultados_detalhados.values():
                if isinstance(lista, list):
                    for item in lista:
                        if isinstance(item, dict) and 'codigo_sisbov' in item:
                            codigos_importados.add(item['codigo_sisbov'])
            
            # Buscar animais do sistema que não estão na lista de importados
            animais_sistema_todos = AnimalIndividual.objects.filter(
                propriedade=propriedade,
                status='ATIVO'
            ).select_related('categoria')
            
            for animal in animais_sistema_todos:
                if animal.codigo_sisbov:
                    codigo_normalizado = _normalizar_codigo_sisbov(animal.codigo_sisbov)
                    if codigo_normalizado and codigo_normalizado not in codigos_importados:
                        # Verificar se já não foi adicionado (evitar duplicatas)
                        ja_adicionado = any(
                            item.get('codigo_sisbov') == codigo_normalizado 
                            for item in resultados_detalhados['animais_nao_conformes_lista']
                        )
                        if not ja_adicionado:
                            resultados_detalhados['animais_nao_conformes_lista'].append({
                                'animal_id': animal.id,
                                'codigo_sisbov': codigo_normalizado,
                                'numero_brinco': animal.numero_brinco or '',
                                'sexo': animal.sexo or 'I',
                                'raca': animal.raca or '',
                            })
            
            # Salvar resultados na sessão para exibir na página de resultados
            request.session['resultados_importacao_bnd'] = {
                'propriedade_id': propriedade_id,
                'animais_criados': animais_criados,
                'animais_atualizados': animais_atualizados,
                'total_processados': linhas_processadas,
                'erros': animais_erros[:20],  # Limitar a 20 erros
                'resultados_detalhados': resultados_detalhados,
                'timestamp': datetime.now().isoformat(),
            }
            
            # Redirecionar para página de resultados da importação
            return redirect('resultado_importacao_bnd_sisbov', propriedade_id=propriedade_id)
        
        except Exception as e:
            import traceback
            messages.error(request, f'❌ Erro ao importar arquivo: {str(e)}')
            import logging
            logging.error(f"Erro geral ao importar BND/SISBOV: {traceback.format_exc()}")
    
    context = {
        'propriedade': propriedade,
    }

    return render(request, 'gestao_rural/importar_bnd_sisbov.html', context)


@login_required
def animais_individuais_lista(request, propriedade_id):
    """Lista de animais individuais"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Filtros
    status_filtro = request.GET.get('status', '')
    categoria_filtro = request.GET.get('categoria', '')
    busca = request.GET.get('busca', '')
    
    animais = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).select_related('categoria', 'propriedade_origem', 'lote_atual', 'lote_atual__sessao')
    
    # Aplicar filtros
    if status_filtro:
        animais = animais.filter(status=status_filtro)
    
    if categoria_filtro:
        animais = animais.filter(categoria_id=categoria_filtro)
    
    if busca:
        # Normalizar código de busca (remover espaços, traços, etc.)
        busca_normalizada = _normalizar_codigo(busca)
        
        # PRIORIDADE: Buscar por SISBOV primeiro (ID principal do animal)
        # Busca por: SISBOV, número de manejo, número do brinco, raça e observações
        filtros_busca = (
            Q(numero_brinco__icontains=busca) |
            Q(raca__icontains=busca) |
            Q(observacoes__icontains=busca)
        )
        
        # Se a busca normalizada tiver 8+ dígitos, pode ser SISBOV
        if len(busca_normalizada) >= 8:
            # Busca por SISBOV completo ou parcial
            filtros_busca |= (
                Q(codigo_sisbov=busca_normalizada) |
                Q(codigo_sisbov__icontains=busca_normalizada) |
                Q(codigo_sisbov__endswith=busca_normalizada) |
                Q(numero_brinco__icontains=busca_normalizada)
            )
        # Se a busca normalizada tiver 6-7 dígitos, pode ser número de manejo ou SISBOV parcial
        elif len(busca_normalizada) >= 6:
            filtros_busca |= (
                Q(numero_manejo=busca_normalizada) |
                Q(codigo_sisbov__icontains=busca_normalizada) |
                Q(codigo_sisbov__endswith=busca_normalizada) |
                Q(numero_brinco__icontains=busca_normalizada)
            )
        # Para códigos menores, também tenta buscar por SISBOV que contém
        elif len(busca_normalizada) >= 3:
            filtros_busca |= (
                Q(codigo_sisbov__icontains=busca_normalizada) |
                Q(numero_brinco__icontains=busca_normalizada)
            )
        
        animais = animais.filter(filtros_busca)
    
    animais = animais.order_by('-data_cadastro')
    
    # Categorias para filtro
    categorias = CategoriaAnimal.objects.filter(ativo=True).order_by('nome')
    
    context = {
        'propriedade': propriedade,
        'animais': animais,
        'categorias': categorias,
        'status_filtro': status_filtro,
        'categoria_filtro': categoria_filtro,
        'busca': busca,
    }
    
    return render(request, 'gestao_rural/animais_individuais_lista.html', context)


@login_required
def animal_individual_novo(request, propriedade_id):
    """Cadastro de novo animal individual"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    if request.method == 'POST':
        try:
            def _parse_decimal_field(valor):
                if valor in (None, '', 'null'):
                    return None
                try:
                    return Decimal(str(valor))
                except (InvalidOperation, TypeError, ValueError):
                    return None
            
            def _parse_date_field(valor):
                if not valor:
                    return None
                try:
                    return datetime.strptime(valor, '%Y-%m-%d').date()
                except ValueError:
                    return None

            # Obter dados do formulário
            numero_brinco = request.POST.get('numero_brinco', '').strip()
            codigo_sisbov = request.POST.get('codigo_sisbov', '').strip()
            codigo_eletronico = request.POST.get('codigo_eletronico', '').strip()
            tipo_brinco = request.POST.get('tipo_brinco', 'VISUAL')
            categoria_id = request.POST.get('categoria')
            data_nascimento = _parse_date_field(request.POST.get('data_nascimento'))
            data_identificacao = _parse_date_field(request.POST.get('data_identificacao'))
            propriedade_origem_id = request.POST.get('propriedade_origem') or None
            sexo = request.POST.get('sexo')
            raca = request.POST.get('raca', '').strip()
            peso_atual_kg = _parse_decimal_field(request.POST.get('peso_atual_kg'))
            observacoes = request.POST.get('observacoes', '').strip()
            status_sanitario = request.POST.get('status_sanitario', 'INDEFINIDO')
            lote_atual_id = request.POST.get('lote_atual') or None
            apelido = request.POST.get('apelido', '').strip()
            foto = request.FILES.get('foto')
            mae_id = request.POST.get('mae') or None
            pai_id = request.POST.get('pai') or None
            classificacao_zootecnica = request.POST.get('classificacao_zootecnica', '').strip()
            grupo_producao = request.POST.get('grupo_producao', '').strip()
            sistema_criacao = request.POST.get('sistema_criacao') or None
            nivel_confinamento = request.POST.get('nivel_confinamento') or None
            status_reprodutivo = request.POST.get('status_reprodutivo', 'INDEFINIDO')
            data_ultima_cobertura = _parse_date_field(request.POST.get('data_ultima_cobertura'))
            data_prevista_parto = _parse_date_field(request.POST.get('data_prevista_parto'))
            produtividade_leite_dia = _parse_decimal_field(request.POST.get('produtividade_leite_dia'))
            registro_vacinal_em_dia = request.POST.get('registro_vacinal_em_dia') == 'on'
            proxima_vacinacao_obrigatoria = _parse_date_field(request.POST.get('proxima_vacinacao_obrigatoria'))
            carencia_produtos_ate = _parse_date_field(request.POST.get('carencia_produtos_ate'))
            tipo_origem = request.POST.get('tipo_origem') or 'NASCIMENTO'
            custo_aquisicao = _parse_decimal_field(request.POST.get('custo_aquisicao'))
            data_aquisicao = _parse_date_field(request.POST.get('data_aquisicao'))
            valor_atual_estimado = _parse_decimal_field(request.POST.get('valor_atual_estimado'))
            reprodutor_origem = request.POST.get('reprodutor_origem', '').strip()
            
            # Validações
            if not numero_brinco:
                messages.error(request, 'O número do brinco é obrigatório!')
                return redirect('animal_individual_novo', propriedade_id=propriedade_id)
            
            if AnimalIndividual.objects.filter(numero_brinco=numero_brinco).exists():
                messages.error(request, f'Já existe um animal com o brinco {numero_brinco}!')
                return redirect('animal_individual_novo', propriedade_id=propriedade_id)
            
            if codigo_sisbov and AnimalIndividual.objects.filter(codigo_sisbov=codigo_sisbov).exists():
                messages.error(request, f'Já existe um animal com o código SISBOV {codigo_sisbov}!')
                return redirect('animal_individual_novo', propriedade_id=propriedade_id)
            
            if codigo_eletronico and AnimalIndividual.objects.filter(codigo_eletronico=codigo_eletronico).exists():
                messages.error(request, f'Já existe um animal com o identificador eletrônico {codigo_eletronico}!')
                return redirect('animal_individual_novo', propriedade_id=propriedade_id)
            
            categoria = get_object_or_404(CategoriaAnimal, pk=categoria_id)
            
            if not data_identificacao:
                data_identificacao = date.today()
            
            tipo_origem_valores = {valor for valor, _ in AnimalIndividual.TIPO_ORIGEM_CHOICES}
            if tipo_origem not in tipo_origem_valores:
                tipo_origem = 'NASCIMENTO'

            status_reprodutivo_valores = {valor for valor, _ in AnimalIndividual.STATUS_REPRODUTIVO_CHOICES}
            if status_reprodutivo not in status_reprodutivo_valores:
                status_reprodutivo = 'INDEFINIDO'

            sistema_criacao_valores = {valor for valor, _ in AnimalIndividual.SISTEMA_CRIACAO_CHOICES}
            if sistema_criacao not in sistema_criacao_valores:
                sistema_criacao = None

            nivel_confinamento_valores = {valor for valor, _ in AnimalIndividual.NIVEL_CONFINAMENTO_CHOICES}
            if nivel_confinamento not in nivel_confinamento_valores:
                nivel_confinamento = None

            if mae_id and not AnimalIndividual.objects.filter(id=mae_id, propriedade=propriedade).exists():
                mae_id = None
            if pai_id and not AnimalIndividual.objects.filter(id=pai_id, propriedade=propriedade).exists():
                pai_id = None

            data_movimentacao_base = data_nascimento or data_identificacao or date.today()

            # Criar animal
            animal = AnimalIndividual.objects.create(
                numero_brinco=numero_brinco,
                codigo_sisbov=codigo_sisbov or numero_brinco,
                codigo_eletronico=codigo_eletronico or None,
                tipo_brinco=tipo_brinco,
                propriedade=propriedade,
                categoria=categoria,
                data_nascimento=data_nascimento,
                data_identificacao=data_identificacao,
                propriedade_origem_id=propriedade_origem_id,
                sexo=sexo,
                raca=raca or None,
                peso_atual_kg=peso_atual_kg,
                observacoes=observacoes or None,
                status='ATIVO',
                status_sanitario=status_sanitario,
                lote_atual_id=lote_atual_id if lote_atual_id else None,
                responsavel_tecnico=request.user if request.user.is_authenticated else None,
                apelido=apelido or None,
                foto=foto,
                mae_id=mae_id,
                pai_id=pai_id,
                classificacao_zootecnica=classificacao_zootecnica or None,
                grupo_producao=grupo_producao or None,
                sistema_criacao=sistema_criacao,
                nivel_confinamento=nivel_confinamento,
                status_reprodutivo=status_reprodutivo,
                data_ultima_cobertura=data_ultima_cobertura,
                data_prevista_parto=data_prevista_parto,
                produtividade_leite_dia=produtividade_leite_dia,
                registro_vacinal_em_dia=registro_vacinal_em_dia,
                proxima_vacinacao_obrigatoria=proxima_vacinacao_obrigatoria,
                carencia_produtos_ate=carencia_produtos_ate,
                tipo_origem=tipo_origem,
                custo_aquisicao=custo_aquisicao,
                data_aquisicao=data_aquisicao,
                valor_atual_estimado=valor_atual_estimado or custo_aquisicao,
                reprodutor_origem=reprodutor_origem or None,
                data_ultima_movimentacao=data_movimentacao_base,
            )
            
            # Garantir que data_identificacao seja igual a data_cadastro.date() (por enquanto)
            if not animal.data_identificacao:
                animal.data_identificacao = animal.data_cadastro.date()
                animal.save(update_fields=['data_identificacao'])
            
            # Criar movimentação de nascimento ou compra
            if tipo_origem == 'NASCIMENTO':
                tipo_mov = 'NASCIMENTO'
            elif tipo_origem == 'COMPRA':
                tipo_mov = 'COMPRA'
            elif tipo_origem == 'TRANSFERENCIA':
                tipo_mov = 'TRANSFERENCIA_ENTRADA'
            else:
                tipo_mov = 'OUTROS'
            
            MovimentacaoIndividual.objects.create(
                animal=animal,
                tipo_movimentacao=tipo_mov,
                data_movimentacao=data_movimentacao_base,
                propriedade_origem_id=propriedade_origem_id,
                peso_kg=peso_atual_kg,
                valor=custo_aquisicao if tipo_mov == 'COMPRA' else None,
                observacoes='Cadastro inicial do animal',
                documento_tipo='OUTROS',
                responsavel=request.user if request.user.is_authenticated else None,
                quantidade_animais=1
            )
            
            # Atualizar status do brinco se existir
            brinco = BrincoAnimal.objects.filter(
                numero_brinco=numero_brinco,
                propriedade=propriedade
            ).first()
            
            if brinco:
                brinco.status = 'EM_USO'
                brinco.animal = animal
                brinco.data_utilizacao = date.today()
                brinco.save()
            
            messages.success(request, f'Animal {numero_brinco} cadastrado com sucesso!')
            return redirect('animal_individual_detalhes', propriedade_id=propriedade_id, animal_id=animal.id)
            
        except Exception as e:
            messages.error(request, f'Erro ao cadastrar animal: {str(e)}')
    
    # Obter dados para formulário
    categorias = CategoriaAnimal.objects.filter(ativo=True).order_by('nome')
    propriedades = Propriedade.objects.filter(
        produtor=propriedade.produtor
    ).order_by('nome_propriedade')
    
    # Sugerir próximo número de brinco
    ultimo_brinco = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).order_by('-id').first()
    
    proximo_numero = 1
    if ultimo_brinco:
        try:
            # Tentar extrair número do último brinco
            import re
            match = re.search(r'(\d+)', ultimo_brinco.numero_brinco)
            if match:
                proximo_numero = int(match.group(1)) + 1
        except (AttributeError, ValueError, TypeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.debug(f"Erro ao extrair número do brinco: {e}")
            pass
    
    lotes = CurralLote.objects.filter(
        sessao__propriedade=propriedade
    ).select_related('sessao').order_by('-sessao__data_inicio', 'nome')
    animais_maes = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        sexo='F'
    ).order_by('numero_brinco')
    animais_pais = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        sexo='M'
    ).order_by('numero_brinco')
    
    context = {
        'propriedade': propriedade,
        'categorias': categorias,
        'propriedades': propriedades,
        'proximo_numero': proximo_numero,
        'status_sanitario_choices': AnimalIndividual.STATUS_SANITARIO_CHOICES,
        'status_reprodutivo_choices': AnimalIndividual.STATUS_REPRODUTIVO_CHOICES,
        'tipo_origem_choices': AnimalIndividual.TIPO_ORIGEM_CHOICES,
        'sistema_criacao_choices': AnimalIndividual.SISTEMA_CRIACAO_CHOICES,
        'nivel_confinamento_choices': AnimalIndividual.NIVEL_CONFINAMENTO_CHOICES,
        'lotes': lotes,
        'animais_maes': animais_maes,
        'animais_pais': animais_pais,
    }
    
    return render(request, 'gestao_rural/animal_individual_novo.html', context)


@login_required
def animal_individual_detalhes(request, propriedade_id, animal_id):
    """Detalhes de um animal individual"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    animal = get_object_or_404(AnimalIndividual, pk=animal_id, propriedade=propriedade)
    
    # Histórico de movimentações
    movimentacoes = MovimentacaoIndividual.objects.filter(
        animal=animal
    ).select_related(
        'propriedade_origem', 'propriedade_destino',
        'categoria_anterior', 'categoria_nova'
    ).order_by('-data_movimentacao')
    
    context = {
        'propriedade': propriedade,
        'animal': animal,
        'movimentacoes': movimentacoes,
    }
    
    return render(request, 'gestao_rural/animal_individual_detalhes.html', context)


@login_required
def animal_individual_editar(request, propriedade_id, animal_id):
    """Edição de animal individual"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    animal = get_object_or_404(AnimalIndividual, pk=animal_id, propriedade=propriedade)
    
    if request.method == 'POST':
        try:
            def _parse_decimal_field(valor):
                if valor in (None, '', 'null'):
                    return None
                try:
                    return Decimal(str(valor))
                except (InvalidOperation, TypeError, ValueError):
                    return None

            def _parse_date_field(valor):
                if not valor:
                    return None
                try:
                    return datetime.strptime(valor, '%Y-%m-%d').date()
                except ValueError:
                    return None

            # Dados principais
            categoria_id = request.POST.get('categoria')
            categoria = get_object_or_404(CategoriaAnimal, pk=categoria_id)

            data_nascimento = _parse_date_field(request.POST.get('data_nascimento'))
            data_identificacao = _parse_date_field(request.POST.get('data_identificacao'))
            propriedade_origem_id = request.POST.get('propriedade_origem') or None
            raca = request.POST.get('raca', '').strip()
            peso_atual_kg = _parse_decimal_field(request.POST.get('peso_atual_kg'))
            observacoes = request.POST.get('observacoes', '').strip()
            status = request.POST.get('status', animal.status)
            status_sanitario = request.POST.get('status_sanitario', animal.status_sanitario)
            lote_atual_id = request.POST.get('lote_atual') or None
            codigo_sisbov = request.POST.get('codigo_sisbov', '').strip()
            codigo_eletronico = request.POST.get('codigo_eletronico', '').strip()
            data_saida = _parse_date_field(request.POST.get('data_saida'))
            motivo_saida = request.POST.get('motivo_saida', '').strip()

            apelido = request.POST.get('apelido', '').strip()
            foto = request.FILES.get('foto')
            mae_id = request.POST.get('mae') or None
            pai_id = request.POST.get('pai') or None
            classificacao_zootecnica = request.POST.get('classificacao_zootecnica', '').strip()
            grupo_producao = request.POST.get('grupo_producao', '').strip()
            sistema_criacao = request.POST.get('sistema_criacao') or None
            nivel_confinamento = request.POST.get('nivel_confinamento') or None
            status_reprodutivo = request.POST.get('status_reprodutivo', animal.status_reprodutivo)
            data_ultima_cobertura = _parse_date_field(request.POST.get('data_ultima_cobertura'))
            data_prevista_parto = _parse_date_field(request.POST.get('data_prevista_parto'))
            produtividade_leite_dia = _parse_decimal_field(request.POST.get('produtividade_leite_dia'))
            registro_vacinal_em_dia = request.POST.get('registro_vacinal_em_dia') == 'on'
            proxima_vacinacao_obrigatoria = _parse_date_field(request.POST.get('proxima_vacinacao_obrigatoria'))
            carencia_produtos_ate = _parse_date_field(request.POST.get('carencia_produtos_ate'))
            tipo_origem = request.POST.get('tipo_origem', animal.tipo_origem)
            custo_aquisicao = _parse_decimal_field(request.POST.get('custo_aquisicao'))
            data_aquisicao = _parse_date_field(request.POST.get('data_aquisicao'))
            valor_atual_estimado = _parse_decimal_field(request.POST.get('valor_atual_estimado'))
            reprodutor_origem = request.POST.get('reprodutor_origem', '').strip()
            documento_saida = request.POST.get('documento_saida', '').strip()
            data_ultima_movimentacao = _parse_date_field(request.POST.get('data_ultima_movimentacao'))

            if codigo_sisbov and AnimalIndividual.objects.exclude(id=animal.id).filter(codigo_sisbov=codigo_sisbov).exists():
                messages.error(request, f'Já existe um animal com o código SISBOV {codigo_sisbov}!')
                return redirect('animal_individual_editar', propriedade_id=propriedade_id, animal_id=animal_id)

            if codigo_eletronico and AnimalIndividual.objects.exclude(id=animal.id).filter(codigo_eletronico=codigo_eletronico).exists():
                messages.error(request, f'Já existe um animal com o identificador eletrônico {codigo_eletronico}!')
                return redirect('animal_individual_editar', propriedade_id=propriedade_id, animal_id=animal_id)

            # Controle de mudança de categoria
            if animal.categoria != categoria:
                MovimentacaoIndividual.objects.create(
                    animal=animal,
                    tipo_movimentacao='MUDANCA_CATEGORIA',
                    data_movimentacao=date.today(),
                    categoria_anterior=animal.categoria,
                    categoria_nova=categoria,
                    observacoes='Mudança de categoria'
                )
                animal.categoria = categoria

            # Atualização dos campos principais
            animal.data_nascimento = data_nascimento
            animal.data_identificacao = data_identificacao or animal.data_identificacao
            animal.propriedade_origem_id = propriedade_origem_id
            animal.raca = raca or None
            animal.peso_atual_kg = peso_atual_kg
            animal.observacoes = observacoes or None
            animal.status = status
            animal.status_sanitario = status_sanitario
            animal.lote_atual_id = lote_atual_id if lote_atual_id else None
            animal.codigo_sisbov = codigo_sisbov or animal.codigo_sisbov or animal.numero_brinco
            animal.codigo_eletronico = codigo_eletronico or None
            animal.data_saida = data_saida
            animal.motivo_saida = motivo_saida or None
            animal.apelido = apelido or None
            if foto:
                animal.foto = foto

            if mae_id and AnimalIndividual.objects.filter(id=mae_id, propriedade=propriedade).exclude(id=animal.id).exists():
                animal.mae_id = int(mae_id)
            else:
                animal.mae_id = None

            if pai_id and AnimalIndividual.objects.filter(id=pai_id, propriedade=propriedade).exclude(id=animal.id).exists():
                animal.pai_id = int(pai_id)
            else:
                animal.pai_id = None

            animal.classificacao_zootecnica = classificacao_zootecnica or None
            animal.grupo_producao = grupo_producao or None

            if sistema_criacao in {valor for valor, _ in AnimalIndividual.SISTEMA_CRIACAO_CHOICES}:
                animal.sistema_criacao = sistema_criacao
            else:
                animal.sistema_criacao = None

            if nivel_confinamento in {valor for valor, _ in AnimalIndividual.NIVEL_CONFINAMENTO_CHOICES}:
                animal.nivel_confinamento = nivel_confinamento
            else:
                animal.nivel_confinamento = None

            if status_reprodutivo in {valor for valor, _ in AnimalIndividual.STATUS_REPRODUTIVO_CHOICES}:
                animal.status_reprodutivo = status_reprodutivo

            if tipo_origem in {valor for valor, _ in AnimalIndividual.TIPO_ORIGEM_CHOICES}:
                animal.tipo_origem = tipo_origem

            animal.data_ultima_cobertura = data_ultima_cobertura
            animal.data_prevista_parto = data_prevista_parto
            animal.produtividade_leite_dia = produtividade_leite_dia
            animal.registro_vacinal_em_dia = registro_vacinal_em_dia
            animal.proxima_vacinacao_obrigatoria = proxima_vacinacao_obrigatoria
            animal.carencia_produtos_ate = carencia_produtos_ate
            animal.custo_aquisicao = custo_aquisicao
            animal.data_aquisicao = data_aquisicao
            animal.valor_atual_estimado = valor_atual_estimado
            animal.reprodutor_origem = reprodutor_origem or None
            animal.documento_saida = documento_saida or None
            if data_ultima_movimentacao:
                animal.data_ultima_movimentacao = data_ultima_movimentacao

            animal.responsavel_tecnico = request.user if request.user.is_authenticated else animal.responsavel_tecnico
            animal.save()

            messages.success(request, f'Animal {animal.numero_brinco} atualizado com sucesso!')
            return redirect('animal_individual_detalhes', propriedade_id=propriedade_id, animal_id=animal.id)

        except Exception as e:
            messages.error(request, f'Erro ao atualizar animal: {str(e)}')
    
    categorias = CategoriaAnimal.objects.filter(ativo=True).order_by('nome')
    propriedades = Propriedade.objects.filter(
        produtor=propriedade.produtor
    ).order_by('nome_propriedade')
    lotes = CurralLote.objects.filter(
        sessao__propriedade=propriedade
    ).select_related('sessao').order_by('-sessao__data_inicio', 'nome')
    animais_maes = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        sexo='F'
    ).exclude(id=animal.id).order_by('numero_brinco')
    animais_pais = AnimalIndividual.objects.filter(
        propriedade=propriedade,
        sexo='M'
    ).exclude(id=animal.id).order_by('numero_brinco')
    
    context = {
        'propriedade': propriedade,
        'animal': animal,
        'categorias': categorias,
        'propriedades': propriedades,
        'status_choices': AnimalIndividual.STATUS_CHOICES,
        'status_sanitario_choices': AnimalIndividual.STATUS_SANITARIO_CHOICES,
        'status_reprodutivo_choices': AnimalIndividual.STATUS_REPRODUTIVO_CHOICES,
        'tipo_origem_choices': AnimalIndividual.TIPO_ORIGEM_CHOICES,
        'sistema_criacao_choices': AnimalIndividual.SISTEMA_CRIACAO_CHOICES,
        'nivel_confinamento_choices': AnimalIndividual.NIVEL_CONFINAMENTO_CHOICES,
        'lotes': lotes,
        'animais_maes': animais_maes,
        'animais_pais': animais_pais,
    }
    
    return render(request, 'gestao_rural/animal_individual_editar.html', context)


@login_required
def movimentacao_individual_nova(request, propriedade_id, animal_id):
    """Cadastro de nova movimentação individual"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    animal = get_object_or_404(AnimalIndividual, pk=animal_id, propriedade=propriedade)
    
    if request.method == 'POST':
        try:
            tipo_movimentacao = request.POST.get('tipo_movimentacao')
            data_movimentacao = request.POST.get('data_movimentacao')
            propriedade_origem_id = request.POST.get('propriedade_origem') or None
            propriedade_destino_id = request.POST.get('propriedade_destino') or None
            categoria_nova_id = request.POST.get('categoria_nova') or None
            peso_kg = request.POST.get('peso_kg') or None
            valor = request.POST.get('valor') or None
            quantidade_animais = request.POST.get('quantidade_animais') or 1
            observacoes = request.POST.get('observacoes', '').strip()
            motivo_detalhado = request.POST.get('motivo_detalhado', '').strip()
            status_sanitario = request.POST.get('status_sanitario', animal.status_sanitario)
            lote_atual_id = request.POST.get('lote_atual') or None
            
            try:
                quantidade_animais_int = int(quantidade_animais)
                if quantidade_animais_int <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                quantidade_animais_int = 1
            
            # Criar movimentação
            numero_documento = request.POST.get('numero_documento') or None
            documento_emissor = request.POST.get('documento_emissor') or None
            data_documento = request.POST.get('data_documento') or None
            if data_documento:
                try:
                    data_documento = datetime.strptime(data_documento, "%Y-%m-%d").date()
                except ValueError:
                    data_documento = None

            movimentacao = MovimentacaoIndividual.objects.create(
                animal=animal,
                tipo_movimentacao=tipo_movimentacao,
                data_movimentacao=data_movimentacao or datetime.today().date(),
                quantidade_animais=quantidade_animais_int,
                propriedade_origem_id=propriedade_origem_id,
                propriedade_destino_id=propriedade_destino_id,
                categoria_nova_id=categoria_nova_id,
                peso_kg=Decimal(str(peso_kg)) if peso_kg else None,
                valor=Decimal(str(valor)) if valor else None,
                numero_documento=numero_documento,
                documento_emissor=documento_emissor,
                data_documento=data_documento,
                responsavel=request.user if request.user.is_authenticated else None,
                motivo_detalhado=motivo_detalhado or None,
                observacoes=observacoes or None,
            )
            
            # Atualizar animal conforme movimentação
            if tipo_movimentacao == 'VENDA':
                animal.status = 'VENDIDO'
                if propriedade_destino_id:
                    animal.propriedade_id = propriedade_destino_id
                animal.data_saida = data_movimentacao
                animal.motivo_saida = motivo_detalhado or 'Venda registrada'
            elif tipo_movimentacao == 'MORTE':
                animal.status = 'MORTO'
                animal.data_saida = data_movimentacao
                animal.motivo_saida = motivo_detalhado or 'Morte registrada'
            elif tipo_movimentacao == 'TRANSFERENCIA_SAIDA':
                animal.status = 'TRANSFERIDO'
                if propriedade_destino_id:
                    animal.propriedade_id = propriedade_destino_id
                animal.data_saida = data_movimentacao
                animal.motivo_saida = motivo_detalhado or 'Transferência registrada'
            elif tipo_movimentacao == 'MUDANCA_CATEGORIA' and categoria_nova_id:
                animal.categoria_id = categoria_nova_id
            
            if peso_kg:
                animal.peso_atual_kg = Decimal(str(peso_kg))
            
            animal.status_sanitario = status_sanitario
            if lote_atual_id:
                animal.lote_atual_id = lote_atual_id
            animal.responsavel_tecnico = request.user if request.user.is_authenticated else animal.responsavel_tecnico
            animal.save()
            
            messages.success(request, 'Movimentação registrada com sucesso!')
            return redirect('animal_individual_detalhes', propriedade_id=propriedade_id, animal_id=animal.id)
            
        except Exception as e:
            messages.error(request, f'Erro ao registrar movimentação: {str(e)}')
    
    categorias = CategoriaAnimal.objects.filter(ativo=True).order_by('nome')
    propriedades = Propriedade.objects.filter(
        produtor=propriedade.produtor
    ).order_by('nome_propriedade')
    lotes = CurralLote.objects.filter(
        sessao__propriedade=propriedade
    ).select_related('sessao').order_by('-sessao__data_inicio', 'nome')
    
    context = {
        'propriedade': propriedade,
        'animal': animal,
        'categorias': categorias,
        'propriedades': propriedades,
        'status_sanitario_choices': AnimalIndividual.STATUS_SANITARIO_CHOICES,
        'lotes': lotes,
        'documento_tipo_choices': MovimentacaoIndividual.DOCUMENTO_TIPO_CHOICES,
    }
    
    return render(request, 'gestao_rural/movimentacao_individual_nova.html', context)


@login_required
def brincos_lista(request, propriedade_id):
    """Lista de brincos da propriedade"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Filtros
    status_filtro = request.GET.get('status', '')
    tipo_filtro = request.GET.get('tipo', '')
    busca = request.GET.get('busca', '')
    
    brincos = BrincoAnimal.objects.filter(propriedade=propriedade).select_related('animal')
    
    if status_filtro:
        brincos = brincos.filter(status=status_filtro)
    
    if tipo_filtro:
        brincos = brincos.filter(tipo_brinco=tipo_filtro)
    
    if busca:
        # Normalizar código de busca (remover espaços, traços, etc.)
        busca_normalizada = _normalizar_codigo(busca)
        
        # PRIORIDADE: Buscar por SISBOV primeiro (ID principal do animal)
        # Busca por: SISBOV no numero_brinco (pode ser SISBOV salvo como brinco), 
        # e também busca no animal relacionado
        filtros_busca = Q(numero_brinco__icontains=busca)
        
        # Se a busca normalizada tiver 8+ dígitos, pode ser SISBOV
        if len(busca_normalizada) >= 8:
            # Busca por SISBOV no numero_brinco (pode ser SISBOV completo salvo como brinco)
            filtros_busca |= (
                Q(numero_brinco=busca_normalizada) |
                Q(numero_brinco__icontains=busca_normalizada) |
                Q(numero_brinco__endswith=busca_normalizada) |
                Q(animal__codigo_sisbov=busca_normalizada) |
                Q(animal__codigo_sisbov__icontains=busca_normalizada) |
                Q(animal__codigo_sisbov__endswith=busca_normalizada)
            )
        # Se a busca normalizada tiver 6-7 dígitos, pode ser número de manejo ou SISBOV parcial
        elif len(busca_normalizada) >= 6:
            filtros_busca |= (
                Q(numero_brinco__icontains=busca_normalizada) |
                Q(numero_brinco__endswith=busca_normalizada) |
                Q(animal__numero_manejo=busca_normalizada) |
                Q(animal__codigo_sisbov__icontains=busca_normalizada) |
                Q(animal__codigo_sisbov__endswith=busca_normalizada)
            )
        # Para códigos menores, também tenta buscar por SISBOV que contém
        elif len(busca_normalizada) >= 3:
            filtros_busca |= (
                Q(numero_brinco__icontains=busca_normalizada) |
                Q(animal__codigo_sisbov__icontains=busca_normalizada)
            )
        
        brincos = brincos.filter(filtros_busca)
    
    brincos = brincos.order_by('numero_brinco')

    total_brincos = BrincoAnimal.objects.filter(propriedade=propriedade)
    total_disponiveis = total_brincos.filter(status='DISPONIVEL').count()
    total_em_uso = total_brincos.filter(status='EM_USO').count()
    total_saldo = total_disponiveis
    ultima_compra = total_brincos.exclude(data_aquisicao__isnull=True).order_by('-data_aquisicao').values_list('data_aquisicao', flat=True).first()

    lotes = (
        total_brincos
        .values('codigo_lote', 'fornecedor')
        .annotate(
            quantidade=Count('id'),
            utilizados=Count('id', filter=Q(status='EM_USO')),
            saldo=Count('id', filter=Q(status='DISPONIVEL')),
            data_aquisicao=Max('data_aquisicao'),
        )
        .order_by('-data_aquisicao')
    )

    lotes_resumo = []
    hoje = datetime.today().date()
    vencendo_em_dois_anos = 0
    for lote in lotes:
        data_aquisicao = lote.get('data_aquisicao')
        vencimento = data_aquisicao + timedelta(days=730) if data_aquisicao else None
        if vencimento and (vencimento - hoje).days <= 60 and (vencimento - hoje).days >= 0:
            status_vencimento = 'vencendo'
        elif vencimento and vencimento < hoje:
            status_vencimento = 'vencido'
            vencendo_em_dois_anos += lote['saldo']
        else:
            status_vencimento = 'ok'
        lotes_resumo.append({
            'codigo_lote': lote.get('codigo_lote') or 'Lote não informado',
            'fornecedor': lote.get('fornecedor') or '-',
            'quantidade': lote['quantidade'],
            'utilizados': lote['utilizados'],
            'saldo': lote['saldo'],
            'data_aquisicao': data_aquisicao,
            'vencimento': vencimento,
            'vencido': status_vencimento == 'vencido',
            'vencendo': status_vencimento == 'vencendo',
        })

    context = {
        'propriedade': propriedade,
        'brincos': brincos,
        'status_filtro': status_filtro,
        'tipo_filtro': tipo_filtro,
        'busca': busca,
        'totais': {
            'total': total_brincos.count(),
            'disponiveis': total_disponiveis,
            'em_uso': total_em_uso,
            'saldo': total_saldo,
            'ultima_compra': ultima_compra,
            'vencendo_em_dois_anos': vencendo_em_dois_anos,
            'lotes': lotes_resumo,
        },
    }
    
    return render(request, 'gestao_rural/brincos_lista.html', context)


@login_required
@require_http_methods(["POST"])
def brinco_excluir_lote(request, propriedade_id):
    """Exclui todos os brincos de um determinado lote da propriedade.

    Utilizado para correção de lançamentos incorretos de compra de brincos.
    """
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    codigo_lote = (request.POST.get('codigo_lote') or '').strip()
    fornecedor = (request.POST.get('fornecedor') or '').strip()
    data_aquisicao_raw = (request.POST.get('data_aquisicao') or '').strip()
    sem_codigo = request.POST.get('sem_codigo') == 'true'

    # Se não há código de lote e não é uma exclusão de lote sem código, retornar erro
    if not codigo_lote and not sem_codigo:
        messages.error(request, 'Código do lote não informado.')
        return redirect('brincos_lista', propriedade_id=propriedade_id)

    # Filtrar brincos
    if sem_codigo:
        # Excluir brincos sem código de lote (None ou string vazia)
        # Não aplicar filtros de fornecedor e data, pois são agregações visuais
        queryset = BrincoAnimal.objects.filter(
            propriedade=propriedade,
        ).filter(
            Q(codigo_lote__isnull=True) | Q(codigo_lote='')
        )
    else:
        queryset = BrincoAnimal.objects.filter(
            propriedade=propriedade,
            codigo_lote=codigo_lote,
        )
        
        # Aplicar filtros adicionais apenas para lotes com código
        if fornecedor and fornecedor != '-':
            queryset = queryset.filter(fornecedor=fornecedor)

        if data_aquisicao_raw:
            try:
                data_aquisicao = datetime.strptime(data_aquisicao_raw, '%Y-%m-%d').date()
                queryset = queryset.filter(data_aquisicao=data_aquisicao)
            except ValueError:
                # Se a data vier em formato inesperado, apenas ignora o filtro de data
                pass

    total_para_excluir = queryset.count()
    
    # Log para debug (remover em produção se necessário)
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f'Exclusão de lote - sem_codigo: {sem_codigo}, total: {total_para_excluir}, propriedade: {propriedade_id}')

    if total_para_excluir == 0:
        messages.warning(
            request,
            'Nenhum brinco foi localizado para o lote informado. Verifique os dados e tente novamente.'
        )
        return redirect('brincos_lista', propriedade_id=propriedade_id)

    # Exclui todos os brincos do lote (inclusive em uso).
    # O número do brinco permanece registrado no animal, evitando perda de histórico.
    # Usar delete() diretamente no queryset para garantir que todos sejam excluídos
    total_excluidos = queryset.delete()[0]

    if sem_codigo:
        messages.success(
            request,
            f'Todos os {total_excluidos} brincos do lote sem código foram excluídos com sucesso.'
        )
    else:
        messages.success(
            request,
            f'Todos os {total_excluidos} brincos do lote {codigo_lote} foram excluídos com sucesso.'
        )
    return redirect('brincos_lista', propriedade_id=propriedade_id)


@login_required
def brinco_cadastrar_lote(request, propriedade_id):
    """Cadastro de brincos em lote"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    MAX_DIGITOS_BRINCO = 15

    if request.method == 'POST':
        try:
            tipo_brinco = request.POST.get('tipo_brinco', 'VISUAL')
            prefixo = (request.POST.get('prefixo') or '').strip()
            quantidade = int(request.POST.get('quantidade', 0))
            numero_inicial_raw = (request.POST.get('numero_inicial') or '1').strip()
            data_aquisicao = request.POST.get('data_aquisicao') or None
            codigo_lote = (request.POST.get('codigo_lote') or '').strip()
            fornecedor = (request.POST.get('fornecedor') or '').strip()
            valor_unitario = request.POST.get('valor_unitario') or None
            valor_unitario_decimal = None
            if valor_unitario:
                try:
                    valor_unitario_decimal = Decimal(str(valor_unitario))
                except (InvalidOperation, ValueError):
                    messages.warning(request, 'Valor unitário inválido. O campo será ignorado.')
                    valor_unitario_decimal = None

            if quantidade <= 0:
                messages.error(request, 'A quantidade deve ser maior que zero!')
                return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

            if prefixo and not prefixo.isdigit():
                messages.error(request, 'O prefixo deve conter apenas números (padrão MAPA).')
                return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

            if len(prefixo) >= MAX_DIGITOS_BRINCO:
                messages.error(request, 'O prefixo excede o limite permitido de 15 dígitos.')
                return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

            digitos_restantes = MAX_DIGITOS_BRINCO - len(prefixo)

            if not numero_inicial_raw.isdigit():
                messages.error(request, 'Informe um número inicial válido (somente dígitos).')
                return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

            numero_inicial = int(numero_inicial_raw)
            if numero_inicial < 0:
                messages.error(request, 'O número inicial deve ser maior ou igual a zero.')
                return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

            limite_sequencia = 10 ** digitos_restantes
            if numero_inicial + quantidade > limite_sequencia:
                messages.error(
                    request,
                    'A sequência ultrapassa o limite de 15 dígitos. Ajuste o prefixo, número inicial ou quantidade.',
                )
                return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

            if data_aquisicao:
                try:
                    data_aquisicao = datetime.strptime(data_aquisicao, "%Y-%m-%d").date()
                except ValueError:
                    data_aquisicao = None

            brincos_criados = 0
            for i in range(quantidade):
                numero = numero_inicial + i
                numero_formatado = f"{numero:0{digitos_restantes}d}"
                numero_brinco = f"{prefixo}{numero_formatado}"

                if len(numero_brinco) != MAX_DIGITOS_BRINCO:
                    messages.error(request, 'Falha ao gerar número do brinco. Verifique prefixo e sequência.')
                    return redirect('brinco_cadastrar_lote', propriedade_id=propriedade_id)

                if BrincoAnimal.objects.filter(numero_brinco=numero_brinco).exists():
                    continue

                BrincoAnimal.objects.create(
                    numero_brinco=numero_brinco,
                    tipo_brinco=tipo_brinco,
                    propriedade=propriedade,
                    status='DISPONIVEL',
                    data_aquisicao=data_aquisicao,
                    codigo_lote=codigo_lote or None,
                    fornecedor=fornecedor or None,
                    valor_unitario=valor_unitario_decimal,
                )
                brincos_criados += 1

            if brincos_criados:
                messages.success(request, f'{brincos_criados} brincos cadastrados com sucesso!')
            else:
                messages.warning(request, 'Nenhum brinco foi cadastrado (verifique se já existiam).')
            return redirect('brincos_lista', propriedade_id=propriedade_id)

        except Exception as e:
            messages.error(request, f'Erro ao cadastrar brincos: {str(e)}')

    context = {
        'propriedade': propriedade,
    }

    return render(request, 'gestao_rural/brinco_cadastrar_lote.html', context)


@login_required
def brinco_importar_lista(request, propriedade_id):
    """Importa brincos de uma lista de texto (arquivo ou colar) - VERSÃO CORRIGIDA"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    if request.method == 'POST':
        try:
            tipo_brinco = request.POST.get('tipo_brinco', 'VISUAL')
            data_aquisicao = request.POST.get('data_aquisicao') or None
            codigo_lote = (request.POST.get('codigo_lote') or '').strip()
            fornecedor = (request.POST.get('fornecedor') or '').strip()
            valor_unitario = request.POST.get('valor_unitario') or None
            valor_unitario_decimal = None
            
            # Validar valor unitário
            if valor_unitario:
                try:
                    valor_unitario_decimal = Decimal(str(valor_unitario).replace(',', '.'))
                    if valor_unitario_decimal < 0:
                        raise ValueError('Valor não pode ser negativo')
                except (InvalidOperation, ValueError) as e:
                    messages.warning(request, f'Valor unitário inválido ({valor_unitario}). O campo será ignorado.')
                    valor_unitario_decimal = None
            
            # Validar data de aquisição
            if data_aquisicao:
                try:
                    data_aquisicao = datetime.strptime(data_aquisicao, "%Y-%m-%d").date()
                    if data_aquisicao > date.today():
                        messages.warning(request, 'Data de aquisição não pode ser futura. Campo será ignorado.')
                        data_aquisicao = None
                except ValueError:
                    messages.warning(request, 'Data de aquisição inválida. Campo será ignorado.')
                    data_aquisicao = None
            
            # Obter lista de números (arquivo ou texto)
            numeros_texto = ''
            nome_arquivo = ''
            
            if 'arquivo' in request.FILES:
                arquivo = request.FILES['arquivo']
                if arquivo:
                    nome_arquivo = arquivo.name
                    # Tentar diferentes encodings
                    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    for encoding in encodings:
                        try:
                            arquivo.seek(0)  # Voltar ao início do arquivo
                            numeros_texto = arquivo.read().decode(encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                    
                    if not numeros_texto:
                        messages.error(request, f'Erro ao ler arquivo {nome_arquivo}. Formato não suportado.')
                        return redirect('brinco_importar_lista', propriedade_id=propriedade_id)
            else:
                numeros_texto = request.POST.get('lista_numeros', '').strip()
            
            if not numeros_texto:
                messages.error(request, 'Informe uma lista de números ou envie um arquivo!')
                return redirect('brinco_importar_lista', propriedade_id=propriedade_id)
            
            # Processar lista de números
            linhas = numeros_texto.replace('\r\n', '\n').replace('\r', '\n').split('\n')
            brincos_validos = []
            brincos_invalidos = []
            brincos_duplicados = []
            brincos_criados = 0
            erros_detalhados = []
            
            for idx, linha in enumerate(linhas, start=1):
                if not linha.strip():
                    continue
                
                # Limpar linha: remover espaços, asteriscos, caracteres especiais
                numero_limpo = linha.strip().rstrip('*').strip()
                # Remover caracteres não numéricos, mas manter apenas números
                numero_limpo = re.sub(r'\D', '', numero_limpo)
                
                if not numero_limpo:
                    continue
                
                # Validar comprimento (aceitar 12 ou 15 dígitos)
                if len(numero_limpo) not in [12, 15]:
                    brincos_invalidos.append(f"Linha {idx}: {linha.strip()[:50]} (tem {len(numero_limpo)} dígitos, esperado 12 ou 15)")
                    continue
                
                # Padronizar para 15 dígitos (adicionar zeros à esquerda se necessário)
                if len(numero_limpo) == 12:
                    numero_limpo = '000' + numero_limpo
                
                # Verificar se já existe na propriedade
                if BrincoAnimal.objects.filter(
                    numero_brinco=numero_limpo,
                    propriedade=propriedade
                ).exists():
                    brincos_duplicados.append(numero_limpo)
                    continue
                
                brincos_validos.append(numero_limpo)
            
            # Criar brincos válidos em lote (mais eficiente)
            from django.db import transaction
            
            with transaction.atomic():
                for numero_brinco in brincos_validos:
                    try:
                        # Verificar novamente antes de criar (evitar race condition)
                        if BrincoAnimal.objects.filter(
                            numero_brinco=numero_brinco,
                            propriedade=propriedade
                        ).exists():
                            brincos_duplicados.append(numero_brinco)
                            continue
                        
                        BrincoAnimal.objects.create(
                            numero_brinco=numero_brinco,
                            tipo_brinco=tipo_brinco,
                            propriedade=propriedade,
                            status='DISPONIVEL',
                            data_aquisicao=data_aquisicao,
                            codigo_lote=codigo_lote or None,
                            fornecedor=fornecedor or None,
                            valor_unitario=valor_unitario_decimal,
                        )
                        brincos_criados += 1
                    except Exception as e:
                        erro_msg = f"{numero_brinco} (erro: {str(e)})"
                        brincos_invalidos.append(erro_msg)
                        erros_detalhados.append(erro_msg)
            
            # Mensagens de resultado detalhadas
            if brincos_criados > 0:
                messages.success(request, f'✅ {brincos_criados} brincos cadastrados com sucesso!')
            
            if brincos_duplicados:
                total_duplicados = len(brincos_duplicados)
                if total_duplicados <= 10:
                    duplicados_msg = ', '.join(brincos_duplicados[:10])
                    messages.warning(request, f'⚠️ {total_duplicados} brincos já existiam e foram ignorados: {duplicados_msg}')
                else:
                    messages.warning(request, f'⚠️ {total_duplicados} brincos já existiam e foram ignorados (mostrando 10 primeiros): {", ".join(brincos_duplicados[:10])}...')
            
            if brincos_invalidos:
                total_invalidos = len(brincos_invalidos)
                if total_invalidos <= 5:
                    invalidos_msg = '; '.join(brincos_invalidos[:5])
                    messages.error(request, f'❌ {total_invalidos} números inválidos: {invalidos_msg}')
                else:
                    messages.error(request, f'❌ {total_invalidos} números inválidos. Verifique o formato dos dados.')
            
            if nome_arquivo:
                messages.info(request, f'📄 Arquivo processado: {nome_arquivo}')
            
            return redirect('brincos_lista', propriedade_id=propriedade_id)
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messages.error(request, f'❌ Erro ao importar brincos: {str(e)}')
            # Log detalhado para debug
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao importar brincos - Propriedade {propriedade_id}: {error_detail}")
    
    context = {
        'propriedade': propriedade,
    }
    
    return render(request, 'gestao_rural/brinco_importar_lista.html', context)


@login_required
def relatorio_rastreabilidade(request, propriedade_id):
    """
    Relatório completo de rastreabilidade conforme IN 51/MAPA
    Inclui todos os anexos exigidos pela Instrução Normativa
    """
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '')
    anexo = request.GET.get('anexo', '')  # Para filtrar por anexo específico
    
    # Anexo I - Identificação Individual dos Animais
    animais = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).select_related('categoria', 'lote_atual').order_by('numero_brinco')
    
    if data_inicio:
        animais = animais.filter(data_identificacao__gte=data_inicio)
    if data_fim:
        animais = animais.filter(data_identificacao__lte=data_fim)
    
    # Anexo II - Movimentação de Animais
    movimentacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade
    ).select_related(
        'animal', 'animal__categoria',
        'propriedade_origem', 'propriedade_destino'
    )
    
    if data_inicio:
        movimentacoes = movimentacoes.filter(data_movimentacao__gte=data_inicio)
    
    if data_fim:
        movimentacoes = movimentacoes.filter(data_movimentacao__lte=data_fim)
    
    if tipo_movimentacao:
        movimentacoes = movimentacoes.filter(tipo_movimentacao=tipo_movimentacao)
    
    movimentacoes = movimentacoes.order_by('-data_movimentacao', 'animal__numero_brinco')
    
    # Anexo III - Sanidade (Vacinações e Tratamentos)
    from .models_manejo import Manejo
    vacinacoes = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='VACINACAO'
    ).select_related('animal', 'animal__categoria')
    
    tratamentos = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao='TRATAMENTO'
    ).select_related('animal', 'animal__categoria')
    
    if data_inicio:
        vacinacoes = vacinacoes.filter(data_movimentacao__gte=data_inicio)
        tratamentos = tratamentos.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        vacinacoes = vacinacoes.filter(data_movimentacao__lte=data_fim)
        tratamentos = tratamentos.filter(data_movimentacao__lte=data_fim)
    
    # Anexo IV - Inventário de Animais
    animais_por_categoria = animais.values('categoria__nome').annotate(
        total=Count('id')
    ).order_by('categoria__nome')
    
    animais_por_status = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).values('status').annotate(
        total=Count('id')
    ).order_by('status')
    
    # Estatísticas gerais
    total_animais = AnimalIndividual.objects.filter(propriedade=propriedade).count()
    animais_ativos = AnimalIndividual.objects.filter(propriedade=propriedade, status='ATIVO').count()
    
    # Movimentações por tipo
    movimentacoes_por_tipo = movimentacoes.values('tipo_movimentacao').annotate(
        total=Count('id')
    ).order_by('tipo_movimentacao')
    
    context = {
        'propriedade': propriedade,
        'movimentacoes': movimentacoes,
        'animais': animais,
        'vacinacoes': vacinacoes,
        'tratamentos': tratamentos,
        'animais_por_categoria': animais_por_categoria,
        'animais_por_status': animais_por_status,
        'movimentacoes_por_tipo': movimentacoes_por_tipo,
        'total_animais': total_animais,
        'animais_ativos': animais_ativos,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'tipo_movimentacao': tipo_movimentacao,
        'anexo': anexo,
        'data_emissao': date.today(),
        'status_labels': dict(AnimalIndividual.STATUS_CHOICES),
    }
    
    return render(request, 'gestao_rural/relatorio_rastreabilidade.html', context)


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


@login_required
def relatorio_dia_barcodes(request, propriedade_id):
    """Emissão de Documentos de Identificação Animal (DIA) com código de barras"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    status_filtro = request.GET.get('status', 'ATIVO')
    categoria_filtro = request.GET.get('categoria', '')
    busca = request.GET.get('busca', '')
    limite = int(request.GET.get('limite', 200))

    animais = AnimalIndividual.objects.filter(propriedade=propriedade)

    if status_filtro:
        animais = animais.filter(status=status_filtro)
    if categoria_filtro:
        animais = animais.filter(categoria_id=categoria_filtro)
    if busca:
        animais = animais.filter(numero_brinco__icontains=busca)

    animais = animais.select_related('categoria').order_by('numero_brinco')[:limite]

    categorias = CategoriaAnimal.objects.filter(ativo=True).order_by('nome')

    context = {
        'propriedade': propriedade,
        'animais': animais,
        'categorias': categorias,
        'status_filtro': status_filtro,
        'categoria_filtro': categoria_filtro,
        'busca': busca,
        'limite': limite,
        'total_emitidos': animais.count(),
    }

    return render(request, 'gestao_rural/relatorio_dia_barcodes.html', context)


@login_required
def relatorio_inventario_sisbov(request, propriedade_id):
    """Inventário oficial SISBOV"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    status_filtro = request.GET.get('status', 'ATIVO')
    categoria_filtro = request.GET.get('categoria', '')

    animais = AnimalIndividual.objects.filter(propriedade=propriedade)
    if status_filtro:
        animais = animais.filter(status=status_filtro)
    if categoria_filtro:
        animais = animais.filter(categoria_id=categoria_filtro)

    animais = animais.select_related('categoria').order_by('categoria__nome', 'numero_brinco')

    categorias = CategoriaAnimal.objects.filter(ativo=True).order_by('nome')
    total_animais = animais.count()
    totais_por_categoria = animais.values('categoria__nome').annotate(total=Count('id')).order_by('categoria__nome')
    totais_por_status = AnimalIndividual.objects.filter(propriedade=propriedade).values('status').annotate(total=Count('id')).order_by('status')

    context = {
        'propriedade': propriedade,
        'animais': animais,
        'categorias': categorias,
        'status_filtro': status_filtro,
        'categoria_filtro': categoria_filtro,
        'total_animais': total_animais,
        'totais_por_categoria': totais_por_categoria,
        'totais_por_status': totais_por_status,
        'data_relatorio': date.today(),
    }

    return render(request, 'gestao_rural/relatorio_inventario_sisbov.html', context)


@login_required
def relatorio_movimentacoes_sisbov(request, propriedade_id):
    """Livro oficial de movimentações SISBOV"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    data_inicio = _parse_date(request.GET.get('data_inicio'))
    data_fim = _parse_date(request.GET.get('data_fim'))
    tipo_filtro = request.GET.get('tipo_movimentacao', '')
    escopo = request.GET.get('escopo', '').lower()
    tipos_por_escopo = []
    if escopo == 'transferencias':
        tipos_por_escopo = ['TRANSFERENCIA_ENTRADA', 'TRANSFERENCIA_SAIDA']
    busca = request.GET.get('busca', '')

    movimentacoes = MovimentacaoIndividual.objects.filter(animal__propriedade=propriedade).select_related(
        'animal', 'animal__categoria', 'propriedade_origem', 'propriedade_destino'
    )

    if data_inicio:
        movimentacoes = movimentacoes.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        movimentacoes = movimentacoes.filter(data_movimentacao__lte=data_fim)
    if tipos_por_escopo:
        movimentacoes = movimentacoes.filter(tipo_movimentacao__in=tipos_por_escopo)
        tipo_filtro = ''  # garante que o select permaneça em "Todos" quando usar escopo
    elif tipo_filtro:
        movimentacoes = movimentacoes.filter(tipo_movimentacao=tipo_filtro)
    if busca:
        movimentacoes = movimentacoes.filter(animal__numero_brinco__icontains=busca)

    movimentacoes = movimentacoes.order_by('-data_movimentacao', 'animal__numero_brinco')

    context = {
        'propriedade': propriedade,
        'movimentacoes': movimentacoes,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'tipo_movimentacao': tipo_filtro,
        'busca': busca,
        'total_movimentacoes': movimentacoes.count(),
        'escopo': escopo,
    }

    return render(request, 'gestao_rural/relatorio_movimentacoes_sisbov.html', context)


@login_required
def relatorio_entradas_sisbov(request, propriedade_id):
    """Relatório de nascimentos e entradas (compras/transferências)"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    data_inicio = _parse_date(request.GET.get('data_inicio'))
    data_fim = _parse_date(request.GET.get('data_fim'))
    busca = request.GET.get('busca', '')
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '').upper()
    tipos_disponiveis = ['NASCIMENTO', 'COMPRA', 'TRANSFERENCIA_ENTRADA']
    if tipo_movimentacao in tipos_disponiveis:
        tipos = [tipo_movimentacao]
    else:
        tipos = tipos_disponiveis
        tipo_movimentacao = ''

    entradas = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao__in=tipos
    ).select_related('animal', 'animal__categoria', 'propriedade_origem')

    if data_inicio:
        entradas = entradas.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        entradas = entradas.filter(data_movimentacao__lte=data_fim)
    if busca:
        entradas = entradas.filter(animal__numero_brinco__icontains=busca)

    entradas = entradas.order_by('-data_movimentacao', 'animal__numero_brinco')

    context = {
        'propriedade': propriedade,
        'entradas': entradas,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'busca': busca,
        'total_entradas': entradas.count(),
        'tipo_movimentacao': tipo_movimentacao,
    }

    return render(request, 'gestao_rural/relatorio_entradas_sisbov.html', context)


@login_required
def relatorio_saidas_sisbov(request, propriedade_id):
    """Relatório de saídas (vendas, transferências, óbitos)"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    data_inicio = _parse_date(request.GET.get('data_inicio'))
    data_fim = _parse_date(request.GET.get('data_fim'))
    busca = request.GET.get('busca', '')
    tipo_movimentacao = request.GET.get('tipo_movimentacao', '').upper()
    tipos_disponiveis = ['VENDA', 'TRANSFERENCIA_SAIDA', 'MORTE']
    if tipo_movimentacao in tipos_disponiveis:
        tipos = [tipo_movimentacao]
    else:
        tipos = tipos_disponiveis
        tipo_movimentacao = ''

    saidas = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao__in=tipos
    ).select_related('animal', 'animal__categoria', 'propriedade_destino')

    if data_inicio:
        saidas = saidas.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        saidas = saidas.filter(data_movimentacao__lte=data_fim)
    if busca:
        saidas = saidas.filter(animal__numero_brinco__icontains=busca)

    saidas = saidas.order_by('-data_movimentacao', 'animal__numero_brinco')

    context = {
        'propriedade': propriedade,
        'saidas': saidas,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'busca': busca,
        'total_saidas': saidas.count(),
        'tipo_movimentacao': tipo_movimentacao,
    }

    return render(request, 'gestao_rural/relatorio_saidas_sisbov.html', context)


@login_required
def relatorio_sanitario_sisbov(request, propriedade_id):
    """Relatório de vacinação e tratamentos"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)

    data_inicio = _parse_date(request.GET.get('data_inicio'))
    data_fim = _parse_date(request.GET.get('data_fim'))
    busca = request.GET.get('busca', '')
    tipo_sanitario = request.GET.get('tipo_sanitario', '')

    tipos = ['VACINACAO', 'TRATAMENTO']
    if tipo_sanitario in tipos:
        tipos_filtrados = [tipo_sanitario]
    else:
        tipos_filtrados = tipos

    registros = MovimentacaoIndividual.objects.filter(
        animal__propriedade=propriedade,
        tipo_movimentacao__in=tipos_filtrados
    ).select_related('animal', 'animal__categoria')

    if data_inicio:
        registros = registros.filter(data_movimentacao__gte=data_inicio)
    if data_fim:
        registros = registros.filter(data_movimentacao__lte=data_fim)
    if busca:
        registros = registros.filter(animal__numero_brinco__icontains=busca)

    registros = registros.order_by('-data_movimentacao', 'animal__numero_brinco')

    context = {
        'propriedade': propriedade,
        'registros': registros,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'busca': busca,
        'tipo_sanitario': tipo_sanitario,
        'total_registros': registros.count(),
    }

    return render(request, 'gestao_rural/relatorio_sanitario_sisbov.html', context)


@login_required
def api_gerar_numero_brinco(request, propriedade_id):
    """API para gerar sugestão de número de brinco"""
    propriedade = get_object_or_404(Propriedade, pk=propriedade_id)
    
    prefixo = request.GET.get('prefixo', 'BR')
    ultimo_numero = AnimalIndividual.objects.filter(
        propriedade=propriedade
    ).count()
    
    proximo_numero = ultimo_numero + 1
    numero_brinco = f"{prefixo}{proximo_numero:08d}"
    
    # Verificar se já existe
    while AnimalIndividual.objects.filter(numero_brinco=numero_brinco).exists():
        proximo_numero += 1
        numero_brinco = f"{prefixo}{proximo_numero:08d}"
    
    return JsonResponse({
        'numero_brinco': numero_brinco,
        'proximo_numero': proximo_numero
    })

